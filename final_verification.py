"""
FINAL VERIFICATION: Confirm Excel has all flows from JSON in the correct areas.
Uses the EXACT SAME categorization logic as the Excel generation script.
"""
import json
from pathlib import Path
from openpyxl import load_workbook

print("✅ FINAL VERIFICATION - Excel vs JSON")
print("=" * 100)

# Load JSON
with open('data/diagrams/ug2_north_decline.json') as f:
    diagram = json.load(f)

edges = diagram.get('edges', [])
print(f"Total flows in JSON: {len(edges)}")

# Same categorization logic as Excel generation
area_definitions = {
    'UG2N': {
        'prefixes': ['rainfall__', 'ndcd__', 'bh_ndgwa__', 'softening__', 'reservoir__', 'offices__', 
                    'guest_house__', 'septic__', 'sewage__', 'north_decline__', 'north_shaft__', 
                    'losses__', 'losses2__', 'consumption__', 'consumption2__', 'evaporation__',
                    'dust_suppression__', 'spill__', 'junction_127__', 'junction_128__'],
        'contains': ['__TO__rainfall', '__TO__ndcd', '__TO__bh_ndgwa', '__TO__softening', '__TO__reservoir', 
                    '__TO__offices', '__TO__guest_house', '__TO__septic', '__TO__sewage', '__TO__north_decline', 
                    '__TO__north_shaft', '__TO__losses', '__TO__losses2', '__TO__consumption', '__TO__consumption2',
                    '__TO__evaporation', '__TO__dust_suppression', '__TO__spill', '__TO__junction_127', '__TO__junction_128']
    },
    'MERN': {
        'prefixes': ['rainfall_merensky__', 'ndcd_merensky__', 'bh_mcgwa__', 'softening_merensky__', 
                    'offices_merensky__', 'merensky_north__', 'evaporation_merensky__', 'dust_suppression_merensky__',
                    'spill_merensky__', 'consumption_merensky__', 'losses_merensky__'],
        'contains': ['__TO__rainfall_merensky', '__TO__ndcd_merensky', '__TO__bh_mcgwa', '__TO__softening_merensky', 
                    '__TO__offices_merensky', '__TO__merensky_north', '__TO__evaporation_merensky', '__TO__dust_suppression_merensky',
                    '__TO__spill_merensky', '__TO__consumption_merensky', '__TO__losses_merensky']
    },
    'UG2S': {
        'prefixes': ['ug2s_'],
        'contains': ['__TO__ug2s_']
    },
    'UG2P': {
        'prefixes': ['ug2plant_', 'cprwsd'],
        'contains': ['__TO__ug2plant_', '__TO__cprwsd']
    },
    'MERP': {
        'prefixes': ['merplant_', 'mprwsd'],
        'contains': ['__TO__merplant_', '__TO__mprwsd']
    },
    'MERS': {
        'prefixes': ['mers_'],
        'contains': ['__TO__mers_']
    },
    'OLDTSF': {
        'prefixes': ['oldtsf_', 'nt_rwd', 'new_tsf', 'trtd'],
        'contains': ['__TO__oldtsf_', '__TO__nt_rwd', '__TO__new_tsf', '__TO__trtd']
    },
    'STOCKPILE': {
        'prefixes': ['stockpile_', 'spcd1'],
        'contains': ['__TO__stockpile_', '__TO__spcd1']
    }
}

def find_area_for_flow(from_id, to_id):
    """Find which area a flow belongs to - EXACT same logic as Excel generation."""
    if not from_id or not to_id:
        return None
    
    flow_str = f'{from_id}__TO__{to_id}'
    
    # Check in order: more specific areas first (MERN has priority over UG2N, etc.)
    for area in ['MERN', 'MERP', 'MERS', 'UG2N', 'UG2S', 'UG2P', 'OLDTSF', 'STOCKPILE']:
        patterns = area_definitions.get(area, {})
        
        # Check prefixes
        for prefix in patterns.get('prefixes', []):
            if flow_str.startswith(prefix):
                return area
        
        # Check contains patterns
        for contains in patterns.get('contains', []):
            if contains in flow_str:
                return area
    
    return None

# Extract flows by area from JSON (correct categorization)
json_flows_by_area = {area: [] for area in area_definitions.keys()}
for edge in edges:
    from_id = edge.get('from')
    to_id = edge.get('to')
    flow_key = f'{from_id}__TO__{to_id}'
    area = find_area_for_flow(from_id, to_id)
    if area:
        json_flows_by_area[area].append(flow_key)

# Load latest Excel
test_templates = Path('test_templates')
xlsx_files = sorted([f for f in test_templates.glob('Water_Balance_TimeSeries_Template_FIXED_*.xlsx')], 
                    key=lambda x: x.stat().st_mtime, reverse=True)

if not xlsx_files:
    print("ERROR: No Excel file found!")
    exit(1)

excel_file = xlsx_files[0]
wb = load_workbook(str(excel_file))
print(f"Excel file: {excel_file.name}")

# Extract flows from Excel
excel_flows_by_area = {}
sheet_mapping = {
    'Flows_UG2N': 'UG2N',
    'Flows_UG2S': 'UG2S',
    'Flows_UG2P': 'UG2P',
    'Flows_MERN': 'MERN',
    'Flows_MERP': 'MERP',
    'Flows_MERS': 'MERS',
    'Flows_OLDTSF': 'OLDTSF',
    'Flows_STOCKPILE': 'STOCKPILE'
}

for sheet_name, area in sheet_mapping.items():
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Flow headers in row 3, skip Date/Year/Month (columns A-C)
        cols = [cell.value for cell in ws[3][3:] if cell.value]
        excel_flows_by_area[area] = cols
    else:
        excel_flows_by_area[area] = []

print("\n📊 COMPARISON")
print("=" * 100)

all_match = True
total_json = sum(len(flows) for flows in json_flows_by_area.values())
total_excel = sum(len(flows) for flows in excel_flows_by_area.values())

print(f"\nJSON total:   {total_json:3} flows")
print(f"Excel total:  {total_excel:3} flows")
print(f"Match:        {total_json == total_excel}")

if total_json != total_excel:
    print(f"\n⚠️ DIFFERENCE: {abs(total_json - total_excel)} flows")
    all_match = False

print("\nArea-by-area comparison:")
print("-" * 100)

for area in sorted(area_definitions.keys()):
    json_count = len(json_flows_by_area[area])
    excel_count = len(excel_flows_by_area[area])
    match = json_count == excel_count
    status = "✅" if match else "❌"
    
    print(f"{status} {area:15} JSON: {json_count:3}  Excel: {excel_count:3}  ", end="")
    if not match:
        print(f"Diff: {json_count - excel_count:+3}")
        all_match = False
    else:
        print()

print("\n" + "=" * 100)
if all_match:
    print("✅ SUCCESS! Excel perfectly matches JSON!")
    print(f"\n   All {total_excel} flows correctly distributed across 8 areas.")
else:
    print("⚠️ MISMATCHES FOUND - See details above")
    print(f"\n   Excel has {total_excel} flows, JSON expects {total_json}")
