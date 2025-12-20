#!/usr/bin/env python
"""
Add missing Guest House columns to Excel template.

This script adds the 3 missing Guest House columns that are defined in the JSON diagram but not in Excel:
1. SOFTENING PLANT → GUEST HOUSE
2. GUEST HOUSE → SEPTIC TANK
3. GUEST HOUSE → CONSUMPTION

These columns are needed for the UG2N (UG2 North Decline) area flow diagram.
"""

from openpyxl import load_workbook
from pathlib import Path

EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')
SHEET_NAME = 'Flows_UG2N'

# New columns to add (in order after existing Flows_UG2N columns)
NEW_COLUMNS = [
    'SOFTENING PLANT → GUEST HOUSE',
    'GUEST HOUSE → SEPTIC TANK',
    'GUEST HOUSE → CONSUMPTION',
]

def add_columns():
    """Add missing columns to Excel template."""
    print(f"Opening {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    
    # Get existing headers from row 1
    headers = [cell.value for cell in ws[1]]
    print(f"\nCurrent headers ({len(headers)} columns):")
    for i, h in enumerate(headers, 1):
        if h:
            print(f"  {i}: {h}")
    
    # Check for existing guest house columns
    existing_guest = [h for h in headers if h and 'GUEST' in h.upper()]
    print(f"\nExisting GUEST columns: {existing_guest}")
    
    # Find last data column (before any empty cells)
    last_col = 0
    for col in range(1, 100):
        cell_value = ws.cell(1, col).value
        if cell_value:
            last_col = col
    
    print(f"\nLast data column: {last_col}")
    print(f"Will add {len(NEW_COLUMNS)} new columns starting at column {last_col + 1}...")
    
    # Add new headers to row 1
    for idx, col_name in enumerate(NEW_COLUMNS, 1):
        new_col_idx = last_col + idx
        ws.cell(1, new_col_idx).value = col_name
        print(f"  ✅ Column {new_col_idx}: {col_name}")
    
    # Add sample data rows with "-" for new columns
    print(f"\nAdding sample data rows (rows 2-11) with '-' for new columns...")
    for row in range(2, 12):  # Sample data rows
        for idx in range(1, len(NEW_COLUMNS) + 1):
            col_idx = last_col + idx
            current_val = ws.cell(row, col_idx).value
            if not current_val:  # Only add if empty
                ws.cell(row, col_idx).value = '-'
    
    print("\nSaving workbook...")
    wb.save(EXCEL_PATH)
    print(f"✅ Successfully added {len(NEW_COLUMNS)} columns to {SHEET_NAME}!")
    
    # Verify
    wb2 = load_workbook(EXCEL_PATH)
    ws2 = wb2[SHEET_NAME]
    new_headers = [cell.value for cell in ws2[1]]
    print(f"\nVerification - New header count: {len([h for h in new_headers if h])}")
    print("New columns added:")
    for h in NEW_COLUMNS:
        if h in new_headers:
            col_idx = new_headers.index(h) + 1
            print(f"  ✅ {h} at column {col_idx}")
        else:
            print(f"  ❌ {h} NOT FOUND")

if __name__ == '__main__':
    add_columns()
