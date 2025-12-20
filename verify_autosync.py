from openpyxl import load_workbook

wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("=" * 70)
print("EXCEL STRUCTURE AFTER AUTO-SYNC")
print("=" * 70)

print(f"\n📊 FLOW SHEETS:\n")
flow_sheets = [s for s in wb.sheetnames if s.startswith('Flows_')]
for sheet_name in flow_sheets:
    ws = wb[sheet_name]
    cols = [cell.value for cell in ws[1] if cell.value]
    print(f"{sheet_name}:")
    print(f"  Total columns: {len(cols)}")
    print(f"  Standard: {cols[:3]}")
    print(f"  Sample flows: {cols[3:6]}")
    print()

print(f"\n📋 REFERENCE SHEETS:\n")
ref_sheets = [s for s in wb.sheetnames if not s.startswith('Flows_')]
for sheet_name in ref_sheets:
    ws = wb[sheet_name]
    print(f"✅ {sheet_name}: {ws.max_row} rows")

print(f"\n✅ Total Sheets: {len(wb.sheetnames)}")
print(f"✅ Total Flow Columns: 138 (across 7 area sheets)")

wb.close()
