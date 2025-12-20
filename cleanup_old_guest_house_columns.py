#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Clean up old GUEST_HOUSE columns from Excel (kept from old conversion)."""

from openpyxl import load_workbook
from pathlib import Path

EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("="*70)
print("Cleaning up old GUEST_HOUSE columns from Excel")
print("="*70)

wb = load_workbook(EXCEL_PATH)
ws = wb['Flows_UG2N']

# Get current headers from row 3
current_headers = [c.value for c in ws[3]]
print(f"\nCurrent Flows_UG2N headers before cleanup: {len([h for h in current_headers if h])} columns")

# Find and remove old GUEST_HOUSE columns
columns_to_remove = []
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(3, col_idx).value
    if header and 'GUEST_HOUSE' in str(header):
        print(f"  Found old column at {col_idx}: {header}")
        columns_to_remove.append(col_idx)

# Remove columns (in reverse order to avoid index shifting)
for col_idx in sorted(columns_to_remove, reverse=True):
    print(f"  Removing column {col_idx}...")
    ws.delete_cols(col_idx)

wb.save(EXCEL_PATH)
print(f"\n✅ Removed {len(columns_to_remove)} old GUEST_HOUSE columns")

# Verify
wb2 = load_workbook(EXCEL_PATH)
ws2 = wb2['Flows_UG2N']
headers = [c.value for c in ws2[3] if c.value]
print(f"\nAfter cleanup:")
print(f"  Total headers: {len(headers)}")
print(f"  Last 5 headers:")
for h in headers[-5:]:
    print(f"    - {h}")

wb2.close()
