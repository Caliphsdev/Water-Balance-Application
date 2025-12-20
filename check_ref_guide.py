from openpyxl import load_workbook

# Check the Reference Guide content for old vs new codes
wb = load_workbook("test_templates/Water_Balance_TimeSeries_Template.xlsx")
ref_sheet = wb["Reference Guide"]

print("=== REFERENCE GUIDE - CHECKING FOR OLD CODES ===\n")

# Collect all node IDs
node_ids = set()
for row in range(4, ref_sheet.max_row+1):
    node_id = ref_sheet.cell(row, 1).value
    if node_id:
        node_ids.add(str(node_id).strip())

# Sample of entries
print("Sample of Node IDs in Reference Guide:")
for i, nid in enumerate(sorted(node_ids)[:15]):
    print(f"  {i+1}. {nid}")

# Check for old naming patterns
old_patterns = ["MERENSKY", "MERSN", "MERSO", "_OLD", "OLDFORMAT"]
print(f"\n\nSearching for OLD NAMING PATTERNS: {old_patterns}")

found_old = {}
for row in range(4, ref_sheet.max_row+1):
    node_id = ref_sheet.cell(row, 1).value
    if node_id:
        for pattern in old_patterns:
            if pattern in str(node_id).upper():
                if pattern not in found_old:
                    found_old[pattern] = []
                found_old[pattern].append(str(node_id))

if found_old:
    print("\n❌ OLD PATTERNS FOUND:")
    for pattern, nodes in found_old.items():
        print(f"  {pattern}: {len(nodes)} occurrences")
        for node in nodes[:3]:
            print(f"    - {node}")
else:
    print("\n✓ No old naming patterns found - Reference Guide is current")

# Check actual sheet names used in flow data
print("\n\n=== ACTUAL SHEET DATA CHECK ===")
print("Total unique node IDs in Reference Guide:", len(node_ids))
