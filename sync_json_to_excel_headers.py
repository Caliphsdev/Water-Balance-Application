"""
Sync JSON Diagram Mappings to Current Excel Headers (Future-Proof)

This script makes the Excel file the source of truth for column naming.
It automatically updates the JSON diagram's excel_mapping.column values
to match the actual headers currently in the Excel workbook.

Run this anytime:
- You rename Excel columns
- You regenerate Excel headers
- The naming convention changes

This ensures validation never fails due to stale mappings.
"""

import json
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, Set, Tuple

# Configuration
DIAGRAM_PATH = Path('data/diagrams/ug2_north_decline.json')
EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("=" * 75)
print("SYNC JSON DIAGRAM - EXCEL HEADERS (Future-Proof)")
print("=" * 75)

# Validate files
if not DIAGRAM_PATH.exists():
    print(f"❌ Diagram not found: {DIAGRAM_PATH}")
    exit(1)
if not EXCEL_PATH.exists():
    print(f"❌ Excel not found: {EXCEL_PATH}")
    exit(1)

# Load Excel and extract all headers by sheet
print(f"\n📄 Reading Excel: {EXCEL_PATH.name}")
wb = load_workbook(EXCEL_PATH)
excel_headers_by_sheet: Dict[str, Set[str]] = {}

for sheet_name in wb.sheetnames:
    if sheet_name.startswith('Flows_'):
        ws = wb[sheet_name]
        headers = {cell.value for cell in ws[3] if cell.value}  # Headers are in row 3
        excel_headers_by_sheet[sheet_name] = headers
        print(f"   ✅ {sheet_name}: {len(headers)} headers")

wb.close()
print(f"\n   Total sheets with headers: {len(excel_headers_by_sheet)}")

# Load diagram
print(f"\n📊 Reading diagram: {DIAGRAM_PATH.name}")
with open(DIAGRAM_PATH, 'r', encoding='utf-8') as f:
    diagram = json.load(f)

edges = diagram.get('edges', [])
print(f"   Found {len(edges)} flow edges")

# Track changes
stats = {
    'total_edges': len(edges),
    'updated': 0,
    'not_found': 0,
    'skipped_disabled': 0,
    'skipped_no_mapping': 0,
    'failed': [],
}

print(f"\n🔍 Processing edges and updating mappings...\n")

for edge_idx, edge in enumerate(edges):
    edge_id = f"{edge.get('from', '')} → {edge.get('to', '')}"
    
    mapping = edge.get('excel_mapping', {})
    
    # Skip disabled or unmapped edges
    if not mapping:
        stats['skipped_no_mapping'] += 1
        continue
    
    if not mapping.get('enabled'):
        stats['skipped_disabled'] += 1
        continue
    
    sheet = mapping.get('sheet', '')
    old_column = mapping.get('column', '')
    
    if not sheet or sheet not in excel_headers_by_sheet:
        print(f"   ⚠️  {edge_id}")
        print(f"       Sheet '{sheet}' not found in Excel")
        stats['not_found'] += 1
        continue
    
    # Find matching header in Excel
    available_headers = excel_headers_by_sheet[sheet]
    
    # Exact match (best case - already correct)
    if old_column in available_headers:
        stats['updated'] += 1  # Count as processed
        continue
    
    # Try to find best match by normalizing
    # Remove common prefixes/suffixes and see if semantically same
    def normalize_header(h: str) -> str:
        """Normalize for fuzzy matching."""
        return h.upper().strip()
    
    norm_old = normalize_header(old_column)
    
    # Look for exact normalized match
    matching_headers = [h for h in available_headers if normalize_header(h) == norm_old]
    
    if matching_headers:
        # Use the actual header from Excel (might have different case/format)
        new_column = matching_headers[0]
        if old_column != new_column:
            mapping['column'] = new_column
            edge['excel_mapping'] = mapping
            print(f"   ✅ {edge_id}")
            print(f"       Old: {old_column}")
            print(f"       New: {new_column}")
            stats['updated'] += 1
        continue
    
    # Try pattern matching: look for headers that contain key parts
    # Remove common prefixes/suffixes to improve matching
    old_normalized = old_column.lower().replace('__to__', ' → ').replace('_', ' ')
    old_parts = set(w for w in old_normalized.split() if w and w != '→')
    
    potential_matches = []
    for h in available_headers:
        h_normalized = h.lower().replace('→', ' → ').replace('_', ' ')
        h_parts = set(w for w in h_normalized.split() if w and w != '→')
        
        # Calculate overlap score
        overlap = old_parts & h_parts
        if overlap:
            # Prefer matches with directional indicator (→)
            has_arrow_old = '→' in old_column.lower().replace('__to__', '→')
            has_arrow_h = '→' in h
            arrow_bonus = 5 if has_arrow_old == has_arrow_h else 0
            
            score = len(overlap) + arrow_bonus
            potential_matches.append((h, score, len(overlap)))
    
    if potential_matches:
        # Sort by score, then by overlap count
        potential_matches.sort(key=lambda x: (x[1], x[2]), reverse=True)
        best_match = potential_matches[0][0]
        
        mapping['column'] = best_match
        edge['excel_mapping'] = mapping
        print(f"   ✅ {edge_id}")
        print(f"       Old: {old_column}")
        print(f"       New: {best_match} (fuzzy match)")
        stats['updated'] += 1
        continue
    
    # No match found
    print(f"   ❌ {edge_id}")
    print(f"       Could not find header matching: {old_column}")
    print(f"       Available in {sheet}: {list(available_headers)[:3]}...")
    stats['not_found'] += 1
    stats['failed'].append({
        'edge': edge_id,
        'sheet': sheet,
        'old_column': old_column,
        'available': list(available_headers)
    })

# Save updated diagram
print(f"\n💾 Saving updated diagram...")
with open(DIAGRAM_PATH, 'w', encoding='utf-8') as f:
    json.dump(diagram, f, indent=2, ensure_ascii=False)

# Summary
print(f"\n{'=' * 75}")
print("SYNC SUMMARY")
print(f"{'=' * 75}")
print(f"Total edges processed:    {stats['total_edges']}")
print(f"✅ Successfully mapped:   {stats['updated']}")
print(f"⚠️  Mappings not found:   {stats['not_found']}")
print(f"⏭️  Disabled/skipped:     {stats['skipped_disabled'] + stats['skipped_no_mapping']}")

if stats['failed']:
    print(f"\n❌ Failed mappings ({len(stats['failed'])}):")
    for fail in stats['failed'][:10]:
        print(f"   - {fail['edge']} in {fail['sheet']}")
        print(f"     Looked for: {fail['old_column']}")

if stats['not_found'] == 0:
    print(f"\n✅ All mappings synchronized successfully!")
    print(f"   The JSON diagram now references actual Excel headers.")
    print(f"   Validation should pass when you run it in the app.")
else:
    print(f"\n⚠️  Some mappings could not be auto-resolved.")
    print(f"   You may need to manually add those columns to Excel,")
    print(f"   or update the JSON mappings manually via the UI.")

print(f"\n{'=' * 75}")
