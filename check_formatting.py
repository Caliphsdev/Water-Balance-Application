from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

wb = load_workbook("test_templates/Water_Balance_TimeSeries_Template.xlsx")

print("=== REFERENCE GUIDE ANALYSIS ===\n")

ref_sheet = wb["Reference Guide"]
print(f"Reference Guide rows: {ref_sheet.max_row}")
print("\nFirst 10 entries:")
for row in range(1, min(11, ref_sheet.max_row+1)):
    node_id = ref_sheet.cell(row, 1).value
    label = ref_sheet.cell(row, 2).value
    area = ref_sheet.cell(row, 3).value
    print(f"  {node_id} | {label} | {area}")

# Check formatting on flow sheets
print("\n\n=== FORMATTING INCONSISTENCY CHECK ===\n")

for sheet_name in ["Flows_OLDTSF", "Flows_UG2P", "Flows_MERS"]:
    ws = wb[sheet_name]
    print(f"{sheet_name}:")
    
    # Check headers (row 3)
    for col in range(1, min(5, ws.max_column+1)):
        cell = ws.cell(3, col)
        has_fill = cell.fill and cell.fill.start_color and cell.fill.start_color.index != "00000000"
        has_font_color = cell.font and cell.font.color
        print(f"  Col {col}: fill={has_fill}, font_color={has_font_color}, value={cell.value[:30] if cell.value else None}...")
    
    # Check data row (row 4)
    print(f"  Data row 4, Col 1 (Date): {ws.cell(4, 1).value}")
    print()
