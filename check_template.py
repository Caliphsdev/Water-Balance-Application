from openpyxl import load_workbook

wb = load_workbook("test_templates/Water_Balance_TimeSeries_Template.xlsx")

print("=== TEMPLATE FILE STRUCTURE ===\n")
print(f"Sheets in template ({len(wb.sheetnames)}):")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n  {sheet}:")
    print(f"    Max row: {ws.max_row}")
    print(f"    Max col: {ws.max_column}")
    # Show first 3 header values
    headers = [ws.cell(3, c).value for c in range(1, min(5, ws.max_column+1))]
    print(f"    First headers: {headers}")

# Check for old areas
print("\n\n=== OLD VS NEW AREAS ===")
print("Sheets with 'Flows_' prefix:")
flows_sheets = [s for s in wb.sheetnames if s.startswith("Flows_")]
for s in flows_sheets:
    print(f"  ✓ {s}")

old_areas = ["MERSN", "MERSO", "MERENSKY_SOUTH"]
print(f"\nLooking for old/deprecated areas: {old_areas}")
for area in old_areas:
    if area in wb.sheetnames:
        print(f"  ❌ FOUND: {area} (should be removed)")
    else:
        print(f"  ✓ Not found: {area}")
