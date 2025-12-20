#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Deduplicate Excel headers across all Flows_* sheets and remove duplicate JSON edges.
- Excel: keeps first occurrence of each header (row 3); deletes later duplicates.
  If only suffixed variants exist (e.g., 'X → Y.1'), first is renamed to base ('X → Y') and others deleted.
- JSON: removes duplicate edges with identical (from, to, sheet, column, enabled) signature.
  Also normalizes mapping columns to base names if they include trailing '.<n>'.

Backups:
- Writes a .bak copy of the Excel workbook and JSON before changes.

Usage:
  python dedupe_excel_and_json.py
"""

import re
import json
import shutil
from pathlib import Path
from typing import Dict, List, Set, Tuple
from openpyxl import load_workbook

# Default locations (align with repo structure)
DEFAULT_JSON = Path('data/diagrams/ug2_north_decline.json')
DEFAULT_XLSX = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

HEADER_ROW = 3
SKIP_HEADERS = {"Date", "Year", "Month"}
SUFFIX_RE = re.compile(r"^(.*)\.(\d+)$")


def backup_file(path: Path) -> Path:
    if not path.exists():
        return path
    bak = path.with_suffix(path.suffix + '.bak')
    shutil.copy2(path, bak)
    return bak


def normalize_header(header: str) -> str:
    """Return header without trailing .<n> duplicate suffix if present."""
    if not isinstance(header, str):
        return header
    m = SUFFIX_RE.match(header)
    return m.group(1) if m else header


def dedupe_excel_headers(xlsx_path: Path) -> Dict[str, Dict[str, int]]:
    """Deduplicate headers in all Flows_* sheets.

    Returns a summary: {sheet_name: {"removed": n_removed, "renamed": n_renamed}}
    """
    summary: Dict[str, Dict[str, int]] = {}

    if not xlsx_path.exists():
        print(f"[SKIP] Excel not found: {xlsx_path}")
        return summary

    backup_file(xlsx_path)

    wb = load_workbook(xlsx_path)

    for sheet in list(wb.sheetnames):
        if not sheet.startswith('Flows_'):
            continue
        ws = wb[sheet]
        # Gather headers from row 3 and group columns by base header
        headers: List[str] = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
        groups: Dict[str, List[int]] = {}
        for col_idx, hdr in enumerate(headers, start=1):
            if hdr is None or hdr in SKIP_HEADERS:
                continue
            base = normalize_header(str(hdr))
            groups.setdefault(base, []).append(col_idx)

        to_delete: List[int] = []
        renamed = 0

        # For each base header having multiple columns, keep the one with most numeric data
        for base, cols in groups.items():
            if len(cols) == 1:
                # Normalize suffix if present
                idx = cols[0]
                current = str(ws.cell(HEADER_ROW, idx).value)
                if current != base:
                    ws.cell(HEADER_ROW, idx).value = base
                    renamed += 1
                continue

            # Score columns by count of numeric values (prefer columns carrying real data)
            scores: List[Tuple[int, int]] = []  # (numeric_count, col_idx)
            max_row = ws.max_row
            for idx in cols:
                numeric_count = 0
                for r in range(HEADER_ROW + 1, max_row + 1):
                    v = ws.cell(r, idx).value
                    if isinstance(v, (int, float)):
                        numeric_count += 1
                        # Early exit if we have sufficient evidence
                        if numeric_count >= 1:
                            break
                scores.append((numeric_count, idx))

            # Keep the column with the highest score; on tie, keep smallest index (left-most)
            scores.sort(key=lambda t: (-t[0], t[1]))
            keep_idx = scores[0][1]
            # Normalize kept header to base
            if str(ws.cell(HEADER_ROW, keep_idx).value) != base:
                ws.cell(HEADER_ROW, keep_idx).value = base
                renamed += 1
            # Mark others for deletion
            for _, idx in scores[1:]:
                to_delete.append(idx)

        # Delete columns from right to left to preserve indices
        removed = 0
        for col_idx in sorted(to_delete, reverse=True):
            ws.delete_cols(col_idx, 1)
            removed += 1

        if removed or renamed:
            summary[sheet] = {"removed": removed, "renamed": renamed}
            print(f"[EXCEL] {sheet}: removed={removed}, renamed={renamed}")

    wb.save(xlsx_path)
    wb.close()

    return summary


def dedupe_json_edges(json_path: Path, xlsx_path: Path) -> Dict[str, int]:
    """Remove duplicate edges and normalize mapping columns without changing semantics.

    - Duplicate edge signature: (from, to, sheet, column, enabled)
    - If mapping column has '.<n>' suffix but base exists in Excel sheet, normalize to base
    """
    summary = {"removed_edges": 0, "normalized_columns": 0}

    if not json_path.exists():
        print(f"[SKIP] JSON not found: {json_path}")
        return summary

    backup_file(json_path)

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    edges: List[Dict] = data.get('edges', [])

    # Build Excel header sets per sheet for normalization
    try:
        wb = load_workbook(xlsx_path, read_only=True)
        excel_headers: Dict[str, Set[str]] = {}
        for sheet in wb.sheetnames:
            if sheet.startswith('Flows_'):
                ws = wb[sheet]
                headers = set()
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(HEADER_ROW, c).value
                    if v and v not in SKIP_HEADERS:
                        headers.add(str(v))
                excel_headers[sheet] = headers
        wb.close()
    except Exception:
        excel_headers = {}

    seen: Set[Tuple[str, str, str, str, bool]] = set()
    new_edges: List[Dict] = []

    for e in edges:
        mapping = e.get('excel_mapping', {})
        sheet = mapping.get('sheet', '')
        column = mapping.get('column', '')
        enabled = bool(mapping.get('enabled'))

        # Normalize column to base if it has a numeric suffix and base exists in Excel
        base = normalize_header(column) if column else column
        if base != column and sheet in excel_headers and base in excel_headers[sheet]:
            mapping['column'] = base
            e['excel_mapping'] = mapping
            column = base
            summary['normalized_columns'] += 1

        sig = (e.get('from', ''), e.get('to', ''), sheet, column, enabled)
        if sig in seen:
            summary['removed_edges'] += 1
            continue
        seen.add(sig)
        new_edges.append(e)

    if summary['removed_edges'] or summary['normalized_columns']:
        data['edges'] = new_edges
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        print(f"[JSON] Removed duplicate edges: {summary['removed_edges']}, normalized columns: {summary['normalized_columns']}")
    else:
        print("[JSON] No duplicate edges or column normalizations required")

    return summary


def main():
    json_path = DEFAULT_JSON
    xlsx_path = DEFAULT_XLSX

    print("== Backups will be created next to the original files (.bak) ==")

    excel_summary = dedupe_excel_headers(xlsx_path)
    json_summary = dedupe_json_edges(json_path, xlsx_path)

    print("\n==== SUMMARY ====")
    if excel_summary:
        for sheet, s in excel_summary.items():
            print(f"  {sheet}: removed={s['removed']}, renamed={s['renamed']}")
    else:
        print("  Excel: no changes")

    print(f"  JSON: removed_edges={json_summary.get('removed_edges', 0)}, normalized_columns={json_summary.get('normalized_columns', 0)}")

    print("\nClean-up complete.")


if __name__ == '__main__':
    main()
