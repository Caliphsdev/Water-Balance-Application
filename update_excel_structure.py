"""
Update Excel structure to match new diagram:
- Remove Merensky North sheet (Flows_MERN)
- Split Old TSF sheet into Old TSF and New TSF (Flows_OLDTSF → keep, add Flows_NEWTSF)
- Update Reference Guide
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
import shutil
from datetime import datetime

excel_path = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

# Create backup
backup_path = excel_path.with_suffix(f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy(excel_path, backup_path)
print(f"✅ Backup created: {backup_path}")

# Load workbook
wb = load_workbook(excel_path)
print(f"\n📊 Current sheets: {wb.sheetnames}")

# 1. Remove Merensky North sheet
if 'Flows_MERN' in wb.sheetnames:
    wb.remove(wb['Flows_MERN'])
    print("🗑️  Removed: Flows_MERN (Merensky North Area - deleted from diagram)")
else:
    print("⚠️  Flows_MERN sheet not found")

# 2. Add New TSF sheet (copy structure from Old TSF)
if 'Flows_NEWTSF' not in wb.sheetnames:
    if 'Flows_OLDTSF' in wb.sheetnames:
        # Create new sheet after Old TSF
        old_tsf_idx = wb.sheetnames.index('Flows_OLDTSF')
        new_tsf = wb.create_sheet('Flows_NEWTSF', old_tsf_idx + 1)
        
        # Copy header structure from Old TSF
        old_tsf = wb['Flows_OLDTSF']
        new_tsf['A1'] = 'Date'
        new_tsf['B1'] = 'Year'
        new_tsf['C1'] = 'Month'
        
        # Copy header formatting
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        
        for cell in ['A1', 'B1', 'C1']:
            new_tsf[cell].font = header_font
            new_tsf[cell].fill = header_fill
            new_tsf[cell].alignment = header_align
        
        # Set column widths
        new_tsf.column_dimensions['A'].width = 12
        new_tsf.column_dimensions['B'].width = 8
        new_tsf.column_dimensions['C'].width = 10
        
        print("➕ Added: Flows_NEWTSF (New TSF facility from split)")
    else:
        print("⚠️  Cannot add Flows_NEWTSF - Flows_OLDTSF not found")
else:
    print("ℹ️  Flows_NEWTSF already exists")

# 3. Update Reference Guide
if 'Reference Guide' in wb.sheetnames:
    ref_sheet = wb['Reference Guide']
    
    # Clear existing content
    ref_sheet.delete_rows(1, ref_sheet.max_row)
    
    # Title
    ref_sheet['A1'] = 'Water Balance Excel Structure - Reference Guide'
    ref_sheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ref_sheet['A1'].fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
    ref_sheet.merge_cells('A1:D1')
    
    # Last Updated
    ref_sheet['A2'] = f'Last Updated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ref_sheet['A2'].font = Font(italic=True, size=9)
    
    # Section: Overview
    ref_sheet['A4'] = 'OVERVIEW'
    ref_sheet['A4'].font = Font(bold=True, size=12)
    ref_sheet['A4'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    ref_sheet['A5'] = 'This Excel file contains flow volume data for the Water Balance Application.'
    ref_sheet['A5'].font = Font(size=10)
    
    ref_sheet['A6'] = 'Each sheet represents flows for a specific area in the mine.'
    ref_sheet['A6'].font = Font(size=10)
    
    # Section: Sheet Structure
    ref_sheet['A8'] = 'SHEET STRUCTURE'
    ref_sheet['A8'].font = Font(bold=True, size=12)
    ref_sheet['A8'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    ref_sheet['A10'] = 'Sheet Name'
    ref_sheet['B10'] = 'Area'
    ref_sheet['C10'] = 'Description'
    ref_sheet['A10'].font = Font(bold=True)
    ref_sheet['B10'].font = Font(bold=True)
    ref_sheet['C10'].font = Font(bold=True)
    
    sheets_info = [
        ('Reference Guide', 'N/A', 'This guide sheet'),
        ('Flows_UG2N', 'UG2 North Decline', 'Underground 2 North mining area flows'),
        ('Flows_STOCKPILE', 'Stockpile Area', 'Ore stockpile facility flows'),
        ('Flows_UG2S', 'UG2 South Decline', 'Underground 2 South mining area flows'),
        ('Flows_MERS', 'Merensky South', 'Merensky South mining area flows'),
        ('Flows_OLDTSF', 'Old TSF', 'Old Tailings Storage Facility (original facility)'),
        ('Flows_NEWTSF', 'New TSF', 'New Tailings Storage Facility (split from Old TSF)'),
        ('Flows_UG2P', 'UG2 Plant', 'Underground 2 processing plant flows'),
        ('Flows_MERP', 'Merensky Plant', 'Merensky processing plant flows'),
    ]
    
    row = 11
    for sheet_name, area, desc in sheets_info:
        ref_sheet[f'A{row}'] = sheet_name
        ref_sheet[f'B{row}'] = area
        ref_sheet[f'C{row}'] = desc
        row += 1
    
    # Section: Important Notes
    ref_sheet[f'A{row+1}'] = 'IMPORTANT NOTES'
    ref_sheet[f'A{row+1}'].font = Font(bold=True, size=12)
    ref_sheet[f'A{row+1}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    notes_row = row + 3
    notes = [
        '✓ Standard Columns: All sheets have Date, Year, Month as the first three columns',
        '✓ Flow Columns: Additional columns represent flow volumes (m³) between components',
        '✓ Column Naming: Use format COMPONENT1_COMPONENT2 (e.g., SOFTENING_RESERVOIR)',
        '✓ Data Format: Enter numeric values only (no text in flow columns)',
        '✓ Missing Data: Leave cells empty or use 0 for no flow',
        '',
        '⚠️ CHANGES FROM PREVIOUS VERSION:',
        '  • Merensky North Area removed (area deleted from operations)',
        '  • Old TSF split into two facilities: Old TSF and New TSF',
        '  • Flows_MERN sheet removed',
        '  • Flows_NEWTSF sheet added for New TSF facility',
        '',
        '💡 TIPS:',
        '  • Use Excel Manager in Flow Diagram Dashboard to manage sheets/columns',
        '  • Auto-Map feature helps link columns to flow lines automatically',
        '  • Always backup Excel file before major changes',
    ]
    
    for note in notes:
        ref_sheet[f'A{notes_row}'] = note
        ref_sheet[f'A{notes_row}'].font = Font(size=10)
        notes_row += 1
    
    # Set column widths
    ref_sheet.column_dimensions['A'].width = 20
    ref_sheet.column_dimensions['B'].width = 25
    ref_sheet.column_dimensions['C'].width = 50
    ref_sheet.column_dimensions['D'].width = 20
    
    print("📝 Updated: Reference Guide with new structure")
else:
    print("⚠️  Reference Guide sheet not found")

# Save changes
wb.save(excel_path)
print(f"\n✅ Excel structure updated: {excel_path}")
print(f"\n📊 Final sheets: {wb.sheetnames}")
wb.close()

print(f"\n🎯 Summary:")
print(f"  - Removed Flows_MERN (Merensky North)")
print(f"  - Added Flows_NEWTSF (New TSF)")
print(f"  - Updated Reference Guide")
print(f"  - Backup saved: {backup_path.name}")
