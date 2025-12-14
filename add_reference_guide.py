"""
Add a Reference Guide sheet to the Excel workbook with column descriptions.
Makes it easy for users to understand what each flow column represents.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Column reference data: (Sheet Name, Column Name, Description, Type)
COLUMN_REFERENCE = [
    # Flows_UG2N
    ('Flows_UG2N', 'BOREHOLE_ABSTRACTION', 'Water from underground borehole (inflow)', 'Inflow'),
    ('Flows_UG2N', 'RAINFALL_UG2N', 'Rainwater collected in UG2 North area', 'Inflow'),
    ('Flows_UG2N', 'OFFICES', 'Water for offices and ablutions', 'Inflow'),
    ('Flows_UG2N', 'ACCOMMODATION', 'Water for accommodation facilities', 'Inflow'),
    ('Flows_UG2N', 'NDCD1_INFLOW', 'North Decline Shaft 1 inflow from underground', 'Inflow'),
    ('Flows_UG2N', 'NDCD2_INFLOW', 'North Decline Shaft 2 inflow from underground', 'Inflow'),
    ('Flows_UG2N', 'RECOVERY_PLANT', 'Water from recovery/treatment plant (returning)', 'Inflow'),
    ('Flows_UG2N', 'RECYCLED_WATER', 'Recycled water returned to process', 'Inflow'),
    ('Flows_UG2N', 'SPILLAGE_LOSS', 'Water lost through spillage/overflow', 'Loss'),
    ('Flows_UG2N', 'SEEPAGE_LOSS', 'Water lost through seepage into ground', 'Loss'),
    
    # Flows_MERN
    ('Flows_MERN', 'RAINFALL_MERN', 'Rainwater collected in Merensky North area', 'Inflow'),
    ('Flows_MERN', 'MERENSKY_PLANT_INFLOW', 'Inflow to Merensky North processing plant', 'Inflow'),
    ('Flows_MERN', 'TREATMENT_PLANT', 'Water from treatment plant facility', 'Inflow'),
    ('Flows_MERN', 'DISCHARGE_POINT', 'Water discharged from the area', 'Outflow'),
    ('Flows_MERN', 'EVAPORATION_LOSS', 'Water lost through evaporation', 'Loss'),
    ('Flows_MERN', 'CONSERVATION_POOL', 'Water stored in conservation pool', 'Storage'),
    
    # Flows_MERENSKY_SOUTH
    ('Flows_MERENSKY_SOUTH', 'RAINFALL_MERENSKY_SOUTH', 'Rainwater collected in Merensky South area', 'Inflow'),
    ('Flows_MERENSKY_SOUTH', 'STORAGE_VOLUME', 'Water volume in storage facilities', 'Storage'),
    ('Flows_MERENSKY_SOUTH', 'TREATMENT_OUTFLOW', 'Water treated and returning to process', 'Outflow'),
    ('Flows_MERENSKY_SOUTH', 'EVAPORATION_MERENSKY_SOUTH', 'Water lost through evaporation', 'Loss'),
    ('Flows_MERENSKY_SOUTH', 'RECIRCULATION_LOOP', 'Water recirculated internally', 'Recirculation'),
    
    # Flows_UG2S
    ('Flows_UG2S', 'RAINFALL_UG2S', 'Rainwater collected in UG2 South area', 'Inflow'),
    ('Flows_UG2S', 'UG2S_INFLOW', 'Underground inflow from UG2 South shaft', 'Inflow'),
    ('Flows_UG2S', 'OUTFLOW_PLANT', 'Water outflow from processing plant', 'Outflow'),
    ('Flows_UG2S', 'SEEPAGE_UG2S', 'Water lost through seepage', 'Loss'),
    
    # Flows_STOCKPILE
    ('Flows_STOCKPILE', 'RAINFALL_STOCKPILE', 'Rainwater collected on stockpile area', 'Inflow'),
    ('Flows_STOCKPILE', 'STOCKPILE_RUNOFF', 'Runoff water from stockpile surface', 'Outflow'),
    ('Flows_STOCKPILE', 'INFILTRATION_LOSS', 'Water lost through infiltration into ground', 'Loss'),
    ('Flows_STOCKPILE', 'PUMP_EXTRACTION', 'Water pumped from stockpile area', 'Outflow'),
    ('Flows_STOCKPILE', 'TREATMENT_INPUT', 'Water input to treatment process', 'Inflow'),
    
    # Flows_OLDTSF
    ('Flows_OLDTSF', 'RAINFALL_OLDTSF', 'Rainwater collected in Old TSF area', 'Inflow'),
    ('Flows_OLDTSF', 'TSF_INFLOW', 'Inflow to Tailings Storage Facility', 'Inflow'),
    ('Flows_OLDTSF', 'DECANT_OUTFLOW', 'Water decanted (poured off) from TSF', 'Outflow'),
    ('Flows_OLDTSF', 'SEEPAGE_OLDTSF', 'Water lost through seepage from TSF', 'Loss'),
    ('Flows_OLDTSF', 'EVAPORATION_OLDTSF', 'Water lost through evaporation from TSF', 'Loss'),
    ('Flows_OLDTSF', 'SPILLWAY_FLOW', 'Water flowing through emergency spillway', 'Outflow'),
    
    # Flows_UG2PLANT
    ('Flows_UG2PLANT', 'PLANT_INFLOW', 'Total inflow to UG2 processing plant', 'Inflow'),
    ('Flows_UG2PLANT', 'PROCESS_WATER_USAGE', 'Water consumed in processing operations', 'Outflow'),
    ('Flows_UG2PLANT', 'RECYCLED_PROCESS_WATER', 'Water recycled back to process', 'Inflow'),
    ('Flows_UG2PLANT', 'PLANT_OUTFLOW', 'Total outflow from UG2 processing plant', 'Outflow'),
    ('Flows_UG2PLANT', 'EVAPORATION_UG2PLANT', 'Water lost through evaporation at plant', 'Loss'),
    ('Flows_UG2PLANT', 'TREATMENT_DISCHARGE', 'Treated water discharged from plant', 'Outflow'),
    
    # Flows_MERPLANT
    ('Flows_MERPLANT', 'PLANT_INFLOW_MER', 'Total inflow to Merensky processing plant', 'Inflow'),
    ('Flows_MERPLANT', 'PROCESS_WATER_MER', 'Water consumed in Merensky processing', 'Outflow'),
    ('Flows_MERPLANT', 'RECYCLED_MER', 'Water recycled at Merensky plant', 'Inflow'),
    ('Flows_MERPLANT', 'OUTFLOW_MER', 'Total outflow from Merensky plant', 'Outflow'),
    ('Flows_MERPLANT', 'EVAPORATION_MER', 'Water lost through evaporation at Merensky', 'Loss'),
    ('Flows_MERPLANT', 'TAILING_THICKENER', 'Water in tailings thickener process', 'Process'),
]

def add_reference_guide():
    """Add Reference Guide sheet to Excel workbook."""
    excel_path = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')
    
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Remove old Reference Guide if it exists
    if 'Reference Guide' in wb.sheetnames:
        del wb['Reference Guide']
    
    # Create new sheet
    ws = wb.create_sheet('Reference Guide', 0)  # Insert as first sheet
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    
    # Header row
    headers = ['Column Name', 'Flow Sheet', 'Description', 'Type']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = header_border
    
    # Data rows
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    type_colors = {
        'Inflow': 'D4EDDA',      # Light green
        'Outflow': 'F8D7DA',     # Light red
        'Loss': 'FFF3CD',        # Light yellow
        'Storage': 'D1ECF1',     # Light cyan
        'Recirculation': 'E7D4F5',  # Light purple
        'Process': 'E2E3E5',     # Light gray
    }
    
    for row_num, (sheet, col, desc, flow_type) in enumerate(COLUMN_REFERENCE, 2):
        ws.cell(row=row_num, column=1).value = col
        ws.cell(row=row_num, column=2).value = sheet
        ws.cell(row=row_num, column=3).value = desc
        ws.cell(row=row_num, column=4).value = flow_type
        
        # Apply formatting to all cells in row
        for col_num in range(1, 5):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # Color code by type (column D)
            if col_num == 4:
                bg_color = type_colors.get(flow_type, 'FFFFFF')
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(excel_path)
    print(f"✅ Reference Guide sheet added to {excel_path}")
    print(f"📋 Added {len(COLUMN_REFERENCE)} column descriptions")
    print("🎨 Color-coded by flow type: Green (Inflow) | Red (Outflow) | Yellow (Loss) | Cyan (Storage) | Purple (Recirculation)")

if __name__ == '__main__':
    add_reference_guide()
