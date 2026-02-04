"""
Excel Manager (DATA ACCESS FACADE).

Centralizes access to the two Excel sources used by the application:
1) Meter Readings Excel (legacy_excel_path) - used by Calculations + Analytics
2) Flow Diagram Excel (timeseries_excel_path) - used by Flow Diagram dashboard

Design goals:
- Provide a single, testable API for Excel access
- Keep dashboards thin (UI calls manager instead of pandas directly)
- Cache loaded data and invalidate on file changes
- Distinguish Excel sources clearly to avoid mixing data
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import warnings
import re

import pandas as pd

# Suppress openpyxl warning about invalid print areas in Excel files
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

from core.app_logger import logger
from core.config_manager import ConfigManager, get_resource_path


@dataclass(frozen=True)
class MeterReadingsConfig:
    """Configuration for Meter Readings Excel (legacy Excel).

    Attributes:
        file_path: Absolute path to the Meter Readings Excel file.
        sheet_name: Sheet containing monthly meter readings.
        header_row: 1-based row index where column names are located.
        first_data_row: 1-based row index where data begins.
    """

    file_path: Path
    sheet_name: str = "Meter Readings"
    header_row: int = 3
    first_data_row: int = 5


@dataclass(frozen=True)
class FlowDiagramConfig:
    """Configuration for Flow Diagram Excel (timeseries Excel).

    Attributes:
        file_path: Absolute path to the Flow Diagram Excel file.
        default_sheet_prefix: Default prefix for flow diagram sheets.
    """

    file_path: Path
    default_sheet_prefix: str = "Flows_"


class ExcelManager:
    """Excel Manager (SINGLE ENTRY POINT FOR EXCEL DATA).

    Responsibilities:
    - Resolve configured Excel paths from app config
    - Load and cache Meter Readings + Flow Diagram data
    - Provide clean API for analytics/calculations/flow dashboards
    - Invalidate cache when Excel files change on disk
    """

    # Optional convenience mapping for flow diagram sheets
    AREA_CODE_TO_SHEET = {
        "UG2N": "Flows_UG2 North",
        "UG2S": "Flows_UG2 Main",
        "MERS": "Flows_Merensky South",
        "MERP": "Flows_Merensky Plant",
        "UG2P": "Flows_UG2 Plant",
        "STOCKPILE": "Flows_Stockpile1",
        "NEWTSF": "Flows_New TSF",
        "OLDTSF": "Flows_Old TSF",
    }

    def __init__(self, config: Optional[ConfigManager] = None) -> None:
        """Initialize Excel manager and internal caches.

        Args:
            config: Optional ConfigManager instance (dependency injection).
        """
        self._config = config or ConfigManager()

        # Cached Meter Readings DataFrame and its file timestamp
        self._meter_df: Optional[pd.DataFrame] = None
        self._meter_mtime: Optional[float] = None
        self._meter_units: Dict[str, str] = {}  # Map column name → unit from row 4

        # Cached Flow Diagram DataFrames by sheet name
        self._flow_df_cache: Dict[str, pd.DataFrame] = {}
        self._flow_mtime: Optional[float] = None

    # ---------------------------------------------------------------------
    # User-Configured Paths (Persistence)
    # ---------------------------------------------------------------------
    def set_meter_readings_path(self, new_path: str) -> None:
        """Persist user-selected Meter Readings Excel path.

        This writes to config so the app remembers the path across sessions.
        It also clears the meter cache to ensure the next read uses the new file.

        Args:
            new_path: Absolute or relative path provided by user.
        """
        # Persist path selection for future sessions (user-defined location)
        self._config.set("data_sources.legacy_excel_path", new_path)
        # Invalidate cache so next access reloads from new file
        self.clear_meter_cache()

    def set_flow_diagram_path(self, new_path: str) -> None:
        """Persist user-selected Flow Diagram Excel path.

        This writes to config so the app remembers the path across sessions.
        It also clears the flow cache to ensure the next read uses the new file.

        Args:
            new_path: Absolute or relative path provided by user.
        """
        # Persist path selection for future sessions (user-defined location)
        self._config.set("data_sources.timeseries_excel_path", new_path)
        # Invalidate cache so next access reloads from new file
        self.clear_flow_cache()

    def meter_readings_status(self) -> Tuple[bool, Path]:
        """Return whether Meter Readings Excel is configured and exists.

        Returns:
            Tuple of (exists, resolved_path).
        """
        path = self.get_meter_readings_path()
        return path.exists(), path

    def flow_diagram_status(self) -> Tuple[bool, Path]:
        """Return whether Flow Diagram Excel is configured and exists.

        Returns:
            Tuple of (exists, resolved_path).
        """
        path = self.get_flow_diagram_path()
        return path.exists(), path

    # ---------------------------------------------------------------------
    # Path Resolution
    # ---------------------------------------------------------------------
    def get_meter_readings_path(self) -> Path:
        """Resolve Meter Readings Excel path (legacy_excel_path).

        Returns:
            Absolute Path to the Meter Readings Excel file.
        """
        excel_path = self._config.get("data_sources.legacy_excel_path")
        if not excel_path:
            # Default fallback for first-time configuration
            excel_path = "data/New Water Balance  20250930 Oct.xlsx"
        return self._resolve_path(excel_path)

    def get_flow_diagram_path(self) -> Path:
        """Resolve Flow Diagram Excel path (timeseries_excel_path).

        Returns:
            Absolute Path to the Flow Diagram Excel file.
        """
        excel_path = self._config.get("data_sources.timeseries_excel_path")
        if not excel_path:
            # Default fallback for first-time configuration
            excel_path = "test_templates/Water_Balance_TimeSeries_Template.xlsx"
        return self._resolve_path(excel_path)

    @staticmethod
    def _resolve_path(path_value: str) -> Path:
        """Resolve path using config rules (relative paths → app root).

        Args:
            path_value: Config path (absolute or relative).

        Returns:
            Absolute Path for use by pandas.
        """
        cfg_path = Path(path_value)
        if cfg_path.is_absolute():
            return cfg_path
        return get_resource_path(path_value)

    # ---------------------------------------------------------------------
    # Existence Checks
    # ---------------------------------------------------------------------
    def meter_readings_exists(self) -> bool:
        """Check if Meter Readings Excel exists on disk."""
        return self.get_meter_readings_path().exists()

    def flow_diagram_exists(self) -> bool:
        """Check if Flow Diagram Excel exists on disk."""
        return self.get_flow_diagram_path().exists()

    # ---------------------------------------------------------------------
    # Cache Management
    # ---------------------------------------------------------------------
    def clear_meter_cache(self) -> None:
        """Clear Meter Readings cache (force reload on next access)."""
        self._meter_df = None
        self._meter_mtime = None
        self._meter_units = {}  # Clear units cache too
        logger.info("Meter Readings Excel cache cleared")

    def clear_flow_cache(self) -> None:
        """Clear Flow Diagram cache (force reload on next access)."""
        self._flow_df_cache.clear()
        self._flow_mtime = None
        logger.info("Flow Diagram Excel cache cleared")

    def clear_all_caches(self) -> None:
        """Clear all cached Excel data (both sources)."""
        self.clear_meter_cache()
        self.clear_flow_cache()

    # ---------------------------------------------------------------------
    # Meter Readings (legacy Excel) API
    # ---------------------------------------------------------------------
    def load_meter_readings(self) -> pd.DataFrame:
        """Load Meter Readings Excel into a cached DataFrame.

        Returns:
            DataFrame with normalized columns and Date column.
        """
        cfg = MeterReadingsConfig(file_path=self.get_meter_readings_path())
        file_path = cfg.file_path

        if not file_path.exists():
            logger.warning(f"Meter Readings Excel not found: {file_path}")
            return pd.DataFrame()

        # EXCEL OPERATIONS LOGGING: Track file loading for debugging
        logger.debug(f"Loading Meter Readings Excel: {file_path}")
        logger.debug(f"Sheet: {cfg.sheet_name}, Header row: {cfg.header_row}, Data starts: {cfg.first_data_row}")

        # Invalidate cache if file changed on disk
        current_mtime = file_path.stat().st_mtime
        if self._meter_mtime and self._meter_mtime != current_mtime:
            logger.debug(f"EXCEL FILE CHANGED - Meter Readings modified (mtime changed), clearing cache")
            self.clear_meter_cache()
        self._meter_mtime = current_mtime

        if self._meter_df is not None:
            logger.debug(f"EXCEL OPERATIONS - Using cached Meter Readings ({len(self._meter_df)} rows)")
            return self._meter_df

        # Read Excel: header is row 3 (1-based → header=2)
        # Row 4 contains units (we'll read it separately)
        df = pd.read_excel(
            file_path,
            sheet_name=cfg.sheet_name,
            header=cfg.header_row - 1,
            engine="openpyxl",
        )
        
        # Read units row (row 4 is index 3 in 0-based)
        units_df = pd.read_excel(
            file_path,
            sheet_name=cfg.sheet_name,
            header=None,
            skiprows=cfg.header_row,  # Skip to units row
            nrows=1,  # Read only units row
            engine="openpyxl",
        )
        # Store units mapping: {column_name: unit}
        column_names = [str(col).strip() for col in df.columns]
        self._meter_units = {}
        for i, col_name in enumerate(column_names):
            if i < len(units_df.columns):
                unit = str(units_df.iloc[0, i]).strip() if pd.notna(units_df.iloc[0, i]) else ""
                self._meter_units[col_name] = unit

        # Normalize column names (trim whitespace)
        df.columns = [str(col).strip() for col in df.columns]

        # First column should be date; enforce consistent name
        date_col = df.columns[0]
        df["Date"] = pd.to_datetime(df[date_col], errors="coerce")

        # Drop rows without valid date
        df = df.dropna(subset=["Date"]).reset_index(drop=True)

        # Skip rows before first_data_row
        extra_skip = max(0, cfg.first_data_row - cfg.header_row - 1)
        if extra_skip > 0:
            df = df.iloc[extra_skip:].reset_index(drop=True)

        self._meter_df = df
        logger.info(
            f"Loaded Meter Readings: {len(df)} rows, {len(df.columns)} columns"
        )
        
        # EXCEL OPERATIONS LOGGING: Track what was loaded with details
        logger.debug(f"Columns loaded: {', '.join(df.columns[:10])}{'...' if len(df.columns) > 10 else ''}")
        if not df.empty and 'Date' in df.columns:
            min_date = df['Date'].min()
            max_date = df['Date'].max()
            logger.debug(f"Date range: {min_date} to {max_date}")
        logger.debug(f"Units mapping: {len(self._meter_units)} columns have units defined")
        
        return df

    def list_meter_readings_sources(self) -> List[str]:
        """List numeric source columns available for charting.

        Automatically excludes:
        - Date/time columns (Date, DateTime, Month, Year, etc.)
        - Unnamed columns (pandas default for empty Excel columns)
        - Columns with no numeric data

        Returns:
            List of numeric column names suitable for charting.
        """
        df = self.load_meter_readings()
        if df.empty:
            return []

        # Common date/time column patterns to exclude
        date_patterns = [
            "date", "time", "datetime", "timestamp",
            "month", "year", "day", "period"
        ]

        sources: List[str] = []
        for col in df.columns:
            col_lower = col.lower()
            
            # Skip unnamed columns (empty Excel columns)
            if col_lower.startswith("unnamed"):
                continue
            
            # Skip if column name matches date patterns
            if any(pattern in col_lower for pattern in date_patterns):
                continue
            
            # Check if column has numeric data
            series = pd.to_numeric(df[col], errors="coerce")
            if series.notna().any():
                sources.append(col)
        return sources

    def get_meter_readings_date_range(self) -> Tuple[Optional[date], Optional[date]]:
        """Return min/max dates in Meter Readings data.

        Returns:
            (min_date, max_date) tuple or (None, None) if data missing.
        """
        df = self.load_meter_readings()
        if df.empty or "Date" not in df.columns:
            return None, None
        return df["Date"].min().date(), df["Date"].max().date()

    def get_meter_readings_series(
        self,
        source_name: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> List[Tuple[date, float]]:
        """Get time series data for a single Meter Readings column.

        Args:
            source_name: Column name to extract (e.g., 'Tonnes Milled').
            start_date: Optional start date filter.
            end_date: Optional end date filter.

        Returns:
            List of (date, value) tuples in ascending date order.
        """
        df = self.load_meter_readings()
        if df.empty or source_name not in df.columns:
            return []

        # Filter by date range if provided
        data = df[["Date", source_name]].copy()
        if start_date:
            data = data[data["Date"] >= pd.to_datetime(start_date)]
        if end_date:
            data = data[data["Date"] <= pd.to_datetime(end_date)]

        # Convert to numeric, drop NaN
        data[source_name] = pd.to_numeric(data[source_name], errors="coerce")
        data = data.dropna(subset=[source_name])

        return [(d.date(), float(v)) for d, v in zip(data["Date"], data[source_name])]

    def get_source_unit(self, source_name: str) -> str:
        """Get the unit of measurement for a source column from Excel row 4.
        
        Args:
            source_name: Column name (e.g., 'Tonnes Milled').
        
        Returns:
            Unit string (e.g., 't', 'm³', '%') or empty string if not found.
        """
        # Ensure data is loaded to populate _meter_units
        if not self._meter_units:
            self.load_meter_readings()
        return self._meter_units.get(source_name, "")

    # ---------------------------------------------------------------------
    # Flow Diagram (timeseries Excel) API
    # ---------------------------------------------------------------------
    def list_flow_sheets(self) -> List[str]:
        """List available flow diagram sheet names (Flows_*).

        Returns:
            List of sheet names in the Flow Diagram Excel file.
        """
        file_path = self.get_flow_diagram_path()
        if not file_path.exists():
            return []

        try:
            excel_file = pd.ExcelFile(file_path)
            return [s for s in excel_file.sheet_names if s.startswith("Flows_")]
        except Exception as exc:
            logger.error(f"Failed to list flow sheets: {exc}")
            return []

    def _resolve_flow_sheet(self, area_code_or_sheet: str) -> str:
        """Resolve flow sheet name from area code or sheet name."""
        if area_code_or_sheet.startswith("Flows_"):
            return area_code_or_sheet
        return self.AREA_CODE_TO_SHEET.get(area_code_or_sheet, f"Flows_{area_code_or_sheet}")

    def load_flow_sheet(self, area_code_or_sheet: str) -> pd.DataFrame:
        """Load a Flow Diagram sheet into cache.

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or explicit sheet name.

        Returns:
            DataFrame with normalized time columns if detected.
        """
        file_path = self.get_flow_diagram_path()
        sheet_name = self._resolve_flow_sheet(area_code_or_sheet)

        if not file_path.exists():
            # Silent return - don't log warnings (called too many times during startup)
            logger.debug(f"Flow Diagram Excel not found: {file_path}")
            return pd.DataFrame()

        # EXCEL OPERATIONS LOGGING: Track Flow Diagram file access
        logger.debug(f"Loading Flow Diagram sheet: {sheet_name} from {file_path.name}")

        # Invalidate cache if file changed on disk
        current_mtime = file_path.stat().st_mtime
        if self._flow_mtime and self._flow_mtime != current_mtime:
            logger.debug(f"EXCEL FILE CHANGED - Flow Diagram modified (mtime changed), clearing cache")
            self.clear_flow_cache()
        self._flow_mtime = current_mtime

        if sheet_name in self._flow_df_cache:
            logger.debug(f"EXCEL OPERATIONS - Using cached Flow Diagram sheet '{sheet_name}' ({len(self._flow_df_cache[sheet_name])} rows)")
            return self._flow_df_cache[sheet_name]

        try:
            # Flow Diagram Excel has headers in row 3 (skip rows 1-2)
            # header=2 means use row 3 as column names (0-based indexing)
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=2, engine="openpyxl")
            df.columns = [str(col).strip() for col in df.columns]
            df = self._normalize_time_columns(df)
            self._flow_df_cache[sheet_name] = df
            logger.info(
                f"Loaded Flow Diagram sheet '{sheet_name}': "
                f"{len(df)} rows, {len(df.columns)} columns"
            )
            # EXCEL OPERATIONS LOGGING: Track detailed sheet info
            logger.debug(f"Flow columns: {', '.join([c for c in df.columns if c not in ['Date', 'Year', 'Month']][:10])}...")
            return df
        except Exception as exc:
            logger.error(f"Failed to load flow sheet '{sheet_name}': {exc}")
            return pd.DataFrame()

    def list_flow_columns(self, area_code_or_sheet: str) -> List[str]:
        """List usable flow columns for a given sheet.
        
        Returns all columns from the Excel sheet except for metadata columns
        (Date, Year, Month which are used for time indexing).
        
        Recirculation columns and all other flow data columns are included,
        even if they contain words like "Date", "Year", or "Month" (e.g., 
        "Recirculation_LastUpdate" will NOT be filtered out).

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or sheet name (e.g., 'Flows_UG2N').

        Returns:
            List of column names, excluding ONLY the exact metadata columns:
            "Date", "Year", "Month".
            
        Note:
            Recirculation columns are fully included. To see all available columns
            for mapping, check Excel sheet directly or use load_flow_sheet().
        """
        df = self.load_flow_sheet(area_code_or_sheet)
        if df.empty:
            return []

        # IMPORTANT: Block ONLY exact metadata column names (case-sensitive)
        # Do NOT filter out columns that CONTAIN these words (e.g., "Recirculation_MonthlyAvg")
        # This ensures recirculation and other flow columns are available for mapping
        blocked = {"Date", "Year", "Month"}
        return [c for c in df.columns if c not in blocked]

    def resolve_flow_sheet_name(self, area_code_or_sheet: str) -> str:
        """Resolve a flow diagram sheet name for public callers.

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or explicit sheet name.

        Returns:
            Resolved sheet name (e.g., 'Flows_UG2 North').
        """
        return self._resolve_flow_sheet(area_code_or_sheet)

    def get_flow_volume(
        self,
        area_code_or_sheet: str,
        column_name: str,
        year: int,
        month: int,
    ) -> Optional[float]:
        """Return a single flow volume for a sheet/column and year/month.
        
        If exact year/month not found, falls back to most recent date in sheet.

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or explicit sheet name.
            column_name: Flow column name in the sheet.
            year: Year to filter (will fallback to latest if not available).
            month: Month to filter (will fallback to latest if not available).

        Returns:
            Float volume in m³ if found, otherwise None.
        """
        df = self.load_flow_sheet(area_code_or_sheet)
        if df.empty or column_name not in df.columns:
            return None

        if "Year" not in df.columns or "Month" not in df.columns:
            return None

        # First try exact year/month match
        matching = df[(df["Year"] == year) & (df["Month"] == month)]
        
        # If not found, use most recent date in sheet
        if matching.empty:
            logger.debug(f"No data for {year}/{month}, using most recent date from sheet")
            # Sort by Year DESC, then Month DESC to get most recent
            try:
                df_sorted = df.sort_values(by=["Year", "Month"], ascending=[False, False])
                matching = df_sorted.head(1)
            except Exception:
                return None
        
        if matching.empty:
            return None

        value = matching.iloc[0][column_name]
        if pd.isna(value):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def auto_map_flow_column(
        self,
        from_id: str,
        to_id: str,
        area_code_or_sheet: Optional[str] = None,
    ) -> Optional[Dict[str, str]]:
        """Auto-detect an Excel sheet/column for a flowline using heuristics.

        This matches common naming patterns such as:
        - {from}_to_{to}
        - {from}_{to}
        - {from}{to}

        Args:
            from_id: Source node ID (e.g., 'BH_NDGWA').
            to_id: Destination node ID (e.g., 'Sump').
            area_code_or_sheet: Optional area code or explicit sheet name to limit search.

        Returns:
            Mapping dict {'sheet': sheet_name, 'column': column_name} or None.
        """
        if not from_id or not to_id:
            return None

        candidates = self._build_flow_candidates(from_id, to_id)
        sheets = (
            [self.resolve_flow_sheet_name(area_code_or_sheet)]
            if area_code_or_sheet
            else self.list_flow_sheets()
        )

        for sheet in sheets:
            columns = self.list_flow_columns(sheet)
            if not columns:
                continue

            # Exact matches (case-sensitive) first
            for candidate in candidates:
                if candidate in columns:
                    return {"sheet": sheet, "column": candidate}

            # Normalized, case-insensitive matching
            normalized = {self._normalize_name(c): c for c in columns}
            for candidate in candidates:
                key = self._normalize_name(candidate)
                if key in normalized:
                    return {"sheet": sheet, "column": normalized[key]}

        return None

    def create_flow_column(self, area_code_or_sheet: str, column_name: str) -> bool:
        """Create a new flow column in the Flow Diagram Excel workbook.

        This writes a new column header into the Flow Diagram Excel file and
        ensures the column exists for future volume entry.

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or explicit sheet name.
            column_name: Column header to create (e.g., 'BH_to_Sump').

        Returns:
            True if the column was created successfully, otherwise False.
        """
        file_path = self.get_flow_diagram_path()
        sheet_name = self._resolve_flow_sheet(area_code_or_sheet)

        if not file_path.exists():
            # Show user-friendly error instead of just logging
            from PySide6.QtWidgets import QMessageBox, QApplication
            app = QApplication.instance()
            if app and hasattr(app, 'activeWindow') and app.activeWindow():
                QMessageBox.critical(
                    app.activeWindow(),
                    "Excel File Not Found",
                    "Cannot create column - Flow Diagram Excel file not found.\n\nPlease configure the Excel file path first."
                )
            logger.error("Flow Diagram Excel file not found. Cannot create column.")
            return False
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in Excel")
                return False

            ws = wb[sheet_name]
            header_row = self._detect_flow_header_row(ws)

            existing_headers = [
                str(cell.value).strip() for cell in ws[header_row] if cell.value is not None
            ]
            if column_name in existing_headers:
                logger.info(f"Column already exists: {sheet_name}:{column_name}")
                return True

            next_col = ws.max_column + 1
            ws.cell(row=header_row, column=next_col, value=column_name)

            # Preserve existing row structure by writing empty cells in data rows.
            for row_idx in range(header_row + 1, max(ws.max_row, header_row + 1) + 1):
                ws.cell(row=row_idx, column=next_col, value=None)

            wb.save(file_path)
            wb.close()

            self.clear_flow_cache()
            logger.info(f"Created flow column '{column_name}' in '{sheet_name}'")
            return True
        except (OSError, ValueError, KeyError) as exc:
            logger.error(f"Failed to create flow column '{column_name}': {exc}")
            return False

    def rename_flow_column(
        self, area_code_or_sheet: str, old_column_name: str, new_column_name: str
    ) -> bool:
        """Rename a flow column in the Flow Diagram Excel workbook (COLUMN MANAGEMENT).

        Renames an existing column header and updates all mappings that reference it.

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or explicit sheet name.
            old_column_name: Current column header name.
            new_column_name: New column header name.

        Returns:
            True if the column was renamed successfully, otherwise False.

        Raises:
            ValueError: If old column name does not exist or new name already exists.
        """
        file_path = self.get_flow_diagram_path()
        sheet_name = self._resolve_flow_sheet(area_code_or_sheet)

        if not file_path.exists():
            logger.error(
                f"Flow Diagram Excel file not found: {file_path}. Cannot rename column."
            )
            return False

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in Excel")
                return False

            ws = wb[sheet_name]
            header_row = self._detect_flow_header_row(ws)

            # Find the column with old_column_name
            old_col_idx = None
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value and str(cell.value).strip() == old_column_name:
                    old_col_idx = col_idx
                    break

            if old_col_idx is None:
                logger.error(f"Column '{old_column_name}' not found in sheet '{sheet_name}'")
                return False

            # Check that new_column_name doesn't already exist
            existing_headers = [
                str(cell.value).strip() for cell in ws[header_row] if cell.value is not None
            ]
            if new_column_name in existing_headers:
                logger.error(
                    f"Column '{new_column_name}' already exists in sheet '{sheet_name}'"
                )
                return False

            # Rename the column header
            ws.cell(row=header_row, column=old_col_idx, value=new_column_name)

            wb.save(file_path)
            wb.close()

            # UPDATE MAPPINGS: Rename column in excel_flow_links.json
            # All flow IDs that referenced old_column_name must now point to new_column_name
            self._update_column_mappings(sheet_name, old_column_name, new_column_name)

            self.clear_flow_cache()
            logger.info(
                f"Renamed column '{old_column_name}' → '{new_column_name}' in '{sheet_name}'"
            )
            return True

        except (OSError, ValueError, KeyError) as exc:
            logger.error(f"Failed to rename flow column '{old_column_name}': {exc}")
            return False

    def delete_flow_column(self, area_code_or_sheet: str, column_name: str) -> bool:
        """Delete a flow column from the Flow Diagram Excel workbook (COLUMN MANAGEMENT).

        Removes the column and shifts remaining columns left. Updates Excel file and
        clears cache to force reload on next access.

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or explicit sheet name.
            column_name: Column header to delete.

        Returns:
            True if the column was deleted successfully, otherwise False.

        Raises:
            ValueError: If column name does not exist.
        """
        file_path = self.get_flow_diagram_path()
        sheet_name = self._resolve_flow_sheet(area_code_or_sheet)

        if not file_path.exists():
            logger.error(
                f"Flow Diagram Excel file not found: {file_path}. Cannot delete column."
            )
            return False

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in Excel")
                return False

            ws = wb[sheet_name]
            header_row = self._detect_flow_header_row(ws)

            # Find the column with column_name
            col_idx = None
            for col_idx_temp, cell in enumerate(ws[header_row], start=1):
                if cell.value and str(cell.value).strip() == column_name:
                    col_idx = col_idx_temp
                    break

            if col_idx is None:
                logger.error(f"Column '{column_name}' not found in sheet '{sheet_name}'")
                return False

            # Delete the column (openpyxl shifts remaining columns left)
            ws.delete_cols(col_idx)

            wb.save(file_path)
            wb.close()

            # REMOVE MAPPINGS: Delete all flow IDs that referenced the deleted column
            # This prevents dangling references to non-existent columns
            self._remove_column_mappings(sheet_name, column_name)

            self.clear_flow_cache()
            logger.info(f"Deleted column '{column_name}' from '{sheet_name}'")
            return True

        except (OSError, ValueError, KeyError) as exc:
            logger.error(f"Failed to delete flow column '{column_name}': {exc}")
            return False

    def suggest_flow_column_name(self, from_id: str, to_id: str) -> str:
        """Suggest a standardized column name for a flowline.

        Args:
            from_id: Source node ID (e.g., 'BH_NDGWA').
            to_id: Destination node ID (e.g., 'SUMP_1').

        Returns:
            Suggested column name (e.g., 'bh_ndgwa_to_sump_1'), or empty string.
        """
        candidates = self._build_flow_candidates(from_id, to_id)
        return candidates[0] if candidates else ""

    def get_flow_series(
        self,
        area_code_or_sheet: str,
        column_name: str,
        year: Optional[int] = None,
        month: Optional[int] = None,
    ) -> List[Tuple[Tuple[int, int], float]]:
        """Get flow volume series for a column with optional year/month filtering.

        Args:
            area_code_or_sheet: Area code (e.g., 'UG2N') or explicit sheet name.
            column_name: Flow column to extract.
            year: Optional year filter.
            month: Optional month filter (1-12).

        Returns:
            List of ((year, month), value) tuples.
        """
        df = self.load_flow_sheet(area_code_or_sheet)
        if df.empty or column_name not in df.columns:
            return []

        data = df.copy()
        if year is not None:
            data = data[data.get("Year") == year]
        if month is not None:
            data = data[data.get("Month") == month]

        data[column_name] = pd.to_numeric(data[column_name], errors="coerce")
        data = data.dropna(subset=[column_name])

        results: List[Tuple[Tuple[int, int], float]] = []
        for _, row in data.iterrows():
            y = int(row.get("Year")) if pd.notna(row.get("Year")) else 0
            m = int(row.get("Month")) if pd.notna(row.get("Month")) else 0
            results.append(((y, m), float(row[column_name])))
        return results

    # ---------------------------------------------------------------------
    # Utilities
    # ---------------------------------------------------------------------
    @staticmethod
    def _normalize_time_columns(df: pd.DataFrame) -> pd.DataFrame:
        """Normalize time columns in Flow Diagram sheets.

        Attempts to ensure Year and Month columns exist. If the sheet
        provides a Date column, Year/Month are derived from it.
        """
        col_lower = {str(c).lower().strip(): c for c in df.columns}

        # Case 1: Year + Month already exist
        if "year" in col_lower and "month" in col_lower:
            df.rename(
                columns={col_lower["year"]: "Year", col_lower["month"]: "Month"},
                inplace=True,
            )
            df["Year"] = pd.to_numeric(df["Year"], errors="coerce")
            df["Month"] = pd.to_numeric(df["Month"], errors="coerce")
            return df

        # Case 2: Date column exists
        date_candidates = [c for c in df.columns if "date" in str(c).lower()]
        if date_candidates:
            dt = pd.to_datetime(df[date_candidates[0]], errors="coerce")
            df["Year"] = dt.dt.year
            df["Month"] = dt.dt.month
            return df

        # Case 3: Attempt to infer Year/Month from first two columns
        if len(df.columns) >= 2:
            df["Year"] = pd.to_numeric(df.iloc[:, 0], errors="coerce")
            df["Month"] = pd.to_numeric(df.iloc[:, 1], errors="coerce")
            return df

        return df

    @staticmethod
    def _normalize_name(value: str) -> str:
        """Normalize a string for robust column matching.

        Args:
            value: Raw string value.

        Returns:
            Normalized string (lowercase, alphanumeric + underscores).
        """
        value = value.strip().lower()
        value = re.sub(r"[^a-z0-9]+", "_", value)
        value = re.sub(r"_+", "_", value).strip("_")
        return value

    def _build_flow_candidates(self, from_id: str, to_id: str) -> List[str]:
        """Build candidate column names for a flowline.

        Args:
            from_id: Source node ID.
            to_id: Destination node ID.

        Returns:
            List of candidate column names ordered by preference.
        """
        from_token = self._normalize_name(from_id)
        to_token = self._normalize_name(to_id)

        if not from_token or not to_token:
            return []

        return [
            f"{from_token}_to_{to_token}",
            f"{from_token}_{to_token}",
            f"{from_token}to{to_token}",
            f"{from_token}-{to_token}",
        ]

    @staticmethod
    def _detect_flow_header_row(ws) -> int:
        """Detect the most likely header row in a flow diagram sheet.

        Strategy:
        - Scan first 5 rows
        - Prefer rows containing "Year"/"Month"/"Date"
        - Fallback to row 1

        Args:
            ws: openpyxl worksheet object.

        Returns:
            1-based row index for header row.
        """
        best_row = 1
        for row_idx in range(1, min(ws.max_row, 5) + 1):
            values = [str(c.value).strip() for c in ws[row_idx] if c.value is not None]
            if not values:
                continue
            lower = {v.lower() for v in values}
            if {"year", "month"}.issubset(lower) or "date" in lower:
                return row_idx
            # Heuristic: prefer rows with many string values
            if len(values) >= 5:
                best_row = row_idx
        return best_row

    def _update_column_mappings(
        self, sheet_name: str, old_column_name: str, new_column_name: str
    ) -> None:
        """Update excel_flow_links.json when a column is renamed (MAPPING SYNCHRONIZATION).

        Scans excel_flow_links.json for any flow IDs that reference old_column_name
        and updates them to reference new_column_name instead. This ensures that
        flow diagram edges continue to point to the correct Excel column after rename.

        Args:
            sheet_name: Sheet where column was renamed (e.g., 'Flows_UG2N').
            old_column_name: Previous column header name.
            new_column_name: New column header name.
        """
        from pathlib import Path

        mappings_file = Path("data/excel_flow_links.json")
        if not mappings_file.exists():
            logger.debug(f"Mappings file not found: {mappings_file}. Skipping mapping update.")
            return

        try:
            import json

            with open(mappings_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Find and update all mappings in this sheet with old_column_name
            if "links" in data:
                updated_count = 0
                for flow_id, mapping in data["links"].items():
                    if (
                        mapping.get("sheet") == sheet_name
                        and mapping.get("column") == old_column_name
                    ):
                        mapping["column"] = new_column_name
                        updated_count += 1

                if updated_count > 0:
                    # Write back updated mappings
                    with open(mappings_file, "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=2, ensure_ascii=False)

                    logger.info(
                        f"Updated {updated_count} flow mapping(s): "
                        f"'{old_column_name}' → '{new_column_name}' in '{sheet_name}'"
                    )
                else:
                    logger.debug(
                        f"No mappings found for '{old_column_name}' in '{sheet_name}'. "
                        "No remapping needed."
                    )

        except (OSError, json.JSONDecodeError, KeyError, TypeError) as exc:
            logger.warning(f"Failed to update column mappings: {exc}")

    def _remove_column_mappings(self, sheet_name: str, column_name: str) -> None:
        """Remove mappings for deleted column from excel_flow_links.json (DATA CLEANUP).

        When a column is deleted, all flow IDs that reference it become invalid.
        This method removes those mappings to prevent broken references.

        Args:
            sheet_name: Sheet where column was deleted (e.g., 'Flows_UG2N').
            column_name: Column header that was deleted.
        """
        from pathlib import Path

        mappings_file = Path("data/excel_flow_links.json")
        if not mappings_file.exists():
            logger.debug(f"Mappings file not found: {mappings_file}. Skipping mapping cleanup.")
            return

        try:
            import json

            with open(mappings_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            # Find and remove all mappings in this sheet with column_name
            if "links" in data:
                removed_flow_ids = []
                for flow_id in list(data["links"].keys()):
                    mapping = data["links"][flow_id]
                    if (
                        mapping.get("sheet") == sheet_name
                        and mapping.get("column") == column_name
                    ):
                        removed_flow_ids.append(flow_id)
                        del data["links"][flow_id]

                if removed_flow_ids:
                    # Write back updated mappings
                    with open(mappings_file, "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=2, ensure_ascii=False)

                    logger.info(
                        f"Removed {len(removed_flow_ids)} dangling flow mapping(s) "
                        f"for deleted column '{column_name}' in '{sheet_name}': "
                        f"{removed_flow_ids}"
                    )
                else:
                    logger.debug(
                        f"No mappings found for '{column_name}' in '{sheet_name}'. "
                        "Nothing to clean up."
                    )

        except (OSError, json.JSONDecodeError, KeyError, TypeError) as exc:
            logger.warning(f"Failed to remove column mappings: {exc}")



# Singleton-style helper
_excel_manager_instance: Optional[ExcelManager] = None


def get_excel_manager() -> ExcelManager:
    """Get shared ExcelManager instance (singleton style)."""
    global _excel_manager_instance
    if _excel_manager_instance is None:
        _excel_manager_instance = ExcelManager()
    return _excel_manager_instance
