"""
On-demand Excel-based flow volume loader for flow diagram segments.

Reads monthly flow volumes directly from Excel without database storage.
Each flow segment is registered in Excel and volumes are fetched on-demand.
"""

import sys
from pathlib import Path
from datetime import date
from typing import Dict, List, Optional, Tuple
import warnings

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from utils.app_logger import logger
from utils.config_manager import config

# Suppress openpyxl warnings about print areas and defined names
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class FlowVolumeLoader:
    """Load flow volumes from Excel on-demand for specific months."""
    
    # Mapping from legacy area codes to new descriptive sheet names
    AREA_CODE_TO_SHEET = {
        'UG2N': 'Flows_UG2 North',
        'UG2S': 'Flows_UG2 Main',
        'MERS': 'Flows_Merensky South',
        'MERP': 'Flows_Merensky Plant',
        'UG2P': 'Flows_UG2 Plant',
        'STOCKPILE': 'Flows_Stockpile1',
        'NEWTSF': 'Flows_New TSF',
        'OLDTSF': 'Flows_Old TSF',
    }
    
    def __init__(self, excel_path: Optional[Path] = None):
        """
        Initialize the flow volume loader.
        
        Args:
            excel_path: Path to Excel file containing flow data.
                       If None, uses config value.
        """
        # Use provided path or fallback to config
        self._df_cache: Dict[str, pd.DataFrame] = {}  # Cache for each sheet
        self.excel_path = self._resolve_excel_path(excel_path)

    def _resolve_excel_path(self, override_path: Optional[Path]) -> Path:
        """Resolve the Excel path, honoring an override and the latest config value."""
        if override_path:
            return Path(override_path)

        # Reload config to pick up latest Settings changes
        try:
            config.load_config()
        except Exception:
            pass

        # Prefer explicit timeseries path; fall back to template_excel_path if set
        excel_config_path = config.get(
            'data_sources.timeseries_excel_path',
            config.get('data_sources.template_excel_path',
                       'test_templates/Water_Balance_TimeSeries_Template.xlsx')
        )

        cfg_path = Path(excel_config_path)
        if cfg_path.is_absolute():
            return cfg_path

        base_dir = Path(__file__).resolve().parents[2]
        return base_dir / cfg_path
    
    def _get_sheet_name(self, area_code: str) -> str:
        """
        Convert area code to Excel sheet name using mapping.
        
        Args:
            area_code: Legacy area code (e.g., 'UG2N', 'MERS')
        
        Returns:
            Excel sheet name (e.g., 'Flows_UG2 North')
        """
        # Check if it's already a full sheet name
        if area_code.startswith('Flows_'):
            return area_code
        
        # Use mapping for legacy area codes
        return self.AREA_CODE_TO_SHEET.get(area_code, f'Flows_{area_code}')
    
    def _load_sheet(self, sheet_name: str) -> pd.DataFrame:
        """Load a sheet from Excel with caching and resilient header detection."""
        # Refresh path from config so Settings changes are picked up without restart
        new_path = self._resolve_excel_path(None)
        if new_path != self.excel_path:
            self.excel_path = new_path
            self.clear_cache()  # Path changed, drop cached sheets

        if sheet_name in self._df_cache:
            return self._df_cache[sheet_name]

        if not self.excel_path.exists():
            logger.error(f"❌ Excel file not found: {self.excel_path}")
            return pd.DataFrame()

        def _flatten_columns(cols: pd.Index) -> List[str]:
            if isinstance(cols, pd.MultiIndex):
                flattened = []
                for tup in cols:
                    parts = [str(x).strip() for x in tup if str(x).strip() and not str(x).startswith('Unnamed:')]
                    # Take the last meaningful part (actual column name)
                    if parts:
                        flattened.append(parts[-1])
                    else:
                        flattened.append(" - ".join([str(x).strip() for x in tup if str(x).strip()]))
                return flattened
            return [str(c).strip() for c in cols]

        def _normalize_time(df: pd.DataFrame) -> pd.DataFrame:
            """Ensure Year/Month columns exist using Date or datetime columns."""
            col_lower = {str(c).lower().strip(): c for c in df.columns}
            
            # Direct Year/Month columns
            if 'year' in col_lower and 'month' in col_lower:
                df.rename(columns={col_lower['year']: 'Year',
                                   col_lower['month']: 'Month'}, inplace=True)
                df['Year'] = pd.to_numeric(df['Year'], errors='coerce')
                df['Month'] = pd.to_numeric(df['Month'], errors='coerce')
                return df

            # Look for date-related columns (case-insensitive, partial match)
            date_candidates = [c for c in df.columns if 'date' in str(c).lower() or 'time' in str(c).lower()]
            if date_candidates:
                dt = pd.to_datetime(df[date_candidates[0]], errors='coerce')
                df['Year'] = dt.dt.year
                df['Month'] = dt.dt.month
                return df

            # Check for datetime-typed columns
            for c in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[c]):
                    dt = pd.to_datetime(df[c], errors='coerce')
                    df['Year'] = dt.dt.year
                    df['Month'] = dt.dt.month
                    return df
            
            # Scan first few data rows for year-like values (YYYY pattern)
            for c in df.columns:
                sample = df[c].dropna().astype(str).head(5)
                if any(s.isdigit() and 1900 <= int(s) <= 2100 for s in sample if len(s) == 4):
                    # Found a year-like column
                    year_col = c
                    # Look for adjacent month column
                    col_idx = df.columns.get_loc(c)
                    if col_idx + 1 < len(df.columns):
                        month_col = df.columns[col_idx + 1]
                        month_sample = df[month_col].dropna().astype(str).head(5)
                        if any(s.isdigit() and 1 <= int(s) <= 12 for s in month_sample):
                            df['Year'] = pd.to_numeric(df[year_col], errors='coerce')
                            df['Month'] = pd.to_numeric(df[month_col], errors='coerce')
                            return df
            return df

        header_options = [
            [0, 1, 2],
            [0, 1],
            [1, 2],
            [2, 3],
            3,
            2,
            1,
            0,
            4,
        ]

        for header in header_options:
            try:
                df = pd.read_excel(
                    self.excel_path,
                    sheet_name=sheet_name,
                    header=header,
                    engine='openpyxl'
                )
            except Exception:
                continue

            if df is None:
                continue

            df.columns = _flatten_columns(df.columns)
            
            # Check if we have Year/Month directly in columns after flattening
            col_lower = {str(c).lower().strip(): c for c in df.columns}
            has_year_month_cols = 'year' in col_lower and 'month' in col_lower
            
            if has_year_month_cols:
                # Rename to standardized names
                df.rename(columns={col_lower['year']: 'Year',
                                   col_lower['month']: 'Month'}, inplace=True)
                df['Year'] = pd.to_numeric(df['Year'], errors='coerce')
                df['Month'] = pd.to_numeric(df['Month'], errors='coerce')
                self._df_cache[sheet_name] = df
                logger.info(
                    f"📥 Loaded sheet '{sheet_name}': "
                    f"{len(df)} rows, {len(df.columns)} columns"
                )
                return df
            
            # Skip empty DataFrames for data-based time normalization
            if df.empty:
                continue
            
            # Try normalizing time columns from data
            df = _normalize_time(df)

            if 'Year' in df.columns and 'Month' in df.columns:
                df['Year'] = pd.to_numeric(df['Year'], errors='coerce')
                df['Month'] = pd.to_numeric(df['Month'], errors='coerce')
                self._df_cache[sheet_name] = df
                logger.info(
                    f"📥 Loaded sheet '{sheet_name}': "
                    f"{len(df)} rows, {len(df.columns)} columns"
                )
                return df

        logger.error(
            f"❌ Sheet '{sheet_name}' missing Date or Year/Month columns after "
            f"trying headers {header_options}"
        )
        return pd.DataFrame()
    
    def get_monthly_volume(self, 
                          area_code: str,
                          flow_id: str,
                          year: int,
                          month: int) -> Optional[float]:
        """
        Get monthly volume for a specific flow.
        
        Args:
            area_code: Mine area code (e.g., 'UG2N', 'MERN', 'MERENSKY_NORTH')
            flow_id: Flow identifier/name (matches column in Excel)
            year: Year
            month: Month (1-12)
        
        Returns:
            Volume in m³, or None if not found
        """
        sheet_name = self._get_sheet_name(area_code)
        df = self._load_sheet(sheet_name)
        
        if df.empty:
            logger.debug(f"⚠️ No data for sheet '{sheet_name}'")
            return None
        
        # Find row matching year and month
        matching_rows = df[(df['Year'] == year) & (df['Month'] == month)]
        
        if matching_rows.empty:
            logger.debug(f"⚠️ No data for {area_code}/{flow_id} in {year}-{month:02d}")
            return None
        
        row = matching_rows.iloc[0]
        
        # Get volume from column named after flow_id
        if flow_id not in row.index:
            logger.debug(f"⚠️ Column '{flow_id}' not found in sheet '{sheet_name}'")
            return None
        
        try:
            volume = float(row[flow_id])
            # Allow negative values; only skip NaN
            if pd.isna(volume):
                return None
            return volume
        except (TypeError, ValueError):
            logger.debug(f"⚠️ Invalid volume for {flow_id}: {row[flow_id]}")
            return None
    
    def get_all_volumes_for_month(self,
                                  area_code: str,
                                  year: int,
                                  month: int) -> Dict[str, float]:
        """
        Get all flow volumes for an area for a specific month.
        
        Args:
            area_code: Mine area code
            year: Year
            month: Month (1-12)
        
        Returns:
            Dictionary mapping flow_id -> volume (m³)
        """
        sheet_name = self._get_sheet_name(area_code)
        df = self._load_sheet(sheet_name)
        
        if df.empty:
            return {}
        
        # Find row matching year and month
        matching_rows = df[(df['Year'] == year) & (df['Month'] == month)]
        
        if matching_rows.empty:
            return {}
        
        row = matching_rows.iloc[0]
        volumes = {}
        
        # Extract all numeric columns (skip Date, Year, Month)
        skip_cols = {'Date', 'Year', 'Month'}
        for col in df.columns:
            if col not in skip_cols:
                try:
                    vol = float(row[col])
                    # Allow negative values; only skip NaN
                    if not pd.isna(vol):
                        volumes[col] = vol
                except (TypeError, ValueError):
                    pass
        
        logger.info(f"📊 Loaded {len(volumes)} flows for {area_code} ({year}-{month:02d})")
        return volumes

    def get_all_volumes_for_month_from_sheet(self,
                                             sheet_name: str,
                                             year: int,
                                             month: int) -> Dict[str, float]:
        """Return all volumes for a given sheet and month, keyed by column.

        This is sheet-aware and avoids cross-sheet column collisions by only
        returning data from the specified sheet.

        Args:
            sheet_name: Exact Excel sheet name (e.g., 'Flows_UG2N')
            year: Year to filter
            month: Month to filter (1-12)

        Returns:
            Dict mapping column name to volume for the requested (year, month).
        """
        df = self._load_sheet(sheet_name)
        if df.empty:
            return {}

        # Find row matching year and month
        matching_rows = df[(df['Year'] == year) & (df['Month'] == month)]
        if matching_rows.empty:
            return {}

        row = matching_rows.iloc[0]
        volumes: Dict[str, float] = {}
        skip_cols = {'Date', 'Year', 'Month'}
        for col in df.columns:
            if col in skip_cols:
                continue
            try:
                vol = float(row[col])
                if not pd.isna(vol):
                    volumes[col] = vol
            except (TypeError, ValueError):
                # Non-numeric or missing values are ignored
                continue

        logger.info(
            f"📊 Loaded {len(volumes)} flows from '{sheet_name}' ({year}-{month:02d})"
        )
        return volumes
    
    def get_flow_volume(self,
                        area_code: str,
                        excel_mapping: Dict,
                        year: int,
                        month: int) -> Optional[float]:
        """Get a single flow volume using an edge's Excel mapping.

        The mapping may specify a `sheet` and a `column`. If `sheet` is omitted,
        it defaults to the area sheet derived from `area_code`.

        Args:
            area_code: Legacy area code (e.g., 'UG2N', 'UG2S'). Used when the
                mapping does not provide an explicit sheet.
            excel_mapping: Dictionary containing at least `column`, optionally
                `sheet`, and `enabled` flags.
            year: Year of interest.
            month: Month of interest (1-12).

        Returns:
            Volume in m³ if found, otherwise None.
        """
        if not isinstance(excel_mapping, dict):
            return None

        if not excel_mapping.get('enabled', True):
            return None

        column = excel_mapping.get('column')
        if not column:
            return None

        sheet = excel_mapping.get('sheet') or self._get_sheet_name(area_code)

        # Load all volumes for the specific sheet and month, then pick the column
        sheet_vols = self.get_all_volumes_for_month_from_sheet(sheet, year, month)
        return sheet_vols.get(column)
    
    def update_diagram_edges(self,
                            area_data: Dict,
                            area_code: str,
                            year: int,
                            month: int) -> Dict:
        """
        Update all edges in diagram with volumes from Excel.
        
        Args:
            area_data: Diagram JSON data with nodes and edges
            area_code: Mine area code (default sheet to load)
            year: Year
            month: Month
        
        Returns:
            Updated area_data with volumes refreshed
        """
        # Collect all unique sheets referenced by enabled edges (sheet-aware)
        edges = area_data.get('edges', [])
        sheets_to_load: set[str] = set()

        for edge in edges:
            excel_mapping = edge.get('excel_mapping', {})
            if not excel_mapping.get('enabled'):
                continue
            # If a sheet is explicitly set, use exactly that sheet; otherwise default
            sheet = excel_mapping.get('sheet') or self._get_sheet_name(area_code)
            sheets_to_load.add(sheet)

        # Load volumes per sheet to avoid cross-sheet collisions
        volumes_by_sheet: Dict[str, Dict[str, float]] = {}
        for sheet in sheets_to_load:
            volumes_by_sheet[sheet] = self.get_all_volumes_for_month_from_sheet(
                sheet, year, month
            )

        # Update edges with volumes using (sheet, column) mapping
        updated_count = 0
        empty_sheets = set()  # Track sheets that returned no data for this month
        
        for sheet, vols in volumes_by_sheet.items():
            if not vols:
                empty_sheets.add(sheet)
        
        for edge in edges:
            excel_mapping = edge.get('excel_mapping', {})
            if not excel_mapping.get('enabled'):
                continue
            
            sheet = excel_mapping.get('sheet') or self._get_sheet_name(area_code)
            flow_id = excel_mapping.get('column')
            if not flow_id:
                logger.debug(f"⚠️ Skipping edge without 'column' mapping: {edge}")
                continue

            sheet_vols = volumes_by_sheet.get(sheet, {})
            if flow_id not in sheet_vols:
                # Clear any stale labels/volumes to reflect missing data deterministically
                # If the entire sheet is empty for this month, display dash to signal emptiness
                if sheet in empty_sheets:
                    edge['volume'] = None
                    edge['label'] = "—"  # Dash indicates sheet had no data for this month
                    logger.debug(
                        f"⚠️ Sheet '{sheet}' empty for {year}-{month:02d}; edge ({flow_id}) shows '—'"
                    )
                else:
                    edge.pop('label', None)
                    edge['volume'] = None
                    logger.debug(
                        f"⚠️ No volume for ({sheet}, {flow_id}) in {year}-{month:02d}"
                    )
                continue

            # Update edge with sheet-specific volume
            new_volume = sheet_vols[flow_id]
            edge['volume'] = new_volume
            edge['label'] = f"{new_volume:,.2f}"
            updated_count += 1
            logger.debug(
                f"✅ Updated {edge.get('from')} → {edge.get('to')} ({sheet}:{flow_id}): {new_volume:,.2f} m³"
            )

        logger.info(
            f"📈 Updated {updated_count} flow volumes from Excel (sheets: {', '.join(sorted(sheets_to_load))})"
        )
        return area_data
    
    def get_available_months(self, area_code: str) -> List[Tuple[int, int]]:
        """
        Get list of available months (year, month) for an area.
        
        Returns:
            List of (year, month) tuples
        """
        sheet_name = self._get_sheet_name(area_code)
        df = self._load_sheet(sheet_name)
        
        if df.empty or 'Year' not in df.columns or 'Month' not in df.columns:
            return []
        
        # Remove rows with missing Year/Month
        valid = df.dropna(subset=['Year', 'Month'])
        
        # Get unique (year, month) pairs
        months = valid[['Year', 'Month']].drop_duplicates().values.tolist()
        months = [(int(y), int(m)) for y, m in months]
        months.sort()
        
        return months
    
    def clear_cache(self):
        """Clear the sheet cache to force reload from Excel."""
        self._df_cache.clear()
        logger.info("🧹 Flow volume cache cleared")

    def list_sheets(self) -> List[str]:
        """List available sheet names in the Excel workbook."""
        if not self.excel_path.exists():
            return []
        try:
            # Use context manager to avoid leaving the workbook handle open,
            # which can block saving the Excel file while the app is running.
            with pd.ExcelFile(self.excel_path, engine='openpyxl') as xls:
                return list(xls.sheet_names)
        except Exception:
            return []

    def list_sheet_columns(self, sheet_name: str) -> List[str]:
        """List column names for a given sheet (excluding Date/Year/Month).

        Robust to varying header rows. Tries openpyxl row 1 first, then row 3,
        and finally falls back to pandas-based detection.
        """
        # Ensure Excel path is up to date
        new_path = self._resolve_excel_path(None)
        if new_path != self.excel_path:
            self.excel_path = new_path
        if not self.excel_path.exists():
            logger.error(f"❌ Excel file not found for columns: {self.excel_path}")
            return []

        skip = {'Date', 'Year', 'Month'}
        cols: List[str] = []

        # Helper: clean, flatten and filter header values
        def _clean(values: List[object]) -> List[str]:
            cleaned: List[str] = []
            for v in values:
                if v is None:
                    continue
                if isinstance(v, tuple):
                    # For multi-index, only use the LAST element (actual column header)
                    # Skip intermediate "Unnamed" levels and title rows
                    valid_parts = [str(x).strip() for x in v if x is not None and str(x).strip() and not str(x).startswith('Unnamed:')]
                    if valid_parts:
                        # Use the last valid part as the column name
                        flat = valid_parts[-1]
                    else:
                        continue
                else:
                    flat = str(v).strip()
                if not flat or flat in skip:
                    continue
                cleaned.append(flat)
            # Deduplicate while preserving order
            seen = set()
            unique: List[str] = []
            for c in cleaned:
                if c not in seen:
                    unique.append(c)
                    seen.add(c)
            return unique

        # Attempt openpyxl for direct header read with heuristics
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                candidates: List[List[object]] = []
                for r in (1, 2, 3):
                    if ws.max_row >= r:
                        vals = [cell.value for cell in ws[r] if cell.value is not None]
                        str_count = sum(1 for v in vals if isinstance(v, str))
                        # Heuristic: at least 5 string-like values suggests a header row
                        if str_count >= 5:
                            candidates.append(vals)
                if candidates:
                    cols = _clean(candidates[0])
            wb.close()
        except Exception as e:
            logger.debug(f"ℹ️ openpyxl header read failed for '{sheet_name}': {e}")

        # If openpyxl yielded nothing, try to infer header row by scanning first 10 rows
        if not cols:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(self.excel_path, read_only=True, data_only=True)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    best_row_values: List[object] = []
                    best_score = -1.0
                    max_scan = min(ws.max_row or 0, 10)
                    for r in range(1, max_scan + 1):
                        vals = [cell.value for cell in ws[r] if cell.value is not None]
                        if not vals:
                            continue
                        str_count = sum(1 for v in vals if isinstance(v, str))
                        total = len(vals)
                        ratio = (str_count / total) if total else 0.0
                        # Score favors more strings and more total values
                        score = str_count + ratio
                        if str_count >= 5 and score > best_score:
                            best_row_values = vals
                            best_score = score
                    if best_row_values:
                        cols = _clean(best_row_values)
                wb.close()
            except Exception as e:
                logger.debug(f"ℹ️ openpyxl row scan failed for '{sheet_name}': {e}")

        # Fallback to pandas if still nothing detected
        if not cols:
            try:
                # Try multi-row headers first, then single-row
                for header_rows in ([0, 1, 2], [0, 1]):
                    try:
                        df = pd.read_excel(
                            self.excel_path,
                            sheet_name=sheet_name,
                            header=list(header_rows),
                            engine='openpyxl'
                        )
                        if df is not None and not df.empty:
                            if isinstance(df.columns, pd.MultiIndex):
                                flattened = [tuple(map(lambda x: str(x).strip() if x is not None else "", tup)) for tup in df.columns]
                                cols = _clean(flattened)
                            else:
                                cols = _clean(list(df.columns))
                            if cols:
                                break
                    except Exception:
                        continue
                if not cols:
                    for header_row in [0, 1, 2, 3, 4]:
                        try:
                            df = pd.read_excel(
                                self.excel_path,
                                sheet_name=sheet_name,
                                header=header_row,
                                engine='openpyxl'
                            )
                            if df is not None and not df.empty:
                                detected = _clean(list(df.columns))
                                if detected:
                                    cols = detected
                                    break
                        except Exception:
                            continue
            except Exception as e:
                logger.debug(f"ℹ️ pandas header detection failed for '{sheet_name}': {e}")

        logger.info(f"📄 Columns for '{sheet_name}': {len(cols)} found")
        return cols


# Singleton instance
_loader_instance: Optional[FlowVolumeLoader] = None


def get_flow_volume_loader() -> FlowVolumeLoader:
    """Get singleton instance of flow volume loader."""
    global _loader_instance
    if _loader_instance is None:
        _loader_instance = FlowVolumeLoader()
    return _loader_instance


def reset_flow_volume_loader():
    """Reset singleton instance to force path reload from config."""
    global _loader_instance
    _loader_instance = None
    logger.info("🔄 Flow volume loader reset")
