import os
from datetime import date, timedelta
from typing import Optional, Tuple
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle, FancyArrowPatch

from .water_balance_calculator import WaterBalanceCalculator
from .config_manager import config, get_resource_path

class ReportGenerator:
    """Generates PDF and Excel water balance reports."""
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.calculator = WaterBalanceCalculator()
        # A4 landscape dimensions (width, height) in inches
        self.A4_LANDSCAPE = (11.69, 8.27)

    def _month_iter(self, start: date, end: date):
        current = date(start.year, start.month, 1)
        last = date(end.year, end.month, 1)
        while current <= last:
            yield current
            # increment month
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)

    def aggregate_range(self, start: date, end: date) -> dict:
        """Aggregate water balance across month range (inclusive)."""
        totals = {
            'inflows': 0.0,
            'outflows': 0.0,
            'net': 0.0,
            'plant_consumption_gross': 0.0,
            'tsf_return': 0.0,
            'evaporation': 0.0,
            'discharge': 0.0,
            'surface_water': 0.0,
            'groundwater': 0.0,
            'underground': 0.0,
            'rainfall': 0.0,
            'ore_moisture': 0.0,
            'seepage_gain': 0.0,
            'dust_suppression': 0.0,
            'mining_consumption': 0.0,
            'domestic_consumption': 0.0,
        }
        months = []
        for m in self._month_iter(start, end):
            bal = self.calculator.calculate_water_balance(m)
            months.append((m, bal))
            totals['inflows'] += bal['inflows']['total']
            totals['outflows'] += bal['outflows']['total']
            totals['net'] += bal['inflows']['total'] - bal['outflows']['total']
            # breakdowns
            totals['plant_consumption_gross'] += bal['outflows'].get('plant_consumption_gross', 0)
            totals['tsf_return'] += bal['inflows'].get('tsf_return', 0)  # TSF return is an inflow
            totals['evaporation'] += bal['outflows'].get('evaporation', 0)
            totals['discharge'] += bal['outflows'].get('discharge', 0)
            totals['dust_suppression'] += bal['outflows'].get('dust_suppression', 0)
            totals['mining_consumption'] += bal['outflows'].get('mining_consumption', 0)
            totals['domestic_consumption'] += bal['outflows'].get('domestic_consumption', 0)
            totals['surface_water'] += bal['inflows'].get('surface_water', 0)
            totals['groundwater'] += bal['inflows'].get('groundwater', 0)
            totals['underground'] += bal['inflows'].get('underground', 0)
            totals['rainfall'] += bal['inflows'].get('rainfall', 0)
            totals['ore_moisture'] += bal['inflows'].get('ore_moisture', 0)
            totals['seepage_gain'] += bal['inflows'].get('seepage_gain', 0)
        return {'months': months, 'totals': totals}

    def generate_excel_report(self, start: date, end: date, filename: Optional[str] = None) -> Path:
        data = self.aggregate_range(start, end)
        if not filename:
            filename = f"water_balance_{start.strftime('%Y%m')}_{end.strftime('%Y%m')}.xlsx"
        path = self.output_dir / filename

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Summary'

        bold = Font(bold=True)
        
        # Add fixed TRP logo if available (centralized in logo/)
        fixed_logo = get_resource_path('logo/Company Logo.png')
        if fixed_logo.exists():
            try:
                img = XLImage(str(fixed_logo))
                img.width = 120
                img.height = 48
                ws_summary.add_image(img, 'F1')
            except:
                pass

        company = config.get_company_name()
        ws_summary.append([company])
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.append(["Water Balance Summary", f"Range: {start} to {end}"])
        ws_summary.append([])
        t = data['totals']
        ws_summary.append(["Total Inflows", t['inflows']]); ws_summary['A5'].font = bold
        ws_summary.append(["Total Outflows", t['outflows']]); ws_summary['A6'].font = bold
        ws_summary.append(["Net Balance", t['net']]); ws_summary['A7'].font = bold
        ws_summary.append([])
        ws_summary.append(["Plant Gross", t['plant_consumption_gross']])
        ws_summary.append(["TSF Return", t['tsf_return']])
        ws_summary.append(["Evaporation", t['evaporation']])
        ws_summary.append(["Discharge", t['discharge']])
        
        # Add calculation metadata sheet
        ws_meta = wb.create_sheet('Calculation Info')
        ws_meta.append(['CALCULATION METADATA']); ws_meta['A1'].font = Font(bold=True, size=12)
        ws_meta.append([])
        ws_meta.append(['Report Generated:', str(date.today())])
        ws_meta.append(['Period:', f'{start} to {end}'])
        ws_meta.append([])
        ws_meta.append(['FORMULAS']); ws_meta['A6'].font = bold
        constants = self.calculator.constants
        ws_meta.append(['Plant Consumption', f"Ore × {constants.get('mining_water_per_tonne', 0.18)} m³/t"])
        ws_meta.append(['TSF Return', f"Plant × {constants.get('tsf_return_percentage', 56)}%"])
        ws_meta.append(['Evaporation', 'Evap Rate (mm/1000) × Area m²'])
        ws_meta.append([])
        ws_meta.append(['CONSTANTS SNAPSHOT']); ws_meta['A11'].font = bold
        for key, val in sorted(constants.items()):
            ws_meta.append([key, val])

        ws_inflows = wb.create_sheet('Inflows')
        ws_inflows.append(["Month", "Surface", "Groundwater", "Underground", "Rainfall", "Total"])
        for m, bal in data['months']:
            i = bal['inflows']
            ws_inflows.append([
                m.strftime('%Y-%m'), i['surface_water'], i['groundwater'], i['underground'], i['rainfall'], i['total']
            ])

        ws_outflows = wb.create_sheet('Outflows')
        ws_outflows.append(["Month", "Plant Gross", "TSF Return", "Net Plant", "Evaporation", "Discharge", "Total"])
        for m, bal in data['months']:
            o = bal['outflows']
            ws_outflows.append([
                m.strftime('%Y-%m'), o['plant_consumption_gross'], o['tsf_return'], o['plant_consumption_net'], o['evaporation'], o['discharge'], o['total']
            ])

        wb.save(path)
        return path

    def generate_pdf_report(self, start: date, end: date, filename: Optional[str] = None) -> Path:
        data = self.aggregate_range(start, end)
        if not filename:
            filename = f"water_balance_{start.strftime('%Y%m')}_{end.strftime('%Y%m')}.pdf"
        path = self.output_dir / filename

        with PdfPages(path) as pdf:
            page_num = 1
            # Summary page
            # Summary Page (styled header + table) - landscape
            fig, ax = plt.subplots(figsize=self.A4_LANDSCAPE)
            ax.axis('off')
            # Header ribbon
            ax.add_patch(Rectangle((0, 0.94), 1, 0.06, facecolor='#0D47A1', edgecolor='none'))
            company = config.get_company_name()
            ax.text(0.02, 0.967, company, fontsize=15, weight='bold', va='center', color='white')
            ax.text(0.70, 0.967, 'Water Balance Report', fontsize=13, weight='bold', va='center', color='white')
            # Logo (fixed)
            fixed_logo = get_resource_path('logo/Company Logo.png')
            if fixed_logo.exists():
                try:
                    logo_img = plt.imread(str(fixed_logo))
                    ax_logo = fig.add_axes([0.88, 0.945, 0.10, 0.05])
                    ax_logo.imshow(logo_img)
                    ax_logo.axis('off')
                except:
                    pass
            t = data['totals']
            # Summary metrics table
            table_data = [
                ['Metric', 'Value (m³)'],
                ['Total Inflows', f"{t['inflows']:,.0f}"],
                ['Total Outflows', f"{t['outflows']:,.0f}"],
                ['Net Balance', f"{t['net']:,.0f}"],
                ['Plant Gross', f"{t['plant_consumption_gross']:,.0f}"],
                ['TSF Return', f"{t['tsf_return']:,.0f}"],
                ['Evaporation', f"{t['evaporation']:,.0f}"],
                ['Discharge', f"{t['discharge']:,.0f}"],
            ]
            col_colors = ['#BBDEFB', '#E3F2FD']
            the_table = ax.table(cellText=table_data, colLabels=None, loc='upper left', colWidths=[0.30, 0.20])
            for (row, col), cell in the_table.get_celld().items():
                if row == 0:
                    cell.set_facecolor('#1976D2')
                    cell.set_text_props(color='white', weight='bold')
                else:
                    cell.set_facecolor(col_colors[row % 2])
                cell.set_edgecolor('white')
                cell.set_height(0.035)
            the_table.scale(1, 1.4)
            ax.text(0.02, 0.88, f"Period: {start} to {end}", fontsize=11, weight='bold')
            ax.text(0.02, 0.84, "Key Metrics", fontsize=12)
            ax.text(0.02, 0.08, "Confidential - Internal Use Only", fontsize=8, color='#757575')
            ax.text(0.5, 0.02, f"Page {page_num}", ha='center', va='center', fontsize=8, color='#616161')
            ax.text(0.98, 0.02, "Confidential", ha='right', va='center', fontsize=7, color='#9E9E9E')
            pdf.savefig(fig); plt.close(fig)
            page_num += 1
            
            # Calculation Metadata Page
            fig_meta, ax_meta = plt.subplots(figsize=self.A4_LANDSCAPE)
            ax_meta.axis('off')
            ax_meta.text(0.05, 0.98, 'CALCULATION METADATA', fontsize=14, weight='bold', va='top')
            
            # Get constants snapshot
            constants = self.calculator.constants
            meta_lines = [
                "",
                f"Report Generated: {date.today()}",
                f"Calculation Period: {start} to {end}",
                "",
                "FORMULAS USED:",
                "─" * 50,
                f"Plant Water Consumption = Ore Tonnes × {constants.get('mining_water_per_tonne', 0.18)} m³/tonne",
                f"TSF Return Water = Plant Consumption × {constants.get('tsf_return_percentage', 56)}%",
                "Net Plant Consumption = Plant Consumption - TSF Return",
                "Evaporation Loss = (Monthly Evap Rate mm ÷ 1000) × Surface Area m²",
                "Water Balance = Total Inflows - Total Outflows",
                "",
                "CONSTANTS APPLIED:",
                "─" * 50,
            ]
            for key, val in sorted(constants.items()):
                meta_lines.append(f"{key}: {val}")
            
            meta_lines.extend([
                "",
                "ASSUMPTIONS:",
                "─" * 50,
                "• Evaporation rates from S-pan measurements",
                "• TSF return based on historical averages",
                "• Source flows from average flow rates",
                "• Rainfall contribution calculated on storage surface area"
            ])
            
            y = 0.90
            for line in meta_lines:
                ax_meta.text(0.05, y, line, fontsize=9, va='top', family='monospace' if '=' in line or ':' in line else 'sans-serif')
                y -= 0.025
                if y < 0.05: break
            ax_meta.text(0.5, 0.02, f"Page {page_num}", ha='center', va='center', fontsize=8, color='#616161')
            ax_meta.text(0.98, 0.02, "Confidential", ha='right', va='center', fontsize=7, color='#9E9E9E')
            pdf.savefig(fig_meta); plt.close(fig_meta)
            page_num += 1

            # Inflows/Outflows Comparison Chart
            fig2, ax2 = plt.subplots(figsize=(8, 5))
            labels = ['Inflows', 'Outflows']
            values = [t['inflows'], t['outflows']]
            colors = ['#2196F3', '#F44336']
            ax2.bar(labels, values, color=colors)
            ax2.set_title('Total Inflows vs Outflows')
            ax2.set_ylabel('Volume (m³)')
            for i, v in enumerate(values):
                ax2.text(i, v, f"{v:,.0f}", ha='center', va='bottom')
            ax2.text(0.5, -0.08, f"Page {page_num}", ha='center', va='center', fontsize=8, color='#616161', transform=ax2.transAxes)
            ax2.text(0.98, -0.08, "Confidential", ha='right', va='center', fontsize=7, color='#9E9E9E', transform=ax2.transAxes)
            pdf.savefig(fig2); plt.close(fig2)
            page_num += 1

            # Breakdown Pie Chart (Inflows)
            fig3, ax3 = plt.subplots(figsize=(8,5))
            inflow_parts = [t['surface_water'], t['groundwater'], t['underground'], t['rainfall']]
            labels = ["Surface", "Groundwater", "Underground", "Rainfall"]
            ax3.pie(inflow_parts, labels=labels, autopct='%1.1f%%')
            ax3.set_title('Inflows Breakdown')
            ax3.text(0.5, -0.08, f"Page {page_num}", ha='center', va='center', fontsize=8, color='#616161', transform=ax3.transAxes)
            ax3.text(0.98, -0.08, "Confidential", ha='right', va='center', fontsize=7, color='#9E9E9E', transform=ax3.transAxes)
            pdf.savefig(fig3); plt.close(fig3)
            page_num += 1

            # Schematic Diagram Page
            fig4, ax4 = plt.subplots(figsize=self.A4_LANDSCAPE)
            ax4.axis('off')
            ax4.set_title('Water Balance Schematic', fontsize=14, pad=20)

            # Positions (x,y,width,height) in normalized figure coords
            boxes = {
                'Surface Water': (0.05, 0.75, 0.25, 0.08),
                'Groundwater': (0.05, 0.60, 0.25, 0.08),
                'Underground': (0.05, 0.45, 0.25, 0.08),
                'Plant': (0.40, 0.60, 0.20, 0.10),
                'TSF Return': (0.75, 0.60, 0.20, 0.08),
                'Storage': (0.40, 0.30, 0.20, 0.10),
                'Evaporation': (0.75, 0.80, 0.20, 0.08),
                'Discharge': (0.75, 0.40, 0.20, 0.08)
            }

            def draw_box(label, spec, color='#ECEFF1'):
                x, y, w, h = spec
                ax4.add_patch(Rectangle((x, y), w, h, facecolor=color, edgecolor='#455A64'))
                ax4.text(x + w/2, y + h/2, label, ha='center', va='center', fontsize=10)

            for lbl, spec in boxes.items():
                draw_box(lbl, spec)

            def arrow(start_box, end_box, color='#1E88E5', text=None):
                sx, sy, sw, sh = boxes[start_box]
                ex, ey, ew, eh = boxes[end_box]
                start = (sx + sw, sy + sh/2)
                end = (ex, ey + eh/2)
                arr = FancyArrowPatch(start, end, arrowstyle='->', mutation_scale=12, color=color, linewidth=2)
                ax4.add_patch(arr)
                if text:
                    mx = (start[0] + end[0]) / 2
                    my = (start[1] + end[1]) / 2
                    ax4.text(mx, my + 0.02, text, ha='center', va='bottom', fontsize=9, color=color)
            
            # Inflow arrows
            arrow('Surface Water', 'Plant', text='Surface inflow')
            arrow('Groundwater', 'Plant', text='Groundwater inflow', color='#43A047')
            arrow('Underground', 'Plant', text='Underground inflow', color='#6D4C41')
            # Plant to storage (net consumption handled internally)
            arrow('Plant', 'Storage', text='Net to storage', color='#FB8C00')
            # TSF return loop
            arrow('TSF Return', 'Plant', text='Return water', color='#8E24AA')
            # Losses
            arrow('Storage', 'Evaporation', text='Evaporation loss', color='#F4511E')
            arrow('Storage', 'Discharge', text='Controlled discharge', color='#039BE5')

            ax4.text(0.05, 0.15, 'Arrows show principal water movements between sources, plant, storage and losses.', fontsize=9)
            ax4.text(0.5, 0.02, f"Page {page_num}", ha='center', va='center', fontsize=8, color='#616161')
            ax4.text(0.98, 0.02, "Confidential", ha='right', va='center', fontsize=7, color='#9E9E9E')
            pdf.savefig(fig4); plt.close(fig4)
            page_num += 1
            
            # Linear Flow Diagram (Simplified structure for easy understanding)
            try:
                fig5 = self.create_linear_flow_figure(data)
                pdf.savefig(fig5); plt.close(fig5)
                page_num += 1
            except Exception as ex:
                fig_err, ax_err = plt.subplots(figsize=self.A4_LANDSCAPE)
                ax_err.axis('off')
                ax_err.text(0.05, 0.95, 'Flow Diagram Unavailable', fontsize=14, weight='bold')
                ax_err.text(0.05, 0.90, f'Error: {ex}', fontsize=9)
                ax_err.text(0.5, 0.02, f"Page {page_num}", ha='center', va='center', fontsize=8, color='#616161')
                ax_err.text(0.98, 0.02, "Confidential", ha='right', va='center', fontsize=7, color='#9E9E9E')
                pdf.savefig(fig_err); plt.close(fig_err)
        return path

    # ============ Simplified Linear Flow Figure ============
    def create_linear_flow_figure(self, data: dict) -> plt.Figure:
        """Create comprehensive water balance flow diagram matching actual calculations.
        
        Clear breakdown of all components with flow directions:
        FRESH WATER SOURCES → PLANT OPERATIONS → STORAGE SYSTEM → OUTFLOWS/RECYCLING
        """
        # Use data from single month if available, otherwise use totals/average
        months_count = len(data['months'])
        
        # Get date range for title
        if months_count > 0:
            first_month, _ = data['months'][0]
            last_month, _ = data['months'][-1]
            if months_count == 1:
                period_str = first_month.strftime('%B %Y')
                period_type = 'MONTHLY'
            else:
                period_str = f"{first_month.strftime('%b %Y')} - {last_month.strftime('%b %Y')}"
                period_type = 'AVERAGE MONTHLY'
        else:
            period_str = 'No Data'
            period_type = ''
        
        if months_count == 1:
            # Single month - use exact values from that month
            _, bal = data['months'][0]
            surface = bal['inflows'].get('surface_water', 0)
            groundwater = bal['inflows'].get('groundwater', 0)
            underground = bal['inflows'].get('underground', 0)
            rainfall = bal['inflows'].get('rainfall', 0)
            ore_moisture = bal['inflows'].get('ore_moisture', 0)
            seepage_gain = bal['inflows'].get('seepage_gain', 0)
            plant_gross = bal['outflows'].get('plant_consumption_gross', 0)
            tsf_return = bal['inflows'].get('tsf_return', 0)
            evaporation = bal['outflows'].get('evaporation', 0)
            discharge = bal['outflows'].get('discharge', 0)
            dust_suppression = bal['outflows'].get('dust_suppression', 0)
            mining_consumption = bal['outflows'].get('mining_consumption', 0)
            domestic = bal['outflows'].get('domestic_consumption', 0)
            storage_change = bal['storage_change'].get('net_storage_change', 0)
            closure_error = bal.get('closure_error_m3', 0)
            closure_error_pct = bal.get('closure_error_percent', 0)
        else:
            # Multiple months - calculate averages
            t = data['totals']
            surface = t.get('surface_water', 0) / months_count
            groundwater = t.get('groundwater', 0) / months_count
            underground = t.get('underground', 0) / months_count
            rainfall = t.get('rainfall', 0) / months_count
            ore_moisture = t.get('ore_moisture', 0) / months_count
            seepage_gain = t.get('seepage_gain', 0) / months_count
            plant_gross = t.get('plant_consumption_gross', 0) / months_count
            tsf_return = t.get('tsf_return', 0) / months_count
            evaporation = t.get('evaporation', 0) / months_count
            discharge = t.get('discharge', 0) / months_count
            dust_suppression = t.get('dust_suppression', 0) / months_count
            mining_consumption = t.get('mining_consumption', 0) / months_count
            domestic = t.get('domestic_consumption', 0) / months_count
            # Calculate average storage change and closure error
            total_storage_change = sum(m[1]['storage_change'].get('net_storage_change', 0) for m in data['months'])
            total_closure_error = sum(m[1].get('closure_error_m3', 0) for m in data['months'])
            storage_change = total_storage_change / months_count
            closure_error = total_closure_error / months_count
            closure_error_pct = (abs(closure_error) / total_fresh_inflow * 100) if total_fresh_inflow > 0 else 0
        
        # Calculate derived values
        fresh_to_plant = surface + groundwater + underground
        total_fresh_inflow = fresh_to_plant + rainfall + ore_moisture + seepage_gain
        net_plant = plant_gross - tsf_return
        
        # VALIDATION: Calculate total water balance
        # NOTE: Evaporation is NOT included in total outflows per water_balance_calculator.py
        # Evaporation is captured in storage change, not as a direct outflow
        # Total outflows = net_plant + auxiliary uses + discharge
        total_outflows = net_plant + discharge + dust_suppression + mining_consumption + domestic
        # Water balance equation: Fresh IN = OUT + ΔStorage + Closure Error
        # Therefore: Closure Error = Fresh IN - OUT - ΔStorage
        # (closure_error already calculated above from balance data)
        
        # Create figure with more space for better layout
        fig, ax = plt.subplots(figsize=(16, 11))
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.axis('off')
        
        # Title
        ax.text(0.5, 0.97, 'Water Balance System Flow Diagram', 
               ha='center', va='top', fontsize=18, weight='bold')
        ax.text(0.5, 0.94, f'{period_str} - Component Breakdown with Flow Directions',
               ha='center', va='top', fontsize=11, color='#666', style='italic')
        
        # Color scheme
        colors = {
            'fresh_source': '#1E88E5',      # Blue - fresh water
            'recycled': '#8E24AA',          # Purple - recycled
            'plant': '#FB8C00',             # Orange - plant operations
            'storage': '#43A047',           # Green - storage
            'outflow': '#F4511E',           # Red - losses
            'auxiliary': '#00ACC1',         # Cyan - auxiliary uses
        }
        
        def draw_box(x, y, width, height, label, value, color, show_value=True, fontsize=11):
            """Draw a component box with label and value - professional alignment"""
            rect = Rectangle((x - width/2, y - height/2), width, height,
                           facecolor=color, edgecolor='white', linewidth=3, alpha=0.9,
                           zorder=2)
            ax.add_patch(rect)
            
            # Label - clear and bold
            label_y = y + 0.012 if show_value else y
            ax.text(x, label_y, label, 
                   ha='center', va='center', fontsize=fontsize, weight='bold', 
                   color='white', zorder=3)
            
            # Value - below label with clear units
            if show_value and value > 0:
                ax.text(x, y - 0.012, f'{value:,.0f} m³/mo', 
                       ha='center', va='center', fontsize=fontsize-2, 
                       color='white', style='italic', zorder=3)
        
        def draw_arrow(x1, y1, x2, y2, label='', color='#1976D2', linewidth=3, curve=0, label_offset=0.02):
            """Draw arrow - cleaner without labels"""
            style = f'arc3,rad={curve}' if curve != 0 else 'arc3,rad=0'
            arrow = FancyArrowPatch((x1, y1), (x2, y2),
                                   arrowstyle='->', mutation_scale=25,
                                   linewidth=linewidth, color=color, alpha=0.8,
                                   connectionstyle=style, zorder=1)
            ax.add_patch(arrow)
        
        # ==================== SECTION 1: FRESH WATER SOURCES (Left) ====================
        ax.text(0.12, 0.91, '① FRESH WATER', fontsize=12, weight='bold', 
               color=colors['fresh_source'], ha='center',
               bbox=dict(boxstyle='round,pad=0.4', facecolor='#E3F2FD', 
                        edgecolor=colors['fresh_source'], linewidth=2, alpha=0.9))
        
        # Source boxes - clean labels
        draw_box(0.12, 0.83, 0.15, 0.055, 'Surface Water', surface, colors['fresh_source'])
        draw_box(0.12, 0.755, 0.15, 0.055, 'Rainfall', rainfall, colors['fresh_source'])
        draw_box(0.12, 0.68, 0.15, 0.055, 'Underground', underground, colors['fresh_source'])
        draw_box(0.12, 0.605, 0.15, 0.055, 'Groundwater', groundwater, colors['fresh_source'])
        
        # Total indicator - compact
        ax.text(0.12, 0.52, f'TO PLANT: {fresh_to_plant:,.0f} m³/mo',
               ha='center', fontsize=8, weight='bold', color='white',
               bbox=dict(boxstyle='round,pad=0.3', facecolor=colors['fresh_source'], 
                        edgecolor='white', linewidth=2, alpha=0.95))
        
        # ==================== SECTION 2: PLANT OPERATIONS & TSF CYCLE (Center-Left) ====================
        ax.text(0.37, 0.91, '② PLANT & TSF CYCLE', fontsize=13, weight='bold',
               color=colors['plant'], ha='center',
               bbox=dict(boxstyle='round,pad=0.5', facecolor='#FFF3E0', 
                        edgecolor=colors['plant'], linewidth=2, alpha=0.9))
        
        # Plant processing box
        draw_box(0.37, 0.80, 0.16, 0.10, 'PLANT\nPROCESSING', plant_gross, colors['plant'], fontsize=11)
        
        # Tailings Storage Facility (TSF) - moved right to avoid overlap
        draw_box(0.28, 0.64, 0.13, 0.055, 'TSF', 0, colors['plant'], 
                show_value=False, fontsize=10)
        
        # Return Water Dam (RWD) - moved right to avoid overlap
        draw_box(0.28, 0.565, 0.13, 0.055, 'RWD', tsf_return, colors['recycled'], fontsize=10)
        
        return_pct = (tsf_return/plant_gross*100) if plant_gross > 0 else 0
        ax.text(0.37, 0.48, f'♻ {return_pct:.0f}%', 
               ha='center', fontsize=9, weight='bold', color=colors['recycled'],
               bbox=dict(boxstyle='round,pad=0.3', facecolor='#F3E5F5', 
                        edgecolor=colors['recycled'], linewidth=2, alpha=0.95))
        
        # Net consumption indicator
        ax.text(0.50, 0.80, f'{net_plant:,.0f} m³/mo', 
               ha='center', fontsize=8, weight='bold', color='white',
               bbox=dict(boxstyle='round,pad=0.3', facecolor=colors['storage'], 
                        edgecolor='white', linewidth=2, alpha=0.95))
        
        # ==================== SECTION 3: STORAGE SYSTEM (Center-Right) ====================
        ax.text(0.63, 0.90, '③ STORAGE SYSTEM', fontsize=13, weight='bold',
               color=colors['storage'], ha='center',
               bbox=dict(boxstyle='round,pad=0.5', facecolor='#E8F5E9', 
                        edgecolor=colors['storage'], linewidth=2, alpha=0.9))
        
        # Main storage box - larger for importance
        # NOTE: Storage receives net plant consumption + rainfall, minus outflows
        draw_box(0.63, 0.70, 0.18, 0.14, 'WATER STORAGE\nDAMS & PONDS', 0, colors['storage'], 
                show_value=False, fontsize=12)
        
        ax.text(0.63, 0.60, 'Central accumulation point',
               ha='center', fontsize=9, color='white', weight='bold', style='italic')
        ax.text(0.63, 0.565, '━━━━━━━━━━━━━━━',
               ha='center', fontsize=10, color='white')
        ax.text(0.63, 0.53, '• Receives net plant water\n• Collects rainfall\n• Supplies operations\n• Allows discharge',
               ha='center', fontsize=8, color='white', weight='bold')
        
        # ==================== SECTION 4: OUTFLOWS & STORAGE LOSSES (Right) ====================
        ax.text(0.86, 0.90, '④ OUTFLOWS', fontsize=13, weight='bold',
               color=colors['outflow'], ha='center',
               bbox=dict(boxstyle='round,pad=0.5', facecolor='#FFEBEE', 
                        edgecolor=colors['outflow'], linewidth=2, alpha=0.9))
        
        # Environmental Discharge (primary outflow)
        draw_box(0.86, 0.82, 0.14, 0.055, 'Environmental\nDischarge', discharge, colors['outflow'], fontsize=10)
        
        # Auxiliary uses (included in total outflows)
        if dust_suppression > 0 or mining_consumption > 0 or domestic > 0:
            ax.text(0.86, 0.745, 'Auxiliary Uses', fontsize=10, weight='bold',
                   color=colors['auxiliary'], ha='center',
                   bbox=dict(boxstyle='round,pad=0.3', facecolor='#E0F7FA', 
                            edgecolor=colors['auxiliary'], linewidth=1.5, alpha=0.9))
            y_pos = 0.68
            if dust_suppression > 0:
                draw_box(0.86, y_pos, 0.14, 0.055, 'Dust\nSuppression', dust_suppression, colors['auxiliary'], fontsize=10)
                y_pos -= 0.065
            if mining_consumption > 0:
                draw_box(0.86, y_pos, 0.14, 0.055, 'Mining\nOperations', mining_consumption, colors['auxiliary'], fontsize=10)
                y_pos -= 0.065
            if domestic > 0:
                draw_box(0.86, y_pos, 0.14, 0.055, 'Domestic\nUse', domestic, colors['auxiliary'], fontsize=10)
        
        # Total outflows indicator (net plant + discharge + auxiliary)
        ax.text(0.86, 0.50, f'TOTAL OUTFLOWS', 
               ha='center', va='top', fontsize=10, weight='bold', color=colors['outflow'])
        ax.text(0.86, 0.47, f'{total_outflows:,.0f} m³/mo', 
               ha='center', va='top', fontsize=11, weight='bold', color='white',
               bbox=dict(boxstyle='round,pad=0.5', facecolor=colors['outflow'], 
                        edgecolor='white', linewidth=3, alpha=0.95))
        
        # Storage losses (shown separately - NOT included in total outflows)
        ax.text(0.86, 0.39, 'Storage Losses', fontsize=9, weight='bold',
               color='#666', ha='center', style='italic')
        draw_box(0.86, 0.33, 0.14, 0.055, 'Evaporation\nLoss', evaporation, '#999', fontsize=9)
        
        # ==================== FLOW ARROWS (Clean - No Labels) ====================
        
        # Fresh sources → Plant
        if surface > 0:
            draw_arrow(0.195, 0.83, 0.29, 0.83, '', colors['fresh_source'], 2.5, 0.05)
        if underground > 0:
            draw_arrow(0.195, 0.68, 0.29, 0.795, '', colors['fresh_source'], 2.5, -0.05)
        if groundwater > 0:
            draw_arrow(0.195, 0.605, 0.29, 0.76, '', colors['fresh_source'], 2.5, -0.10)
        
        # Rainfall → Storage
        if rainfall > 0:
            draw_arrow(0.195, 0.755, 0.545, 0.73, '', colors['fresh_source'], 2.5, 0.25)
        
        # Plant → TSF → RWD → Plant
        if plant_gross > 0:
            draw_arrow(0.37, 0.75, 0.28, 0.665, '', colors['plant'], 3, -0.15)
        if tsf_return > 0:
            draw_arrow(0.28, 0.615, 0.28, 0.590, '', colors['recycled'], 2.5, 0)
            draw_arrow(0.35, 0.565, 0.37, 0.75, '', colors['recycled'], 3.5, 0.4)
        
        # Net → Storage
        if net_plant > 0:
            draw_arrow(0.54, 0.76, 0.545, 0.76, '', colors['storage'], 3, 0)
        
        # Storage → Evaporation (loss from storage - shown with gray)
        if evaporation > 0:
            draw_arrow(0.72, 0.73, 0.79, 0.33, '', '#999', 2, 0.20)
        
        # Storage → Discharge
        if discharge > 0:
            draw_arrow(0.72, 0.68, 0.79, 0.82, '', colors['outflow'], 2.5, 0)
        
        # ==================== LEGEND (Bottom Left) ====================
        legend_y = 0.25
        ax.text(0.5, legend_y + 0.06, 'LEGEND', fontsize=12, weight='bold', ha='center',
               bbox=dict(boxstyle='round,pad=0.4', facecolor='#FAFAFA', 
                        edgecolor='#BDBDBD', linewidth=2))
        
        # Legend items - two rows for better spacing
        legend_items_row1 = [
            (colors['fresh_source'], 'Fresh Water Sources'),
            (colors['recycled'], 'Recycled Water'),
            (colors['plant'], 'Plant Operations'),
        ]
        legend_items_row2 = [
            (colors['storage'], 'Storage System'),
            (colors['outflow'], 'Water Losses'),
            (colors['auxiliary'], 'Auxiliary Uses'),
        ]
        
        # Row 1
        legend_x_start = 0.18
        spacing = 0.21
        for i, (color, label) in enumerate(legend_items_row1):
            x = legend_x_start + (i * spacing)
            rect = Rectangle((x - 0.018, legend_y - 0.015), 0.036, 0.030,
                           facecolor=color, edgecolor='white', linewidth=2, alpha=0.9)
            ax.add_patch(rect)
            ax.text(x + 0.05, legend_y, label, fontsize=9, va='center', ha='left', weight='bold')
        
        # Row 2
        legend_y -= 0.045
        for i, (color, label) in enumerate(legend_items_row2):
            x = legend_x_start + (i * spacing)
            rect = Rectangle((x - 0.018, legend_y - 0.015), 0.036, 0.030,
                           facecolor=color, edgecolor='white', linewidth=2, alpha=0.9)
            ax.add_patch(rect)
            ax.text(x + 0.05, legend_y, label, fontsize=9, va='center', ha='left', weight='bold')
        
        # ==================== WATER BALANCE EQUATION (Bottom Center) ====================
        eq_y = 0.12
        ax.text(0.5, eq_y + 0.05, 'WATER BALANCE VALIDATION', fontsize=12, weight='bold', 
               ha='center', color='#333')
        
        eq_box_style = dict(boxstyle='round,pad=0.5', facecolor='#E8EAF6', 
                           edgecolor='#3F51B5', linewidth=2.5, alpha=0.95)
        
        # Determine balance status based on closure error percentage
        if abs(closure_error_pct) < 5.0:
            balance_status = 'EXCELLENT'
            balance_color = '#2E7D32'  # Green
        elif abs(closure_error_pct) < 10.0:
            balance_status = 'GOOD'
            balance_color = '#43A047'  # Light green
        elif abs(closure_error_pct) < 20.0:
            balance_status = 'ACCEPTABLE'
            balance_color = '#F57C00'  # Orange
        else:
            balance_status = 'NEEDS REVIEW'
            balance_color = '#D32F2F'  # Red
        
        eq_text = f'Fresh IN: {total_fresh_inflow:,.0f} m³  =  OUT: {total_outflows:,.0f} m³  +  ΔStorage: {storage_change:,.0f} m³  +  Error: {closure_error:,.0f} m³'
        ax.text(0.5, eq_y + 0.01, eq_text,
               fontsize=9, ha='center', weight='bold', color='#1565C0',
               bbox=eq_box_style)
        
        ax.text(0.5, eq_y - 0.03, f'Closure Error: {closure_error_pct:.1f}% | Status: {balance_status}',
               fontsize=10, ha='center', weight='bold', color=balance_color,
               bbox=dict(boxstyle='round,pad=0.4', facecolor='white', 
                        edgecolor=balance_color, linewidth=2, alpha=0.95))
        
        # ==================== FLOW SUMMARY (Bottom) ====================
        summary_y = 0.03
        summary_title = f'{period_type} FLOW SUMMARY' if period_type else 'FLOW SUMMARY'
        ax.text(0.5, summary_y + 0.025, summary_title, 
               fontsize=11, weight='bold', ha='center', color='#333',
               bbox=dict(boxstyle='round,pad=0.4', facecolor='#FAFAFA', 
                        edgecolor='#9E9E9E', linewidth=1.5))
        
        summary_text = (
            f'Fresh Inflow: {total_fresh_inflow:,.0f} m³  |  '
            f'Plant Gross: {plant_gross:,.0f} m³  |  '
            f'TSF Return: {tsf_return:,.0f} m³ ({return_pct:.0f}%)  |  '
            f'Net Plant: {net_plant:,.0f} m³  |  '
            f'Total Outflows: {total_outflows:,.0f} m³'
        )
        
        ax.text(0.5, summary_y - 0.005, summary_text, 
               fontsize=8, ha='center', color='#424242', weight='bold')
        
        plt.tight_layout()
        return fig
    
    # ============ Legacy Dynamic Flow Figure (retained for compatibility) ============
    def create_dynamic_flow_figure(self, data: dict) -> plt.Figure:
        """Create dynamic flow diagram figure (average monthly volumes)."""
        t = data['totals']
        months_count = max(len(data['months']), 1)
        surface = t['surface_water'] / months_count
        groundwater = t['groundwater'] / months_count
        underground = t['underground'] / months_count
        rainfall = t['rainfall'] / months_count
        plant_gross = t['plant_consumption_gross'] / months_count
        tsf_return = t['tsf_return'] / months_count
        net_plant = plant_gross - tsf_return
        evaporation = t['evaporation'] / months_count
        discharge = t['discharge'] / months_count

        def width(v, base=800.0):
            return max(1.0, (v / base) * 8.0)

        fig, ax = plt.subplots(figsize=(9, 6))
        ax.axis('off')
        ax.set_title('Dynamic Flow Diagram (Avg Monthly Volumes)', fontsize=14, pad=15)

        # Auto vertical spacing for inflow nodes (prevents crowding)
        inflow_labels = ['Surface', 'Groundwater', 'Underground', 'Rainfall']
        top_y, bottom_y = 0.85, 0.40
        step = (top_y - bottom_y) / (len(inflow_labels) - 1)
        nodes = {}
        for i, lbl in enumerate(inflow_labels):
            nodes[lbl] = (0.06, top_y - i * step)
        nodes.update({
            'Plant': (0.48, 0.68),
            'Storage': (0.73, 0.50),
            'TSF Return': (0.73, 0.82),
            'Evaporation': (0.94, 0.82),
            'Discharge': (0.94, 0.42)
        })

        def node(label, x, y):
            ax.add_patch(Rectangle((x-0.035, y-0.025), 0.07, 0.05, facecolor='#FAFAFA', edgecolor='#455A64'))
            ax.text(x, y, label, ha='center', va='center', fontsize=9)

        for lbl, (nx, ny) in nodes.items():
            node(lbl, nx, ny)

        def flow(a, b, value, color='#1976D2', label=None, y_offset=0.0):
            x1, y1 = nodes[a]
            x2, y2 = nodes[b]
            lw = width(value)
            arr = FancyArrowPatch((x1+0.035, y1 + y_offset), (x2-0.035, y2 + y_offset),
                                   arrowstyle='->', mutation_scale=12, linewidth=lw,
                                   color=color, alpha=0.78, connectionstyle='arc3,rad=0.08')
            ax.add_patch(arr)
            if label:
                mx = (x1 + x2) / 2
                my = (y1 + y2) / 2 + 0.045 + y_offset
                ax.text(mx, my, f"{label}\n{value:,.0f} m³", ha='center', va='bottom', fontsize=8, color=color)

        # Inflows with slight vertical offsets to avoid overlap
        flow('Surface', 'Plant', surface, '#1E88E5', 'Surface', y_offset=0.01)
        flow('Groundwater', 'Plant', groundwater, '#43A047', 'Groundwater', y_offset=0.00)
        flow('Underground', 'Plant', underground, '#6D4C41', 'Underground', y_offset=-0.01)
        # Rainfall direct to storage
        flow('Rainfall', 'Storage', rainfall, '#00ACC1', 'Rainfall')
        # Net plant to storage and TSF return loop (slight separation)
        flow('Plant', 'Storage', net_plant, '#FB8C00', 'Net Plant', y_offset=0.00)
        flow('TSF Return', 'Plant', tsf_return, '#8E24AA', 'TSF Return', y_offset=0.02)
        # Losses from storage
        flow('Storage', 'Evaporation', evaporation, '#F4511E', 'Evaporation')
        flow('Storage', 'Discharge', discharge, '#039BE5', 'Discharge')

        ax.text(0.05, 0.08, 'Arrow thickness proportional to average monthly volume.', fontsize=9)
        return fig

    def generate_schematic_png(self, filename: Optional[str] = None) -> Path:
        """Export standalone schematic diagram as PNG."""
        if not filename:
            filename = 'water_balance_schematic.png'
        path = self.output_dir / filename

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.axis('off')
        ax.set_title('Water Balance Schematic', fontsize=16, pad=20)

        boxes = {
            'Surface Water': (0.05, 0.70, 0.25, 0.12),
            'Groundwater': (0.05, 0.50, 0.25, 0.12),
            'Underground': (0.05, 0.30, 0.25, 0.12),
            'Plant': (0.45, 0.55, 0.20, 0.14),
            'TSF Return': (0.80, 0.55, 0.15, 0.12),
            'Storage': (0.45, 0.25, 0.20, 0.14),
            'Evaporation': (0.80, 0.75, 0.15, 0.12),
            'Discharge': (0.80, 0.35, 0.15, 0.12)
        }

        def draw_box(label, spec, color='#FAFAFA'):
            x, y, w, h = spec
            ax.add_patch(Rectangle((x, y), w, h, facecolor=color, edgecolor='#37474F', linewidth=1.5))
            ax.text(x + w/2, y + h/2, label, ha='center', va='center', fontsize=11)

        for lbl, spec in boxes.items():
            draw_box(lbl, spec)

        def arrow(start_box, end_box, color='#1976D2', text=None):
            sx, sy, sw, sh = boxes[start_box]
            ex, ey, ew, eh = boxes[end_box]
            start = (sx + sw, sy + sh/2)
            end = (ex, ey + eh/2)
            arr = FancyArrowPatch(start, end, arrowstyle='->', mutation_scale=14, color=color, linewidth=2)
            ax.add_patch(arr)
            if text:
                mx = (start[0] + end[0]) / 2
                my = (start[1] + end[1]) / 2
                ax.text(mx, my + 0.015, text, ha='center', va='bottom', fontsize=9, color=color)

        arrow('Surface Water', 'Plant', text='Surface inflow')
        arrow('Groundwater', 'Plant', text='Groundwater inflow', color='#2E7D32')
        arrow('Underground', 'Plant', text='Underground inflow', color='#5D4037')
        arrow('Plant', 'Storage', text='Net to storage', color='#EF6C00')
        arrow('TSF Return', 'Plant', text='Return water', color='#6A1B9A')
        arrow('Storage', 'Evaporation', text='Evaporation loss', color='#D84315')
        arrow('Storage', 'Discharge', text='Controlled discharge', color='#0288D1')

        ax.text(0.05, 0.05, 'Diagram illustrates primary water flows: inflows to plant, return water loop, storage and losses.', fontsize=10)
        fig.savefig(path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return path

__all__ = ["ReportGenerator"]
