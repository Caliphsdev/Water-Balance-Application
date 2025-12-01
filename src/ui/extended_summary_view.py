"""Extended Summary View
Displays aggregated extended inflow/outflow breakdown and quality flags.
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import date, timedelta
from pathlib import Path
from database.db_manager import db
from utils.config_manager import config

class ExtendedSummaryView(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        # Use shared global db instance to avoid accidental separate empty database
        self.db = db
        self.start_date = date.today() - timedelta(days=6)
        self.end_date = date.today()

    def load(self):
        self._build_header()
        self._build_date_range()
        self._build_tree()
        self._populate()

    def _build_header(self):
        title = ttk.Label(self, text="Extended Water Balance Summary", font=config.get_font('heading_large'))
        title.pack(anchor='w', pady=(10,5), padx=15)
        subtitle = ttk.Label(self, text="Detailed inflows, outflows, and quality distribution", font=config.get_font('body'))
        subtitle.pack(anchor='w', padx=15, pady=(0,10))

    def _build_date_range(self):
        frame = ttk.Frame(self)
        frame.pack(fill='x', padx=15, pady=5)
        ttk.Label(frame, text="Start:").pack(side='left')
        self.start_var = tk.StringVar(value=self.start_date.isoformat())
        self.end_var = tk.StringVar(value=self.end_date.isoformat())
        start_entry = ttk.Entry(frame, textvariable=self.start_var, width=12)
        start_entry.pack(side='left', padx=5)
        ttk.Label(frame, text="End:").pack(side='left')
        end_entry = ttk.Entry(frame, textvariable=self.end_var, width=12)
        end_entry.pack(side='left', padx=5)
        ttk.Button(frame, text="Refresh", command=self._refresh).pack(side='left', padx=10)
        ttk.Button(frame, text="Export PDF", command=self._export_pdf).pack(side='left', padx=5)
        ttk.Button(frame, text="Export Excel", command=self._export_excel).pack(side='left', padx=5)

    def _build_tree(self):
        columns = (
            'metric','value'
        )
        self.tree = ttk.Treeview(self, columns=columns, show='headings', height=28)
        self.tree.heading('metric', text='Metric')
        self.tree.heading('value', text='Total (m³)')
        self.tree.column('metric', width=260)
        self.tree.column('value', width=160, anchor='e')
        self.tree.pack(fill='both', expand=True, padx=15, pady=10)
        self.quality_frame = ttk.Frame(self)
        self.quality_frame.pack(fill='x', padx=15, pady=5)

    def _refresh(self):
        try:
            self.start_date = date.fromisoformat(self.start_var.get())
            self.end_date = date.fromisoformat(self.end_var.get())
        except ValueError:
            return
        for i in self.tree.get_children():
            self.tree.delete(i)
        for w in self.quality_frame.winfo_children():
            w.destroy()
        self._populate()

    def _populate(self):
        # Broaden to all calc types; most entries are 'daily'
        rows = self.db.get_calculations(self.start_date, self.end_date, calc_type=None)
        if not rows:
            self.tree.insert('', 'end', values=("No calculations found for selected range", ""))
            return
        agg = {}
        numeric_cols = [
            ('total_inflows','Total Inflows'),('total_outflows','Total Outflows'),('storage_change','Net Storage Change'),
            ('closure_error_m3','Closure Error (m³)'),('surface_water_inflow','Surface Water'),('groundwater_inflow','Groundwater'),
            ('underground_inflow','Underground'),('rainfall_inflow','Rainfall'),('ore_moisture_inflow','Ore Moisture'),
            ('tsf_return_inflow','TSF Return (Inflow)'),('plant_returns_inflow','Plant Returns'),('returns_to_pit_inflow','Returns to Pit'),
            ('seepage_gain_inflow','Seepage Gain'),('plant_consumption_gross','Plant Consumption Gross'),('plant_consumption_net','Plant Consumption Net'),
            ('tailings_retention_loss','Tailings Retention'),('evaporation_loss','Evaporation'),('seepage_loss','Seepage Loss'),
            ('dust_suppression','Dust Suppression'),('domestic_consumption','Domestic'),('product_moisture','Product Moisture'),
            ('tsf_return_volume','TSF Return (Credit)')
        ]
        for key,_ in numeric_cols:
            agg[key] = 0.0
        quality_flags = {}
        for r in rows:
            for key,_ in numeric_cols:
                agg[key] += r.get(key,0) or 0
            qf = r.get('calc_quality_flag','unknown')
            quality_flags[qf] = quality_flags.get(qf,0) + 1
        for key,label in numeric_cols:
            self.tree.insert('', 'end', values=(label, f"{agg[key]:,.2f}"))
        ttk.Label(self.quality_frame, text="Quality Flags", font=config.get_font('heading_small')).pack(anchor='w')
        for flag,count in quality_flags.items():
            ttk.Label(self.quality_frame, text=f"{flag}: {count}").pack(anchor='w')
    
    def _export_pdf(self):
        """Export extended summary to PDF report"""
        try:
            from matplotlib.backends.backend_pdf import PdfPages
            import matplotlib.pyplot as plt
            from matplotlib.patches import Rectangle
        except ImportError:
            messagebox.showerror("Export Error", "Matplotlib not available for PDF export")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension='.pdf',
            filetypes=[('PDF Document', '*.pdf'), ('All Files', '*.*')],
            initialfile=f'extended_summary_{self.start_date}_{self.end_date}.pdf'
        )
        if not filepath:
            return
        
        try:
            rows = self.db.get_calculations(self.start_date, self.end_date, calc_type=None)
            if not rows:
                messagebox.showwarning("No Data", "No calculations found for selected period")
                return
            
            # Aggregate data
            agg = {}
            numeric_cols = [
                ('total_inflows','Total Inflows'),('total_outflows','Total Outflows'),('storage_change','Net Storage Change'),
                ('closure_error_m3','Closure Error (m³)'),('surface_water_inflow','Surface Water'),('groundwater_inflow','Groundwater'),
                ('underground_inflow','Underground'),('rainfall_inflow','Rainfall'),('ore_moisture_inflow','Ore Moisture'),
                ('tsf_return_inflow','TSF Return (Inflow)'),('plant_returns_inflow','Plant Returns'),('returns_to_pit_inflow','Returns to Pit'),
                ('seepage_gain_inflow','Seepage Gain'),('plant_consumption_gross','Plant Consumption Gross'),('plant_consumption_net','Plant Consumption Net'),
                ('tailings_retention_loss','Tailings Retention'),('evaporation_loss','Evaporation'),('seepage_loss','Seepage Loss'),
                ('dust_suppression','Dust Suppression'),('domestic_consumption','Domestic'),('product_moisture','Product Moisture'),
                ('tsf_return_volume','TSF Return (Credit)')
            ]
            for key,_ in numeric_cols:
                agg[key] = 0.0
            quality_flags = {}
            for r in rows:
                for key,_ in numeric_cols:
                    agg[key] += r.get(key,0) or 0
                qf = r.get('calc_quality_flag','unknown')
                quality_flags[qf] = quality_flags.get(qf,0) + 1
            
            with PdfPages(filepath) as pdf:
                # Page 1: Summary table
                # A4 landscape
                fig, ax = plt.subplots(figsize=(11.69, 8.27))
                ax.axis('off')
                # Header ribbon
                ax.add_patch(Rectangle((0, 0.94), 1, 0.06, facecolor='#0D47A1', edgecolor='none'))
                company = config.get_company_name()
                ax.text(0.02, 0.967, company, fontsize=14, weight='bold', va='center', ha='left', color='white')
                ax.text(0.50, 0.967, 'Extended Water Balance Summary', fontsize=12, weight='bold', va='center', ha='center', color='white')
                # Optional logo on right
                try:
                    logo_path = config.get_logo_path()
                    if logo_path and Path(logo_path).exists():
                        import matplotlib.pyplot as _plt
                        logo_img = _plt.imread(logo_path)
                        ax_logo = fig.add_axes([0.90, 0.945, 0.075, 0.05])
                        ax_logo.imshow(logo_img)
                        ax_logo.axis('off')
                except Exception:
                    pass
                
                ax.text(0.02, 0.88, f"Period: {self.start_date} to {self.end_date}", fontsize=11, weight='bold')
                ax.text(0.02, 0.84, "Detailed Inflow & Outflow Breakdown", fontsize=10)
                
                # Build table data
                table_data = [['Metric', 'Volume (m³)']]
                for key, label in numeric_cols:
                    table_data.append([label, f"{agg[key]:,.2f}"])
                
                col_colors = ['#BBDEFB', '#E3F2FD']
                the_table = ax.table(cellText=table_data, colLabels=None, loc='upper left', 
                                    bbox=[0.02, 0.20, 0.96, 0.60], colWidths=[0.60, 0.36])
                for (row, col), cell in the_table.get_celld().items():
                    if row == 0:
                        cell.set_facecolor('#1976D2')
                        cell.set_text_props(color='white', weight='bold', fontsize=9)
                    else:
                        cell.set_facecolor(col_colors[row % 2])
                        cell.set_text_props(fontsize=8)
                    cell.set_edgecolor('white')
                    cell.set_height(0.025)
                
                ax.text(0.02, 0.15, "Quality Flags Distribution", fontsize=11, weight='bold')
                y_qual = 0.12
                for flag, count in quality_flags.items():
                    ax.text(0.02, y_qual, f"• {flag}: {count} calculations", fontsize=9)
                    y_qual -= 0.02
                
                ax.text(0.02, 0.03, f"Generated on: {date.today()}", fontsize=8, color='#616161')
                ax.text(0.5, 0.02, "Page 1", ha='center', va='center', fontsize=8, color='#616161')
                ax.text(0.98, 0.02, "Confidential", ha='right', va='center', fontsize=7, color='#9E9E9E')
                pdf.savefig(fig); plt.close(fig)
            
            messagebox.showinfo("Export Success", f"Extended summary exported to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export PDF:\n{str(e)}")
    
    def _export_excel(self):
        """Export extended summary to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showerror("Export Error", "openpyxl not available for Excel export")
            return
        
        filepath = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel Workbook', '*.xlsx'), ('All Files', '*.*')],
            initialfile=f'extended_summary_{self.start_date}_{self.end_date}.xlsx'
        )
        if not filepath:
            return
        
        try:
            rows = self.db.get_calculations(self.start_date, self.end_date, calc_type=None)
            if not rows:
                messagebox.showwarning("No Data", "No calculations found for selected period")
                return
            
            # Aggregate data
            agg = {}
            numeric_cols = [
                ('total_inflows','Total Inflows'),('total_outflows','Total Outflows'),('storage_change','Net Storage Change'),
                ('closure_error_m3','Closure Error (m³)'),('surface_water_inflow','Surface Water'),('groundwater_inflow','Groundwater'),
                ('underground_inflow','Underground'),('rainfall_inflow','Rainfall'),('ore_moisture_inflow','Ore Moisture'),
                ('tsf_return_inflow','TSF Return (Inflow)'),('plant_returns_inflow','Plant Returns'),('returns_to_pit_inflow','Returns to Pit'),
                ('seepage_gain_inflow','Seepage Gain'),('plant_consumption_gross','Plant Consumption Gross'),('plant_consumption_net','Plant Consumption Net'),
                ('tailings_retention_loss','Tailings Retention'),('evaporation_loss','Evaporation'),('seepage_loss','Seepage Loss'),
                ('dust_suppression','Dust Suppression'),('domestic_consumption','Domestic'),('product_moisture','Product Moisture'),
                ('tsf_return_volume','TSF Return (Credit)')
            ]
            for key,_ in numeric_cols:
                agg[key] = 0.0
            quality_flags = {}
            for r in rows:
                for key,_ in numeric_cols:
                    agg[key] += r.get(key,0) or 0
                qf = r.get('calc_quality_flag','unknown')
                quality_flags[qf] = quality_flags.get(qf,0) + 1
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'Extended Summary'
            
            # Header
            bold = Font(bold=True, size=12)
            header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
            
            ws['A1'] = config.get_company_name()
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = 'Extended Water Balance Summary'
            ws['A2'].font = bold
            ws['A3'] = f'Period: {self.start_date} to {self.end_date}'
            ws.append([])  # blank row
            
            # Table header
            ws.append(['Metric', 'Volume (m³)'])
            ws['A5'].font = bold
            ws['A5'].fill = header_fill
            ws['B5'].font = bold
            ws['B5'].fill = header_fill
            
            # Data rows
            for key, label in numeric_cols:
                ws.append([label, agg[key]])
            
            # Quality flags section
            ws.append([])
            ws.append(['Quality Distribution'])
            ws[f'A{ws.max_row}'].font = bold
            for flag, count in quality_flags.items():
                ws.append([flag, count])
            
            # Auto-size columns
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 18
            
            wb.save(filepath)
            messagebox.showinfo("Export Success", f"Extended summary exported to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export Excel:\n{str(e)}")
