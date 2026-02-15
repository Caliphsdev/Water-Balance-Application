"""
Excel Preview Widget (FLOW DIAGRAM EXCEL VIEWER).

Provides a lightweight, in-app spreadsheet view for Flow Diagram Excel sheets.
Users can:
- View flow volume columns for a selected sheet
- Edit values directly in a table
- Add blank rows for new months
- Save changes back to the Flow Diagram Excel file

IMPORTANT:
- This widget operates on the Flow Diagram Excel (timeseries_excel_path)
- It does NOT touch the Meter Readings Excel file
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Optional

import pandas as pd
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QComboBox,
    QAbstractItemView,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QBrush

from services.excel_manager import get_excel_manager


class ExcelPreviewWidget(QWidget):
    """Spreadsheet-like preview and editor for Flow Diagram Excel data."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        """Initialize the preview widget and table controls.

        Args:
            parent: Parent widget (dialog container).
        """
        super().__init__(parent)
        self._excel_manager = get_excel_manager()
        self._current_sheet: Optional[str] = None
        self._highlight_column: Optional[str] = None
        self._editable_column: Optional[str] = None
        self._add_row_enabled: bool = True

        # Explanatory label - what is this preview for?
        self._label_help = QLabel("<b>Excel Preview</b> - View Excel data (read-only reference)")
        self._label_help.setStyleSheet("color: #666; font-size: 10pt; padding: 5px;")
        
        # Sheet selector dropdown
        self._combo_sheet = QComboBox()
        self._combo_sheet.setPlaceholderText("Select Excel Sheet...")
        self._combo_sheet.currentTextChanged.connect(self._on_sheet_changed)
        
        self._label_title = QLabel()
        self._label_title.setStyleSheet("font-weight: bold;")
        self._label_title.hide()  # Hidden by default until Excel data loads

        self._table = QTableWidget()
        self._table.setAlternatingRowColors(True)
        self._table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.SelectedClicked)

        self._btn_add_row = QPushButton("Add Row")
        self._btn_save = QPushButton("Save Changes")
        self._btn_save.setToolTip("Save edits back to Excel file (use cautiously!)")
        self._btn_add_row.setObjectName("btn_preview_add_row")
        self._btn_save.setObjectName("btn_preview_save_changes")
        self._btn_add_row.setMinimumHeight(32)
        self._btn_save.setMinimumHeight(32)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self._btn_add_row)
        button_layout.addWidget(self._btn_save)
        button_layout.addStretch(1)

        layout = QVBoxLayout(self)
        layout.addWidget(self._label_help)  # Explanatory text
        layout.addWidget(self._combo_sheet)  # Sheet selector dropdown
        layout.addWidget(self._label_title)
        layout.addWidget(self._table)
        layout.addLayout(button_layout)

        self._btn_add_row.clicked.connect(self._on_add_row)
        self._btn_save.clicked.connect(self._on_save)

    def set_sheet(self, sheet_name: str) -> None:
        """Load and display the specified sheet.

        Args:
            sheet_name: Flow Diagram Excel sheet name (e.g., "Flows_UG2 North").
        """
        if not sheet_name:
            self._clear_table()
            self._label_title.hide()
            return

        if not self._excel_manager.flow_diagram_exists():
            # Excel file not configured - show warning to user
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(
                self,
                "Excel File Not Found",
                "Flow Diagram Excel file not configured.\n\nPlease use the Excel Setup dialog to configure the Excel file path."
            )
            self._clear_table()
            self._label_title.hide()
            return

        # Populate sheet dropdown if empty
        if self._combo_sheet.count() == 0:
            self._populate_sheet_list()
        
        # Set current sheet in dropdown (without triggering reload)
        self._combo_sheet.blockSignals(True)
        index = self._combo_sheet.findText(sheet_name)
        if index >= 0:
            self._combo_sheet.setCurrentIndex(index)
        self._combo_sheet.blockSignals(False)
        
        self._current_sheet = sheet_name
        df = self._excel_manager.load_flow_sheet(sheet_name)
        if df.empty:
            # Sheet is empty - hide title
            self._clear_table()
            self._label_title.hide()
            return

        # Data loaded successfully - show title and populate table
        self._label_title.setText(f"{sheet_name}")
        self._label_title.setStyleSheet("font-weight: bold;")
        self._label_title.show()
        self._populate_table(df)
        self._apply_highlight()
    
    def _populate_sheet_list(self) -> None:
        """Populate the sheet selector dropdown with available sheets from Excel."""
        self._combo_sheet.clear()
        flow_sheets = self._excel_manager.list_flow_sheets()
        if flow_sheets:
            self._combo_sheet.addItems(flow_sheets)
    
    def _on_sheet_changed(self, sheet_name: str) -> None:
        """Handle sheet selection from dropdown.
        
        Args:
            sheet_name: Selected sheet name from dropdown.
        """
        if sheet_name:
            self.set_sheet(sheet_name)

    def set_highlight_column(self, column_name: Optional[str]) -> None:
        """Highlight a column in the preview for easier mapping.

        Args:
            column_name: Column name to highlight, or None to clear highlight.
        """
        self._highlight_column = column_name
        self._apply_highlight()
        self._apply_edit_permissions()

    def set_editable_column(self, column_name: Optional[str]) -> None:
        """Restrict edits to a single column.

        Args:
            column_name: Exact column header that remains editable.
                If None/empty, all table cells become read-only.
        """
        cleaned = (column_name or "").strip()
        self._editable_column = cleaned or None
        self._apply_edit_permissions()

    def set_add_row_enabled(self, enabled: bool) -> None:
        """Show/hide Add Row control based on caller mode."""
        self._add_row_enabled = enabled
        self._btn_add_row.setVisible(enabled)
        self._btn_add_row.setEnabled(enabled)

    def set_help_text(self, html: str) -> None:
        """Update helper copy above the preview table."""
        self._label_help.setText(html)

    def has_loaded_column(self, column_name: Optional[str]) -> bool:
        """Return True when the column exists in the currently loaded table."""
        if not column_name:
            return False
        return self._get_header_index(column_name) is not None

    def _clear_table(self) -> None:
        """Clear the table contents."""
        self._table.clear()
        self._table.setRowCount(0)
        self._table.setColumnCount(0)

    def _get_header_index(self, header_name: str) -> Optional[int]:
        """Find the column index for a given header name.

        Args:
            header_name: Column header to look for.

        Returns:
            Column index if found, otherwise None.
        """
        for col in range(self._table.columnCount()):
            item = self._table.horizontalHeaderItem(col)
            if item and item.text().strip().lower() == header_name.lower():
                return col
        return None

    def _populate_table(self, df: pd.DataFrame) -> None:
        """Populate the table widget with DataFrame content.

        The table displays data starting from row index 0 (first data row),
        with column headers shown separately in the table header.

        Args:
            df: DataFrame loaded from Flow Diagram Excel sheet.
        """
        self._table.clear()
        self._table.setColumnCount(len(df.columns))
        self._table.setRowCount(len(df))
        self._table.setHorizontalHeaderLabels([str(col) for col in df.columns])

        # Populate table rows with actual data (not headers)
        # df.iterrows() returns (index, row) - we use enumerate to get sequential row numbers
        for display_row_idx, (df_index, row) in enumerate(df.iterrows()):
            for col_idx, col_name in enumerate(df.columns):
                value = row[col_name]
                display = "" if pd.isna(value) else str(value)
                item = QTableWidgetItem(display)
                item.setFlags(item.flags() | Qt.ItemIsEditable)
                # Use display_row_idx (0, 1, 2...) instead of df_index
                self._table.setItem(display_row_idx, col_idx, item)

        self._table.resizeColumnsToContents()
        self._apply_edit_permissions()

    def _apply_highlight(self) -> None:
        """Apply column highlight styling if requested."""
        if not self._highlight_column:
            self._clear_column_highlight()
            self._apply_edit_permissions()
            return

        headers = [self._table.horizontalHeaderItem(i).text()
                   for i in range(self._table.columnCount())]
        if self._highlight_column not in headers:
            self._clear_column_highlight()
            self._apply_edit_permissions()
            return

        self._clear_column_highlight()
        col_index = headers.index(self._highlight_column)
        for row in range(self._table.rowCount()):
            item = self._table.item(row, col_index)
            if item:
                item.setBackground(QBrush(QColor("#fff6a4")))

        # Ensure the highlighted column is visible for the user.
        # UX rule: keep mapping column in view so the yellow highlight is obvious.
        first_item = self._table.item(0, col_index) if self._table.rowCount() else None
        if first_item:
            self._table.scrollToItem(first_item, QAbstractItemView.PositionAtCenter)

        self._apply_edit_permissions()

    def _clear_column_highlight(self) -> None:
        """Remove any existing column highlight styling."""
        for row in range(self._table.rowCount()):
            for col in range(self._table.columnCount()):
                item = self._table.item(row, col)
                if item:
                    item.setBackground(QBrush(QColor("#ffffff")))

    def _apply_edit_permissions(self) -> None:
        """Apply editable/read-only flags with visual affordance by column."""
        editable_index: Optional[int] = None
        if self._editable_column:
            editable_index = self._get_header_index(self._editable_column)

        for row in range(self._table.rowCount()):
            for col in range(self._table.columnCount()):
                item = self._table.item(row, col)
                if not item:
                    continue

                is_editable_col = editable_index is not None and col == editable_index
                base_flags = item.flags() | Qt.ItemIsSelectable | Qt.ItemIsEnabled
                if is_editable_col:
                    item.setFlags(base_flags | Qt.ItemIsEditable)
                    item.setForeground(QBrush(QColor("#0f2747")))
                    item.setBackground(QBrush(QColor("#fff6a4")))
                else:
                    item.setFlags(base_flags & ~Qt.ItemIsEditable)
                    item.setForeground(QBrush(QColor("#667085")))
                    item.setBackground(QBrush(QColor("#eef2f7")))

    def _on_add_row(self) -> None:
        """Insert a blank row at the end of the table."""
        if not self._add_row_enabled:
            return

        row_idx = self._table.rowCount()
        self._table.insertRow(row_idx)

        year_col = self._get_header_index("Year")
        month_col = self._get_header_index("Month")

        if year_col is not None and month_col is not None and row_idx > 0:
            prev_year_item = self._table.item(row_idx - 1, year_col)
            prev_month_item = self._table.item(row_idx - 1, month_col)

            try:
                prev_year = int(prev_year_item.text()) if prev_year_item else 0
                prev_month = int(prev_month_item.text()) if prev_month_item else 0
            except ValueError:
                prev_year = 0
                prev_month = 0

            if prev_year and prev_month:
                next_month = prev_month + 1
                next_year = prev_year
                if next_month > 12:
                    next_month = 1
                    next_year += 1

                self._table.setItem(row_idx, year_col, QTableWidgetItem(str(next_year)))
                self._table.setItem(row_idx, month_col, QTableWidgetItem(str(next_month)))

        self._apply_edit_permissions()

    def _on_save(self) -> None:
        """Save table contents back to the Flow Diagram Excel sheet."""
        if not self._current_sheet:
            QMessageBox.warning(self, "Save", "No sheet selected for saving.")
            return

        file_path = self._excel_manager.get_flow_diagram_path()
        if not file_path.exists():
            QMessageBox.warning(self, "Save", "Flow Diagram Excel file not found.")
            return

        headers = [self._table.horizontalHeaderItem(i).text()
                   for i in range(self._table.columnCount())]
        data: List[List[object]] = []
        for row in range(self._table.rowCount()):
            row_values = []
            for col in range(self._table.columnCount()):
                item = self._table.item(row, col)
                row_values.append(item.text() if item else "")
            data.append(row_values)

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            if self._current_sheet not in wb.sheetnames:
                QMessageBox.warning(self, "Save", "Selected sheet not found in Excel.")
                return

            ws = wb[self._current_sheet]
            header_row = self._detect_header_row(ws)
            editable_index = (
                self._get_header_index(self._editable_column)
                if self._editable_column else None
            )

            if editable_index is not None:
                # Flowline-only safeguard: persist only the selected mapped column.
                excel_col = editable_index + 1
                for row_offset, row_values in enumerate(data, start=1):
                    ws.cell(
                        row=header_row + row_offset,
                        column=excel_col,
                        value=row_values[editable_index],
                    )
            else:
                # Default behavior for unrestricted modes.
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=header_row, column=col_idx, value=header)

                for row_offset, row_values in enumerate(data, start=1):
                    for col_idx, value in enumerate(row_values, start=1):
                        ws.cell(row=header_row + row_offset, column=col_idx, value=value)

            wb.save(file_path)
            wb.close()

            # Refresh cached data
            self._excel_manager.clear_flow_cache()
            QMessageBox.information(self, "Save", "Excel data saved successfully.")
        except (OSError, ValueError, KeyError) as exc:
            QMessageBox.warning(self, "Save", f"Failed to save Excel data: {exc}")

    @staticmethod
    def _detect_header_row(ws) -> int:
        """Detect header row using basic heuristics.

        Args:
            ws: openpyxl worksheet object.

        Returns:
            1-based row index for header row.
        """
        best_row = 1
        for row_idx in range(1, min(ws.max_row, 5) + 1):
            values = [str(c.value).strip() for c in ws[row_idx] if c.value is not None]
            if not values:
                continue
            lower = {v.lower() for v in values}
            if {"year", "month"}.issubset(lower) or "date" in lower:
                return row_idx
            if len(values) >= 5:
                best_row = row_idx
        return best_row
