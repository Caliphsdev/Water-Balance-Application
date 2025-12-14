"""
Settings & Configuration UI - Revamped
Modern, clean interface with better organization and visual hierarchy
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path

from database.db_manager import DatabaseManager
from utils.backup_manager import BackupManager
from utils.config_manager import config
from utils.alert_manager import alert_manager
from utils.app_logger import logger


class SettingsModule:
    """Revamped Settings & Configuration UI"""
    
    def __init__(self, parent, initial_tab=None):
        self.parent = parent
        self.container = ttk.Frame(parent)
        self.db = DatabaseManager()
        self.backup = BackupManager()
        self.initial_tab = initial_tab
        self.notebook = None

    def load(self):
        self.container.pack(fill='both', expand=True)
        self._build_ui()

    def _build_ui(self):
        """Build modern settings interface"""
        # Header with icon and description
        header = ttk.Frame(self.container)
        header.pack(fill='x', padx=20, pady=(20, 10))
        
        title_label = ttk.Label(
            header, 
            text="⚙️ Settings & Configuration",
            font=('Segoe UI', 20, 'bold')
        )
        title_label.pack(anchor='w')
        
        subtitle = ttk.Label(
            header,
            text="Manage application settings, constants, backups, and system configuration",
            font=('Segoe UI', 10),
            foreground='#666'
        )
        subtitle.pack(anchor='w', pady=(5, 0))
        
        # Separator
        ttk.Separator(self.container, orient='horizontal').pack(fill='x', padx=20, pady=10)
        
        # Notebook with modern tabs
        self.notebook = ttk.Notebook(self.container)
        self.notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Create tab frames
        self.branding_frame = ttk.Frame(self.notebook)
        self.constants_frame = ttk.Frame(self.notebook)
        self.data_sources_frame = ttk.Frame(self.notebook)
        self.backup_frame = ttk.Frame(self.notebook)
        
        # Add tabs with icons
        self.notebook.add(self.branding_frame, text='  🎨 Branding  ')
        self.notebook.add(self.constants_frame, text='  📊 Constants  ')
        self.notebook.add(self.data_sources_frame, text='  📂 Data Sources  ')
        self.notebook.add(self.backup_frame, text='  💾 Backup  ')
        
        # Build each tab
        self._build_branding_tab()
        self._build_constants_tab()
        self._build_data_sources_tab()
        self._build_backup_tab()
        


    def _build_branding_tab(self):
        """Build branding configuration with modern card layout"""
        scroll_container = ttk.Frame(self.branding_frame)
        scroll_container.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Company Information Card
        company_card = ttk.LabelFrame(scroll_container, text='  📝 Company Information  ', padding=20)
        company_card.pack(fill='x', pady=(0, 15))
        
        # Company name with better layout
        name_frame = ttk.Frame(company_card)
        name_frame.pack(fill='x', pady=(0, 15))
        
        ttk.Label(name_frame, text='Company Name', font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        ttk.Label(name_frame, text='This name will appear on all reports and exports', 
                 font=('Segoe UI', 9), foreground='#666').pack(anchor='w', pady=(2, 5))
        
        name_entry_frame = ttk.Frame(name_frame)
        name_entry_frame.pack(fill='x')
        
        self.company_name_entry = ttk.Entry(name_entry_frame, font=('Segoe UI', 10), width=50)
        self.company_name_entry.insert(0, config.get_company_name())
        self.company_name_entry.pack(side='left', fill='x', expand=True)
        
        ttk.Button(name_entry_frame, text='💾 Save', command=self._save_company_name, 
                  style='Accent.TButton').pack(side='left', padx=(10, 0))
        
        # Logo Configuration Card
        logo_card = ttk.LabelFrame(scroll_container, text='  🖼️ Company Logo  ', padding=20)
        logo_card.pack(fill='x', pady=(0, 15))
        
        ttk.Label(logo_card, text='Logo Image', font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        ttk.Label(logo_card, text='Upload your company logo to display in the application toolbar and reports',
                 font=('Segoe UI', 9), foreground='#666').pack(anchor='w', pady=(2, 10))
        
        # Logo preview with border
        preview_frame = ttk.Frame(logo_card, relief='solid', borderwidth=1)
        preview_frame.pack(fill='x', pady=(0, 10))
        
        self.logo_preview_label = ttk.Label(preview_frame, text='No logo selected', 
                                           foreground='#999', padding=40)
        self.logo_preview_label.pack()
        self._update_logo_preview()
        
        # Logo path display
        path_frame = ttk.Frame(logo_card)
        path_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(path_frame, text='Current Path:', font=('Segoe UI', 9)).pack(side='left')
        self.logo_path_label = ttk.Label(path_frame, text=config.get_logo_path() or 'None', 
                                        foreground='#666', font=('Segoe UI', 9, 'italic'))
        self.logo_path_label.pack(side='left', padx=(5, 0))
        
        # Logo action buttons
        logo_btn_frame = ttk.Frame(logo_card)
        logo_btn_frame.pack(fill='x')
        
        ttk.Button(logo_btn_frame, text='📁 Choose Logo File', command=self._select_logo).pack(side='left', padx=(0, 10))
        ttk.Button(logo_btn_frame, text='❌ Remove Logo', command=self._remove_logo).pack(side='left')
        
        # Recommendations Card
        tips_card = ttk.LabelFrame(scroll_container, text='  💡 Recommendations  ', padding=15)
        tips_card.pack(fill='x')
        
        tips = [
            "• Use PNG format with transparent background for best results",
            "• Recommended dimensions: 200×80 pixels (or similar 2.5:1 ratio)",
            "• Logo will be automatically resized to fit the toolbar",
            "• Supported formats: PNG, JPG, JPEG"
        ]
        
        for tip in tips:
            ttk.Label(tips_card, text=tip, font=('Segoe UI', 9), foreground='#555').pack(anchor='w', pady=2)
    
    def _save_company_name(self):
        """Save company name with validation"""
        name = self.company_name_entry.get().strip()
        if not name:
            messagebox.showwarning('Validation Error', 'Company name cannot be empty')
            return
        
        config.set_company_name(name)
        messagebox.showinfo('✓ Saved', f'Company name updated to: {name}\n\nChanges will appear in reports and exports.')
    
    def _select_logo(self):
        """Select and copy logo file"""
        filename = filedialog.askopenfilename(
            title='Select Company Logo',
            filetypes=[
                ('PNG Images', '*.png'),
                ('JPEG Images', '*.jpg *.jpeg'),
                ('All Images', '*.png *.jpg *.jpeg')
            ]
        )
        if filename:
            try:
                import shutil
                dest_dir = Path(__file__).parent.parent.parent / 'config' / 'branding'
                dest_dir.mkdir(parents=True, exist_ok=True)
                dest_file = dest_dir / Path(filename).name
                shutil.copy2(filename, dest_file)
                
                config.set_logo_path(str(dest_file))
                self.logo_path_label.config(text=str(dest_file))
                self._update_logo_preview()
                messagebox.showinfo('✓ Logo Updated', 
                                   f'Logo successfully uploaded: {dest_file.name}\n\n'
                                   'Restart the application to see it in the toolbar.')
            except Exception as ex:
                messagebox.showerror('Upload Error', f'Failed to copy logo file:\n{str(ex)}')
    
    def _remove_logo(self):
        """Remove current logo"""
        if not config.get_logo_path():
            messagebox.showinfo('No Logo', 'No logo is currently set.')
            return
        
        confirm = messagebox.askyesno('Confirm Removal', 
                                     'Remove the current logo?\n\n'
                                     'This will remove it from the toolbar and reports.')
        if not confirm:
            return
        
        config.set_logo_path('')
        self.logo_path_label.config(text='None')
        self._update_logo_preview()
        messagebox.showinfo('✓ Logo Removed', 'Logo has been removed.')
    
    def _update_logo_preview(self):
        """Update logo preview with better styling"""
        logo_path = config.get_logo_path()
        if logo_path and Path(logo_path).exists():
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path)
                
                # Create thumbnail maintaining aspect ratio
                img.thumbnail((200, 100), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                
                self.logo_preview_label.config(image=photo, text='')
                self.logo_preview_label.image = photo
            except Exception as ex:
                self.logo_preview_label.config(
                    text=f'⚠️ Preview unavailable\n{str(ex)}',
                    foreground='#dc3545'
                )
        else:
            self.logo_preview_label.config(
                text='📷\nNo logo uploaded',
                foreground='#999',
                image=''
            )

    def _build_constants_tab(self):
        """Build constants management with enhanced UI"""
        container = ttk.Frame(self.constants_frame)
        container.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Filter and search bar
        filter_bar = ttk.Frame(container)
        filter_bar.pack(fill='x', pady=(0, 15))
        
        ttk.Label(filter_bar, text='Category:', font=('Segoe UI', 10)).pack(side='left', padx=(0, 5))
        self.category_filter = ttk.Combobox(
            filter_bar,
            values=['All', 'Plant', 'Evaporation', 'Operating', 'calculation', 'threshold'],
            state='readonly',
            width=15
        )
        self.category_filter.set('All')
        self.category_filter.pack(side='left', padx=(0, 20))
        self.category_filter.bind('<<ComboboxSelected>>', lambda e: self._load_constants())
        
        ttk.Button(filter_bar, text='🔄 Refresh', command=self._load_constants).pack(side='left', padx=(0, 10))
        ttk.Button(filter_bar, text='📊 Export to Excel', command=self._export_constants).pack(side='left', padx=(0, 10))
        ttk.Button(filter_bar, text='📜 View History', command=self._view_history).pack(side='left')
        
        # Constants table with improved styling
        table_frame = ttk.LabelFrame(container, text='  System Constants  ', padding=10)
        table_frame.pack(fill='both', expand=True, pady=(0, 15))
        
        cols = ('Category', 'Key', 'Value', 'Unit', 'Range', 'Description')
        self.tree = ttk.Treeview(table_frame, columns=cols, show='headings', height=12)
        
        widths = {'Category': 100, 'Key': 180, 'Value': 90, 'Unit': 90, 'Range': 120, 'Description': 250}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=widths[c], anchor='w')
        
        scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self._load_constants()
        
        # Edit panel with modern card design
        edit_card = ttk.LabelFrame(container, text='  ✏️ Edit Selected Constant  ', padding=20)
        edit_card.pack(fill='x')
        
        # Selected constant info
        info_frame = ttk.Frame(edit_card)
        info_frame.pack(fill='x', pady=(0, 15))
        
        ttk.Label(info_frame, text='Constant:', font=('Segoe UI', 9)).grid(row=0, column=0, sticky='w', pady=2)
        self.selected_key = ttk.Label(info_frame, text='(None selected)', 
                                      font=('Segoe UI', 10, 'bold'), foreground='#0066cc')
        self.selected_key.grid(row=0, column=1, sticky='w', padx=10, pady=2)
        
        ttk.Label(info_frame, text='Description:', font=('Segoe UI', 9)).grid(row=1, column=0, sticky='nw', pady=2)
        self.selected_desc = ttk.Label(info_frame, text='', wraplength=500, 
                                       font=('Segoe UI', 9), foreground='#666')
        self.selected_desc.grid(row=1, column=1, sticky='w', padx=10, pady=2)
        
        # Value editor
        editor_frame = ttk.Frame(edit_card)
        editor_frame.pack(fill='x')
        
        ttk.Label(editor_frame, text='New Value:', font=('Segoe UI', 10, 'bold')).pack(side='left', padx=(0, 10))
        self.new_value = ttk.Entry(editor_frame, font=('Segoe UI', 11), width=20)
        self.new_value.pack(side='left', padx=(0, 10))
        
        self.value_range_label = ttk.Label(editor_frame, text='', foreground='#17a2b8', 
                                           font=('Segoe UI', 9))
        self.value_range_label.pack(side='left', padx=(0, 15))
        
        ttk.Button(editor_frame, text='💾 Update Constant', command=self._update_constant,
                  style='Accent.TButton').pack(side='left')
        
        # Bind selection event
        self.tree.bind('<<TreeviewSelect>>', self._on_constant_select)
        self.current_constant = None

    def _on_constant_select(self, event):
        """Handle constant selection"""
        sel = self.tree.selection()
        if not sel:
            return
        
        item = self.tree.item(sel[0])
        # We ignore range limits now; range_str kept for column compatibility
        cat, key, value, unit, _range_str, desc = item['values']
        
        self.selected_key.config(text=key)
        self.selected_desc.config(text=desc if desc and desc != '—' else 'No description available')
        # Limits removed – clear any previous range display
        self.value_range_label.config(text='')
        
        self.new_value.delete(0, 'end')
        self.new_value.insert(0, str(value))
        
        # Load metadata for validation
        self.current_constant = self._get_constant_metadata(key)

    def _get_constant_metadata(self, key):
        """Get full constant metadata"""
        result = self.db.execute_query(
            'SELECT * FROM system_constants WHERE constant_key = ?', (key,)
        )
        return result[0] if result else None

    def _load_constants(self):
        """Load constants into table"""
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        cat_filter = self.category_filter.get()
        
        query = '''SELECT constant_key, constant_value, unit, category, description, 
                   min_value, max_value FROM system_constants ORDER BY category, constant_key'''
        results = self.db.execute_query(query)
        
        for row in results:
            key = row['constant_key']
            value = row['constant_value']
            unit = row.get('unit') or '—'
            category = row.get('category') or 'Other'
            desc = row.get('description') or '—'
            # Previously used for range validation; now ignored
            # min_val = row.get('min_value')
            # max_val = row.get('max_value')
            
            if cat_filter != 'All' and category != cat_filter:
                continue
            
            # Always show em dash since limits removed
            range_str = '—'
            
            self.tree.insert('', 'end', values=(category, key, value, unit, range_str, desc))

    def _update_constant(self):
        """Update selected constant with validation"""
        key = self.selected_key.cget('text')
        if key == '(None selected)':
            messagebox.showwarning('No Selection', 'Please select a constant to update.')
            return
        
        try:
            value = float(self.new_value.get())
        except ValueError:
            messagebox.showerror('Invalid Value', 'Please enter a valid numeric value.')
            return
        
        # Range validation removed per request – any numeric value allowed
        
        # Critical constant warning
        critical = ['TSF_RETURN_RATE', 'MINING_WATER_RATE', 'SLURRY_DENSITY']
        if key in critical:
            confirm = messagebox.askyesno('⚠️ Critical Constant',
                f'You are modifying a critical system constant: {key}\n\n'
                f'Current value: {self.current_constant["constant_value"]}\n'
                f'New value: {value}\n\n'
                f'This will affect all future water balance calculations.\n\n'
                'Do you want to proceed?')
            if not confirm:
                return
        
        self.db.update_constant(key, value, user='user')
        self._load_constants()
        messagebox.showinfo('✓ Updated', f'Constant {key} has been updated to {value}.')

    def _export_constants(self):
        """Export constants to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from datetime import datetime
            
            filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile=f'water_balance_constants_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
            if not filename:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'System Constants'
            
            # Headers with styling
            headers = ['Category', 'Constant Key', 'Value', 'Unit', 'Min', 'Max', 'Description']
            ws.append(headers)
            
            header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
            
            # Data
            query = '''SELECT category, constant_key, constant_value, unit, min_value, max_value, description 
                      FROM system_constants ORDER BY category, constant_key'''
            results = self.db.execute_query(query)
            
            for row in results:
                ws.append([
                    row.get('category') or '',
                    row['constant_key'],
                    row['constant_value'],
                    row.get('unit') or '',
                    row.get('min_value') or '',
                    row.get('max_value') or '',
                    row.get('description') or ''
                ])
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo('✓ Export Complete', f'Constants exported successfully to:\n\n{filename}')
        except Exception as ex:
            messagebox.showerror('Export Error', f'Failed to export constants:\n{str(ex)}')

    def _view_history(self):
        """Show audit log of constant changes"""
        history_window = tk.Toplevel(self.parent)
        history_window.title('Constants Change History')
        history_window.geometry('900x500')
        
        frame = ttk.Frame(history_window, padding=15)
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text='📜 Recent Constant Changes', 
                 font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
        
        cols = ('Timestamp', 'Constant', 'Old Value', 'New Value', 'User')
        tree = ttk.Treeview(frame, columns=cols, show='headings', height=18)
        
        widths = {'Timestamp': 180, 'Constant': 200, 'Old Value': 120, 'New Value': 120, 'User': 100}
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=widths[c])
        
        tree.pack(fill='both', expand=True, pady=(0, 10))
        
        # Load history
        try:
            import json
            logs = self.db.execute_query(
                "SELECT * FROM audit_log WHERE table_name = 'system_constants' "
                "ORDER BY changed_at DESC LIMIT 100"
            )
            
            for log in logs:
                timestamp = log['changed_at']
                old_vals = json.loads(log['old_values']) if log['old_values'] else {}
                new_vals = json.loads(log['new_values']) if log['new_values'] else {}
                
                for key in new_vals:
                    old_val = old_vals.get(key, '—')
                    new_val = new_vals[key]
                    tree.insert('', 'end', values=(timestamp, key, old_val, new_val, log['changed_by']))
        except Exception as ex:
            ttk.Label(frame, text=f'⚠️ Error loading history: {ex}', 
                     foreground='#dc3545').pack()
        
        ttk.Button(frame, text='Close', command=history_window.destroy).pack(pady=(10, 0))


    def _build_backup_tab(self):
        """Build backup management with modern UI"""
        container = ttk.Frame(self.backup_frame)
        container.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Info card
        info_card = ttk.LabelFrame(container, text='  💾 Database Backup & Restore  ', padding=20)
        info_card.pack(fill='x', pady=(0, 15))
        
        ttk.Label(info_card, text='Regular backups protect your water balance data',
                 font=('Segoe UI', 10), foreground='#666').pack(anchor='w', pady=(0, 10))
        
        # Actions
        actions = ttk.Frame(info_card)
        actions.pack(fill='x')
        
        ttk.Button(actions, text='💾 Create Backup Now', command=self._create_backup,
                  style='Accent.TButton').pack(side='left', padx=(0, 10))
        ttk.Button(actions, text='🔄 Refresh List', command=self._load_backups).pack(side='left')
        
        # Backups list
        list_card = ttk.LabelFrame(container, text='  📂 Available Backups  ', padding=15)
        list_card.pack(fill='both', expand=True, pady=(0, 15))
        
        self.backups_list = tk.Listbox(list_card, height=12, font=('Consolas', 9))
        self.backups_list.pack(fill='both', expand=True)
        
        self._load_backups()
        
        # Restore section
        restore_card = ttk.LabelFrame(container, text='  ⚠️ Restore from Backup  ', padding=15)
        restore_card.pack(fill='x')
        
        ttk.Label(restore_card, text='⚠️ Warning: Restoring will overwrite the current database',
                 font=('Segoe UI', 9), foreground='#dc3545').pack(anchor='w', pady=(0, 10))
        
        ttk.Button(restore_card, text='🔙 Restore Selected Backup', 
                  command=self._restore_selected).pack(anchor='w')

    def _build_data_sources_tab(self):
        """Build data sources configuration tab for Excel file paths"""
        container = ttk.Frame(self.data_sources_frame)
        container.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Info card
        info_card = ttk.LabelFrame(container, text='  📂 Excel Data Source Configuration  ', padding=20)
        info_card.pack(fill='x', pady=(0, 15))
        
        ttk.Label(info_card, text='Configure paths to Excel files used by the application',
                 font=('Segoe UI', 10), foreground='#666').pack(anchor='w', pady=(0, 10))
        
        ttk.Label(info_card, text='💡 If you move Excel files to a different location, update the paths here.',
                 font=('Segoe UI', 9), foreground='#0066cc').pack(anchor='w')
        
        # Legacy Excel file
        legacy_card = ttk.LabelFrame(container, text='  📊 TRP Water Balance Excel  ', padding=20)
        legacy_card.pack(fill='x', pady=(0, 15))
        
        ttk.Label(legacy_card, text='TRP historical water balance data (optional)',
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        ttk.Label(legacy_card, text='Used for historical analysis and legacy charts. App works without this file.',
                 font=('Segoe UI', 9), foreground='#666').pack(anchor='w', pady=(2, 10))
        
        legacy_path_frame = ttk.Frame(legacy_card)
        legacy_path_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(legacy_path_frame, text='Current Path:', font=('Segoe UI', 9)).pack(side='left')
        
        legacy_path = config.get('data_sources.legacy_excel_path', 'data/New Water Balance  20250930 Oct.xlsx')
        self.legacy_path_label = ttk.Label(legacy_path_frame, text=legacy_path,
                                          foreground='#0066cc', font=('Segoe UI', 9, 'italic'))
        self.legacy_path_label.pack(side='left', padx=(5, 10))
        
        # Status indicator
        base_dir = Path(__file__).parent.parent.parent
        legacy_full = base_dir / legacy_path if not Path(legacy_path).is_absolute() else Path(legacy_path)
        status_text = '✓ Found' if legacy_full.exists() else '❌ Missing'
        status_color = '#28a745' if legacy_full.exists() else '#dc3545'
        self.legacy_status_label = ttk.Label(legacy_path_frame, text=status_text,
                                             foreground=status_color, font=('Segoe UI', 9, 'bold'))
        self.legacy_status_label.pack(side='left')
        
        legacy_btn_frame = ttk.Frame(legacy_card)
        legacy_btn_frame.pack(fill='x')
        
        ttk.Button(legacy_btn_frame, text='📁 Browse...', 
                  command=self._select_legacy_excel).pack(side='left', padx=(0, 10))
        ttk.Button(legacy_btn_frame, text='🔄 Reset to Default',
                  command=self._reset_legacy_path).pack(side='left')
        
        # Template Excel file
        template_card = ttk.LabelFrame(container, text='  📝 Application Inputs Excel  ', padding=20)
        template_card.pack(fill='x', pady=(0, 15))
        
        ttk.Label(template_card, text='Primary data input file (required)',
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        ttk.Label(template_card, text='This file is monitored for changes and auto-reloaded. App requires this file to function.',
                 font=('Segoe UI', 9), foreground='#666').pack(anchor='w', pady=(2, 10))
        
        template_path_frame = ttk.Frame(template_card)
        template_path_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(template_path_frame, text='Current Path:', font=('Segoe UI', 9)).pack(side='left')
        
        template_path = config.get('data_sources.template_excel_path', 'templates/Water_Balance_TimeSeries_Template.xlsx')
        self.template_path_label = ttk.Label(template_path_frame, text=template_path,
                                            foreground='#0066cc', font=('Segoe UI', 9, 'italic'))
        self.template_path_label.pack(side='left', padx=(5, 10))
        
        # Status indicator
        base_dir = Path(__file__).parent.parent.parent
        template_full = base_dir / template_path if not Path(template_path).is_absolute() else Path(template_path)
        status_text = '✓ Found' if template_full.exists() else '❌ Missing'
        status_color = '#28a745' if template_full.exists() else '#dc3545'
        self.template_status_label = ttk.Label(template_path_frame, text=status_text,
                                              foreground=status_color, font=('Segoe UI', 9, 'bold'))
        self.template_status_label.pack(side='left')
        
        template_btn_frame = ttk.Frame(template_card)
        template_btn_frame.pack(fill='x')
        
        ttk.Button(template_btn_frame, text='📁 Browse...', 
                  command=self._select_template_excel).pack(side='left', padx=(0, 10))
        ttk.Button(template_btn_frame, text='🔄 Reset to Default',
                  command=self._reset_template_path).pack(side='left', padx=(0, 10))
        ttk.Button(template_btn_frame, text='📂 Open Folder',
                  command=self._open_template_folder).pack(side='left')
        
        # Warning card
        warning_card = ttk.LabelFrame(container, text='  ⚠️ Important Notes  ', padding=15)
        warning_card.pack(fill='x')
        
        warnings = [
            "• Changes take effect immediately but require app restart for full reload",
            "• Ensure Excel files exist at the specified paths",
            "• Use absolute paths or paths relative to application root",
            "• Backup your Excel files before making path changes",
            "• Template file should be accessible for read/write operations"
        ]
        
        for warning in warnings:
            ttk.Label(warning_card, text=warning, font=('Segoe UI', 9), foreground='#666').pack(anchor='w', pady=2)
    
    def _select_legacy_excel(self):
        """Select legacy Excel file"""
        current_path = config.get('data_sources.legacy_excel_path', 'data/New Water Balance  20250930 Oct.xlsx')
        initial_dir = Path(current_path).parent if Path(current_path).exists() else Path.cwd()
        
        filename = filedialog.askopenfilename(
            title='Select TRP Water Balance Excel File',
            initialdir=initial_dir,
            filetypes=[
                ('Excel Files', '*.xlsx *.xls'),
                ('All Files', '*.*')
            ]
        )
        
        if filename:
            # Convert to relative path if inside project
            try:
                rel_path = Path(filename).relative_to(Path.cwd())
                path_str = str(rel_path)
            except ValueError:
                path_str = filename
            
            config.set('data_sources.legacy_excel_path', path_str)
            self.legacy_path_label.config(text=path_str)
            
            # Update status indicator
            base_dir = Path.cwd()
            full_path = base_dir / path_str if not Path(path_str).is_absolute() else Path(path_str)
            if full_path.exists():
                self.legacy_status_label.config(text='✓ Found', foreground='#28a745')
            else:
                self.legacy_status_label.config(text='❌ Missing', foreground='#dc3545')
            
            # Trigger automatic reload
            self._reload_excel_data()
            
            messagebox.showinfo('✓ Path Updated', 
                               f'TRP Excel path updated to:\n{path_str}\n\n'
                               'Data has been reloaded from the new location.')
    
    def _select_template_excel(self):
        """Select template Excel file"""
        current_path = config.get('data_sources.template_excel_path', 'templates/Water_Balance_TimeSeries_Template.xlsx')
        initial_dir = Path(current_path).parent if Path(current_path).exists() else Path.cwd()
        
        filename = filedialog.askopenfilename(
            title='Select Application Inputs Excel File',
            initialdir=initial_dir,
            filetypes=[
                ('Excel Files', '*.xlsx *.xls'),
                ('All Files', '*.*')
            ]
        )
        
        if filename:
            # Convert to relative path if inside project
            try:
                rel_path = Path(filename).relative_to(Path.cwd())
                path_str = str(rel_path)
            except ValueError:
                path_str = filename
            
            config.set('data_sources.template_excel_path', path_str)
            # Also update timeseries_excel_path so flow loader picks it up
            config.set('data_sources.timeseries_excel_path', path_str)
            
            # Reset flow volume loader singleton to force path reload
            try:
                from utils.flow_volume_loader import reset_flow_volume_loader
                reset_flow_volume_loader()
                logger.info(f"🔄 Flow loader reset after path change to: {path_str}")
            except Exception as e:
                logger.warning(f"Could not reset flow volume loader: {e}")
            
            self.template_path_label.config(text=path_str)
            
            # Update status indicator
            base_dir = Path.cwd()
            full_path = base_dir / path_str if not Path(path_str).is_absolute() else Path(path_str)
            if full_path.exists():
                self.template_status_label.config(text='✓ Found', foreground='#28a745')
            else:
                self.template_status_label.config(text='❌ Missing', foreground='#dc3545')
            
            # Trigger automatic reload
            self._reload_excel_data()
            
            messagebox.showinfo('✓ Path Updated', 
                               f'Application Inputs Excel path updated to:\n{path_str}\n\n'
                               'Data has been reloaded from the new location.')
    
    def _reset_legacy_path(self):
        """Reset legacy path to default"""
        default_path = 'data/New Water Balance  20250930 Oct.xlsx'
        config.set('data_sources.legacy_excel_path', default_path)
        self.legacy_path_label.config(text=default_path)
        
        # Update status
        base_dir = Path.cwd()
        full_path = base_dir / default_path
        if full_path.exists():
            self.legacy_status_label.config(text='✓ Found', foreground='#28a745')
        else:
            self.legacy_status_label.config(text='❌ Missing', foreground='#dc3545')
        
        self._reload_excel_data()
        messagebox.showinfo('✓ Reset', f'TRP Excel path reset to default:\n{default_path}\n\nData reloaded.')
    
    def _reset_template_path(self):
        """Reset template path to default"""
        default_path = 'templates/Water_Balance_TimeSeries_Template.xlsx'
        config.set('data_sources.template_excel_path', default_path)
        self.template_path_label.config(text=default_path)
        
        # Update status
        base_dir = Path.cwd()
        full_path = base_dir / default_path
        if full_path.exists():
            self.template_status_label.config(text='✓ Found', foreground='#28a745')
        else:
            self.template_status_label.config(text='❌ Missing', foreground='#dc3545')
        
        self._reload_excel_data()
        messagebox.showinfo('✓ Reset', f'Application Inputs Excel path reset to default:\n{default_path}\n\nData reloaded.')
    
    def _open_template_folder(self):
        """Open the folder containing the template file"""
        import subprocess
        import platform
        
        template_path = config.get('data_sources.template_excel_path', 'templates/Water_Balance_TimeSeries_Template.xlsx')
        folder_path = Path(template_path).parent
        
        try:
            if platform.system() == 'Windows':
                subprocess.run(['explorer', str(folder_path.resolve())])
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', str(folder_path.resolve())])
            else:  # Linux
                subprocess.run(['xdg-open', str(folder_path.resolve())])
        except Exception as ex:
            messagebox.showerror('Error', f'Failed to open folder:\n{str(ex)}')
    
    def _reload_excel_data(self):
        """Force reload Excel data after path change"""
        try:
            from utils.app_logger import logger
            logger.info("🔄 Excel path changed - forcing data reload...")
            
            # Clear singleton instances to force reload
            from utils.excel_timeseries_extended import ExcelTimeSeriesExtended
            
            # Reset singleton
            if ExcelTimeSeriesExtended._instance is not None:
                ExcelTimeSeriesExtended._instance._loaded = False
                ExcelTimeSeriesExtended._instance._initialized = False
                ExcelTimeSeriesExtended._instance = None
                logger.info("✓ Excel singleton reset")
            
            # Recreate with new path
            _ = ExcelTimeSeriesExtended()
            logger.info("✓ Excel data reloaded from new path")
            
            # Notify user
            from utils.ui_notify import notifier
            notifier.success("Excel data reloaded successfully", title="Data Refresh")
            
        except Exception as ex:
            from utils.app_logger import logger
            logger.error(f"Failed to reload Excel data: {ex}")
            messagebox.showerror('Reload Error', 
                               f'Failed to reload Excel data:\n{str(ex)}\n\n'
                               'You may need to restart the application.')

    def _create_backup(self):
        """Create database backup"""
        try:
            path = self.backup.create_backup()
            messagebox.showinfo('✓ Backup Created', 
                               f'Database backup created successfully:\n\n{path}')
            self._load_backups()
        except Exception as ex:
            messagebox.showerror('Backup Error', f'Failed to create backup:\n{str(ex)}')

    def _load_backups(self):
        """Load available backups"""
        self.backups_list.delete(0, 'end')
        backups = self.backup.list_backups()
        
        if not backups:
            self.backups_list.insert('end', '(No backups found)')
        else:
            for f in backups:
                self.backups_list.insert('end', f.name)

    def _restore_selected(self):
        """Restore from selected backup"""
        sel = self.backups_list.curselection()
        if not sel:
            messagebox.showwarning('No Selection', 'Please select a backup file to restore.')
            return
        
        name = self.backups_list.get(sel[0])
        if name == '(No backups found)':
            return
        
        backup_path = self.backup.backup_dir / name
        
        confirm = messagebox.askyesno('⚠️ Confirm Restore',
            f'Restore database from backup:\n\n{name}\n\n'
            'This will OVERWRITE the current database.\n'
            'Current data will be lost unless backed up.\n\n'
            'Continue with restore?')
        
        if not confirm:
            return
        
        try:
            self.backup.restore_backup(backup_path)
            messagebox.showinfo('✓ Restore Complete', 
                               f'Database successfully restored from:\n{name}\n\n'
                               'Please restart the application for changes to take effect.')
        except Exception as ex:
            messagebox.showerror('Restore Error', f'Failed to restore backup:\n{str(ex)}')


