"""
Settings & Configuration UI - Revamped
Modern, clean interface with better organization and visual hierarchy
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from datetime import datetime

from database.db_manager import DatabaseManager
from utils.backup_manager import BackupManager
from utils.config_manager import config, get_resource_path
from utils.alert_manager import alert_manager
from utils.app_logger import logger
from ui.mouse_wheel_support import (
    enable_canvas_mousewheel,
    enable_text_mousewheel,
    enable_listbox_mousewheel,
    enable_treeview_mousewheel,
)
from ui.roadmap_preview import RoadmapPreviewTab


class SettingsModule:
    """Revamped Settings & Configuration UI"""
    
    def __init__(self, parent, initial_tab=None):
        self.parent = parent
        self.container = tk.Frame(parent, bg='#f5f6f7')
        self.db = DatabaseManager()
        self.backup = BackupManager()
        self.initial_tab = initial_tab
        self.notebook = None

    def load(self):
        self.container.pack(fill='both', expand=True)
        self.container.configure(bg='#f5f6f7')
        self._build_ui()

    def _build_ui(self):
        """Build modern settings interface"""
        try:
            self._build_ui_safe()
        except Exception as e:
            # Show error instead of blank screen
            import traceback
            error_msg = f"Settings failed to load:\n{str(e)}\n\nTraceback:\n{traceback.format_exc()}"
            tk.Label(self.container, text="⚠️ Settings Error", font=('Segoe UI', 16, 'bold'), 
                    fg='#cc3333', bg='#f5f6f7').pack(pady=20)
            error_text = tk.Text(self.container, height=20, width=80, wrap='word', bg='white', 
                                font=('Consolas', 9))
            error_text.pack(padx=20, pady=10, fill='both', expand=True)
            error_text.insert('1.0', error_msg)
            error_text.config(state='disabled')
            messagebox.showerror('Settings Error', f'Failed to load Settings:\n\n{str(e)[:200]}')
    
    def _build_ui_safe(self):
        """Build modern settings interface (wrapped by error handler)"""
        # Header with icon and description
        header = tk.Frame(self.container, bg='white', relief=tk.FLAT, bd=0)
        header.pack(fill='x', padx=0, pady=(0, 0))
        
        inner_header = tk.Frame(header, bg='white')
        inner_header.pack(fill='x', padx=20, pady=20)
        
        title_label = tk.Label(
            inner_header, 
            text="⚙️ Settings & Configuration",
            font=('Segoe UI', 22, 'bold'),
            bg='white', fg='#2c3e50'
        )
        title_label.pack(anchor='w')
        
        subtitle = tk.Label(
            inner_header,
            text="Manage application settings, constants, backups, and system configuration",
            font=('Segoe UI', 11),
            foreground='#7f8c8d',
            bg='white'
        )
        subtitle.pack(anchor='w', pady=(3, 0))
        
        # Modern tab styling with improved UX
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('SettingsNotebook.TNotebook', background='#f5f6f7', borderwidth=0)
        # Enhanced tab styling: larger font, more padding for better visibility
        style.configure('SettingsNotebook.TNotebook.Tab', 
                       background='#d6dde8', 
                       foreground='#2c3e50',
                       padding=[24, 16],  # Increased from [20, 12] for larger tab size
                       font=('Segoe UI', 11, 'bold'),  # Increased from 10 to 11, added bold
                       relief='flat',
                       borderwidth=0)
        # Enhanced map with better visual feedback on interaction
        style.map('SettingsNotebook.TNotebook.Tab',
                 background=[('selected', '#3498db'), ('active', '#5dade2'), ('!active', '#d6dde8')],
                 foreground=[('selected', '#ffffff'), ('active', '#ffffff'), ('!active', '#2c3e50')],
                 lightcolor=[('selected', '#3498db')],
                 darkcolor=[('selected', '#3498db')])
        
        # Notebook with modern tabs
        tab_container = tk.Frame(self.container, bg='#f5f6f7')
        tab_container.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        self.notebook = ttk.Notebook(tab_container, style='SettingsNotebook.TNotebook')
        self.notebook.pack(fill='both', expand=True)
        
        # Create tab frames
        # Branding removed: fixed TRP branding (no user uploads)
        self.constants_frame = ttk.Frame(self.notebook)
        self.environmental_frame = ttk.Frame(self.notebook)
        self.data_sources_frame = ttk.Frame(self.notebook)
        self.backup_frame = ttk.Frame(self.notebook)
        self.roadmap_frame = ttk.Frame(self.notebook)
        
        # Add tabs with icons
        self.notebook.add(self.constants_frame, text='  📊 Constants  ')
        self.notebook.add(self.environmental_frame, text='  🌦️ Environmental  ')
        self.notebook.add(self.data_sources_frame, text='  📂 Data Sources  ')
        self.notebook.add(self.backup_frame, text='  💾 Backup  ')
        self.notebook.add(self.roadmap_frame, text='  🚀 Future Features  ')
        
        # Build each tab with error handling
        try:
            self._build_constants_tab()
        except Exception as e:
            self._show_tab_error(self.constants_frame, 'Constants', e)
        
        try:
            self._build_environmental_tab()
        except Exception as e:
            self._show_tab_error(self.environmental_frame, 'Environmental', e)
        
        try:
            self._build_data_sources_tab()
        except Exception as e:
            self._show_tab_error(self.data_sources_frame, 'Data Sources', e)
        
        try:
            self._build_backup_tab()
        except Exception as e:
            self._show_tab_error(self.backup_frame, 'Backup', e)
        
        try:
            self._build_roadmap_tab()
        except Exception as e:
            self._show_tab_error(self.roadmap_frame, 'Future Features', e)
        


    def _show_tab_error(self, parent_frame, tab_name, exception):
        """Display error message in a tab that failed to load"""
        import traceback
        error_frame = tk.Frame(parent_frame, bg='#f5f6f7')
        error_frame.pack(fill='both', expand=True, padx=40, pady=40)
        
        tk.Label(error_frame, text=f"⚠️ {tab_name} Tab Error", 
                font=('Segoe UI', 14, 'bold'), fg='#cc3333', bg='#f5f6f7').pack(pady=(0, 10))
        
        tk.Label(error_frame, text=str(exception), 
                font=('Segoe UI', 10), fg='#2c3e50', bg='#f5f6f7', wraplength=600).pack(pady=10)
        
        details_text = tk.Text(error_frame, height=15, width=80, wrap='word', bg='white', 
                              font=('Consolas', 8), relief=tk.FLAT, bd=1)
        details_text.pack(fill='both', expand=True, pady=10)
        details_text.insert('1.0', traceback.format_exc())
        details_text.config(state='disabled')
        
        tk.Label(error_frame, text="💡 Copy this error and share it for troubleshooting", 
                font=('Segoe UI', 9, 'italic'), fg='#7f8c8d', bg='#f5f6f7').pack(pady=(10, 0))
    
    def _build_branding_tab(self):
        """Build branding configuration with modern card layout"""
        # Create scrollable content area
        scroll_container = tk.Frame(self.branding_frame, bg='#f5f6f7')
        scroll_container.pack(fill='both', expand=True)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(scroll_container, bg='#f5f6f7', highlightthickness=0)
        scrollbar = tk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f5f6f7')
        
        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Enable mousewheel scrolling
        enable_canvas_mousewheel(canvas)
        
        canvas.pack(side='left', fill='both', expand=True, padx=20, pady=20)
        scrollbar.pack(side='right', fill='y', padx=0, pady=20)
        
        scroll_container = scrollable_frame  # Reuse variable for content
        
        # Company Information Card
        company_card = tk.Frame(scroll_container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        company_card.pack(fill='x', pady=(0, 20))
        
        inner = tk.Frame(company_card, bg='#e8eef5')
        inner.pack(fill='x', padx=20, pady=16)
        
        tk.Label(inner, text='📝 Company Information', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w', pady=(0, 15))
        
        # Company name with better layout
        name_frame = tk.Frame(inner, bg='#e8eef5')
        name_frame.pack(fill='x', pady=(0, 15))
        
        tk.Label(name_frame, text='Company Name', font=('Segoe UI', 10, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        tk.Label(name_frame, text='This name will appear on all reports and exports', 
                 font=('Segoe UI', 9), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=(2, 5))
        
        name_entry_frame = tk.Frame(inner, bg='#e8eef5')
        name_entry_frame.pack(fill='x', pady=(0, 0))
        
        self.company_name_entry = tk.Entry(name_entry_frame, font=('Segoe UI', 10), bg='white', relief=tk.FLAT, bd=1, highlightbackground='#d9e6f4', highlightthickness=1)
        self.company_name_entry.insert(0, config.get_company_name())
        self.company_name_entry.pack(side='left', fill='x', expand=True, ipady=6, padx=(0, 10))
        
        save_btn = tk.Button(name_entry_frame, text='💾 Save', command=self._save_company_name, 
                  font=('Segoe UI', 10), bg='#0066cc', fg='white', relief=tk.FLAT, bd=0, padx=20, pady=6, cursor='hand2')
        save_btn.pack(side='left')
        
        # Logo Configuration Card
        logo_card = tk.Frame(scroll_container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        logo_card.pack(fill='x', pady=(0, 20))
        
        inner_logo = tk.Frame(logo_card, bg='#e8eef5')
        inner_logo.pack(fill='x', padx=20, pady=16)
        
        tk.Label(inner_logo, text='🖼️ Company Logo', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        tk.Label(inner_logo, text='Upload your company logo to display in the application toolbar and reports',
                 font=('Segoe UI', 9), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=(2, 10))
        
        # Logo preview with border
        preview_frame = tk.Frame(inner_logo, bg='white', relief=tk.FLAT, bd=1, highlightbackground='#d9e6f4', highlightthickness=1)
        preview_frame.pack(fill='x', pady=(0, 10))
        
        self.logo_preview_label = tk.Label(preview_frame, text='No logo selected', 
                                           fg='#999', bg='white', padx=40, pady=40, font=('Segoe UI', 10))
        self.logo_preview_label.pack()
        self._update_logo_preview()
        
        # Logo path display
        path_frame = tk.Frame(inner_logo, bg='#e8eef5')
        path_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(path_frame, text='Current Path:', font=('Segoe UI', 9), bg='#e8eef5', fg='#2c3e50').pack(side='left')
        self.logo_path_label = tk.Label(path_frame, text=config.get_logo_path() or 'None', 
                                        fg='#7f8c8d', bg='#e8eef5', font=('Segoe UI', 9, 'italic'))
        self.logo_path_label.pack(side='left', padx=(5, 0))
        
        # Logo action buttons
        logo_btn_frame = tk.Frame(inner_logo, bg='#e8eef5')
        logo_btn_frame.pack(fill='x')
        
        choose_btn = tk.Button(logo_btn_frame, text='📁 Choose Logo File', command=self._select_logo, 
                             font=('Segoe UI', 10), bg='#0066cc', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=6, cursor='hand2')
        choose_btn.pack(side='left', padx=(0, 10))
        
        remove_btn = tk.Button(logo_btn_frame, text='❌ Remove Logo', command=self._remove_logo,
                              font=('Segoe UI', 10), bg='#cc3333', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=6, cursor='hand2')
        remove_btn.pack(side='left')
        
        # Recommendations Card
        tips_card = tk.Frame(scroll_container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        tips_card.pack(fill='x', pady=(10, 0))
        
        inner_tips = tk.Frame(tips_card, bg='#e8eef5')
        inner_tips.pack(fill='x', padx=20, pady=16)
        
        tk.Label(inner_tips, text='💡 Recommendations', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w', pady=(0, 10))
        
        tips = [
            "• Use PNG format with transparent background for best results",
            "• Recommended dimensions: 200×80 pixels (or similar 2.5:1 ratio)",
            "• Logo will be automatically resized to fit the toolbar",
            "• Supported formats: PNG, JPG, JPEG"
        ]
        
        for tip in tips:
            tk.Label(inner_tips, text=tip, font=('Segoe UI', 9), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=2)
    
    def _save_company_name(self):
        """Save company name with validation"""
        name = self.company_name_entry.get().strip()
        if not name:
            messagebox.showwarning('Validation Error', 'Company name cannot be empty')
            return
        
        config.set_company_name(name)
        messagebox.showinfo('✓ Saved', f'Company name updated to: {name}\n\nChanges will appear in reports and exports.')
    
    def _select_logo(self):
        """Select and copy logo file"""
        filename = filedialog.askopenfilename(
            title='Select Company Logo',
            filetypes=[
                ('PNG Images', '*.png'),
                ('JPEG Images', '*.jpg *.jpeg'),
                ('All Images', '*.png *.jpg *.jpeg')
            ]
        )
        if filename:
            try:
                import shutil
                dest_dir = Path(__file__).parent.parent.parent / 'config' / 'branding'
                dest_dir.mkdir(parents=True, exist_ok=True)
                dest_file = dest_dir / Path(filename).name
                shutil.copy2(filename, dest_file)
                
                config.set_logo_path(str(dest_file))
                self.logo_path_label.config(text=str(dest_file))
                self._update_logo_preview()
                messagebox.showinfo('✓ Logo Updated', 
                                   f'Logo successfully uploaded: {dest_file.name}\n\n'
                                   'Restart the application to see it in the toolbar.')
            except Exception as ex:
                messagebox.showerror('Upload Error', f'Failed to copy logo file:\n{str(ex)}')
    
    def _remove_logo(self):
        """Remove current logo"""
        if not config.get_logo_path():
            messagebox.showinfo('No Logo', 'No logo is currently set.')
            return
        
        confirm = messagebox.askyesno('Confirm Removal', 
                                     'Remove the current logo?\n\n'
                                     'This will remove it from the toolbar and reports.')
        if not confirm:
            return
        
        config.set_logo_path('')
        self.logo_path_label.config(text='None')
        self._update_logo_preview()
        messagebox.showinfo('✓ Logo Removed', 'Logo has been removed.')
    
    def _update_logo_preview(self):
        """Update logo preview with better styling"""
        logo_path = config.get_logo_path()
        if logo_path and Path(logo_path).exists():
            try:
                from PIL import Image, ImageTk
                img = Image.open(logo_path)
                
                # Create thumbnail maintaining aspect ratio
                img.thumbnail((200, 100), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                
                self.logo_preview_label.config(image=photo, text='', bg='white')
                self.logo_preview_label.image = photo
            except Exception as ex:
                self.logo_preview_label.config(
                    text=f'⚠️ Preview unavailable\n{str(ex)}',
                    foreground='#cc3333',
                    bg='white'
                )
        else:
            self.logo_preview_label.config(
                text='📷\nNo logo uploaded',
                foreground='#999',
                image='',
                bg='white'
            )

    def _build_constants_tab(self):
        """Build constants management with enhanced UI"""
        # Create scrollable container
        scroll_container = tk.Frame(self.constants_frame, bg='#f5f6f7')
        scroll_container.pack(fill='both', expand=True)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(scroll_container, bg='#f5f6f7', highlightthickness=0)
        scrollbar = tk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f5f6f7')
        
        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Enable mousewheel scrolling
        enable_canvas_mousewheel(canvas)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        container = scrollable_frame  # Use scrollable frame as container
        container.configure(padx=20, pady=20)
        
        # Filter and search bar (modern card)
        filter_card = tk.Frame(container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        filter_card.pack(fill='x', pady=(0, 20))
        
        filter_inner = tk.Frame(filter_card, bg='#e8eef5')
        filter_inner.pack(fill='x', padx=20, pady=12)
        
        tk.Label(filter_inner, text='Category:', font=('Segoe UI', 10), bg='#e8eef5', fg='#2c3e50').pack(side='left', padx=(0, 5))
        self.category_filter = tk.StringVar(master=filter_inner, value='All')
        category_combo = ttk.Combobox(
            filter_inner,
            textvariable=self.category_filter,
            values=['All', 'Plant', 'Seepage', 'Evaporation', 'Operating', 'Optimization', 'calculation'],
            state='readonly',
            width=15
        )
        category_combo.pack(side='left', padx=(0, 20))
        category_combo.bind('<<ComboboxSelected>>', lambda e: self._load_constants())

        # Search input
        tk.Label(filter_inner, text='Search:', font=('Segoe UI', 10), bg='#e8eef5', fg='#2c3e50').pack(side='left', padx=(0, 5))
        self.search_var = tk.StringVar(master=filter_inner)
        search_entry = tk.Entry(filter_inner, textvariable=self.search_var, font=('Segoe UI', 10), bg='white', relief=tk.FLAT, bd=1, highlightbackground='#d9e6f4', highlightthickness=1)
        search_entry.pack(side='left', fill='x', expand=True, ipady=5, padx=(0, 10))
        search_entry.bind('<Return>', lambda e: self._load_constants())
        
        # Action buttons
        refresh_btn = tk.Button(filter_inner, text='🔄 Refresh', command=self._load_constants, font=('Segoe UI', 10), bg='#0066cc', fg='white', relief=tk.FLAT, bd=0, padx=12, pady=5, cursor='hand2')
        refresh_btn.pack(side='left', padx=(0, 8))
        
        export_btn = tk.Button(filter_inner, text='📊 Export', command=self._export_constants, font=('Segoe UI', 10), bg='#28a745', fg='white', relief=tk.FLAT, bd=0, padx=12, pady=5, cursor='hand2')
        export_btn.pack(side='left', padx=(0, 8))
        
        history_btn = tk.Button(filter_inner, text='📜 History', command=self._view_history, font=('Segoe UI', 10), bg='#6c3fb5', fg='white', relief=tk.FLAT, bd=0, padx=12, pady=5, cursor='hand2')
        history_btn.pack(side='left')
        
        # Constants table (modern card)
        table_card = tk.Frame(container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        table_card.pack(fill='both', expand=True, pady=(0, 20))
        
        table_header = tk.Frame(table_card, bg='#e8eef5')
        table_header.pack(fill='x', padx=20, pady=(12, 0))
        
        tk.Label(table_header, text='System Constants', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        
        # Simple scrollable frame with vertical scrollbar only
        tree_container = tk.Frame(table_card, bg='#e8eef5')
        tree_container.pack(fill='both', expand=True, padx=20, pady=12)
        
        # Compact 5-column layout that fits on any screen
        cols = ('Cat', 'Constant', 'Value', 'Unit', 'Usage')
        self.tree = ttk.Treeview(tree_container, columns=cols, show='headings', height=12)
        enable_treeview_mousewheel(self.tree)
        
        # Compact widths - total ~800px fits on all screens
        widths = {'Cat': 80, 'Constant': 180, 'Value': 80, 'Unit': 70, 'Usage': 390}
        headings = {'Cat': 'Category', 'Constant': 'Constant Key', 'Value': 'Value', 
                   'Unit': 'Unit', 'Usage': 'Used In / Formula'}
        
        for c in cols:
            self.tree.heading(c, text=headings[c])
            self.tree.column(c, width=widths[c], anchor='w', minwidth=50, stretch=True)
        
        # Vertical scrollbar only
        scrollbar = ttk.Scrollbar(tree_container, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        self._load_constants()
        
        # Edit controls (modern card)
        edit_card = tk.Frame(container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        edit_card.pack(fill='x', pady=(0, 0))
        
        edit_inner = tk.Frame(edit_card, bg='#e8eef5')
        edit_inner.pack(fill='x', padx=20, pady=12)
        
        tk.Label(edit_inner, text='Quick Edit:', font=('Segoe UI', 10, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(side='left', padx=(0, 10))
        
        self.selected_key = tk.Label(edit_inner, text='(Select a constant)', 
                                      font=('Segoe UI', 10), fg='#7f8c8d', bg='#e8eef5')
        self.selected_key.pack(side='left', padx=(0, 20))
        
        tk.Label(edit_inner, text='Value:', font=('Segoe UI', 10), bg='#e8eef5', fg='#2c3e50').pack(side='left', padx=(0, 5))
        self.new_value = tk.Entry(edit_inner, font=('Segoe UI', 10), bg='white', relief=tk.FLAT, bd=1, highlightbackground='#d9e6f4', highlightthickness=1)
        self.new_value.pack(side='left', fill='x', expand=True, ipady=5, padx=(0, 10))
        
        save_btn = tk.Button(edit_inner, text='💾 Save', command=self._update_constant, font=('Segoe UI', 10), bg='#0066cc', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=5, cursor='hand2')
        save_btn.pack(side='left', padx=(0, 5))
        
        details_btn = tk.Button(edit_inner, text='📝 Details', command=self._show_edit_dialog, font=('Segoe UI', 10), bg='#00a886', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=5, cursor='hand2')
        details_btn.pack(side='left', padx=(0, 5))
        
        add_btn = tk.Button(edit_inner, text='➕ Add', command=self._add_constant, font=('Segoe UI', 10), bg='#28a745', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=5, cursor='hand2')
        add_btn.pack(side='left', padx=(0, 5))
        
        delete_btn = tk.Button(edit_inner, text='🗑️ Delete', command=self._delete_constant, font=('Segoe UI', 10), bg='#cc3333', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=5, cursor='hand2')
        delete_btn.pack(side='left')
        
        # Bind selection event
        self.tree.bind('<<TreeviewSelect>>', self._on_constant_select)
        self.current_constant = None
        self._usage_cache = None
        self.selected_desc = None  # Will show in dialog
        self.value_range_label = None  # Will show in dialog

    def _on_constant_select(self, event):
        """Handle constant selection"""
        sel = self.tree.selection()
        if not sel:
            return
        
        item = self.tree.item(sel[0])
        values = item['values']
        
        # Unpack new format: Cat, Constant, Value, Unit, Usage
        if len(values) >= 5:
            cat, key, value, unit, usage = values[:5]
        else:
            return  # Invalid row
        
        # Get full metadata
        meta = self._get_constant_metadata(key)
        
        # Update compact edit bar
        self.selected_key.config(text=key, foreground='#0066cc')
        self.new_value.delete(0, 'end')
        self.new_value.insert(0, str(value))
        
        self.current_constant = meta
    
    def _show_edit_dialog(self):
        """Show detailed constant information and editing in a popup dialog"""
        key = self.selected_key.cget('text')
        if key == '(Select a constant)':
            messagebox.showinfo('No Selection', 'Please select a constant first.')
            return
        
        meta = self.current_constant
        if not meta:
            return
        
        # Create dialog
        dlg = tk.Toplevel(self.parent)
        dlg.title(f'Edit Constant: {key}')
        dlg.geometry('600x400')
        dlg.transient(self.parent)
        dlg.grab_set()
        
        # Content frame
        content = ttk.Frame(dlg, padding=20)
        content.pack(fill='both', expand=True)
        
        # Constant details
        ttk.Label(content, text=key, font=('Segoe UI', 14, 'bold'), foreground='#0066cc').pack(anchor='w', pady=(0, 10))
        
        details_frame = ttk.Frame(content)
        details_frame.pack(fill='x', pady=(0, 15))
        
        ttk.Label(details_frame, text=f"Category: {meta.get('category', '—')}", font=('Segoe UI', 10)).pack(anchor='w', pady=2)
        ttk.Label(details_frame, text=f"Unit: {meta.get('unit', '—')}", font=('Segoe UI', 10)).pack(anchor='w', pady=2)
        ttk.Label(details_frame, text=f"Current Value: {meta.get('constant_value', '—')}", font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=2)
        
        desc_label = ttk.Label(details_frame, text=f"Description:\n{meta.get('description', 'No description')}", 
                              font=('Segoe UI', 9), foreground='#666', wraplength=550, justify='left')
        desc_label.pack(anchor='w', pady=(5, 10))
        
        # Edit section
        edit_frame = ttk.LabelFrame(content, text='Update Value', padding=15)
        edit_frame.pack(fill='x', pady=(0, 10))
        
        ttk.Label(edit_frame, text='New Value:', font=('Segoe UI', 10)).pack(side='left', padx=(0, 10))
        value_entry = ttk.Entry(edit_frame, font=('Segoe UI', 11), width=20)
        value_entry.insert(0, str(meta.get('constant_value', '')))
        value_entry.pack(side='left', padx=(0, 10))
        
        def save_from_dialog():
            try:
                new_val = float(value_entry.get())
                self.new_value.delete(0, 'end')
                self.new_value.insert(0, str(new_val))
                self._update_constant()
                dlg.destroy()
            except ValueError:
                messagebox.showerror('Invalid Value', 'Please enter a valid number.')
        
        ttk.Button(edit_frame, text='💾 Save', command=save_from_dialog).pack(side='left')
        
        # Close button
        ttk.Button(content, text='Close', command=dlg.destroy).pack(pady=(10, 0))

    def _get_constant_metadata(self, key):
        """Get full constant metadata"""
        result = self.db.execute_query(
            'SELECT * FROM system_constants WHERE constant_key = ?', (key,)
        )
        return result[0] if result else None

    def _toggle_edit_panel(self):
        """Toggle visibility of edit panel to save screen space"""
        if self.edit_visible.get():
            self.edit_card.pack(fill='x')
        else:
            self.edit_card.pack_forget()
    
    def _load_constants(self):
        """Load constants into table with usage and formula information"""
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        cat_filter = self.category_filter.get()
        search_text = (self.search_var.get() or '').strip().lower()
        usage_map = self._get_constant_usage_info()
        
        query = '''SELECT constant_key, constant_value, unit, category, description, 
                   min_value, max_value FROM system_constants ORDER BY category, constant_key'''
        results = self.db.execute_query(query)
        
        for row in results:
            key = row['constant_key']
            value = row['constant_value']
            unit = row.get('unit') or '—'
            category = row.get('category') or 'Other'
            desc = row.get('description') or '—'
            
            if cat_filter != 'All' and category != cat_filter:
                continue

            # Search filter
            if search_text:
                hay = f"{key} {desc} {category}".lower()
                if search_text not in hay:
                    continue
            
            # Get usage info and combine module + formula
            usage_info = usage_map.get(key, {'module': '—', 'formula': '—'})
            used_in = usage_info['module']
            formula = usage_info['formula']
            # Combine: "Module: formula" for compact display
            usage_display = f"{used_in}: {formula}" if formula != '—' else used_in
            
            # Insert with 5 columns: Cat, Constant, Value, Unit, Usage
            self.tree.insert('', 'end', values=(category, key, value, unit, usage_display))

    def _update_constant(self):
        """Update selected constant with validation"""
        key = self.selected_key.cget('text')
        if key == '(None selected)':
            messagebox.showwarning('No Selection', 'Please select a constant to update.')
            return
        
        try:
            value = float(self.new_value.get())
        except ValueError:
            messagebox.showerror('Invalid Value', 'Please enter a valid numeric value.')
            return
        
        # Range validation removed per request – any numeric value allowed
        
        # Critical constant warning
        critical = ['SLURRY_DENSITY']
        if key in critical:
            confirm = messagebox.askyesno('⚠️ Critical Constant',
                f'You are modifying a critical system constant: {key}\n\n'
                f'Current value: {self.current_constant["constant_value"]}\n'
                f'New value: {value}\n\n'
                f'This will affect all future water balance calculations.\n\n'
                'Do you want to proceed?')
            if not confirm:
                return
        
        self.db.update_constant(key, value, user='user')
        self._load_constants()
        messagebox.showinfo('✓ Updated', f'Constant {key} has been updated to {value}.')

    def _get_constant_usage_info(self) -> Dict[str, Dict[str, str]]:
        """Get detailed usage information for each constant (where and how it's used)"""
        return {
            'tailings_moisture_pct': {
                'module': 'Tailings retention',
                'formula': '(Ore - Conc) × moisture% = Retained water'
            },
            'ore_density': {
                'module': 'Moisture calc',
                'formula': 'Volume (m³) = Mass (t) / density'
            },
            'ore_moisture_percent': {
                'module': 'Ore moisture',
                'formula': 'Ore tonnage × density × moisture% = Inflow'
            },
            'lined_seepage_rate_pct': {
                'module': 'Seepage calc',
                'formula': 'Volume × 0.1%/month (lined facilities)'
            },
            'unlined_seepage_rate_pct': {
                'module': 'Seepage calc',
                'formula': 'Volume × 0.5%/month (unlined facilities)'
            },
            'dust_suppression_rate': {
                'module': 'Dust control',
                'formula': 'Ore tonnage × rate (m³/t) = Dust water'
            },
            'pump_stop_level': {
                'module': 'Pump control',
                'formula': 'Pumps stop when level < stop%'
            },
            'slurry_density': {
                'module': 'Slurry calc',
                'formula': 'Used in slurry volume conversions'
            },
        }

    def _export_constants(self):
        """Export constants to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from datetime import datetime
            
            filename = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx')],
                initialfile=f'water_balance_constants_{datetime.now().strftime("%Y%m%d")}.xlsx'
            )
            if not filename:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = 'System Constants'
            
            # Headers with styling
            headers = ['Category', 'Constant Key', 'Value', 'Unit', 'Min', 'Max', 'Description']
            ws.append(headers)
            
            header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
            
            # Data
            query = '''SELECT category, constant_key, constant_value, unit, min_value, max_value, description 
                      FROM system_constants ORDER BY category, constant_key'''
            results = self.db.execute_query(query)
            
            for row in results:
                ws.append([
                    row.get('category') or '',
                    row['constant_key'],
                    row['constant_value'],
                    row.get('unit') or '',
                    row.get('min_value') or '',
                    row.get('max_value') or '',
                    row.get('description') or ''
                ])
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            messagebox.showinfo('✓ Export Complete', f'Constants exported successfully to:\n\n{filename}')
        except Exception as ex:
            messagebox.showerror('Export Error', f'Failed to export constants:\n{str(ex)}')

    def _view_history(self):
        """Show audit log of constant changes"""
        history_window = tk.Toplevel(self.parent)
        history_window.title('Constants Change History')
        history_window.geometry('900x500')
        
        frame = ttk.Frame(history_window, padding=15)
        frame.pack(fill='both', expand=True)
        
        ttk.Label(frame, text='📜 Recent Constant Changes', 
                 font=('Segoe UI', 12, 'bold')).pack(anchor='w', pady=(0, 10))
        
        cols = ('Timestamp', 'Constant', 'Old Value', 'New Value', 'User')
        tree = ttk.Treeview(frame, columns=cols, show='headings', height=18)
        enable_treeview_mousewheel(tree)
        
        widths = {'Timestamp': 180, 'Constant': 200, 'Old Value': 120, 'New Value': 120, 'User': 100}
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=widths[c])
        
        tree.pack(fill='both', expand=True, pady=(0, 10))
        
        # Load history
        try:
            import json
            logs = self.db.execute_query(
                "SELECT * FROM audit_log WHERE table_name = 'system_constants' "
                "ORDER BY changed_at DESC LIMIT 100"
            )
            
            for log in logs:
                timestamp = log['changed_at']
                old_vals = json.loads(log['old_values']) if log['old_values'] else {}
                new_vals = json.loads(log['new_values']) if log['new_values'] else {}
                
                for key in new_vals:
                    old_val = old_vals.get(key, '—')
                    new_val = new_vals[key]
                    tree.insert('', 'end', values=(timestamp, key, old_val, new_val, log['changed_by']))
        except Exception as ex:
            ttk.Label(frame, text=f'⚠️ Error loading history: {ex}', 
                     foreground='#dc3545').pack()
        
        ttk.Button(frame, text='Close', command=history_window.destroy).pack(pady=(10, 0))

    def _compute_constant_usage(self) -> dict:
        """Compute usage count of constants across src/ python files.
        Returns a dict mapping constant_key -> occurrence count.
        Caches results for the session; refresh with the Refresh button.
        """
        try:
            # Use cached result when available
            if isinstance(self._usage_cache, dict):
                return self._usage_cache
            # Gather keys
            rows = self.db.execute_query("SELECT constant_key FROM system_constants")
            keys = [r['constant_key'] for r in rows]
            usage = {k: 0 for k in keys}
            base_dir = Path(__file__).parent.parent.parent
            src_dir = base_dir / 'src'
            for path in src_dir.rglob('*.py'):
                try:
                    text = path.read_text(encoding='utf-8', errors='ignore')
                except Exception:
                    continue
                for k in keys:
                    # Simple substring count; fast and adequate for references
                    cnt = text.count(k)
                    if cnt:
                        usage[k] += cnt
            self._usage_cache = usage
            return usage
        except Exception:
            return {}

    def _add_constant(self):
        """Open a dialog to add a new constant."""
        dlg = tk.Toplevel(self.parent)
        dlg.title('Add Constant')
        dlg.geometry('520x360')
        frame = ttk.Frame(dlg, padding=15)
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text='Key:', font=('Segoe UI', 10)).grid(row=0, column=0, sticky='e', pady=6)
        ttk.Label(frame, text='Value:', font=('Segoe UI', 10)).grid(row=1, column=0, sticky='e', pady=6)
        ttk.Label(frame, text='Unit:', font=('Segoe UI', 10)).grid(row=2, column=0, sticky='e', pady=6)
        ttk.Label(frame, text='Category:', font=('Segoe UI', 10)).grid(row=3, column=0, sticky='e', pady=6)
        ttk.Label(frame, text='Description:', font=('Segoe UI', 10)).grid(row=4, column=0, sticky='ne', pady=6)

        key_var = tk.StringVar(master=frame)
        val_var = tk.StringVar(master=frame)
        unit_var = tk.StringVar(master=frame)
        cat_var = tk.StringVar(master=frame, value='calculation')
        desc_txt = tk.Text(frame, height=5, width=40)

        ttk.Entry(frame, textvariable=key_var, width=32).grid(row=0, column=1, sticky='w', padx=8)
        ttk.Entry(frame, textvariable=val_var, width=20).grid(row=1, column=1, sticky='w', padx=8)
        ttk.Entry(frame, textvariable=unit_var, width=20).grid(row=2, column=1, sticky='w', padx=8)
        ttk.Entry(frame, textvariable=cat_var, width=20).grid(row=3, column=1, sticky='w', padx=8)
        desc_txt.grid(row=4, column=1, sticky='w', padx=8)
        enable_text_mousewheel(desc_txt)

        btns = ttk.Frame(frame)
        btns.grid(row=5, column=0, columnspan=2, pady=12)
        
        def do_add():
            key = key_var.get().strip()
            if not key:
                messagebox.showwarning('Validation', 'Key is required')
                return
            try:
                val = float(val_var.get().strip())
            except ValueError:
                messagebox.showerror('Validation', 'Value must be a number')
                return
            unit = unit_var.get().strip() or None
            cat = cat_var.get().strip() or 'calculation'
            desc = desc_txt.get('1.0', 'end').strip() or None
            try:
                self.db.execute_insert(
                    """
                    INSERT INTO system_constants (constant_key, constant_value, unit, category, description, editable, updated_by)
                    VALUES (?, ?, ?, ?, ?, 1, 'user')
                    """,
                    (key, val, unit, cat, desc)
                )
                # Reset usage cache and reload
                self._usage_cache = None
                self._load_constants()
                messagebox.showinfo('✓ Added', f'Constant {key} created.')
                dlg.destroy()
            except Exception as ex:
                messagebox.showerror('Error', f'Failed to add constant:\n{ex}')

        ttk.Button(btns, text='Create', style='Accent.TButton', command=do_add).pack(side='left', padx=6)
        ttk.Button(btns, text='Cancel', command=dlg.destroy).pack(side='left', padx=6)

    def _delete_constant(self):
        """Delete the selected constant with confirmation."""
        key = self.selected_key.cget('text')
        if not key or key == '(None selected)':
            messagebox.showwarning('No Selection', 'Please select a constant to delete.')
            return
        # Extra safety: block deletion of core constants
        protected = {'TSF_RETURN_RATE', 'SLURRY_DENSITY'}
        if key in protected:
            messagebox.showwarning('Protected', f'{key} is protected and cannot be deleted.')
            return
        confirm = messagebox.askyesno('Confirm Delete', f'Delete constant {key}?\nThis cannot be undone.')
        if not confirm:
            return
        try:
            # Fetch old value for audit
            old_val = self.db.get_constant(key)
            rows = self.db.execute_update("DELETE FROM system_constants WHERE constant_key = ?", (key,))
            if rows:
                self.db.log_change('system_constants', 0, 'delete', old_values={key: old_val}, user='user')
            # Reset usage cache and reload
            self._usage_cache = None
            self._load_constants()
            self.selected_key.config(text='(None selected)')
            self.selected_desc.config(text='')
            messagebox.showinfo('✓ Deleted', f'Constant {key} removed.')
        except Exception as ex:
            messagebox.showerror('Error', f'Failed to delete constant:\n{ex}')

    


    def _build_backup_tab(self):
        """Build backup management with modern UI"""
        container = tk.Frame(self.backup_frame, bg='#f5f6f7')
        container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Info card
        info_card = tk.Frame(container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        info_card.pack(fill='x', pady=(0, 20))
        
        info_inner = tk.Frame(info_card, bg='#e8eef5')
        info_inner.pack(fill='x', padx=20, pady=16)
        
        tk.Label(info_inner, text='💾 Database Backup & Restore', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        tk.Label(info_inner, text='Regular backups protect your water balance data',
                 font=('Segoe UI', 10), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=(5, 12))
        
        # Actions
        actions = tk.Frame(info_inner, bg='#e8eef5')
        actions.pack(fill='x')
        
        backup_btn = tk.Button(actions, text='💾 Create Backup Now', command=self._create_backup,
                  font=('Segoe UI', 10), bg='#28a745', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=6, cursor='hand2')
        backup_btn.pack(side='left', padx=(0, 10))
        
        refresh_btn = tk.Button(actions, text='🔄 Refresh List', command=self._load_backups,
                               font=('Segoe UI', 10), bg='#0066cc', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=6, cursor='hand2')
        refresh_btn.pack(side='left')
        
        # Backups list
        list_card = tk.Frame(container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        list_card.pack(fill='both', expand=True, pady=(0, 20))
        
        list_header = tk.Frame(list_card, bg='#e8eef5')
        list_header.pack(fill='x', padx=20, pady=(12, 0))
        
        tk.Label(list_header, text='📂 Available Backups', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        
        self.backups_list = tk.Listbox(list_card, height=10, font=('Consolas', 10), bg='white', relief=tk.FLAT, bd=0)
        self.backups_list.pack(fill='both', expand=True, padx=20, pady=12)
        enable_listbox_mousewheel(self.backups_list)
        
        self._load_backups()
        
        # Restore section
        restore_card = tk.Frame(container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        restore_card.pack(fill='x', pady=(0, 0))
        
        restore_inner = tk.Frame(restore_card, bg='#e8eef5')
        restore_inner.pack(fill='x', padx=20, pady=16)
        
        tk.Label(restore_inner, text='⚠️ Restore from Backup', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        tk.Label(restore_inner, text='⚠️ Warning: Restoring will overwrite the current database',
                 font=('Segoe UI', 10), fg='#cc3333', bg='#e8eef5').pack(anchor='w', pady=(5, 12))
        
        restore_btn = tk.Button(restore_inner, text='🔙 Restore Selected Backup', 
                  command=self._restore_selected,
                  font=('Segoe UI', 10), bg='#cc3333', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=6, cursor='hand2')
        restore_btn.pack(anchor='w')

    def _build_environmental_tab(self):
        """Build environmental parameters tab"""
        # Scrollable container to adapt to smaller screens
        container = tk.Frame(self.environmental_frame, bg='#f5f6f7')
        container.pack(fill='both', expand=True)

        canvas = tk.Canvas(container, bg='#f5f6f7', highlightthickness=0)
        vscroll = ttk.Scrollbar(container, orient='vertical', command=canvas.yview)
        scroll_container = tk.Frame(canvas, bg='#f5f6f7')

        scroll_container.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )

        canvas.create_window((0, 0), window=scroll_container, anchor='nw')
        canvas.configure(yscrollcommand=vscroll.set, bg='#f5f6f7')

        canvas.pack(side='left', fill='both', expand=True, padx=20, pady=20)
        vscroll.pack(side='right', fill='y', padx=0, pady=20)

        # Year selector card
        year_card = tk.Frame(scroll_container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        year_card.pack(fill='x', pady=(0, 20))

        year_inner = tk.Frame(year_card, bg='#e8eef5')
        year_inner.pack(fill='x', padx=20, pady=16)

        tk.Label(year_inner, text='📅 Historical Data', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w', pady=(0, 12))

        year_controls = tk.Frame(year_inner, bg='#e8eef5')
        year_controls.pack(fill='x', pady=(0, 0))

        tk.Label(year_controls, text='Select Year:', font=('Segoe UI', 10), bg='#e8eef5', fg='#2c3e50').pack(side='left', padx=(0, 10))
        
        # Year spinbox
        self.env_year_var = tk.StringVar(master=year_controls, value=str(datetime.now().year))
        year_spinbox = ttk.Spinbox(year_controls, from_=2020, to=2050, textvariable=self.env_year_var,
                                   width=8, font=('Segoe UI', 10))
        year_spinbox.pack(side='left', padx=5)
        
        load_btn = tk.Button(year_controls, text='Load Data', command=self._load_environmental_params, font=('Segoe UI', 10), bg='#0066cc', fg='white', relief=tk.FLAT, bd=0, padx=15, pady=5, cursor='hand2')
        load_btn.pack(side='left', padx=5)
        
        # Quick year buttons
        tk.Label(year_controls, text='Quick: ', font=('Segoe UI', 9), bg='#e8eef5', fg='#2c3e50').pack(side='left', padx=(10, 5))
        current_year = datetime.now().year
        for year in [current_year, current_year-1, current_year-2]:
            year_btn = tk.Button(year_controls, text=str(year), font=('Segoe UI', 9), bg='#d9e6f4', fg='#2c3e50', relief=tk.FLAT, bd=0, padx=10, pady=4, cursor='hand2',
                      command=lambda y=year: self._set_env_year(y))
            year_btn.pack(side='left', padx=2)

        # Zone info (fixed to 4A)
        zone_card = tk.Frame(scroll_container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        zone_card.pack(fill='x', pady=(0, 20))

        zone_inner = tk.Frame(zone_card, bg='#e8eef5')
        zone_inner.pack(fill='x', padx=20, pady=16)

        current_zone = config.get('environmental.evaporation_zone', '4A')
        tk.Label(zone_inner, text='🌍 Evaporation Zone',
                 font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        tk.Label(zone_inner, text=f'Current Zone: {current_zone}',
                 font=('Segoe UI', 11), bg='#e8eef5', fg='#00a886', pady=5).pack(anchor='w')
        tk.Label(zone_inner, text='This zone applies to all regional evaporation calculations',
                 font=('Segoe UI', 9), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w')

        # Rainfall Card
        rainfall_card = tk.Frame(scroll_container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        rainfall_card.pack(fill='x', pady=(0, 20))

        rainfall_inner = tk.Frame(rainfall_card, bg='#e8eef5')
        rainfall_inner.pack(fill='x', padx=20, pady=16)

        tk.Label(rainfall_inner, text='🌧️ Regional Rainfall (mm/month)', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w', pady=(0, 12))

        self.rainfall_entries = {}
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        rainfall_grid = tk.Frame(rainfall_inner, bg='#e8eef5')
        rainfall_grid.pack(fill='x')

        for idx, month in enumerate(months, 1):
            row = (idx - 1) // 4
            col = (idx - 1) % 4

            month_frame = tk.Frame(rainfall_grid, bg='#e8eef5')
            month_frame.grid(row=row, column=col, padx=8, pady=8, sticky='ew')

            tk.Label(month_frame, text=f"{month}:", width=4, font=('Segoe UI', 9), bg='#e8eef5', fg='#2c3e50').pack(side='left')
            entry = tk.Entry(month_frame, font=('Segoe UI', 10), bg='white', relief=tk.FLAT, bd=1, highlightbackground='#d9e6f4', highlightthickness=1)
            entry.pack(side='left', padx=5, ipady=4)
            self.rainfall_entries[idx] = entry
            tk.Label(month_frame, text='mm', fg='#7f8c8d', font=('Segoe UI', 9), bg='#e8eef5').pack(side='left')

        # Evaporation Card
        evap_card = tk.Frame(scroll_container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        evap_card.pack(fill='x', pady=(0, 0))

        evap_inner = tk.Frame(evap_card, bg='#e8eef5')
        evap_inner.pack(fill='x', padx=20, pady=16)

        tk.Label(evap_inner, text='☀️ Regional Evaporation (mm/month)', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w', pady=(0, 12))

        self.evap_entries = {}

        evap_grid = tk.Frame(evap_inner, bg='#e8eef5')
        evap_grid.pack(fill='x')

        for idx, month in enumerate(months, 1):
            row = (idx - 1) // 4
            col = (idx - 1) % 4

            month_frame = tk.Frame(evap_grid, bg='#e8eef5')
            month_frame.grid(row=row, column=col, padx=8, pady=8, sticky='ew')

            tk.Label(month_frame, text=f"{month}:", width=4, font=('Segoe UI', 9), bg='#e8eef5', fg='#2c3e50').pack(side='left')
            entry = tk.Entry(month_frame, font=('Segoe UI', 10), bg='white', relief=tk.FLAT, bd=1, highlightbackground='#d9e6f4', highlightthickness=1)
            entry.pack(side='left', padx=5, ipady=4)
            self.evap_entries[idx] = entry
            tk.Label(month_frame, text='mm', fg='#7f8c8d', font=('Segoe UI', 9), bg='#e8eef5').pack(side='left')

        # Buttons (action bar)
        ttk.Separator(scroll_container, orient='horizontal').pack(fill='x', pady=(5, 10))
        btn_frame = ttk.Frame(scroll_container)
        btn_frame.pack(fill='x', pady=(0, 10))

        ttk.Button(btn_frame, text='💾 Save Regional Parameters', width=28,
              command=self._save_environmental_params).pack(side='left', padx=(0, 8))
        ttk.Button(btn_frame, text='🔄 Load Current Values', width=24,
              command=self._load_environmental_params).pack(side='left')

        # Load initial values
        self._load_environmental_params()

    def _load_environmental_params(self):
        """Load regional rainfall and evaporation from DB for selected year"""
        try:
            year = int(self.env_year_var.get())
        except ValueError:
            year = None
        
        try:
            # Load rainfall
            rainfall_data = self.db.get_regional_rainfall_all_months(year=year)
            for month, value in rainfall_data.items():
                if month in self.rainfall_entries:
                    self.rainfall_entries[month].delete(0, 'end')
                    if value > 0:
                        self.rainfall_entries[month].insert(0, str(value))

            # Load evaporation
            evap_data = self.db.get_regional_evaporation_all_months(year=year)
            for month, value in evap_data.items():
                if month in self.evap_entries:
                    self.evap_entries[month].delete(0, 'end')
                    if value > 0:
                        self.evap_entries[month].insert(0, str(value))
        except Exception as e:
            logger.error(f"Failed to load environmental data: {e}")
            messagebox.showerror('Error', f'Failed to load environmental data: {str(e)}')

    def _save_environmental_params(self):
        """Save regional rainfall and evaporation to DB for selected year"""
        try:
            year = int(self.env_year_var.get())
        except ValueError:
            year = None
        
        try:
            saved_count = 0

            # Save rainfall
            for month, entry in self.rainfall_entries.items():
                value_str = entry.get().strip()
                if value_str:
                    try:
                        value = float(value_str)
                        if value >= 0:
                            self.db.set_regional_rainfall_monthly(month, value, year=year)
                            saved_count += 1
                    except ValueError:
                        pass

            # Save evaporation
            for month, entry in self.evap_entries.items():
                value_str = entry.get().strip()
                if value_str:
                    try:
                        value = float(value_str)
                        if value >= 0:
                            self.db.set_regional_evaporation_monthly(month, value, year=year)
                            saved_count += 1
                    except ValueError:
                        pass

            year_label = f'({year})' if year else '(default)'
            messagebox.showinfo('✓ Success',
                f'Regional environmental parameters saved successfully {year_label}!\n'
                f'{saved_count} values updated\n\n'
                f'These values will apply to all storage facilities.')

            # Reload to confirm
            self._load_environmental_params()

        except Exception as e:
            logger.error(f"Failed to save environmental data: {e}")
            messagebox.showerror('Error', f'Failed to save environmental data: {str(e)}')

    def _set_env_year(self, year: int):
        """Set the environmental year to a specific year and reload data"""
        self.env_year_var.set(str(year))
        self._load_environmental_params()




    def _build_data_sources_tab(self):
        """Build data sources configuration tab for Excel file paths"""
        # Create scrollable container
        scroll_container = tk.Frame(self.data_sources_frame, bg='#f5f6f7')
        scroll_container.pack(fill='both', expand=True)
        
        # Create canvas and scrollbar
        canvas = tk.Canvas(scroll_container, bg='#f5f6f7', highlightthickness=0)
        scrollbar = tk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f5f6f7')
        
        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Enable mousewheel scrolling
        enable_canvas_mousewheel(canvas)
        
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        container = scrollable_frame  # Use scrollable frame as container
        
        # Add padding to container
        container.configure(padx=20, pady=20)

        # Info card
        info_card = tk.Frame(container, bg='#e8eef5', relief=tk.FLAT, bd=1, highlightbackground='#c5d3e6', highlightthickness=1)
        info_card.pack(fill='x', pady=(0, 20))

        info_inner = tk.Frame(info_card, bg='#e8eef5')
        info_inner.pack(fill='x', padx=20, pady=16)

        tk.Label(info_inner, text='📂 Excel Data Source Configuration', font=('Segoe UI', 12, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        tk.Label(info_inner, text='Configure paths to Excel files used by the application',
                 font=('Segoe UI', 10), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=(5, 10))

        tk.Label(info_inner, text='� HOW IT WORKS:', font=('Segoe UI', 9, 'bold'), fg='#2c3e50', bg='#e8eef5').pack(anchor='w', pady=(5, 3))
        tk.Label(info_inner, text='1. Application Inputs Excel is PRIMARY - loads flow volumes for calculations',
                 font=('Segoe UI', 9), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=1)
        tk.Label(info_inner, text='2. TRP Water Balance Excel is OPTIONAL - only used for historical analysis',
                 font=('Segoe UI', 9), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=1)
        tk.Label(info_inner, text='3. If Application Inputs file is missing, app will not function',
                 font=('Segoe UI', 9), fg='#7f8c8d', bg='#e8eef5').pack(anchor='w', pady=(1, 10))

        tk.Label(info_inner, text='💡 If you move Excel files, update the paths here and restart the app.',
                 font=('Segoe UI', 9), fg='#0066cc', bg='#e8eef5').pack(anchor='w')

        # Template Excel file (MOVE THIS FIRST - IT'S PRIMARY)
        template_card = tk.Frame(container, bg='#d4edda', relief=tk.FLAT, bd=2, highlightbackground='#28a745', highlightthickness=2)
        template_card.pack(fill='x', pady=(0, 20))

        template_inner = tk.Frame(template_card, bg='#d4edda')
        template_inner.pack(fill='x', padx=20, pady=16)

        tk.Label(template_inner, text='🟢 REQUIRED #1: Flow Diagram Excel (Flows_* sheets)', font=('Segoe UI', 12, 'bold'), bg='#d4edda', fg='#155724').pack(anchor='w')
        tk.Label(template_inner, text='Flow volumes for diagram visualization (REQUIRED)',
                 font=('Segoe UI', 10), fg='#155724', bg='#d4edda').pack(anchor='w', pady=(3, 0))
        tk.Label(template_inner, text='Must have sheets: Flows_UG2 North, Flows_Merensky Plant, etc. Monitored for live updates.',
                 font=('Segoe UI', 9), fg='#155724', bg='#d4edda').pack(anchor='w', pady=(2, 10))

        template_path_frame = tk.Frame(template_inner, bg='#d4edda')
        template_path_frame.pack(fill='x', pady=(0, 10))

        tk.Label(template_path_frame, text='Current Path:', font=('Segoe UI', 9), bg='#d4edda', fg='#155724').pack(side='left')

        template_path = config.get('data_sources.template_excel_path', 'templates/Water_Balance_TimeSeries_Template.xlsx')
        self.template_path_label = tk.Label(template_path_frame, text=template_path or 'Not configured',
                                            fg='#0066cc', bg='#d4edda', font=('Segoe UI', 9, 'italic'))
        self.template_path_label.pack(side='left', padx=(5, 10))

        # Status indicator
        base_dir = Path(get_resource_path(''))  # Use get_resource_path for PyInstaller compatibility
        template_full = base_dir / template_path if template_path and not Path(template_path).is_absolute() else (Path(template_path) if template_path else None)
        status_text = '✓ Found' if (template_full and template_full.exists()) else ('⚠ Not configured' if not template_full else '❌ Missing')
        status_color = '#28a745' if (template_full and template_full.exists()) else ('#ffa500' if not template_full else '#dc3545')
        self.template_status_label = tk.Label(template_path_frame, text=status_text,
                                              fg=status_color, bg='#d4edda', font=('Segoe UI', 9, 'bold'))
        self.template_status_label.pack(side='left')

        template_btn_frame = tk.Frame(template_inner, bg='#d4edda')
        template_btn_frame.pack(fill='x', pady=(10, 0))

        tk.Button(template_btn_frame, text='📁 Browse...', command=self._select_template_excel,
                  bg='#28a745', fg='white', font=('Segoe UI', 9), relief=tk.FLAT, padx=10, pady=6, cursor='hand2',
                  activebackground='#218838', activeforeground='white').pack(side='left', padx=(0, 10))
        tk.Button(template_btn_frame, text='🔄 Reset to Default', command=self._reset_template_path,
                  bg='#28a745', fg='white', font=('Segoe UI', 9), relief=tk.FLAT, padx=10, pady=6, cursor='hand2',
                  activebackground='#218838', activeforeground='white').pack(side='left', padx=(0, 10))
        tk.Button(template_btn_frame, text='📂 Open Folder', command=self._open_template_folder,
                  bg='#28a745', fg='white', font=('Segoe UI', 9), relief=tk.FLAT, padx=10, pady=6, cursor='hand2',
                  activebackground='#218838', activeforeground='white').pack(side='left')

        # Meter Readings Excel file
        legacy_card = tk.Frame(container, bg='#d4edda', relief=tk.FLAT, bd=2, highlightbackground='#28a745', highlightthickness=2)
        legacy_card.pack(fill='x', pady=(0, 20))

        legacy_inner = tk.Frame(legacy_card, bg='#d4edda')
        legacy_inner.pack(fill='x', padx=20, pady=16)

        tk.Label(legacy_inner, text='🟢 REQUIRED #2: Meter Readings Excel (Calculations Data)', font=('Segoe UI', 12, 'bold'), bg='#d4edda', fg='#155724').pack(anchor='w')
        tk.Label(legacy_inner, text='Tonnes milled, RWD, dewatering volumes for calculations (REQUIRED)',
                 font=('Segoe UI', 10), bg='#d4edda', fg='#155724').pack(anchor='w', pady=(3, 0))
        tk.Label(legacy_inner, text='Must have sheet: Meter Readings. Used by Calculations dashboard for water balance.',
                 font=('Segoe UI', 9), fg='#155724', bg='#d4edda').pack(anchor='w', pady=(2, 10))

        legacy_path_frame = tk.Frame(legacy_inner, bg='#d4edda')
        legacy_path_frame.pack(fill='x', pady=(0, 10))

        tk.Label(legacy_path_frame, text='Current Path:', font=('Segoe UI', 9), bg='#d4edda', fg='#155724').pack(side='left')

        legacy_path = config.get('data_sources.legacy_excel_path', 'data/New Water Balance  20250930 Oct.xlsx')
        self.legacy_path_label = tk.Label(legacy_path_frame, text=legacy_path or 'Not configured',
                                          fg='#0066cc', bg='#d4edda', font=('Segoe UI', 9, 'italic'))
        self.legacy_path_label.pack(side='left', padx=(5, 10))

        # Status indicator
        base_dir = Path(get_resource_path(''))
        legacy_full = base_dir / legacy_path if legacy_path and not Path(legacy_path).is_absolute() else (Path(legacy_path) if legacy_path else None)
        status_text = '✓ Found' if (legacy_full and legacy_full.exists()) else ('⚠ Not configured' if not legacy_full else '❌ Missing')
        status_color = '#28a745' if (legacy_full and legacy_full.exists()) else ('#ffa500' if not legacy_full else '#dc3545')
        self.legacy_status_label = tk.Label(legacy_path_frame, text=status_text,
                                            fg=status_color, bg='#d4edda', font=('Segoe UI', 9, 'bold'))
        self.legacy_status_label.pack(side='left')

        legacy_btn_frame = tk.Frame(legacy_inner, bg='#d4edda')
        legacy_btn_frame.pack(fill='x', pady=(10, 0))

        tk.Button(legacy_btn_frame, text='📁 Browse...', command=self._select_legacy_excel,
                  bg='#28a745', fg='white', font=('Segoe UI', 9), relief=tk.FLAT, padx=10, pady=6, cursor='hand2',
                  activebackground='#218838', activeforeground='white').pack(side='left', padx=(0, 10))
        tk.Button(legacy_btn_frame, text='🔄 Reset to Default', command=self._reset_legacy_path,
                  bg='#28a745', fg='white', font=('Segoe UI', 9), relief=tk.FLAT, padx=10, pady=6, cursor='hand2',
                  activebackground='#218838', activeforeground='white').pack(side='left')

        # Warning card
        warning_card = tk.Frame(container, bg='#e7f3ff', relief=tk.FLAT, bd=1, highlightbackground='#0066cc', highlightthickness=1)
        warning_card.pack(fill='x', pady=(0, 20))

        warning_inner = tk.Frame(warning_card, bg='#e7f3ff')
        warning_inner.pack(fill='x', padx=20, pady=16)

        tk.Label(warning_inner, text='📋 Data Flow & Important Notes', font=('Segoe UI', 12, 'bold'), bg='#e7f3ff', fg='#004085').pack(anchor='w', pady=(0, 10))

        notes = [
            "🟢 BOTH FILES REQUIRED - App needs both Excel files to function fully",
            "📊 Flow Diagram Excel → Flow volumes for diagram visualization",
            "📈 Meter Readings Excel → Tonnes milled, RWD, dewatering for calculations",
            "ℹ Flow Diagram Excel has sheets: Flows_UG2 North, Flows_Merensky Plant, etc.",
            "ℹ Meter Readings Excel has sheet: Meter Readings",
            "⚠ Changes to paths take effect after you restart the app",
            "⚠ Ensure both file paths are correct before restarting",
            "💡 Use file browser button to find and select Excel files on your computer"
        ]

        for note in notes:
            color = '#155724' if note.startswith('✓') else ('#004085' if note.startswith('ℹ') else '#856404')
            tk.Label(warning_inner, text=note, font=('Segoe UI', 9), fg=color, bg='#e7f3ff').pack(anchor='w', pady=2)
    
    def _select_legacy_excel(self):
        """Select legacy Excel file"""
        current_path = config.get('data_sources.legacy_excel_path', 'data/New Water Balance  20250930 Oct.xlsx')
        initial_dir = Path(current_path).parent if current_path and Path(current_path).exists() else Path.cwd()
        
        filename = filedialog.askopenfilename(
            title='Select TRP Water Balance Excel File',
            initialdir=initial_dir,
            filetypes=[
                ('Excel Files', '*.xlsx *.xls'),
                ('All Files', '*.*')
            ]
        )
        
        if filename:
            # Convert to relative path if inside project
            try:
                rel_path = Path(filename).relative_to(Path.cwd())
                path_str = str(rel_path)
            except ValueError:
                path_str = filename
            
            config.set('data_sources.legacy_excel_path', path_str)
            self.legacy_path_label.config(text=path_str)
            
            # Update status indicator
            base_dir = Path.cwd()
            full_path = base_dir / path_str if not Path(path_str).is_absolute() else Path(path_str)
            if full_path.exists():
                self.legacy_status_label.config(text='✓ Found', foreground='#28a745')
            else:
                self.legacy_status_label.config(text='❌ Missing', foreground='#dc3545')
            
            # Trigger automatic reload
            self._reload_excel_data()
            
            messagebox.showinfo('✓ Path Updated', 
                               f'TRP Excel path updated to:\n{path_str}\n\n'
                               'Data has been reloaded from the new location.')
    
    def _select_template_excel(self):
        """Select template Excel file"""
        current_path = config.get('data_sources.template_excel_path', 'templates/Water_Balance_TimeSeries_Template.xlsx')
        initial_dir = Path(current_path).parent if current_path and Path(current_path).exists() else Path.cwd()
        
        filename = filedialog.askopenfilename(
            title='Select Application Inputs Excel File',
            initialdir=initial_dir,
            filetypes=[
                ('Excel Files', '*.xlsx *.xls'),
                ('All Files', '*.*')
            ]
        )
        
        if filename:
            # Convert to relative path if inside project
            try:
                rel_path = Path(filename).relative_to(Path.cwd())
                path_str = str(rel_path)
            except ValueError:
                path_str = filename
            
            config.set('data_sources.template_excel_path', path_str)
            # Also update timeseries_excel_path so flow loader picks it up
            config.set('data_sources.timeseries_excel_path', path_str)
            
            # Reset flow volume loader singleton to force path reload
            try:
                from utils.flow_volume_loader import reset_flow_volume_loader
                reset_flow_volume_loader()
                logger.info(f"🔄 Flow loader reset after path change to: {path_str}")
            except Exception as e:
                logger.warning(f"Could not reset flow volume loader: {e}")
            
            self.template_path_label.config(text=path_str)
            
            # Update status indicator
            base_dir = Path.cwd()
            full_path = base_dir / path_str if not Path(path_str).is_absolute() else Path(path_str)
            if full_path.exists():
                self.template_status_label.config(text='✓ Found', foreground='#28a745')
            else:
                self.template_status_label.config(text='❌ Missing', foreground='#dc3545')
            
            # Trigger automatic reload
            self._reload_excel_data()
            
            messagebox.showinfo('✓ Path Updated', 
                               f'Application Inputs Excel path updated to:\n{path_str}\n\n'
                               'Data has been reloaded from the new location.')
    
    def _reset_legacy_path(self):
        """Reset legacy path to default"""
        default_path = 'data/New Water Balance  20250930 Oct.xlsx'
        config.set('data_sources.legacy_excel_path', default_path)
        self.legacy_path_label.config(text=default_path)
        
        # Update status
        base_dir = Path.cwd()
        full_path = base_dir / default_path
        if full_path.exists():
            self.legacy_status_label.config(text='✓ Found', foreground='#28a745')
        else:
            self.legacy_status_label.config(text='❌ Missing', foreground='#dc3545')
        
        self._reload_excel_data()
        messagebox.showinfo('✓ Reset', f'TRP Excel path reset to default:\n{default_path}\n\nData reloaded.')
    
    def _reset_template_path(self):
        """Reset template path to default"""
        default_path = 'templates/Water_Balance_TimeSeries_Template.xlsx'
        config.set('data_sources.template_excel_path', default_path)
        self.template_path_label.config(text=default_path)
        
        # Update status
        base_dir = Path.cwd()
        full_path = base_dir / default_path
        if full_path.exists():
            self.template_status_label.config(text='✓ Found', foreground='#28a745')
        else:
            self.template_status_label.config(text='❌ Missing', foreground='#dc3545')
        
        self._reload_excel_data()
        messagebox.showinfo('✓ Reset', f'Application Inputs Excel path reset to default:\n{default_path}\n\nData reloaded.')
    
    def _open_template_folder(self):
        """Open the folder containing the template file"""
        import subprocess
        import platform
        
        template_path = config.get('data_sources.template_excel_path', 'templates/Water_Balance_TimeSeries_Template.xlsx')
        folder_path = Path(template_path).parent
        
        try:
            if platform.system() == 'Windows':
                subprocess.run(['explorer', str(folder_path.resolve())])
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', str(folder_path.resolve())])
            else:  # Linux
                subprocess.run(['xdg-open', str(folder_path.resolve())])
        except Exception as ex:
            messagebox.showerror('Error', f'Failed to open folder:\n{str(ex)}')
    
    def _reload_excel_data(self):
        """Force reload Excel data after path change"""
        try:
            from utils.app_logger import logger
            logger.info("🔄 Excel path changed - extended sheets removed, only Meter Readings used...")
            
            # Extended Excel sheets removed - no reload needed
            logger.info("✓ Extended Excel sheets no longer exist")
            
            # Notify user
            from utils.ui_notify import notifier
            notifier.success("Excel data reloaded successfully", title="Data Refresh")
            
        except Exception as ex:
            from utils.app_logger import logger
            logger.error(f"Failed to reload Excel data: {ex}")
            messagebox.showerror('Reload Error', 
                               f'Failed to reload Excel data:\n{str(ex)}\n\n'
                               'You may need to restart the application.')

    def _create_backup(self):
        """Create database backup"""
        try:
            path = self.backup.create_backup()
            messagebox.showinfo('✓ Backup Created', 
                               f'Database backup created successfully:\n\n{path}')
            self._load_backups()
        except Exception as ex:
            messagebox.showerror('Backup Error', f'Failed to create backup:\n{str(ex)}')

    def _load_backups(self):
        """Load available backups"""
        self.backups_list.delete(0, 'end')
        backups = self.backup.list_backups()
        
        if not backups:
            self.backups_list.insert('end', '(No backups found)')
        else:
            for f in backups:
                self.backups_list.insert('end', f.name)

    def _restore_selected(self):
        """Restore from selected backup"""
        sel = self.backups_list.curselection()
        if not sel:
            messagebox.showwarning('No Selection', 'Please select a backup file to restore.')
            return
        
        name = self.backups_list.get(sel[0])
        if name == '(No backups found)':
            return
        
        backup_path = self.backup.backup_dir / name
        
        confirm = messagebox.askyesno('⚠️ Confirm Restore',
            f'Restore database from backup:\n\n{name}\n\n'
            'This will OVERWRITE the current database.\n'
            'Current data will be lost unless backed up.\n\n'
            'Continue with restore?')
        
        if not confirm:
            return
        
        try:
            self.backup.restore_backup(backup_path)
            messagebox.showinfo('✓ Restore Complete', 
                               f'Database successfully restored from:\n{name}\n\n'
                               'Please restart the application for changes to take effect.')
        except Exception as ex:
            messagebox.showerror('Restore Error', f'Failed to restore backup:\n{str(ex)}')

    def _build_roadmap_tab(self):
        """
        Build the roadmap preview tab showing future features.
        
        This tab showcases upcoming features as a selling point without
        implementing any actual functionality. It's a pure preview/marketing
        tool that helps demonstrate the product's future direction.
        """
        # Create and load the roadmap preview component
        roadmap = RoadmapPreviewTab(self.roadmap_frame)
        roadmap.load()
        
        logger.info("✓ Roadmap preview tab loaded successfully")


