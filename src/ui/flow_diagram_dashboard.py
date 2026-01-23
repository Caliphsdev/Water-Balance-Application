"""
Flow Diagram Editor - Manual Segment Drawing
- Draw flow lines manually by clicking points (orthogonal segments only)
- Lines independent from components (don't move when components move)
- Click-to-draw: each click creates a segment corner
- Excel-based volume loading on-demand per month
"""

import json
import tkinter as tk
from tkinter import Canvas, Frame, Label, Scrollbar, messagebox, Button, ttk, simpledialog
from pathlib import Path
from datetime import date, datetime
import warnings
from openpyxl import load_workbook
import uuid
import sys
import os

# Suppress openpyxl warnings about print areas
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

sys.path.insert(0, str(Path(__file__).parent.parent))
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from utils.app_logger import logger
from database.db_manager import DatabaseManager
from utils.flow_volume_loader import get_flow_volume_loader
from utils.excel_mapping_registry import (
    get_excel_mapping_registry,
    reset_excel_mapping_registry,
)
from utils.column_alias_resolver import get_column_alias_resolver
from ui.mouse_wheel_support import enable_canvas_mousewheel, enable_listbox_mousewheel


class FlowDiagramBalanceCheckDialog:
    """Balance check dialog for flow diagram - user categorizes flowlines and sees balance calculation"""
    
    def __init__(self, parent, flow_diagram):
        """
        Args:
            parent: Parent window
            flow_diagram: DetailedNetworkFlowDiagram instance
        """
        self.parent = parent
        self.flow_diagram = flow_diagram
        self.dialog = None
        self.categorizations = {}  # {edge_id: 'inflow'|'recirculation'|'outflow'|'ignore'}
        self.category_vars = {}  # {edge_id: tk.StringVar}
        self.result_label = None
        self.category_source_info = ""
        
        # Load saved categorizations
        self._load_categorizations()
        
    def _copy_bundled_categories_if_needed(self, user_dir: str) -> None:
        """Copy bundled categories from app bundle to user directory on first run.
        
        In EXE mode, if user directory doesn't have categories, copy the bundled
        development categories as defaults for the first installation.
        """
        if not user_dir:
            return  # Not in EXE mode
        
        user_config_path = Path(user_dir) / 'data' / 'balance_check_flow_categories.json'
        if user_config_path.exists():
            return  # User already has categories, don't overwrite
        
        # Try to find bundled categories in PyInstaller bundle
        try:
            bundled_path = Path(sys._MEIPASS) / 'data' / 'balance_check_flow_categories.json'
            if bundled_path.exists():
                # Copy bundled categories to user directory
                user_config_path.parent.mkdir(parents=True, exist_ok=True)
                with open(bundled_path, 'r', encoding='utf-8') as src:
                    bundled_data = json.load(src)
                with open(user_config_path, 'w', encoding='utf-8') as dst:
                    json.dump(bundled_data, dst, indent=2)
                logger.info(f"✅ Copied bundled balance check categories to {user_config_path}")
        except (AttributeError, FileNotFoundError, Exception) as e:
            # Not in PyInstaller bundle or bundled file not found - no action needed
            pass
    
    def _load_categorizations(self):
        """Load saved flow categorizations from JSON"""
        try:
            # Use user data directory (EXE mode) or project data folder (dev mode)
            user_dir = os.environ.get('WATERBALANCE_USER_DIR')
            if user_dir:
                # First-run: copy bundled categories if user doesn't have any
                self._copy_bundled_categories_if_needed(user_dir)
                config_path = Path(user_dir) / 'data' / 'balance_check_flow_categories.json'
            else:
                config_path = Path(__file__).parent.parent.parent / 'data' / 'balance_check_flow_categories.json'
            
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Load for current area with sensible fallback to 'default'
                    # Prefer the diagram's area code; if missing, infer from title.
                    inferred_area = self.flow_diagram.area_code or self.flow_diagram._get_area_code_from_title()
                    area_code = inferred_area or 'default'
                    base = data.get('default', {})
                    area_specific = data.get(area_code, {})
                    # Merge: area-specific overrides default baseline
                    merged = dict(base)
                    merged.update(area_specific)
                    self.categorizations = merged
                    # Build source info label text
                    base_has = bool(base)
                    area_has = bool(area_specific)
                    if area_has and base_has:
                        self.category_source_info = f"{area_code} + default baseline"
                    elif area_has:
                        self.category_source_info = f"{area_code}"
                    elif base_has:
                        self.category_source_info = "default baseline"
                    else:
                        self.category_source_info = "none (all ignore by default)"
                    loaded_key = area_code if area_specific else 'default'
                    logger.info(f"✅ Loaded balance check categories for {loaded_key}")
        except Exception as e:
            logger.warning(f"Could not load balance check categories: {e}")
            self.categorizations = {}
            self.category_source_info = "none (all ignore by default)"
    
    def _save_categorizations(self):
        """Save flow categorizations to JSON"""
        try:
            # Use user data directory (EXE mode) or project data folder (dev mode)
            user_dir = os.environ.get('WATERBALANCE_USER_DIR')
            if user_dir:
                config_path = Path(user_dir) / 'data' / 'balance_check_flow_categories.json'
            else:
                config_path = Path(__file__).parent.parent.parent / 'data' / 'balance_check_flow_categories.json'
            
            config_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Load existing data
            all_data = {}
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8') as f:
                    all_data = json.load(f)
            
            # Update for current area
            area_code = self.flow_diagram.area_code or 'default'
            all_data[area_code] = self.categorizations
            
            # Save
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(all_data, f, indent=2)
            
            logger.info(f"💾 Saved balance check categories for {area_code}")
        except Exception as e:
            logger.error(f"Failed to save balance check categories: {e}")
    
    def show(self):
        """Open the balance check dialog"""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title("⚖️ Flow Diagram Balance Check")
        
        # Calculate responsive window size based on screen dimensions
        screen_width = self.dialog.winfo_screenwidth()
        screen_height = self.dialog.winfo_screenheight()
        # Use 85% of screen for safety margin (taskbar, etc.)
        window_width = int(screen_width * 0.85)
        window_height = int(screen_height * 0.85)
        # Clamp to reasonable bounds
        window_width = max(900, min(window_width, 1300))
        window_height = max(700, min(window_height, 1000))
        
        # Center window on screen
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        self.dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.dialog.configure(bg='#2c3e50')
        
        # Make modal
        self.dialog.transient(self.parent)
        self.dialog.grab_set()
        # Auto-save on window close (top-right X)
        self.dialog.protocol("WM_DELETE_WINDOW", self._save_and_close)
        
        # Header
        header_frame = tk.Frame(self.dialog, bg='#34495e', relief=tk.SOLID, borderwidth=1)
        header_frame.pack(fill=tk.X, padx=10, pady=10)
        
        tk.Label(header_frame, text="⚖️ Balance Check", 
                font=('Segoe UI', 16, 'bold'), bg='#34495e', fg='#e8eef5').pack(anchor='w', padx=15, pady=(10, 2))
        tk.Label(header_frame, text="Categorize flowlines and verify water balance using: (Inflows - Recirculation - Outflows) / Inflows × 100", 
                font=('Segoe UI', 10), bg='#34495e', fg='#95a5a6').pack(anchor='w', padx=15, pady=(0, 10))
        
        # Date selection
        date_frame = tk.Frame(self.dialog, bg='#2c3e50')
        date_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        tk.Label(date_frame, text="📅 Period:", font=('Segoe UI', 10, 'bold'), 
                bg='#2c3e50', fg='#e8eef5').pack(side='left', padx=(0, 10))
        
        tk.Label(date_frame, text="Year:", bg='#2c3e50', fg='#e8eef5').pack(side='left', padx=(0, 5))

        # Show which category set is active for clarity
        info_frame = tk.Frame(self.dialog, bg='#2c3e50')
        info_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        tk.Label(
            info_frame,
            text=f"🔖 Active Categories: {self.category_source_info}",
            font=('Segoe UI', 9),
            bg='#2c3e50',
            fg='#95a5a6'
        ).pack(anchor='w', padx=0)
        self.year_var = tk.StringVar(value=str(self.flow_diagram.current_year))
        tk.Spinbox(date_frame, from_=2020, to=2100, textvariable=self.year_var, 
                  width=6, font=('Segoe UI', 9)).pack(side='left', padx=(0, 15))
        
        tk.Label(date_frame, text="Month:", bg='#2c3e50', fg='#e8eef5').pack(side='left', padx=(0, 5))
        self.month_var = tk.StringVar(value=str(self.flow_diagram.current_month))
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        month_combo = ttk.Combobox(date_frame, textvariable=self.month_var, 
                                   values=months, state='readonly', width=11)
        month_combo.current(self.flow_diagram.current_month - 1)
        month_combo.pack(side='left', padx=(0, 15))
        
        tk.Button(date_frame, text="🔄 Recalculate", command=self._calculate_balance,
                 bg='#3498db', fg='white', font=('Segoe UI', 9, 'bold'), padx=15).pack(side='left', padx=5)
        
        # Main content area - scrollable
        main_frame = tk.Frame(self.dialog, bg='#2c3e50')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        canvas = tk.Canvas(main_frame, bg='#2c3e50', highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_frame, orient='vertical', command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#2c3e50')
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Enable mouse wheel scrolling
        enable_canvas_mousewheel(canvas)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Flowlines table
        self._create_flowlines_table(scrollable_frame)
        
        # Results section (centered)
        results_frame = tk.Frame(self.dialog, bg='#34495e', relief=tk.SOLID, borderwidth=1)
        # Center the results card within the dialog
        results_frame.pack(padx=10, pady=(0, 10), anchor='center')
        
        tk.Label(results_frame, text="📊 Balance Calculation Results", 
            font=('Segoe UI', 12, 'bold'), bg='#34495e', fg='#e8eef5').pack(anchor='center', padx=15, pady=(10, 5))
        
        self.result_label = tk.Label(results_frame, text="Click 'Recalculate' to see results", 
                         font=('Segoe UI', 11), bg='#34495e', fg='#95a5a6', justify='left')
        # Center the results text block while keeping left-justified content
        self.result_label.pack(anchor='center', padx=15, pady=(0, 10))
        
        # Action buttons
        btn_frame = tk.Frame(self.dialog, bg='#2c3e50')
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        tk.Button(btn_frame, text="💾 Save Categories", command=self._save_and_close,
                 bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'), padx=20).pack(side='left', padx=5)
        # Close also saves; reduces risk of losing work
        tk.Button(btn_frame, text="✖ Close", command=self._save_and_close,
                 bg='#95a5a6', fg='white', font=('Segoe UI', 10, 'bold'), padx=20).pack(side='right', padx=5)
        
        # Auto-calculate on open
        self.dialog.after(100, self._calculate_balance)
    
    def _create_flowlines_table(self, parent):
        """Create table showing all flowlines with categorization options and filters"""
        # Filters row
        filters_frame = tk.Frame(parent, bg='#2c3e50')
        filters_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Area filter (flows list automatically shows only mapped flows)
        tk.Label(filters_frame, text="Area:", bg='#2c3e50', fg='#e8eef5', font=('Segoe UI', 9)).pack(side='left', padx=(0, 5))
        self.area_filter_var = tk.StringVar(value='all')
        area_options = ['all'] + self._collect_edge_areas()
        area_combo = ttk.Combobox(filters_frame, textvariable=self.area_filter_var,
                      values=area_options, state='readonly', width=18)
        area_combo.current(0)
        area_combo.pack(side='left')
        area_combo.bind('<<ComboboxSelected>>', lambda e: self._refresh_flowlines_table())
        # Keep reference for dynamic refresh
        self.area_combo = area_combo
        
        # Table header
        header_frame = tk.Frame(parent, bg='#34495e', relief=tk.SOLID, borderwidth=1)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Simplified headers (hide technical Flow ID for user-friendly UI)
        headers = [
            ("From → To", 560),
            ("Category", 220)
        ]
        
        for i, (label, width) in enumerate(headers):
            tk.Label(header_frame, text=label, font=('Segoe UI', 9, 'bold'), 
                    bg='#34495e', fg='#e8eef5', width=width//7, anchor='w').grid(row=0, column=i, padx=5, pady=8, sticky='w')
        
        # Flowlines content container (for filterable rebuilds)
        if hasattr(self, 'table_content_frame') and self.table_content_frame.winfo_exists():
            self.table_content_frame.destroy()
        self.table_content_frame = tk.Frame(parent, bg='#2c3e50')
        self.table_content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Initial build
        self._build_flowlines_rows(self.table_content_frame)

    def _get_filtered_edges(self):
        """Return edges filtered by mapping presence and selected area."""
        edges = self.flow_diagram.area_data.get('edges', [])
        area_filter = self.area_filter_var.get() if hasattr(self, 'area_filter_var') else 'all'
        
        filtered = []
        for idx, edge in enumerate(edges):
            edge_id = edge.get('id', f"edge_{idx}")
            excel_mapping = edge.get('excel_mapping', '')
            # Always show only mapped flows
            if not excel_mapping:
                continue
            if isinstance(excel_mapping, dict):
                if not excel_mapping.get('enabled') or not excel_mapping.get('column'):
                    continue
            if area_filter != 'all':
                edge_area = edge.get('area') or edge.get('area_code') or edge.get('areaCode') or edge.get('areaCode'.lower())
                if not edge_area and isinstance(excel_mapping, dict):
                    edge_area = excel_mapping.get('sheet')
                if str(edge_area) != area_filter:
                    continue
            filtered.append((idx, edge))
        return filtered

    def _collect_edge_areas(self):
        """Collect unique area identifiers from edges for filtering."""
        edges = self.flow_diagram.area_data.get('edges', [])
        areas = []
        seen = set()
        for edge in edges:
            area_val = edge.get('area') or edge.get('area_code') or edge.get('areaCode') or edge.get('areaCode'.lower())
            if not area_val:
                mapping = edge.get('excel_mapping', {})
                if isinstance(mapping, dict):
                    area_val = mapping.get('sheet')
            if area_val and area_val not in seen:
                seen.add(area_val)
                areas.append(str(area_val))
        areas.sort()
        return areas

    def _refresh_flowlines_table(self):
        """Refresh table content after filter or category change."""
        # Refresh area options dynamically
        if hasattr(self, 'area_combo') and self.area_combo.winfo_exists():
            # Keep dropdown options synchronized with latest edge metadata so new sheets/areas show up immediately
            options = ['all'] + self._collect_edge_areas()
            current = self.area_filter_var.get() if hasattr(self, 'area_filter_var') else 'all'
            self.area_combo['values'] = options
            if current not in options:
                current = 'all'
                self.area_filter_var.set(current)
            # Update current selection index
            try:
                idx = options.index(current)
                self.area_combo.current(idx)
            except ValueError:
                self.area_combo.current(0)

        if hasattr(self, 'table_content_frame') and self.table_content_frame.winfo_exists():
            self._build_flowlines_rows(self.table_content_frame)

    def _build_flowlines_rows(self, parent):
        """Build rows into the given parent using current filters."""
        # Clear any previous
        for w in parent.winfo_children():
            w.destroy()
        
        nodes = self.flow_diagram.area_data.get('nodes', [])
        nodes_by_id = {n['id']: n for n in nodes}
        
        for idx, edge in self._get_filtered_edges():
            edge_id = edge.get('id', f"edge_{idx}")
            from_id = edge.get('from', '')
            to_id = edge.get('to', '')
            excel_mapping = edge.get('excel_mapping', '')
            
            from_label = nodes_by_id.get(from_id, {}).get('label', from_id)
            to_label = nodes_by_id.get(to_id, {}).get('label', to_id)
            
            row_frame = tk.Frame(parent, bg='#2c3e50', relief=tk.SOLID, borderwidth=1)
            row_frame.pack(fill=tk.X, pady=2)
            
            # From → To
            flow_text = f"{from_label} → {to_label}"
            tk.Label(row_frame, text=flow_text, font=('Segoe UI', 9), 
                     bg='#2c3e50', fg='#e8eef5', width=560//7, anchor='w').grid(row=0, column=0, padx=5, pady=5, sticky='w')
            
            # Category dropdown
            current_category = self.categorizations.get(edge_id, 'ignore')
            category_var = self.category_vars.get(edge_id)
            if category_var is None:
                category_var = tk.StringVar(value=current_category)
                self.category_vars[edge_id] = category_var
            else:
                category_var.set(current_category)
            
            category_combo = ttk.Combobox(row_frame, textvariable=category_var, 
                                         values=['inflow', 'recirculation', 'outflow', 'ignore'],
                                         state='readonly', width=18)
            category_combo.grid(row=0, column=1, padx=5, pady=5, sticky='w')
            category_combo.bind('<<ComboboxSelected>>', 
                                lambda e, eid=edge_id: self._update_category(eid))

    # Tooltip removed per user request
    
    def _update_category(self, edge_id):
        """Update categorization when user changes dropdown"""
        self.categorizations[edge_id] = self.category_vars[edge_id].get()
        # Refresh view if a filter is active
        self._refresh_flowlines_table()
    
    def _calculate_balance(self):
        """Calculate water balance using categorized flows"""
        try:
            # Refresh table to reflect latest mappings/categories
            self._refresh_flowlines_table()
            # Get year/month
            year = int(self.year_var.get())
            months = ['January', 'February', 'March', 'April', 'May', 'June',
                     'July', 'August', 'September', 'October', 'November', 'December']
            month_name = self.month_var.get()
            month = months.index(month_name) + 1 if month_name in months else 1
            
            area_code = self.flow_diagram.area_code or 'default'
            
            # Get flow volumes from Excel
            flow_loader = get_flow_volume_loader()
            flow_loader.clear_cache()
            
            # Calculate totals by category
            inflows_total = 0.0
            recirculation_total = 0.0
            outflows_total = 0.0
            
            inflows_list = []
            recirculation_list = []
            outflows_list = []
            
            edges = self.flow_diagram.area_data.get('edges', [])
            
            for idx, edge in enumerate(edges):
                edge_id = edge.get('id', f"edge_{idx}")
                category = self.categorizations.get(edge_id, 'ignore')
                
                if category == 'ignore':
                    continue
                
                excel_mapping = edge.get('excel_mapping', '')
                if not excel_mapping:
                    continue
                
                # Extract volume from Excel
                try:
                    volume = flow_loader.get_flow_volume(area_code, excel_mapping, year, month)
                    if volume is None:
                        volume = 0.0
                    
                    if category == 'inflow':
                        inflows_total += volume
                        inflows_list.append((edge_id, volume))
                    elif category == 'recirculation':
                        recirculation_total += volume
                        recirculation_list.append((edge_id, volume))
                    elif category == 'outflow':
                        outflows_total += volume
                        outflows_list.append((edge_id, volume))
                
                except Exception as e:
                    logger.warning(f"Could not get volume for {excel_mapping}: {e}")
            
            # Calculate balance
            balance_diff = inflows_total - recirculation_total - outflows_total
            balance_error_pct = (balance_diff / inflows_total * 100) if inflows_total != 0 else 0
            
            # Determine status
            if abs(balance_error_pct) < 5.0:
                status = "✅ CLOSED"
                status_color = '#27ae60'
            elif abs(balance_error_pct) < 10.0:
                status = "⚠️ ACCEPTABLE"
                status_color = '#f39c12'
            else:
                status = "❌ CHECK REQUIRED"
                status_color = '#e74c3c'
            
            # Build result text
            result_text = f"""
📅 Period: {month_name} {year}
🏭 Area: {area_code}

📊 Balance Components:
  • Total Inflows:        {inflows_total:>15,.0f} m³  ({len(inflows_list)} flows)
  • Total Recirculation:  {recirculation_total:>15,.0f} m³  ({len(recirculation_list)} flows)
  • Total Outflows:       {outflows_total:>15,.0f} m³  ({len(outflows_list)} flows)

⚖️ Balance Equation:
  {inflows_total:,.0f} - {recirculation_total:,.0f} - {outflows_total:,.0f} = {balance_diff:,.0f} m³

📈 Balance Error:
  ({balance_diff:,.0f} / {inflows_total:,.0f}) × 100 = {balance_error_pct:.2f}%

Status: {status}
"""
            
            # Update result label
            self.result_label.config(text=result_text, fg='#e8eef5')
            
            # Save results for dashboard display
            self._save_balance_results({
                'timestamp': datetime.now().isoformat(),
                'year': year,
                'month': month,
                'month_name': month_name,
                'area_code': area_code,
                'inflows_total': inflows_total,
                'recirculation_total': recirculation_total,
                'outflows_total': outflows_total,
                'balance_diff': balance_diff,
                'balance_error_pct': balance_error_pct,
                'status': status,
                'inflow_count': len(inflows_list),
                'recirculation_count': len(recirculation_list),
                'outflow_count': len(outflows_list)
            })
            
            logger.info(f"✅ Balance check calculated: {balance_error_pct:.2f}% error")
        
        except Exception as e:
            logger.error(f"Balance calculation error: {e}", exc_info=True)
            self.result_label.config(text=f"❌ Error: {str(e)}", fg='#e74c3c')
    
    def _save_balance_results(self, results):
        """Save balance check results for dashboard display"""
        try:
            # Save to per-user data directory in EXE mode to avoid Program Files permission issues
            user_dir = os.environ.get('WATERBALANCE_USER_DIR')
            if user_dir:
                results_path = Path(user_dir) / 'data' / 'balance_check_last_run.json'
            else:
                results_path = Path(__file__).parent.parent.parent / 'data' / 'balance_check_last_run.json'
            results_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(results_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2)
            
            logger.info(f"💾 Saved balance check results for dashboard")
        except Exception as e:
            logger.error(f"Failed to save balance results: {e}")
    
    def _save_and_close(self):
        """Save categorizations and close dialog"""
        # Update categorizations from vars
        for edge_id, var in self.category_vars.items():
            self.categorizations[edge_id] = var.get()
        
        self._save_categorizations()
        messagebox.showinfo("✅ Saved", "Flow categorizations saved successfully!\n\nYour categories will be remembered the next time you open Balance Check.")
        self.dialog.destroy()


# Flow type and color mapping (clean, evaporation/losses, dirty variants)
def _default_color_for_type(flow_type: str) -> str:
    """Return default hex color for a given flow type.
    Categories:
    - clean: blue
    - evaporation|losses: black
    - dirty and variants: red
    """
    t = (flow_type or '').strip().lower()
    if t in ('clean', 'clean water'):
        return '#3498db'  # blue
    if t in ('evaporation', 'losses', 'evaporation/losses', 'evaporation_losses'):
        return '#000000'  # black
    # Dirty umbrella and variants
    dirty_variants = {
        'dirty', 'effluent', 'runoff', 'ug return', 'ug_return', 'dewatering', 'outflow',
        'inflow', 'drainage', 'return', 'irrigation', 'process_dirty', 'stormwater'
    }
    if t in dirty_variants:
        return '#e74c3c'  # red
    # Fallback
    return '#e74c3c'


class DetailedNetworkFlowDiagram:
    """
    Manual Segment-Based Flow Diagram Editor
    - Draw flow lines by clicking points (orthogonal 90° segments)
    - Lines are independent drawings (don't move with components)
    - Database-locked connections only
    """

    def __init__(self, parent):
        self.parent = parent
        self.canvas = None
        self.area_data = {}
        self.json_file = None
        self.db = DatabaseManager()
        
        # Excel-based volume loading (on-demand, no database)
        self.flow_loader = get_flow_volume_loader()
        self.current_year = date.today().year
        self.current_month = date.today().month
        self.area_code = None  # Will be determined from area_data
        
        # Node tracking
        self.selected_node = None
        self.dragging = False
        self.drag_start_x = 0
        self.drag_start_y = 0
        self.node_items = {}
        self.nodes_by_id = {}
        self.snap_to_grid = False
        self.grid_size = 20
        self.show_grid = True
        self.locked_nodes = {}  # {node_id: True/False}
        
        # Drawing mode
        self.drawing_mode = False
        self.drawing_segments = []  # [(x1,y1), (x2,y2), ...]
        self.drawing_from = None
        self.drawing_to = None
        
        # Label dragging
        self.dragging_label = False
        self.dragged_label_edge_idx = None
        self.label_items = {}  # {canvas_id: edge_index}
        self.edge_to_label_items = {}  # {edge_idx: [box_id, text_id]}
        
        # Recirculation dragging
        self.dragging_recirculation = False
        self.dragged_recirculation_edge_idx = None
        self.recirculation_drag_start_x = 0
        self.recirculation_drag_start_y = 0
        self.recirculation_locked = {}  # {edge_idx: True/False}
        

        
        # Snap to component
        self.snap_distance = 15  # pixels to snap (reduced for precision)
        self.snap_anchor_points = {}  # {node_id: [anchor_points]}
        self.hovered_anchor = None  # Currently hovered anchor point
        
        # Valid connections from database
        self.valid_connections = set()

    def load(self):
        """Load editor"""
        for widget in self.parent.winfo_children():
            widget.destroy()

        self._load_valid_connections()
        self._create_ui()
        self._load_diagram_data()
        logger.info("✅ Manual segment flow editor loaded")

    def _load_valid_connections(self):
        """Load valid connections from database (optional reference only)"""
        try:
            # This is just for reference - we don't enforce validation
            query = '''
                SELECT DISTINCT 
                    fs.structure_code as from_code,
                    ts.structure_code as to_code
                FROM wb_flow_connections fc
                JOIN wb_structures fs ON fc.from_structure_id = fs.structure_id
                JOIN wb_structures ts ON fc.to_structure_id = ts.structure_id
            '''
            result = self.db.execute_query(query)
            self.valid_connections = set()
            for row in result:
                from_code = row.get('from_code') or row.get(0)
                to_code = row.get('to_code') or row.get(1)
                if from_code and to_code:
                    self.valid_connections.add((from_code.lower(), to_code.lower()))
            
            logger.info(f"📚 Loaded {len(self.valid_connections)} known connections (reference only)")
        except Exception as e:
            logger.info(f"ℹ️ Database reference connections not available: {e}")
            self.valid_connections = set()

    def _is_connection_valid(self, from_id, to_id):
        """Connection validation disabled - all connections allowed"""
        # You can draw any connection between any two components
        logger.info(f"✅ Drawing flow: {from_id} → {to_id}")
        return True

    def _create_ui(self):
        """Create UI"""
        controls = Frame(self.parent, bg='#2c3e50', height=120)
        controls.pack(fill='x', padx=0, pady=0)

        title = Label(controls, text='FLOW DIAGRAM - Manual Flow Line Drawing', 
                     font=('Segoe UI', 12, 'bold'), bg='#2c3e50', fg='white')
        title.pack(pady=5)
        # Inline legend banner
        legend = Label(controls, text='Legend: Clean (Blue) • Evaporation/Losses (Black) • Dirty + variants (Red) • Recirculation (Purple)',
                       font=('Segoe UI', 8), bg='#2c3e50', fg='#e8eef5')
        legend.pack(pady=2)

        # Row 1: EDITING TOOLS (Flow Lines + Components combined)
        edit_frame = Frame(controls, bg='#2c3e50')
        edit_frame.pack(fill='x', padx=10, pady=3)
        
        # Flow Lines Section
        Label(edit_frame, text='🎯 FLOWS:', font=('Segoe UI', 8, 'bold'),
              bg='#2c3e50', fg='#3498db').pack(side='left', padx=(0,8))
        Button(edit_frame, text='✏️ Draw', command=self._start_drawing,
               bg='#3498db', fg='white', font=('Segoe UI', 9, 'bold'), padx=10).pack(side='left', padx=2)
        Button(edit_frame, text='🎨 Edit', command=self._edit_flow_unified,
               bg='#9b59b6', fg='white', font=('Segoe UI', 9, 'bold'), padx=10).pack(side='left', padx=2)
        Button(edit_frame, text='🗑️ Delete', command=self._delete_line,
               bg='#e74c3c', fg='white', font=('Segoe UI', 9, 'bold'), padx=10).pack(side='left', padx=2)
        
        # Separator
        Label(edit_frame, text='│', bg='#2c3e50', fg='#7f8c8d',
              font=('Segoe UI', 12)).pack(side='left', padx=10)
        
        # Components Section
        Label(edit_frame, text='🔧 NODES:', font=('Segoe UI', 8, 'bold'),
              bg='#2c3e50', fg='#e67e22').pack(side='left', padx=(0,8))
        Button(edit_frame, text='➕ Add', command=self._add_component,
               bg='#27ae60', fg='white', font=('Segoe UI', 9, 'bold'), padx=10).pack(side='left', padx=2)
        Button(edit_frame, text='✏️ Edit', command=self._edit_node,
               bg='#f39c12', fg='white', font=('Segoe UI', 9, 'bold'), padx=10).pack(side='left', padx=2)
        Button(edit_frame, text='🗑️ Delete', command=self._delete_node,
               bg='#c0392b', fg='white', font=('Segoe UI', 9, 'bold'), padx=10).pack(side='left', padx=2)
        Button(edit_frame, text='🔒 Lock', command=self._toggle_lock_selected,
               bg='#7f8c8d', fg='white', font=('Segoe UI', 9, 'bold'), padx=10).pack(side='left', padx=2)

        # Row 2: VIEW & ACTIONS
        view_frame = Frame(controls, bg='#2c3e50')
        view_frame.pack(fill='x', padx=10, pady=3)
        
        Button(view_frame, text='🔍 Zoom In', command=lambda: self._zoom(1.1),
               bg='#2ecc71', fg='white', font=('Segoe UI', 9, 'bold'), padx=12).pack(side='left', padx=2)
        Button(view_frame, text='🔍 Zoom Out', command=lambda: self._zoom(0.9),
               bg='#8e44ad', fg='white', font=('Segoe UI', 9, 'bold'), padx=12).pack(side='left', padx=2)
        
        # Spacer
        Frame(view_frame, width=20, bg='#2c3e50').pack(side='left')
        
        Button(view_frame, text='💾 Save Diagram', command=self._save_to_json,
               bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'), padx=20).pack(side='left', padx=2)

        # Row 3: DATA LOADING (Excel)
        excel_frame = Frame(controls, bg='#2c3e50')
        excel_frame.pack(fill='x', padx=10, pady=3)
        
        Label(excel_frame, text='📊 LOAD DATA:', font=('Segoe UI', 8, 'bold'),
              bg='#2c3e50', fg='#16a085').pack(side='left', padx=(0,10))
        
        Label(excel_frame, text='Year:', bg='#2c3e50', fg='white',
              font=('Segoe UI', 9)).pack(side='left', padx=(0,4))
        self.year_var = tk.StringVar(master=excel_frame, value=str(self.current_year))
        year_spin = tk.Spinbox(excel_frame, from_=2020, to=2100, textvariable=self.year_var,
                               width=6, font=('Segoe UI', 9))
        year_spin.pack(side='left', padx=(0,10))
        
        Label(excel_frame, text='Month:', bg='#2c3e50', fg='white',
              font=('Segoe UI', 9)).pack(side='left', padx=(0,4))
        self.month_var = tk.StringVar(master=excel_frame, value=str(self.current_month))
        months = ['January', 'February', 'March', 'April', 'May', 'June',
                  'July', 'August', 'September', 'October', 'November', 'December']
        month_combo = ttk.Combobox(excel_frame, textvariable=self.month_var,
                                   values=months, state='readonly', width=11, font=('Segoe UI', 9))
        month_combo.current(self.current_month - 1)
        month_combo.pack(side='left', padx=(0,15))
        
        Button(excel_frame, text='📥 Load Excel', command=self._load_volumes_from_excel,
               bg='#16a085', fg='white', font=('Segoe UI', 10, 'bold'), padx=15).pack(side='left', padx=2)
        Button(excel_frame, text='🔧 Excel Setup', command=self._open_excel_setup_unified,
               bg='#e67e22', fg='white', font=('Segoe UI', 10, 'bold'), padx=15).pack(side='left', padx=2)
        
        # Spacer
        Frame(excel_frame, width=20, bg='#2c3e50').pack(side='left')
        
        Button(excel_frame, text='⚖️ Balance Check', command=self._open_balance_check,
               bg='#9b59b6', fg='white', font=('Segoe UI', 10, 'bold'), padx=15).pack(side='left', padx=2)

        # Simple help text
        info = Label(controls, 
                    text='Drag to move • Click to select • Right-click for options',
                    font=('Segoe UI', 8), bg='#2c3e50', fg='#95a5a6', pady=3)
        info.pack()

        canvas_frame = Frame(self.parent)
        canvas_frame.pack(fill='both', expand=True, padx=5, pady=5)

        self.canvas = Canvas(canvas_frame, bg='#f5f6f7', highlightthickness=0, cursor='hand2')
        vscroll = Scrollbar(canvas_frame, orient='vertical', command=self.canvas.yview)
        hscroll = Scrollbar(canvas_frame, orient='horizontal', command=self.canvas.xview)

        self.canvas.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)
        self.canvas.grid(row=0, column=0, sticky='nsew')
        vscroll.grid(row=0, column=1, sticky='ns')
        hscroll.grid(row=1, column=0, sticky='ew')

        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)
        # Initial scrollregion; will be tightened in _draw_diagram
        self.canvas.configure(scrollregion=(0, 0, 2000, 1200))

        # Event bindings
        self.canvas.bind('<Button-1>', self._on_canvas_click)
        self.canvas.bind('<B1-Motion>', self._on_canvas_drag)
        self.canvas.bind('<ButtonRelease-1>', self._on_canvas_release)
        self.canvas.bind('<Button-3>', self._on_canvas_right_click)
        self.canvas.bind('<Motion>', self._on_canvas_motion)

    def _toggle_snap_grid(self):
        # Feature removed per request
        messagebox.showinfo('Grid Snap', 'Snap to grid feature removed')
    
    def _toggle_show_grid(self):
        self.show_grid = not self.show_grid
        self._draw_diagram()
        state = '✅ Visible' if self.show_grid else '❌ Hidden'
        messagebox.showinfo('Grid Display', f'Grid lines {state}')
    
    def _toggle_lock_selected(self):
        """Toggle lock for selected node or recirculation flow."""
        # Check if a recirculation flow is selected (by clicking on it)
        if hasattr(self, '_selected_recirculation_idx') and self._selected_recirculation_idx is not None:
            edge_idx = self._selected_recirculation_idx
            current_state = self.recirculation_locked.get(edge_idx, False)
            self.recirculation_locked[edge_idx] = not current_state
            
            new_state = '🔒 LOCKED' if self.recirculation_locked[edge_idx] else '🔓 UNLOCKED'
            edge = self.area_data.get('edges', [])[edge_idx]
            from_name = self._format_node_name(edge['from'])
            to_name = self._format_node_name(edge['to'])
            
            logger.info(f"Toggled lock for recirculation flow '{from_name} → {to_name}': {new_state}")
            messagebox.showinfo('Lock Status', f'Recirculation Flow:\n{from_name} → {to_name}\n\n{new_state}')
            self._save_to_json()
            self._draw_diagram()
            return
        
        # Otherwise, handle node locking
        if not self.selected_node:
            messagebox.showwarning('No Selection', 
                                 'Select a component or recirculation flow first.\n\n' +
                                 'Click on a component or flow line to select it.')
            return
        
        # Toggle lock state for node
        current_state = self.locked_nodes.get(self.selected_node, False)
        self.locked_nodes[self.selected_node] = not current_state
        
        new_state = '🔒 LOCKED' if self.locked_nodes[self.selected_node] else '🔓 UNLOCKED'
        logger.info(f"Toggled lock for node '{self.selected_node}': {new_state}")
        messagebox.showinfo('Lock Status', f'Component:\n{self.selected_node}\n\n{new_state}')
        self._save_to_json()
        self._draw_diagram()

    def _align_to_grid(self):
        messagebox.showinfo('Align', 'Align to grid feature removed')

    def _find_nearest_anchor(self, node_id, click_x, click_y):
        """Find nearest anchor point on component, snap if within snap_distance"""
        if node_id not in self.snap_anchor_points:
            return None
        
        anchors = self.snap_anchor_points[node_id]
        min_dist = float('inf')
        best_anchor = None
        
        for anchor_name, (ax, ay) in anchors.items():
            dist = ((click_x - ax)**2 + (click_y - ay)**2)**0.5
            if dist < min_dist:
                min_dist = dist
                best_anchor = (ax, ay)
        
        # Only snap if within snap_distance
        if min_dist <= self.snap_distance:
            logger.debug(f"🔧 Snapped to anchor: distance={min_dist:.1f}px")
            return best_anchor
        
        return None

    def _get_node_area(self, node_id):
        """Determine which area a node belongs to based on its y-position"""
        # Handle junction connections (not real nodes)
        if node_id.startswith('junction_'):
            return "Junction"
        
        if node_id not in self.nodes_by_id:
            return "Unknown"
        
        node = self.nodes_by_id[node_id]
        node_y = node['y']
        
        # Check zone backgrounds to determine area
        zone_bgs = self.area_data.get('zone_bg', [])
        for zone in zone_bgs:
            zone_y = zone.get('y', 0)
            zone_height = zone.get('height', 0)
            if zone_y <= node_y < (zone_y + zone_height):
                return zone.get('name', 'Unknown')
        
        # Fallback: check node_id prefix patterns
        node_id_lower = node_id.lower()
        if 'ug2n' in node_id_lower or node_id_lower.startswith('ug2_'):
            return "UG2 North Decline Area"
        elif 'meren' in node_id_lower and 'north' in node_id_lower:
            return "Merensky North Area"
        elif 'stockpile' in node_id_lower or 'spcd' in node_id_lower:
            return "Stockpile Area"
        elif 'ug2s' in node_id_lower or 'ug2_south' in node_id_lower:
            return "UG2 Main Decline Area"
        elif 'meren' in node_id_lower or 'mers' in node_id_lower:
            return "Merensky South Area"
        elif 'oldtsf' in node_id_lower or 'old_tsf' in node_id_lower:
            return "Old TSF Area"
        
        return "Other"

    def _format_node_name(self, node_id):
        """Format node name for display - remove prefix and make readable"""
        # Friendly name for virtual junctions
        if isinstance(node_id, str) and node_id.startswith('junction_'):
            parts = node_id.split('_')
            if len(parts) >= 4:
                return f"Junction ({parts[-2]}, {parts[-1]})"
            return "Junction"
        if node_id not in self.nodes_by_id:
            return node_id
        
        node = self.nodes_by_id[node_id]
        label = node.get('label', node_id)
        
        # Replace newlines with spaces for better display
        label = label.replace('\n', ' ')
        
        return label

    def _update_connected_edges(self, node_id, dx, dy):
        """Update all edges connected to a moved component"""
        edges = self.area_data.get('edges', [])
        
        for edge in edges:
            updated = False
            segments = edge.get('segments', [])
            
            if not segments or len(segments) < 2:
                continue
            
            # Update first point if this is the FROM node
            if edge.get('from') == node_id:
                old_x, old_y = segments[0]
                segments[0] = [old_x + dx, old_y + dy]
                updated = True
            
            # Update last point if this is the TO node
            if edge.get('to') == node_id:
                old_x, old_y = segments[-1]
                segments[-1] = [old_x + dx, old_y + dy]
                updated = True
            
            if updated:
                # Clear manually-positioned label so it follows the moved line
                if 'label_pos' in edge:
                    del edge['label_pos']
                logger.debug(f"📌 Updated edge {edge.get('from')} → {edge.get('to')}")

    def _straighten_edge(self, edge_idx):
        """Straighten a flow line by creating direct path from start to end"""
        edges = self.area_data.get('edges', [])
        if edge_idx >= len(edges):
            return
        
        edge = edges[edge_idx]
        segments = edge.get('segments', [])
        
        if len(segments) < 2:
            return
        
        # Keep only first and last points
        start_point = segments[0]
        end_point = segments[-1]
        
        # Create straight line
        edge['segments'] = [start_point, end_point]
        
        logger.info(f"📏 Straightened edge {edge.get('from')} → {edge.get('to')}")
        self._draw_diagram()
        messagebox.showinfo("Straightened", f"Flow line from {edge.get('from')} to {edge.get('to')} is now straight")

    def _draw_snap_anchors_in_drawing_mode(self):
        """Draw anchor points for all components when in drawing mode"""
        if not self.drawing_mode:
            return
        
        for node_id, anchors in self.snap_anchor_points.items():
            # Draw anchor points as small circles
            for anchor_name, (ax, ay) in anchors.items():
                # Skip center anchor to reduce clutter
                if anchor_name == 'center':
                    continue
                
                radius = 4
                # Check if this is the hovered anchor
                is_hovered = (anchor_name == self.hovered_anchor.get('name') and 
                             self.hovered_anchor.get('node_id') == node_id) if self.hovered_anchor else False
                
                color = '#f39c12' if is_hovered else '#95a5a6'  # Orange if hovered, gray otherwise
                self.canvas.create_oval(ax-radius, ay-radius, ax+radius, ay+radius, 
                                       fill=color, outline='#2c3e50', width=1)

    def _load_diagram_data(self):
        """Load diagram JSON - loads the master UG2N diagram which contains all areas"""
        import os
        # Resolve diagram path: EXE mode → WATERBALANCE_USER_DIR, Dev mode → project data folder
        user_dir = os.environ.get('WATERBALANCE_USER_DIR')
        if user_dir:
            # EXE/deployment mode: use user data directory
            self.json_file = Path(user_dir) / 'data' / 'diagrams' / 'ug2_north_decline.json'
        else:
            # Development mode: use project source tree
            self.json_file = Path(__file__).parent.parent.parent / 'data' / 'diagrams' / 'ug2_north_decline.json'
        
        if not self.json_file.exists():
            messagebox.showerror("Error", f"Master diagram not found: {self.json_file}")
            return

        try:
            with open(self.json_file, 'r', encoding='utf-8') as f:
                self.area_data = json.load(f)
            
            # Reset all flow volumes to "-" (load from Excel on demand)
            edges = self.area_data.get('edges', [])
            for edge in edges:
                edge['volume'] = '-'
                edge['label'] = '-'
            
            logger.info(f"✅ Loaded: {self.area_data.get('title')} - Master diagram with all areas")
            self._draw_diagram()
        except Exception as e:
            logger.error(f"❌ Load error: {e}")
            messagebox.showerror("Error", f"Failed to load: {e}")

    def _validate_excel_mapping(self):
        """Validate that each edge's excel_mapping references an existing sheet and column."""
        edges = self.area_data.get('edges', [])
        if not edges:
            messagebox.showinfo('Validate', 'No edges to validate')
            return
        excel_path = getattr(self.flow_loader, 'excel_path', None)
        if not excel_path or not Path(excel_path).exists():
            messagebox.showerror(
                'Excel File Missing',
                f"Flow Diagram Excel file not found:\n\n{excel_path}\n\n"
                "This file contains flow volumes for the diagram visualization.\n"
                "Please configure the path in Settings > Data Sources (timeseries_excel_path)."
            )
            return

        def _collect_issues(wb):
            sheets = set(wb.sheetnames)
            print(f"\n[DEBUG] Available sheets: {sheets}")

            # Use the FlowVolumeLoader to detect columns consistently
            loader = get_flow_volume_loader()
            header_cache = {}
            for sheet in sheets:
                try:
                    cols = set(loader.list_sheet_columns(sheet))
                    header_cache[sheet] = cols
                    print(f"[DEBUG] {sheet} has {len(cols)} detected columns (loader)")
                    if sheet.startswith('Flows_'):
                        sample = list(cols)[:3]
                        print(f"[DEBUG] {sheet} sample columns: {sample}")
                except Exception:
                    header_cache[sheet] = set()

            problems_local = []
            missing_local = set()
            for edge in edges:
                mapping = edge.get('excel_mapping', {})
                if not mapping or not mapping.get('enabled'):
                    continue
                sheet = mapping.get('sheet') or f"Flows_{self._get_area_code_from_title()}"
                column = mapping.get('column')
                
                if not sheet or sheet not in sheets:
                    problems_local.append(f"Missing sheet: {sheet} for {edge.get('from')} → {edge.get('to')}")
                    continue
                    
                cols = header_cache.get(sheet, set())
                if not column or column not in cols:
                    print(f"[DEBUG] Missing: sheet={sheet}, column={column}")
                    print(f"[DEBUG]   Looking for: '{column}'")
                    print(f"[DEBUG]   Available columns count: {len(cols)}")
                    problems_local.append(f"Missing column: {column} in {sheet} for {edge.get('from')} → {edge.get('to')}")
                    missing_local.add((sheet, column, edge.get('from'), edge.get('to')))
            return problems_local, missing_local

        def _sync_missing_to_labels(missing_items):
            """Attempt to repair missing column mappings using current component labels."""
            self._load_diagram_data()
            nodes = self.area_data.get('nodes', [])
            id_to_label = {n.get('id'): (n.get('label') or n.get('id')) for n in nodes}

            excel_path = getattr(self.flow_loader, 'excel_path', None)
            if not excel_path or not Path(excel_path).exists():
                messagebox.showerror(
                    'Excel File Missing',
                    f"Flow Diagram Excel file not found:\n\n{excel_path}\n\n"
                    "This file contains flow volumes for the diagram visualization.\n"
                    "Please configure the path in Settings > Data Sources (timeseries_excel_path)."
                )
                return 0

            wb = load_workbook(excel_path)
            headers_cache = {sheet: {cell.value: cell for cell in wb[sheet][3] if cell.value}
                             for sheet in wb.sheetnames}
            repaired = 0

            def _norm(name: str) -> str:
                return ''.join(ch for ch in name.lower().replace('→', '_').replace('-', '_').replace(' ', '_') if ch.isalnum() or ch == '_')

            for edge in edges:
                mapping = edge.get('excel_mapping', {}) or {}
                if not mapping.get('enabled'):
                    continue
                sheet = mapping.get('sheet') or f"Flows_{self._get_area_code_from_title()}"
                if sheet not in headers_cache:
                    continue
                column = mapping.get('column')
                headers = headers_cache[sheet]

                # Only attempt repair if this mapping was missing
                key = (sheet, column, edge.get('from'), edge.get('to'))
                if key not in missing_items:
                    continue

                if column in headers:
                    continue

                from_label = id_to_label.get(edge.get('from', ''), edge.get('from', ''))
                to_label = id_to_label.get(edge.get('to', ''), edge.get('to', ''))
                proposed = f"{from_label} → {to_label}"

                new_col = None
                if proposed in headers:
                    new_col = proposed
                else:
                    norm_target = _norm(proposed)
                    for col_name in headers.keys():
                        if _norm(col_name) == norm_target:
                            new_col = col_name
                            break

                if new_col:
                    mapping['column'] = new_col
                    edge['excel_mapping'] = mapping
                    repaired += 1

            if repaired:
                with open(self.json_file, 'w', encoding='utf-8') as f:
                    json.dump(self.area_data, f, indent=2, ensure_ascii=False)
                wb.save(excel_path)
            wb.close()
            return repaired

        workbook = load_workbook(excel_path)
        problems, missing_items = _collect_issues(workbook)

        if not problems:
            workbook.close()
            messagebox.showinfo('Excel Mapping', 'All mappings look good ✅')
            return

        attempt = messagebox.askyesno(
            'Excel Mapping Issues',
            '\n'.join(problems[:25]) + '\n\nAttempt auto-repair using current labels?'
        )

        if attempt:
            repaired = _sync_missing_to_labels(missing_items)
            workbook.close()

            if repaired:
                workbook = load_workbook(excel_path)
                problems, _ = _collect_issues(workbook)

        if problems:
            messagebox.showwarning('Excel Mapping Issues', '\n'.join(problems[:25]))

        workbook.close()

    def _draw_diagram(self):
        """Draw diagram"""
        if not self.area_data:
            return

        self.canvas.delete('all')
        self.node_items = {}
        self.nodes_by_id = {}
        self.label_items = {}  # Reset label tracking (canvas_item_id -> edge_idx)
        self.edge_to_label_items = {}  # Reverse mapping (edge_idx -> [box_id, text_id])

        # Update scroll region dynamically based on diagram size (adds padding for panning)
        # Tighten scroll region to actual content bounds to avoid large empty space
        nodes = self.area_data.get('nodes', [])
        max_x = max([(n.get('x',0) + n.get('width',0)) for n in nodes] + [0])
        max_y = max([(n.get('y',0) + n.get('height',0)) for n in nodes] + [0])
        pad_x = 150
        pad_y = 150
        scroll_w = max_x + pad_x
        scroll_h = max_y + pad_y
        self.canvas.configure(scrollregion=(0, 0, scroll_w, scroll_h))

        # Ensure outflow area stays near view: adjust canvas size subtly

        # Draw zone backgrounds (multiple areas)
        zone_bgs = self.area_data.get('zone_bg', None)
        if zone_bgs:
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    self.canvas.create_rectangle(
                        zone.get('x', 0), zone.get('y', 0),
                        zone.get('x', 0) + zone.get('width', 1800),
                        zone.get('y', 0) + zone.get('height', 420),
                        fill=zone.get('color', '#f5f6fa'), outline='', tags='zone_bg')
            else:
                # fallback for single zone_bg dict
                zone = zone_bgs
                self.canvas.create_rectangle(
                    zone.get('x', 0), zone.get('y', 0),
                    zone.get('x', 0) + zone.get('width', 1800),
                    zone.get('y', 0) + zone.get('height', 900),
                    fill=zone.get('color', '#f5f6fa'), outline='', tags='zone_bg')

        # Draw grid lines if enabled
        if self.show_grid:
            canvas_width = int(scroll_w)
            canvas_height = int(scroll_h)
            for x in range(0, canvas_width, self.grid_size):
                self.canvas.create_line(x, 0, x, canvas_height, fill='#e0e0e0', width=1, tags='grid')
            for y in range(0, canvas_height, self.grid_size):
                self.canvas.create_line(0, y, canvas_width, y, fill='#e0e0e0', width=1, tags='grid')
            # Thicker lines every 100px
            for x in range(0, canvas_width, 100):
                self.canvas.create_line(x, 0, x, canvas_height, fill='#c5d3e6', width=2, tags='grid')
            for y in range(0, canvas_height, 100):
                self.canvas.create_line(0, y, canvas_width, y, fill='#c5d3e6', width=2, tags='grid')

        # Title for UG2 North Decline Area
        title = self.area_data.get('title', 'Flow Diagram')
        self.canvas.create_text(50, 15, text=title, font=('Segoe UI', 14, 'bold'), 
                               fill='#2c3e50', anchor='nw')

        # Title for Merensky North Area (if present)
        merensky_title = self.area_data.get('merensky_title', None)
        if merensky_title:
            # Find the y-position of the Merensky zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            merensky_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    if zone.get('name', '').lower().startswith('merensky'):
                        merensky_zone = zone
                        break
            if merensky_zone:
                y = merensky_zone.get('y', 470) + 10
            else:
                y = 480
            self.canvas.create_text(50, y, text=merensky_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Title for Stockpile Area (if present)
        stockpile_title = self.area_data.get('stockpile_title', None)
        if stockpile_title:
            # Find the y-position of the Stockpile zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            stockpile_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    if zone.get('name', '').lower().startswith('stockpile'):
                        stockpile_zone = zone
                        break
            if stockpile_zone:
                y = stockpile_zone.get('y', 900) + 10
            else:
                y = 910
            self.canvas.create_text(50, y, text=stockpile_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Title for UG2 Main Decline Area (if present)
        ug2south_title = self.area_data.get('ug2south_title', None)
        if ug2south_title:
            # Find the y-position of the UG2 Main Decline Area zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            ug2south_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    zone_name = zone.get('name', '').lower()
                    if zone_name.startswith('ug2 main decline') or zone_name.startswith('ug2 south'):
                        ug2south_zone = zone
                        break
            if ug2south_zone:
                y = ug2south_zone.get('y', 930) + 10
            else:
                y = 940
            self.canvas.create_text(50, y, text=ug2south_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Title for Merensky South Area (if present)
        merenskysouth_title = self.area_data.get('merenskysouth_title', None)
        if merenskysouth_title:
            # Find the y-position of the Merensky South zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            merenskysouth_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    if zone.get('name', '').lower().startswith('merensky south'):
                        merenskysouth_zone = zone
                        break
            if merenskysouth_zone:
                y = merenskysouth_zone.get('y', 1640) + 10
            else:
                y = 1650
            self.canvas.create_text(50, y, text=merenskysouth_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Title for Old TSF Area (if present)
        oldtsf_title = self.area_data.get('oldtsf_title', None)
        if oldtsf_title:
            # Find the y-position of the Old TSF zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            oldtsf_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    if zone.get('name', '').lower().startswith('old tsf'):
                        oldtsf_zone = zone
                        break
            if oldtsf_zone:
                y = oldtsf_zone.get('y', 2070) + 10
            else:
                y = 2080
            self.canvas.create_text(50, y, text=oldtsf_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Title for New TSF Area (if present)
        newtsf_title = self.area_data.get('newtsf_title', None)
        if newtsf_title:
            # Find the y-position of the New TSF zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            newtsf_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    if zone.get('name', '').lower().startswith('new tsf'):
                        newtsf_zone = zone
                        break
            if newtsf_zone:
                y = newtsf_zone.get('y', 2000) + 10
            else:
                y = 2010
            self.canvas.create_text(50, y, text=newtsf_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Title for UG2 Plant Area (if present)
        ug2plant_title = self.area_data.get('ug2plant_title', None)
        if ug2plant_title:
            # Find the y-position of the UG2 Plant zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            ug2plant_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    if zone.get('name', '').lower().startswith('ug2 plant'):
                        ug2plant_zone = zone
                        break
            if ug2plant_zone:
                y = ug2plant_zone.get('y', 2650) + 10
            else:
                y = 2660
            self.canvas.create_text(50, y, text=ug2plant_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Title for Merensky Plant Area (if present)
        merplant_title = self.area_data.get('merplant_title', None)
        if merplant_title:
            # Find the y-position of the Merensky Plant zone background
            zone_bgs = self.area_data.get('zone_bg', [])
            merplant_zone = None
            if isinstance(zone_bgs, list):
                for zone in zone_bgs:
                    if zone.get('name', '').lower().startswith('merensky plant'):
                        merplant_zone = zone
                        break
            if merplant_zone:
                y = merplant_zone.get('y', 3230) + 10
            else:
                y = 3240
            self.canvas.create_text(50, y, text=merplant_title, font=('Segoe UI', 14, 'bold'),
                                   fill='#2c3e50', anchor='nw')

        # Section labels
        self.canvas.create_text(100, 60, text='INFLOWS', font=('Segoe UI', 11, 'bold'), 
                               fill='#2980b9', anchor='center')
        self.canvas.create_text(1380, 60, text='OUTFLOWS', font=('Segoe UI', 11, 'bold'), 
                               fill='#e74c3c', anchor='center')

        # Build node lookup and load lock states
        nodes = self.area_data.get('nodes', [])
        for node in nodes:
            self.nodes_by_id[node['id']] = node
            # Load lock state from node data
            if 'locked' in node:
                self.locked_nodes[node['id']] = node['locked']

        # Draw nodes first, then edges, then raise edges so arrowheads stay visible
        # Draw nodes
        for node in nodes:
            self._draw_node(node)

        # Draw edges (manual segments - independent from nodes)
        edges = self.area_data.get('edges', [])
        for idx, edge in enumerate(edges):
            self._draw_edge_segments(edge, idx)

        # Ensure flow lines and labels render above nodes (important for dams/ovals)
        self.canvas.tag_raise('flow_line')
        self.canvas.tag_raise('flow_label')

        # Draw current drawing segments if in drawing mode
        if self.drawing_mode and len(self.drawing_segments) > 0:
            for i in range(len(self.drawing_segments) - 1):
                x1, y1 = self.drawing_segments[i]
                x2, y2 = self.drawing_segments[i + 1]
                self.canvas.create_line(x1, y1, x2, y2, fill='#3498db', width=3, 
                                       dash=(5, 3), arrow='last', arrowshape=(12, 15, 6))
        
        # Draw snap anchor points when in drawing mode
        self._draw_snap_anchors_in_drawing_mode()

        logger.info(f"Drew {len(nodes)} components and {len(edges)} flows")
        logger.debug(f"Tracking {len(self.label_items)} label items: {self.label_items}")

    def _draw_node(self, node):
        """Draw component"""
        x, y = node['x'], node['y']
        width, height = node['width'], node['height']
        fill = node.get('fill', '#dfe6ed')
        outline = node.get('outline', '#2c3e50')
        label = node.get('label', '')
        node_id = node['id']

        # Check if locked
        is_locked = self.locked_nodes.get(node_id, False)
        is_selected = self.selected_node == node_id
        
        # Determine outline
        if is_locked:
            outline_color = '#c0392b'  # Red for locked
            outline_width = 4
        elif is_selected:
            outline_color = '#e74c3c'  # Light red for selected
            outline_width = 3
        else:
            outline_color = outline  # Normal
            outline_width = 2

        if node.get('shape') == 'oval':
            item = self.canvas.create_oval(x, y, x+width, y+height, fill=fill, 
                                          outline=outline_color, width=outline_width)
        else:
            item = self.canvas.create_rectangle(x, y, x+width, y+height, fill=fill, 
                                               outline=outline_color, width=outline_width)

        self.node_items[item] = node_id

        # Calculate and store anchor points for snap-to functionality
        # Increased density: quarters along each side + corners + midpoints + center
        cx = x + width / 2
        cy = y + height / 2
        qx1 = x + width * 0.25
        qx3 = x + width * 0.75
        qy1 = y + height * 0.25
        qy3 = y + height * 0.75

        anchors = {
            'center': (cx, cy),
            # Top edge
            'top_left': (x, y),
            'top_q1': (qx1, y),
            'top': (cx, y),
            'top_q3': (qx3, y),
            'top_right': (x + width, y),
            # Bottom edge
            'bottom_left': (x, y + height),
            'bottom_q1': (qx1, y + height),
            'bottom': (cx, y + height),
            'bottom_q3': (qx3, y + height),
            'bottom_right': (x + width, y + height),
            # Left edge
            'left_top_q1': (x, qy1),
            'left': (x, cy),
            'left_bottom_q3': (x, qy3),
            # Right edge
            'right_top_q1': (x + width, qy1),
            'right': (x + width, cy),
            'right_bottom_q3': (x + width, qy3)
        }
        self.snap_anchor_points[node_id] = anchors

        # Get font properties from node (or use defaults)
        font_size = node.get('font_size', None)  # None means use proportional
        font_weight = node.get('font_weight', 'normal')
        
        # Labels - scale font size based on component dimensions
        lines = label.split('\n') if label else []
        
        # Calculate proportional font sizes based on component width and height
        # Use smaller of the two to determine scale factor
        min_dimension = min(width, height)
        
        # Determine font sizes to use
        if font_size is not None:
            # User explicitly set font size - use it (round to integer for Tkinter)
            primary_font_size = round(font_size)
            secondary_font_size = max(5, round(font_size) - 2)
            type_font_size = max(4, round(font_size) - 4)
        else:
            # No explicit font size - calculate proportionally
            # Font size formula: scale between 6 and 10pt based on component size
            # Small components (40px): ~6pt, Large components (150px+): ~10pt
            primary_font_size = max(6, min(10, int(min_dimension / 15)))
            secondary_font_size = max(5, min(8, int(min_dimension / 18)))
            type_font_size = max(4, min(6, int(min_dimension / 25)))
        
        # Adjust line spacing based on height
        line_spacing = max(8, int(height / 6))
        
        for idx, line in enumerate(lines):
            # Truncate text if it's too long for the component width
            max_chars = max(3, int(width / 7))
            if len(line) > max_chars:
                line = line[:max_chars-1] + '…'
            
            # Apply font weight to all lines
            font = ('Segoe UI', primary_font_size, font_weight) if idx == 0 else ('Segoe UI', secondary_font_size, font_weight)
            
            # Adjust vertical position based on number of lines and component height
            if len(lines) == 1:
                text_y = y + height / 2  # Center for single line
            else:
                text_y = y + 10 + (idx * line_spacing)
            
            self.canvas.create_text(x + width/2, text_y, text=line, 
                                   font=font, fill='#000', anchor='center')

        # Add lock icon if locked (scale icon size too)
        if is_locked:
            lock_x = x + width - 10
            lock_y = y + 8
            lock_icon_size = max(8, min(12, int(min_dimension / 12)))
            self.canvas.create_text(lock_x, lock_y, text='🔒', font=('Segoe UI', lock_icon_size),
                                   fill='#c0392b', tags=f'lock_{node_id}')

    def _get_edge_connection_point(self, node, click_x, click_y):
        """Get connection point from user's exact click location on node.
        
        Args:
            node: Node dictionary with x, y, width, height
            click_x: X coordinate where user clicked
            click_y: Y coordinate where user clicked
            
        Returns:
            Tuple of (x, y) coordinates at click location
        """
        # Return exact click point - user has full control
        return click_x, click_y

    def _draw_recirculation_icon(self, edge, edge_idx):
        """Draw recirculation loop as an outline rectangle with label inside - draggable (unless locked)"""
        color = edge.get('color', '#9b59b6')
        label = edge.get('label', '')
        component_id = edge.get('from')
        
        if component_id not in self.nodes_by_id:
            return
        
        node = self.nodes_by_id[component_id]
        
        # Get recirculation position or default to right of component
        loop_data = edge.get('recirculation_pos', {})
        if isinstance(loop_data, dict) and 'x' in loop_data and 'y' in loop_data:
            cx = float(loop_data['x'])
            cy = float(loop_data['y'])
        else:
            # Default position: to the right of the component, middle height
            cx = node['x'] + node['width'] + 35
            cy = node['y'] + node['height'] / 2
        
        # Get customizable box size (default: 40x28 pixels)
        rect_width = edge.get('box_width', 40)
        rect_height = edge.get('box_height', 28)
        x1 = cx - rect_width/2
        y1 = cy - rect_height/2
        x2 = cx + rect_width/2
        y2 = cy + rect_height/2
        
        # Check if this recirculation is locked
        is_locked = self.recirculation_locked.get(edge_idx, False)
        outline_color = color if not is_locked else '#7f8c8d'  # Gray if locked
        dash_pattern = () if not is_locked else (4, 2)  # Dashed if locked
        
        rect_id = self.canvas.create_rectangle(x1, y1, x2, y2, fill='', outline=outline_color, 
                                               width=2, dash=dash_pattern, tags=('flow_line', 'recirculation'))
        
        # Draw label text directly inside the rectangle
        if label:
            # Add lock symbol if locked
            display_label = label if not is_locked else f"🔒 {label}"
            label_font_size = edge.get('label_font_size', 8.0)
            text_id = self.canvas.create_text(cx, cy, text=display_label,
                                             font=('Segoe UI', round(label_font_size), 'bold'), fill=outline_color,
                                             anchor='center', tags=('flow_label', 'recirculation_label'))
            
            # Store for dragging - mark this rectangle as draggable recirculation (if not locked)
            if not is_locked:
                self.label_items[rect_id] = edge_idx
                self.label_items[text_id] = edge_idx
                self.edge_to_label_items[edge_idx] = [rect_id, text_id]


    def _draw_edge_segments(self, edge, edge_idx):
        """Draw flow line with manual control and proper arrow placement"""
        segments = edge.get('segments', [])
        color = edge.get('color', '#3498db')
        label = edge.get('label', '')
        from_id = edge.get('from')
        to_id = edge.get('to')
        is_junction = edge.get('is_junction', False)
        is_junction_start = bool(edge.get('is_junction_start', False))
        
        # Check if this is a recirculation loop - render as special icon instead of line
        is_recirculation = edge.get('is_recirculation', False)
        if is_recirculation:
            self._draw_recirculation_icon(edge, edge_idx)
            return
        
        # Determine start endpoint: component or junction
        from_node = None
        junction_start_pos = None
        if is_junction_start or (isinstance(from_id, str) and from_id.startswith('junction_')):
            # Start from a junction; use stored coordinates
            jpos = edge.get('junction_start_pos')
            if isinstance(jpos, dict) and 'x' in jpos and 'y' in jpos:
                junction_start_pos = (float(jpos['x']), float(jpos['y']))
            else:
                # If segments exist, trust first segment as start
                if len(segments) >= 1:
                    junction_start_pos = (segments[0][0], segments[0][1])
            if junction_start_pos is None:
                return
        else:
            # Validate from component exists
            if from_id not in self.nodes_by_id:
                return
            from_node = self.nodes_by_id[from_id]
        
        # Handle junction connections (line-to-line) vs component connections
        if is_junction:
            # Endpoint is on another flow line, not a component
            junction_pos = edge.get('junction_pos', {})
            if not junction_pos:
                return  # Invalid junction
            to_node = None
        else:
            # Normal component-to-component connection
            if to_id not in self.nodes_by_id:
                return
            to_node = self.nodes_by_id[to_id]
        
        # Use stored segment points directly - full manual control
        if len(segments) >= 2:
            # User-defined start and end points
            from_x, from_y = segments[0]
            to_x, to_y = segments[-1]
        else:
            # Fallback to center if no segments stored
            from_x = from_node['x'] + from_node['width'] / 2
            from_y = from_node['y'] + from_node['height'] / 2
            if is_junction:
                # Junction endpoint uses stored coordinates
                junction_pos = edge.get('junction_pos', {})
                to_x = junction_pos.get('x', from_x)
                to_y = junction_pos.get('y', from_y)
            else:
                to_x = to_node['x'] + to_node['width'] / 2
                to_y = to_node['y'] + to_node['height'] / 2

        # Calculate intersection with component edge for precise arrow head placement
        def get_edge_intersection(x1, y1, x2, y2, node, arrow_pad=10):
            """Calculate where line intersects component boundary and pull back by arrow_pad"""
            # Rectangle bounds
            left = node['x']
            right = node['x'] + node['width']
            top = node['y']
            bottom = node['y'] + node['height']
            
            # Handle oval shapes
            is_oval = node.get('shape') == 'oval'
            
            if is_oval:
                # For ovals, calculate accurate line-ellipse intersection
                cx = (left + right) / 2
                cy = (top + bottom) / 2
                rx = node['width'] / 2
                ry = node['height'] / 2

                dx = x2 - x1
                dy = y2 - y1
                if abs(dx) < 1e-6 and abs(dy) < 1e-6:
                    return x2, y2

                # Solve quadratic for t where (x1+dx*t, y1+dy*t) hits ellipse boundary
                # ((x-cx)^2)/rx^2 + ((y-cy)^2)/ry^2 = 1
                a = (dx*dx)/(rx*rx) + (dy*dy)/(ry*ry)
                b = 2*((x1-cx)*dx/(rx*rx) + (y1-cy)*dy/(ry*ry))
                c = ((x1-cx)*(x1-cx))/(rx*rx) + ((y1-cy)*(y1-cy))/(ry*ry) - 1

                disc = b*b - 4*a*c
                if disc < 0 or abs(a) < 1e-12:
                    return x2, y2

                sqrt_disc = disc ** 0.5
                t1 = (-b - sqrt_disc) / (2*a)
                t2 = (-b + sqrt_disc) / (2*a)
                candidates_t = [t for t in (t1, t2) if t > 0]
                if not candidates_t:
                    return x2, y2

                t_hit = min(candidates_t)
                intersect_x = x1 + dx * t_hit
                intersect_y = y1 + dy * t_hit

                # Pull back by arrow_pad to keep arrowhead outside the oval
                length = (dx*dx + dy*dy) ** 0.5
                if length > 1e-6 and arrow_pad > 0:
                    intersect_x -= dx / length * arrow_pad
                    intersect_y -= dy / length * arrow_pad

                return intersect_x, intersect_y
            else:
                # Rectangle intersection
                # Direction vector
                dx = x2 - x1
                dy = y2 - y1
                
                # Avoid division by zero
                if abs(dx) < 1e-6 and abs(dy) < 1e-6:
                    return x2, y2
                
                # Find intersection with each side
                candidates = []
                if abs(dx) > 1e-6:
                    # Left side
                    t = (left - x1) / dx
                    y = y1 + t * dy
                    if 0 < t <= 1 and top <= y <= bottom:
                        candidates.append((left, y, t))
                    # Right side
                    t = (right - x1) / dx
                    y = y1 + t * dy
                    if 0 < t <= 1 and top <= y <= bottom:
                        candidates.append((right, y, t))
                
                if abs(dy) > 1e-6:
                    # Top side
                    t = (top - y1) / dy
                    x = x1 + t * dx
                    if 0 < t <= 1 and left <= x <= right:
                        candidates.append((x, top, t))
                    # Bottom side
                    t = (bottom - y1) / dy
                    x = x1 + t * dx
                    if 0 < t <= 1 and left <= x <= right:
                        candidates.append((x, bottom, t))
                
                # Pick the closest intersection point
                if candidates:
                    # Sort by t parameter (closest to start point)
                    candidates.sort(key=lambda c: c[2])
                    ix, iy, _ = candidates[-1]
                    length = (dx*dx + dy*dy) ** 0.5
                    if length > 1e-6 and arrow_pad > 0:
                        ix -= dx / length * arrow_pad
                        iy -= dy / length * arrow_pad
                    return ix, iy
                
                return x2, y2

        # Handle all cases with segments
        if len(segments) >= 2:
            # Build path with adjusted endpoint for arrow
            path_points = []
            
            # Add all intermediate points
            for i, seg in enumerate(segments):
                if i < len(segments) - 1:
                    # All points except the last
                    path_points.extend(seg)
                else:
                    # For last segment, calculate intersection with target
                    if len(segments) >= 2:
                        # Get second-to-last point
                        prev_x, prev_y = segments[-2]
                        target_x, target_y = seg
                        
                        if is_junction:
                            # Junction endpoint - use exact coordinates
                            path_points.extend([target_x, target_y])
                        else:
                            # Component endpoint - calculate intersection with edge and keep arrowhead outside
                            intersect_x, intersect_y = get_edge_intersection(
                                prev_x, prev_y, target_x, target_y, to_node, arrow_pad=10
                            )
                            path_points.extend([intersect_x, intersect_y])
                    else:
                        path_points.extend(seg)
            # If starting from a junction, ensure the start point is the junction_start_pos
            if junction_start_pos is not None:
                # Replace the first coordinate pair with junction start to avoid relying on component center
                if len(path_points) >= 2:
                    path_points[0] = junction_start_pos[0]
                    path_points[1] = junction_start_pos[1]
            
            # Draw the polyline with arrow(s). Always show direction; enlarge arrowheads for visibility.
            arrowshape = (12, 15, 6)
            arrow_opt = 'both' if bool(edge.get('bidirectional', False)) else 'last'
            line_id = self.canvas.create_line(*path_points, fill=color, width=1.2,
                                   arrow=arrow_opt, arrowshape=arrowshape, smooth=False, tags=('flow_line',))
            # Ensure arrows render above component fills (especially dams/ovals)
            self.canvas.tag_raise(line_id)
            
            # Draw junction marker if this is a line-to-line connection
            if is_junction and len(path_points) >= 2:
                # Draw a small circle at the junction point to make it visible
                jx, jy = path_points[-2], path_points[-1]
                self.canvas.create_oval(jx-3, jy-3, jx+3, jy+3, fill=color, outline='white', width=1)
            
            # Draw label - prefer manually-positioned labels, otherwise calculate from segments
            if label:
                # Check if label has been manually positioned (dragged by user)
                label_pos = edge.get('label_pos')
                if isinstance(label_pos, dict) and 'x' in label_pos and 'y' in label_pos:
                    # Use stored position from manual drag
                    mid_x = float(label_pos['x'])
                    mid_y = float(label_pos['y'])
                else:
                    # Recalculate from current path_points to follow component movement
                    label_offset = edge.get('label_offset', 0.5)
                    total_segments = len(path_points) // 2 - 1
                    if total_segments > 0:
                        target_segment = int(label_offset * total_segments)
                        target_segment = max(0, min(target_segment, total_segments - 1))
                        seg_idx = target_segment * 2
                        mid_x = (path_points[seg_idx] + path_points[seg_idx + 2]) / 2
                        mid_y = (path_points[seg_idx + 1] + path_points[seg_idx + 3]) / 2
                    else:
                        # Fallback for single segment
                        mid_x = (path_points[0] + path_points[2]) / 2
                        mid_y = (path_points[1] + path_points[3]) / 2
                
                box_id = self.canvas.create_rectangle(mid_x - 30, mid_y - 8, mid_x + 30, mid_y + 8,
                                            fill='white', outline=color, width=1, tags='flow_label')
                label_font_size = edge.get('label_font_size', 7.0)
                text_id = self.canvas.create_text(mid_x, mid_y, text=label, font=('Segoe UI', round(label_font_size)),
                                       fill='#2c3e50', anchor='center', tags='flow_label')
                
                # Store edge index for this label so it can be dragged (both directions)
                self.label_items[box_id] = edge_idx
                self.label_items[text_id] = edge_idx
                self.edge_to_label_items[edge_idx] = [box_id, text_id]
                logger.debug(f"Tracked label for edge {edge_idx}: box={box_id}, text={text_id}")
            return
        
        # Fallback for edges with no segments - use node centers with intersection
        logger.warning(f"Edge {from_id} -> {to_id} has no segments, using node centers")
        
        # Calculate intersection for arrow placement
        intersect_x, intersect_y = get_edge_intersection(from_x, from_y, to_x, to_y, to_node, arrow_pad=10)

        arrowshape = (12, 15, 6)
        arrow_opt = 'both' if bool(edge.get('bidirectional', False)) else 'last'
        line_id = self.canvas.create_line(from_x, from_y, intersect_x, intersect_y,
                       fill=color, width=1.2, arrow=arrow_opt, arrowshape=arrowshape, tags=('flow_line',))
        self.canvas.tag_raise(line_id)
        if label:
            label_offset = edge.get('label_offset', 0.5)
            mid_x = from_x + (intersect_x - from_x) * label_offset
            mid_y = from_y + (intersect_y - from_y) * label_offset
            
            box_id = self.canvas.create_rectangle(mid_x - 30, mid_y - 8, mid_x + 30, mid_y + 8,
                                        fill='white', outline=color, width=1, tags='flow_label')
            label_font_size = edge.get('label_font_size', 7.0)
            text_id = self.canvas.create_text(mid_x, mid_y, text=label, font=('Segoe UI', round(label_font_size)),
                                   fill='#2c3e50', anchor='center', tags='flow_label')
            
            # Store edge index for this label so it can be dragged (both directions)
            self.label_items[box_id] = edge_idx
            self.label_items[text_id] = edge_idx
            self.edge_to_label_items[edge_idx] = [box_id, text_id]
            logger.debug(f"Tracked fallback label for edge {edge_idx}: box={box_id}, text={text_id}")

    def _start_redrawing(self):
        """Deprecated - replaced by _edit_line"""
        self._edit_line()

    def _straighten_line(self):
        """Deprecated - feature removed"""
        messagebox.showinfo("Straighten", "Straighten feature removed. Use Edit Line to manage paths and styles.")

    def _add_component(self):
        """Add a new component (node) to the diagram"""
        # Create add dialog
        dialog = self._create_styled_dialog("Add New Component", 550, 750)
        
        # Title header
        header = tk.Frame(dialog, bg='#27ae60', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="➕ Add New Component",
                 font=('Segoe UI', 14, 'bold'), bg='#27ae60', fg='white').pack(pady=18)
        
        # Form frame
        form = tk.Frame(dialog, bg='white', padx=25, pady=20)
        form.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Auto-generate unique component ID
        def generate_unique_id(base_label='component'):
            """Generate a unique ID based on label and counter"""
            nodes = self.area_data.get('nodes', [])
            existing_ids = {node['id'] for node in nodes}
            
            # Clean base label (remove spaces, special chars, convert to lowercase)
            base = base_label.strip().lower().replace(' ', '_').replace('-', '_')
            base = ''.join(c for c in base if c.isalnum() or c == '_')
            if not base:
                base = 'component'
            
            # Try base name first
            if base not in existing_ids:
                return base
            
            # Add counter suffix
            counter = 1
            while f"{base}_{counter}" in existing_ids:
                counter += 1
            return f"{base}_{counter}"
        
        # Component ID (auto-generated, read-only with refresh button)
        tk.Label(form, text="Component ID:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=8)
        id_var = tk.StringVar(value=generate_unique_id())
        
        id_frame = tk.Frame(form, bg='white')
        id_frame.grid(row=0, column=1, pady=8, padx=5, sticky='ew')
        id_entry = tk.Entry(id_frame, textvariable=id_var, font=('Segoe UI', 10), state='readonly', width=28)
        id_entry.pack(side='left', fill='x', expand=True)
        
        def refresh_id():
            label_text = label_var.get() if 'label_var' in locals() else 'component'
            id_var.set(generate_unique_id(label_text))
        
        refresh_btn = tk.Button(id_frame, text="🔄", command=refresh_id, bg='#3498db', fg='white',
                               font=('Segoe UI', 9), padx=6, relief='flat', cursor='hand2')
        refresh_btn.pack(side='left', padx=(4, 0))
        
        tk.Label(form, text="(auto-generated, click 🔄 to refresh)", font=('Segoe UI', 8, 'italic'), bg='white', fg='#7f8c8d').grid(row=1, column=1, sticky='w', padx=5)
        
        # Label (display name)
        tk.Label(form, text="Label:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=2, column=0, sticky='w', pady=8)
        label_var = tk.StringVar(value='NEW COMPONENT')
        label_entry = tk.Entry(form, textvariable=label_var, font=('Segoe UI', 10), width=35)
        label_entry.grid(row=2, column=1, pady=8, padx=5, sticky='ew')
        
        # Auto-update ID when label changes (with debounce)
        def on_label_change(*args):
            refresh_id()
        label_var.trace_add('write', on_label_change)
        
        # Position X
        tk.Label(form, text="Position X:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=3, column=0, sticky='w', pady=8)
        x_var = tk.IntVar(value=500)
        x_spin = tk.Spinbox(form, from_=0, to=2000, textvariable=x_var, font=('Segoe UI', 10), width=12)
        x_spin.grid(row=3, column=1, sticky='w', pady=8, padx=5)
        
        # Position Y
        tk.Label(form, text="Position Y:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=4, column=0, sticky='w', pady=8)
        y_var = tk.IntVar(value=500)
        y_spin = tk.Spinbox(form, from_=0, to=3500, textvariable=y_var, font=('Segoe UI', 10), width=12)
        y_spin.grid(row=4, column=1, sticky='w', pady=8, padx=5)
        
        # Width
        tk.Label(form, text="Width:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=5, column=0, sticky='w', pady=8)
        width_var = tk.IntVar(value=120)
        width_frame = tk.Frame(form, bg='white')
        width_frame.grid(row=5, column=1, sticky='w', pady=8, padx=5)
        width_spin = tk.Spinbox(width_frame, from_=40, to=400, textvariable=width_var, font=('Segoe UI', 10), width=10)
        width_spin.pack(side='left', padx=2)
        tk.Label(width_frame, text="px", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)
        
        # Height
        tk.Label(form, text="Height:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=6, column=0, sticky='w', pady=8)
        height_var = tk.IntVar(value=40)
        height_frame = tk.Frame(form, bg='white')
        height_frame.grid(row=6, column=1, sticky='w', pady=8, padx=5)
        height_spin = tk.Spinbox(height_frame, from_=20, to=200, textvariable=height_var, font=('Segoe UI', 10), width=10)
        height_spin.pack(side='left', padx=2)
        tk.Label(height_frame, text="px", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)
        
        # Type
        tk.Label(form, text="Type:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=7, column=0, sticky='w', pady=8)
        type_var = tk.StringVar(value='process')
        type_combo = ttk.Combobox(form, textvariable=type_var,
                                  values=['source', 'process', 'storage', 'consumption', 'building', 
                                         'treatment', 'plant', 'tsf', 'reservoir', 'loss', 'discharge'],
                                  state='readonly', font=('Segoe UI', 10), width=15)
        type_combo.grid(row=7, column=1, sticky='w', pady=8, padx=5)
        
        # Shape
        tk.Label(form, text="Shape:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=8, column=0, sticky='w', pady=8)
        shape_var = tk.StringVar(value='rect')
        shape_combo = ttk.Combobox(form, textvariable=shape_var, values=['rect', 'oval', 'diamond'],
                                   state='readonly', font=('Segoe UI', 10), width=15)
        shape_combo.grid(row=8, column=1, sticky='w', pady=8, padx=5)
        
        # Font size (supports decimals: 7.5, 10.25, etc.)
        tk.Label(form, text="Font Size:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=9, column=0, sticky='w', pady=8)
        font_size_var = tk.DoubleVar(value=10.0)
        font_size_frame = tk.Frame(form, bg='white')
        font_size_frame.grid(row=9, column=1, sticky='w', pady=8, padx=5)
        font_size_spin = tk.Spinbox(font_size_frame, from_=4.0, to=36.0, increment=0.5, textvariable=font_size_var, font=('Segoe UI', 10), width=8)
        font_size_spin.pack(side='left', padx=2)
        tk.Label(font_size_frame, text="pt", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)
        
        # Font weight (Bold/Regular)
        tk.Label(form, text="Font Style:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=10, column=0, sticky='w', pady=8)
        font_weight_var = tk.StringVar(value='normal')
        font_weight_frame = tk.Frame(form, bg='white')
        font_weight_frame.grid(row=10, column=1, sticky='w', pady=8, padx=5)
        
        bold_btn = tk.Button(font_weight_frame, text="B  Bold", 
                            bg='#95a5a6', fg='white', font=('Segoe UI', 9, 'bold'), 
                            padx=10, relief='raised')
        regular_btn = tk.Button(font_weight_frame, text="Regular",
                               bg='#3498db', fg='white', font=('Segoe UI', 9),
                               padx=10, relief='sunken')
        
        def toggle_bold():
            current = font_weight_var.get()
            new_weight = 'normal' if current == 'bold' else 'bold'
            font_weight_var.set(new_weight)
            bold_btn.config(bg='#3498db' if new_weight == 'bold' else '#95a5a6', relief='sunken' if new_weight == 'bold' else 'raised')
            regular_btn.config(bg='#95a5a6' if new_weight != 'bold' else '#3498db', relief='sunken' if new_weight != 'bold' else 'raised')
        
        def set_regular():
            font_weight_var.set('normal')
            bold_btn.config(bg='#95a5a6', relief='raised')
            regular_btn.config(bg='#3498db', relief='sunken')
        
        bold_btn.config(command=toggle_bold)
        regular_btn.config(command=set_regular)
        bold_btn.pack(side='left', padx=2)
        regular_btn.pack(side='left', padx=2)
        
        # Fill color with picker
        tk.Label(form, text="Fill Color:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=11, column=0, sticky='w', pady=8)
        fill_var = tk.StringVar(value='#3498db')
        fill_frame = tk.Frame(form, bg='white')
        fill_frame.grid(row=11, column=1, sticky='w', pady=8, padx=5)
        fill_preview = tk.Canvas(fill_frame, width=30, height=25, bg=fill_var.get(), highlightthickness=1, highlightbackground='#95a5a6')
        fill_preview.pack(side='left', padx=2)
        def pick_fill_color():
            from tkinter.colorchooser import askcolor
            color = askcolor(color=fill_var.get(), title="Choose Fill Color")
            if color[1]:
                fill_var.set(color[1])
                fill_preview.config(bg=color[1])
        fill_btn = tk.Button(fill_frame, text="🎨 Pick", command=pick_fill_color, bg='#3498db', fg='white', font=('Segoe UI', 9), padx=8, relief='flat')
        fill_btn.pack(side='left', padx=2)
        fill_entry = tk.Entry(fill_frame, textvariable=fill_var, font=('Segoe UI', 9), width=10)
        fill_entry.pack(side='left', padx=2)
        
        # Outline color with picker
        tk.Label(form, text="Outline Color:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=12, column=0, sticky='w', pady=8)
        outline_var = tk.StringVar(value='#2c3e50')
        outline_frame = tk.Frame(form, bg='white')
        outline_frame.grid(row=12, column=1, sticky='w', pady=8, padx=5)
        outline_preview = tk.Canvas(outline_frame, width=30, height=25, bg=outline_var.get(), highlightthickness=1, highlightbackground='#95a5a6')
        outline_preview.pack(side='left', padx=2)
        def pick_outline_color():
            from tkinter.colorchooser import askcolor
            color = askcolor(color=outline_var.get(), title="Choose Outline Color")
            if color[1]:
                outline_var.set(color[1])
                outline_preview.config(bg=color[1])
        outline_btn = tk.Button(outline_frame, text="🎨 Pick", command=pick_outline_color, bg='#2c3e50', fg='white', font=('Segoe UI', 9), padx=8, relief='flat')
        outline_btn.pack(side='left', padx=2)
        outline_entry = tk.Entry(outline_frame, textvariable=outline_var, font=('Segoe UI', 9), width=10)
        outline_entry.pack(side='left', padx=2)
        
        form.grid_columnconfigure(1, weight=1)
        
        def create_component():
            # Auto-generate ID if empty (shouldn't happen, but safeguard)
            comp_id = id_var.get().strip()
            if not comp_id:
                comp_id = generate_unique_id(label_var.get())
                id_var.set(comp_id)
            
            # Check for duplicate ID (double-check safety)
            nodes = self.area_data.get('nodes', [])
            if any(node['id'] == comp_id for node in nodes):
                # Auto-regenerate with timestamp suffix
                import time
                comp_id = f"{comp_id}_{int(time.time() % 10000)}"
                id_var.set(comp_id)
            
            # Create new node
            new_node = {
                'id': comp_id,
                'label': label_var.get(),
                'type': type_var.get(),
                'shape': shape_var.get(),
                'x': float(x_var.get()),
                'y': float(y_var.get()),
                'width': width_var.get(),
                'height': height_var.get(),
                'fill': fill_var.get(),
                'outline': outline_var.get(),
                'font_size': font_size_var.get(),
                'font_weight': font_weight_var.get(),
                'locked': False
            }
            
            # Add to area data
            self.area_data['nodes'].append(new_node)
            
            # Redraw
            self._draw_diagram()
            
            logger.info(f"✅ Added new component: {comp_id}")
            messagebox.showinfo('Success', f'Component "{comp_id}" created successfully!\n\nRemember to save your changes!')
            dialog.destroy()
        
        # Buttons
        btn_frame = tk.Frame(dialog, bg='#e8eef5')
        btn_frame.pack(fill='x', pady=15, padx=10)
        
        tk.Button(btn_frame, text="✅ Create", command=create_component,
                  bg='#27ae60', fg='white', font=('Segoe UI', 11, 'bold'),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        tk.Button(btn_frame, text="✖ Cancel", command=dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 11),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        dialog.wait_window()

    def _edit_node(self):
        """Edit selected component properties (label, size, color with picker)"""
        if not self.selected_node:
            messagebox.showwarning('No Selection', 'Please click on a component to select it first')
            return
        
        nodes = self.area_data.get('nodes', [])
        node_data = None
        for node in nodes:
            if node['id'] == self.selected_node:
                node_data = node
                break
        
        if not node_data:
            messagebox.showerror('Error', 'Selected component not found in data')
            return
        
        # Create edit dialog
        dialog = self._create_styled_dialog(f"Edit Component: {self.selected_node}", 600, 700)
        
        # Title header with background
        header = tk.Frame(dialog, bg='#3498db', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="✏️ Edit Component Properties",
                 font=('Segoe UI', 14, 'bold'), bg='#3498db', fg='white').pack(pady=18)
        
        info_frame = tk.Frame(dialog, bg='#e8eef5')
        info_frame.pack(fill='x', pady=5)
        tk.Label(info_frame, text=f"Component ID: {self.selected_node}",
                 font=('Segoe UI', 9, 'italic'), fg='#7f8c8d', bg='#e8eef5').pack()
        
        # Form frame with white background
        form = tk.Frame(dialog, bg='white', padx=25, pady=20)
        form.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Label
        tk.Label(form, text="Label:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=8)
        label_var = tk.StringVar(value=node_data.get('label', ''))
        label_entry = tk.Entry(form, textvariable=label_var, font=('Segoe UI', 10), width=35)
        label_entry.grid(row=0, column=1, columnspan=2, pady=8, padx=5, sticky='ew')
        
        # Width
        tk.Label(form, text="Width:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=1, column=0, sticky='w', pady=8)
        width_var = tk.IntVar(value=node_data.get('width', 100))
        width_spin = tk.Spinbox(form, from_=40, to=400, textvariable=width_var,
                                font=('Segoe UI', 10), width=12)
        width_spin.grid(row=1, column=1, sticky='w', pady=8, padx=5)
        tk.Label(form, text="px", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').grid(row=1, column=2, sticky='w', padx=2)
        
        # Height
        tk.Label(form, text="Height:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=2, column=0, sticky='w', pady=8)
        height_var = tk.IntVar(value=node_data.get('height', 40))
        height_spin = tk.Spinbox(form, from_=20, to=200, textvariable=height_var,
                                 font=('Segoe UI', 10), width=12)
        height_spin.grid(row=2, column=1, sticky='w', pady=8, padx=5)
        tk.Label(form, text="px", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').grid(row=2, column=2, sticky='w', padx=2)
        
        # Fill color with picker
        tk.Label(form, text="Fill Color:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=3, column=0, sticky='w', pady=8)
        fill_var = tk.StringVar(value=node_data.get('fill', '#ffffff'))
        
        fill_frame = tk.Frame(form, bg='white')
        fill_frame.grid(row=3, column=1, columnspan=2, sticky='ew', pady=8, padx=5)
        
        fill_entry = tk.Entry(fill_frame, textvariable=fill_var, font=('Segoe UI', 9), width=16)
        fill_entry.pack(side='left', padx=2)
        
        # Color preview box
        fill_preview = tk.Canvas(fill_frame, width=30, height=25, bg=fill_var.get(), relief='sunken', bd=1)
        fill_preview.pack(side='left', padx=5)
        
        def pick_fill_color():
            from tkinter.colorchooser import askcolor
            color = askcolor(color=fill_var.get(), title="Choose Fill Color")
            if color[1]:  # color[1] is the hex code
                fill_var.set(color[1])
                fill_preview.config(bg=color[1])
        
        fill_btn = tk.Button(fill_frame, text="🎨 Pick", command=pick_fill_color,
                            bg='#3498db', fg='white', font=('Segoe UI', 9), padx=8)
        fill_btn.pack(side='left', padx=2)
        
        # Outline color with picker
        tk.Label(form, text="Outline Color:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=4, column=0, sticky='w', pady=8)
        outline_var = tk.StringVar(value=node_data.get('outline', '#000000'))
        
        outline_frame = tk.Frame(form, bg='white')
        outline_frame.grid(row=4, column=1, columnspan=2, sticky='ew', pady=8, padx=5)
        
        outline_entry = tk.Entry(outline_frame, textvariable=outline_var, font=('Segoe UI', 9), width=16)
        outline_entry.pack(side='left', padx=2)
        
        # Color preview box
        outline_preview = tk.Canvas(outline_frame, width=30, height=25, bg=outline_var.get(), relief='sunken', bd=1)
        outline_preview.pack(side='left', padx=5)
        
        def pick_outline_color():
            from tkinter.colorchooser import askcolor
            color = askcolor(color=outline_var.get(), title="Choose Outline Color")
            if color[1]:  # color[1] is the hex code
                outline_var.set(color[1])
                outline_preview.config(bg=color[1])
        
        outline_btn = tk.Button(outline_frame, text="🎨 Pick", command=pick_outline_color,
                               bg='#3498db', fg='white', font=('Segoe UI', 9), padx=8)
        outline_btn.pack(side='left', padx=2)
        
        # Shape
        tk.Label(form, text="Shape:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=5, column=0, sticky='w', pady=8)
        shape_var = tk.StringVar(value=node_data.get('shape', 'rect'))
        shape_combo = ttk.Combobox(form, textvariable=shape_var, values=['rect', 'oval', 'diamond'],
                                   state='readonly', font=('Segoe UI', 10), width=15)
        shape_combo.grid(row=5, column=1, sticky='w', pady=8, padx=5)
        
        # Type
        tk.Label(form, text="Type:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=6, column=0, sticky='w', pady=8)
        type_var = tk.StringVar(value=node_data.get('type', 'process'))
        type_combo = ttk.Combobox(form, textvariable=type_var,
                                  values=['source', 'process', 'storage', 'consumption', 'discharge', 'building', 'treatment', 'plant', 'tsf', 'reservoir', 'loss'],
                                  state='readonly', font=('Segoe UI', 10), width=15)
        type_combo.grid(row=6, column=1, sticky='w', pady=8, padx=5)
        
        # Font size (supports decimals: 7.5, 10.25, etc.)
        tk.Label(form, text="Font Size:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=7, column=0, sticky='w', pady=8)
        font_size_var = tk.DoubleVar(value=node_data.get('font_size', 10.0))
        font_size_frame = tk.Frame(form, bg='white')
        font_size_frame.grid(row=7, column=1, sticky='w', pady=8, padx=5)
        font_size_spin = tk.Spinbox(font_size_frame, from_=4.0, to=36.0, increment=0.5, textvariable=font_size_var, font=('Segoe UI', 10), width=8)
        font_size_spin.pack(side='left', padx=2)
        tk.Label(font_size_frame, text="pt", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)
        
        # Font weight (Bold/Regular)
        tk.Label(form, text="Font Style:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=8, column=0, sticky='w', pady=8)
        font_weight_var = tk.StringVar(value=node_data.get('font_weight', 'normal'))
        font_weight_frame = tk.Frame(form, bg='white')
        font_weight_frame.grid(row=8, column=1, sticky='w', pady=8, padx=5)
        
        bold_btn = tk.Button(font_weight_frame, text="B  Bold", 
                            bg='#3498db' if font_weight_var.get() == 'bold' else '#95a5a6', 
                            fg='white', font=('Segoe UI', 9, 'bold'), 
                            padx=10, relief='sunken' if font_weight_var.get() == 'bold' else 'raised')
        regular_btn = tk.Button(font_weight_frame, text="Regular",
                               bg='#95a5a6' if font_weight_var.get() != 'bold' else '#3498db',
                               fg='white', font=('Segoe UI', 9),
                               padx=10, relief='sunken' if font_weight_var.get() != 'bold' else 'raised')
        
        def toggle_bold():
            current = font_weight_var.get()
            new_weight = 'normal' if current == 'bold' else 'bold'
            font_weight_var.set(new_weight)
            bold_btn.config(bg='#3498db' if new_weight == 'bold' else '#95a5a6', relief='sunken' if new_weight == 'bold' else 'raised')
            regular_btn.config(bg='#95a5a6' if new_weight != 'bold' else '#3498db', relief='sunken' if new_weight != 'bold' else 'raised')
        
        def set_regular():
            font_weight_var.set('normal')
            bold_btn.config(bg='#95a5a6', relief='raised')
            regular_btn.config(bg='#3498db', relief='sunken')
        
        bold_btn.config(command=toggle_bold)
        regular_btn.config(command=set_regular)
        bold_btn.pack(side='left', padx=2)
        regular_btn.pack(side='left', padx=2)
        
        form.grid_columnconfigure(1, weight=1)
        
        def save_changes():
            # Update node data
            node_data['label'] = label_var.get()
            node_data['width'] = width_var.get()
            node_data['height'] = height_var.get()
            node_data['fill'] = fill_var.get()
            node_data['outline'] = outline_var.get()
            node_data['shape'] = shape_var.get()
            node_data['type'] = type_var.get()
            node_data['font_size'] = font_size_var.get()
            node_data['font_weight'] = font_weight_var.get()
            
            self._draw_diagram()
            messagebox.showinfo('Success', f'Component "{self.selected_node}" updated successfully!')
            dialog.destroy()
        
        # Buttons
        btn_frame = tk.Frame(dialog, bg='#e8eef5')
        btn_frame.pack(fill='x', pady=15, padx=10)
        
        tk.Button(btn_frame, text="💾 Save Changes", command=save_changes,
                  bg='#27ae60', fg='white', font=('Segoe UI', 11, 'bold'),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        tk.Button(btn_frame, text="✖ Cancel", command=dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 11),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        dialog.wait_window()
    
    def _delete_node(self):
        """Delete selected component and all connected flow lines"""
        if not self.selected_node:
            messagebox.showwarning('No Selection', 'Please click on a component to select it first')
            return
        
        # Find the node
        nodes = self.area_data.get('nodes', [])
        node_data = None
        node_index = None
        for i, node in enumerate(nodes):
            if node['id'] == self.selected_node:
                node_data = node
                node_index = i
                break
        
        if not node_data:
            messagebox.showerror('Error', 'Selected component not found in data')
            return
        
        # Count connected flows
        edges = self.area_data.get('edges', [])
        connected_edges = []
        for i, edge in enumerate(edges):
            if edge.get('from') == self.selected_node or edge.get('to') == self.selected_node:
                connected_edges.append(i)
        
        # Confirmation dialog
        node_label = node_data.get('label', self.selected_node)
        confirm_msg = f"Delete component: {node_label}?\n\n"
        confirm_msg += f"This will also delete {len(connected_edges)} connected flow line(s).\n\n"
        confirm_msg += "⚠️ This action cannot be undone after saving!"
        
        if not messagebox.askyesno('Confirm Deletion', confirm_msg):
            return
        
        # Delete the node
        del nodes[node_index]
        
        # Delete connected edges (reverse order to maintain indices)
        for edge_idx in sorted(connected_edges, reverse=True):
            del edges[edge_idx]
        
        # Update data
        self.area_data['nodes'] = nodes
        self.area_data['edges'] = edges
        
        # Clear selection and locked state
        if self.selected_node in self.locked_nodes:
            del self.locked_nodes[self.selected_node]
        self.selected_node = None
        
        # Redraw
        self._draw_diagram()
        
        logger.info(f"🗑️ Deleted component '{node_label}' and {len(connected_edges)} flow line(s)")
        messagebox.showinfo('Deleted', f'Component "{node_label}" and {len(connected_edges)} flow line(s) deleted.\n\nRemember to save your changes!')

    def _edit_line(self):
        """Edit flow line properties: type, color, volume, bidirectional"""
        edges = self.area_data.get('edges', [])
        if not edges:
            messagebox.showwarning("No Flows", "No flow lines to edit")
            return

        dialog = self._create_styled_dialog("Edit Flow Line", 950, 650)

        # Header
        header = tk.Frame(dialog, bg='#9b59b6', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="🎨 Edit Flow Line Properties",
                 font=('Segoe UI', 14, 'bold'), bg='#9b59b6', fg='white').pack(pady=18)

        # Instructions
        info_frame = tk.Frame(dialog, bg='white')
        info_frame.pack(fill='x', padx=10, pady=10)
        tk.Label(info_frame, text="Select a flow line to edit (grouped by area):",
                 font=('Segoe UI', 10), bg='white', fg='#2c3e50').pack(pady=5)

        container = tk.Frame(dialog, bg='white')
        container.pack(fill='both', expand=True, padx=10, pady=5)

        # List of edges grouped by area
        list_frame = tk.Frame(container)
        list_frame.pack(side='left', fill='both', expand=True)
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                             font=('Segoe UI', 9), height=20, selectmode='single')
        listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        enable_listbox_mousewheel(listbox)

        # Group edges by area like _delete_line does
        edges_by_area = {}
        for i, edge in enumerate(edges):
            from_area = self._get_node_area(edge['from'])
            to_area = self._get_node_area(edge['to'])
            area = from_area if from_area == to_area else f"{from_area} ↔ {to_area}"
            if area not in edges_by_area:
                edges_by_area[area] = []
            edges_by_area[area].append(i)
        
        edge_index_map = []
        for area in sorted(edges_by_area.keys()):
            listbox.insert(tk.END, f"")
            listbox.itemconfig(tk.END, bg='#e8eef5')
            edge_index_map.append(None)
            
            listbox.insert(tk.END, f"━━━ {area} ━━━")
            listbox.itemconfig(tk.END, bg='#3498db', fg='white')
            edge_index_map.append(None)
            
            for edge_idx in edges_by_area[area]:
                edge = edges[edge_idx]
                from_name = self._format_node_name(edge['from'])
                to_name = self._format_node_name(edge['to'])

                # Safely format volume/label (supports numeric, negative, or text)
                vol_val = edge.get('volume', '')
                if isinstance(vol_val, (int, float)):
                    vol_display = f"{vol_val:,.2f} m³"
                else:
                    vol_display = str(vol_val)

                display_text = f"  {from_name} → {to_name} | {edge.get('flow_type','unknown')} | {vol_display}"
                listbox.insert(tk.END, display_text)
                edge_index_map.append(edge_idx)

        # Edit panel
        edit_frame = tk.Frame(container, relief='groove', borderwidth=1)
        edit_frame.pack(side='right', fill='y', padx=10)

        tk.Label(edit_frame, text="Flow Type:", font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=8, pady=(10,0))
        type_var = tk.StringVar(value='clean')
        # New simplified categories with dirty variants
        tk.OptionMenu(
            edit_frame,
            type_var,
            'clean',
            'evaporation',
            'dirty',
            'effluent',
            'runoff',
            'ug_return',
            'dewatering',
            'outflow',
            'inflow',
            'drainage',
            'return',
            'irrigation',
            'recirculation'
        ).pack(fill='x', padx=8)

        tk.Label(edit_frame, text="Color (hex):", font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=8, pady=(10,0))
        color_var = tk.StringVar(value='#4b78a8')
        tk.Entry(edit_frame, textvariable=color_var, font=('Segoe UI', 10)).pack(fill='x', padx=8)

        tk.Label(edit_frame, text="Volume or Label (m³):", font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=8, pady=(10,0))
        vol_var = tk.StringVar(value='0')
        tk.Entry(edit_frame, textvariable=vol_var, font=('Segoe UI', 10)).pack(fill='x', padx=8)

        bidir_var = tk.BooleanVar(value=False)
        tk.Checkbutton(edit_frame, text="Bidirectional (arrows both ends)", variable=bidir_var).pack(anchor='w', padx=8, pady=8)

        tk.Label(edit_frame, text="Label Font Size:", font=('Segoe UI', 9, 'bold')).pack(anchor='w', padx=8, pady=(10,0))
        label_font_size_var = tk.DoubleVar(value=8.0)
        label_font_frame = tk.Frame(edit_frame)
        label_font_frame.pack(fill='x', padx=8)
        label_font_spin = tk.Spinbox(label_font_frame, from_=4.0, to=36.0, increment=0.5, textvariable=label_font_size_var, font=('Segoe UI', 10), width=8)
        label_font_spin.pack(side='left', padx=2)
        tk.Label(label_font_frame, text="pt", font=('Segoe UI', 9), fg='#7f8c8d').pack(side='left', padx=2)

        def on_select(evt=None):
            sel = listbox.curselection()
            if not sel:
                return
            list_idx = sel[0]
            if list_idx >= len(edge_index_map) or edge_index_map[list_idx] is None:
                return  # Header or separator selected
            edge_idx = edge_index_map[list_idx]
            e = edges[edge_idx]
            type_var.set(e.get('flow_type','clean'))
            # Default color based on type if missing
            color_var.set(e.get('color') or _default_color_for_type(e.get('flow_type')))
            vol_var.set(str(e.get('volume', 0)))
            bidir_var.set(bool(e.get('bidirectional', False)))
            label_font_size_var.set(e.get('label_font_size', 8.0))

        listbox.bind('<<ListboxSelect>>', on_select)

        def on_apply():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("No Selection", "Select a flow line to edit")
                return
            list_idx = sel[0]
            if list_idx >= len(edge_index_map) or edge_index_map[list_idx] is None:
                messagebox.showwarning("Invalid Selection", "Please select a flow line (not a header)")
                return
            edge_idx = edge_index_map[list_idx]
            e = edges[edge_idx]
            e['flow_type'] = type_var.get()
            # If user left color empty, apply default per type
            chosen_color = color_var.get().strip()
            e['color'] = chosen_color if chosen_color else _default_color_for_type(e['flow_type'])
            # Allow any value: numbers (positive/negative), dashes, text, etc.
            vol_str = vol_var.get().strip()
            if not vol_str:
                messagebox.showwarning("Invalid Volume", "Enter a volume or label")
                return
    
            try:
                # Try to format as number if it's numeric
                if vol_str.replace('-', '').replace('.', '').isdigit() or (vol_str.startswith('-') and vol_str[1:].replace('.', '').isdigit()):
                    e['volume'] = float(vol_str)
                    e['label'] = f"{e['volume']:,.2f}"
                else:
                    # Non-numeric value (like "-" or "N/A")
                    e['volume'] = vol_str
                    e['label'] = vol_str
            except ValueError:
                messagebox.showwarning("Invalid Volume", "Error processing volume")
                return
            e['bidirectional'] = bidir_var.get()
            e['label_font_size'] = label_font_size_var.get()
            self._draw_diagram()
            messagebox.showinfo("Updated", "Flow line updated")

        tk.Button(edit_frame, text="Apply", command=on_apply,
                  bg='#27ae60', fg='white', font=('Segoe UI', 9, 'bold')).pack(fill='x', padx=8, pady=12)

        tk.Button(dialog, text="Close", command=dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 9)).pack(pady=8)

    def _delete_line(self):
        """Delete an existing flow line with scrollable list dialog grouped by area"""
        edges = self.area_data.get('edges', [])
        if not edges:
            messagebox.showwarning("No Flows", "No flow lines to delete")
            return
        
        # Create custom dialog with scrollable listbox
        dialog = self._create_styled_dialog("Delete Flow Line", 850, 600)
        
        # Header
        header = tk.Frame(dialog, bg='#e74c3c', height=70)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="\ud83d\uddd1\ufe0f Delete Flow Lines",
                 font=('Segoe UI', 14, 'bold'), bg='#e74c3c', fg='white').pack(pady=10)
        tk.Label(header, text="\u26a0\ufe0f  Warning: Deletion is permanent after saving",
                 font=('Segoe UI', 10), bg='#e74c3c', fg='#fff3cd').pack()
        
        # Instructions
        info_frame = tk.Frame(dialog, bg='white')
        info_frame.pack(fill='x', padx=10, pady=10)
        tk.Label(info_frame, text="Select flow lines to delete (grouped by area) - Use Ctrl+Click for multiple selections",
                 font=('Segoe UI', 10), bg='white', fg='#2c3e50').pack(pady=5)
        
        list_frame = tk.Frame(dialog, bg='white')
        list_frame.pack(fill='both', expand=True, padx=10, pady=5)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                             font=('Segoe UI', 9), height=20, selectmode='extended')
        listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        enable_listbox_mousewheel(listbox)
        
        edges_by_area = {}
        for i, edge in enumerate(edges):
            from_area = self._get_node_area(edge['from'])
            to_area = self._get_node_area(edge['to'])
            area = from_area if from_area == to_area else f"{from_area} ↔ {to_area}"
            if area not in edges_by_area:
                edges_by_area[area] = []
            edges_by_area[area].append(i)
        
        edge_index_map = []
        for area in sorted(edges_by_area.keys()):
            listbox.insert(tk.END, f"")
            listbox.itemconfig(tk.END, bg='#e8eef5')
            edge_index_map.append(None)
            
            listbox.insert(tk.END, f"━━━ {area} ━━━")
            listbox.itemconfig(tk.END, bg='#e74c3c', fg='white')
            edge_index_map.append(None)
            
            for edge_idx in edges_by_area[area]:
                edge = edges[edge_idx]
                from_name = self._format_node_name(edge['from'])
                to_name = self._format_node_name(edge['to'])
                # Format volume (handle non-numeric values like '-')
                volume = edge.get('volume', 0)
                try:
                    volume_str = f"{float(volume):,.2f}"
                except (ValueError, TypeError):
                    volume_str = str(volume)
                display_text = f"  {from_name} → {to_name} | {edge.get('flow_type','unknown')} | {volume_str} m³"
                listbox.insert(tk.END, display_text)
                edge_index_map.append(edge_idx)
        
        selected_indices = []
        
        def on_delete():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select flow lines to delete")
                return
            chosen_edges = []
            for list_idx in selection:
                if list_idx < len(edge_index_map) and edge_index_map[list_idx] is not None:
                    chosen_edges.append(edge_index_map[list_idx])
            if not chosen_edges:
                messagebox.showwarning("Invalid Selection", "Please select flow lines (not headers)")
                return
            selected_indices[:] = sorted(set(chosen_edges))
            dialog.destroy()
        
        btn_frame = tk.Frame(dialog, bg='#e8eef5')
        btn_frame.pack(fill='x', pady=15, padx=10)
        
        tk.Button(btn_frame, text="🗑️ Delete Selected", command=on_delete,
                  bg='#e74c3c', fg='white', font=('Segoe UI', 11, 'bold'),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        tk.Button(btn_frame, text="✖ Cancel", command=dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 11),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        dialog.wait_window()
        
        if selected_indices:
            lines = []
            for idx in selected_indices:
                e = edges[idx]
                # Format volume (handle non-numeric values like '-')
                volume = e.get('volume', 0)
                try:
                    volume_str = f"{float(volume):,.2f}"
                except (ValueError, TypeError):
                    volume_str = str(volume)
                lines.append(f"- {self._format_node_name(e['from'])} → {self._format_node_name(e['to'])} ({e.get('flow_type','unknown')}, {volume_str} m³)")
            if messagebox.askyesno("Confirm Delete", "Delete selected flow lines?\n\n" + "\n".join(lines)):
                for idx in sorted(selected_indices, reverse=True):
                    edge = edges.pop(idx)
                    logger.info(f"🗑️ Deleted flow: {edge['from']} → {edge['to']}")
                self._draw_diagram()
                self._save_to_json()  # Auto-save deletion to prevent flowline from reappearing
                messagebox.showinfo("Deleted", f"Removed {len(selected_indices)} flow line(s)")

    def _add_recirculation(self):
        """Add a recirculation loop symbol to a component"""

        if not self.selected_node:
            messagebox.showwarning("No Selection", "Select a component first (click on it) to add recirculation")
            return
        
        component_id = self.selected_node
        
        # Show dialog for volume and type
        dialog = tk.Toplevel(self.canvas)
        dialog.title("Add Recirculation Loop")
        dialog.transient(self.canvas)
        dialog.grab_set()
        
        dialog_width = 480
        dialog_height = 520  # Increased to ensure button is fully visible
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - dialog_width) // 2
        y = (screen_height - dialog_height) // 2
        dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        dialog.minsize(480, 520)  # Prevent resizing too small
        
        # Header
        header = tk.Label(dialog, text="♻️ Recirculation Loop", 
                         font=('Segoe UI', 13, 'bold'), bg='#16a085', fg='white', pady=12)
        header.pack(fill='x')
        
        # Main scrollable content frame
        main_frame = tk.Frame(dialog)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Component info
        info_frame = tk.Frame(main_frame, bg='#e8eef5', pady=12)
        info_frame.pack(fill='x', padx=0, pady=(0,10))
        tk.Label(info_frame, text=f"Component: {self._format_node_name(component_id)}", 
                font=('Segoe UI', 10, 'bold'), bg='#e8eef5', fg='#2c3e50').pack(anchor='w', padx=10)
        
        # Flow type with more space
        flow_frame = tk.Frame(main_frame)
        flow_frame.pack(fill='x', padx=0, pady=(10,10))
        tk.Label(flow_frame, text="Flow Type:", font=('Segoe UI', 11, 'bold'), fg='#2c3e50').pack(anchor='w')
        
        type_var = tk.StringVar(value='recirculation')
        types = [
            ('♻️  Recirculation (internal loop)', 'recirculation'),
            ('💧 Clean water (blue)', 'clean'),
            ('☀️  Evaporation/Losses (black)', 'evaporation'),
            ('⚠️  Dirty water (red)', 'dirty'),
            ('🧪 Effluent', 'effluent'),
            ('🌧️ Runoff', 'runoff'),
            ('⛏️ UG Return', 'ug_return'),
            ('⛏️ Dewatering', 'dewatering'),
            ('➡️ Outflow', 'outflow'),
            ('⬅️ Inflow', 'inflow'),
            ('🕳️ Drainage', 'drainage'),
            ('↩️ Return', 'return'),
            ('🌿 Irrigation', 'irrigation')
        ]
        for label, value in types:
            tk.Radiobutton(flow_frame, text=label, variable=type_var, value=value,
                          font=('Segoe UI', 10), anchor='w', selectcolor='#16a085',
                          activebackground='#e8eef5').pack(anchor='w', padx=15, pady=6)
        
        # Volume section
        vol_frame = tk.Frame(main_frame)
        vol_frame.pack(fill='x', padx=0, pady=(10,10))
        tk.Label(vol_frame, text="Volume or Label (m³/month):", font=('Segoe UI', 11, 'bold'), 
                fg='#2c3e50').pack(anchor='w')
        vol_var = tk.StringVar(value='0')
        tk.Entry(vol_frame, textvariable=vol_var, font=('Segoe UI', 11), width=20).pack(anchor='w', padx=15, pady=8)
        
        error_label = tk.Label(main_frame, text="", font=('Segoe UI', 9), fg='#e74c3c')
        error_label.pack(pady=(5,10))
        
        def on_ok():
            vol_str = vol_var.get().strip()
            if not vol_str:
                error_label.config(text="Enter a volume or label")
                return
            
            # Allow any value: numbers (positive/negative), dashes, text, etc.
            try:
                # Try to format as number if it's numeric
                if vol_str.replace('-', '').replace('.', '').isdigit() or (vol_str.startswith('-') and vol_str[1:].replace('.', '').isdigit()):
                    volume = float(vol_str)
                    label_str = f"{volume:,.2f}"
                else:
                    # Non-numeric value (like "-" or "N/A")
                    volume = vol_str
                    label_str = vol_str
                
                # Create recirculation edge (from component to itself)
                flow_type = type_var.get()
                color = _default_color_for_type(flow_type) if flow_type != 'recirculation' else '#9b59b6'
                
                # Create a minimal recirculation edge with special flag
                edges = self.area_data.get('edges', [])
                node = self.nodes_by_id[component_id]
                
                # Position loop icon to the right of component, roughly in middle
                loop_x = node['x'] + node['width'] + 15
                loop_y = node['y'] + node['height'] / 2
                
                recirculation_edge = {
                    'from': component_id,
                    'to': component_id,
                    'segments': [
                        (node['x'] + node['width'] / 2, node['y'] + node['height'] / 2),
                        (node['x'] + node['width'] / 2, node['y'] + node['height'] / 2)
                    ],
                    'flow_type': flow_type,
                    'volume': volume,
                    'color': color,
                    'label': label_str,
                    'bidirectional': False,
                    'is_recirculation': True,
                    'recirculation_pos': {'x': loop_x, 'y': loop_y}
                }
                
                edges.append(recirculation_edge)
                self.area_data['edges'] = edges
                self._draw_diagram()
                dialog.destroy()
                messagebox.showinfo("Success", f"♻️ Recirculation loop added to {self._format_node_name(component_id)}\n{label_str} m³ {flow_type}")
            except Exception as e:
                error_label.config(text=f"Error: {str(e)}")

        
        # Button frame fixed at bottom
        button_frame = tk.Frame(dialog)
        button_frame.pack(fill='x', padx=10, pady=(15,10), side='bottom')
        tk.Button(button_frame, text="✅ OK", command=on_ok, bg='#27ae60', fg='white',
                 font=('Segoe UI', 11, 'bold'), padx=30, pady=8, width=15).pack(expand=True)

    def _edit_recirculation(self):
        """Edit existing recirculation loop properties: size, font, color, volume"""
        edges = self.area_data.get('edges', [])
        recirculation_edges = [(i, e) for i, e in enumerate(edges) if e.get('is_recirculation', False)]

        if not recirculation_edges:
            messagebox.showwarning("No Recirculation", "No recirculation loops found in this diagram")
            return

        # Selection dialog
        select_dialog = self._create_styled_dialog("Select Recirculation to Edit", 600, 400)

        # Header
        header = tk.Frame(select_dialog, bg='#9b59b6', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="✏️ Edit Recirculation Loop",
                 font=('Segoe UI', 14, 'bold'), bg='#9b59b6', fg='white').pack(pady=18)

        # Listbox frame
        list_frame = tk.Frame(select_dialog)
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)

        tk.Label(list_frame, text="Select a recirculation loop to edit:",
                font=('Segoe UI', 10)).pack(anchor='w', pady=(0, 5))

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')

        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, font=('Segoe UI', 9), height=15)
        listbox.pack(fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        enable_listbox_mousewheel(listbox)

        # Populate listbox
        for idx, (edge_idx, edge) in enumerate(recirculation_edges):
            from_node = edge.get('from', 'Unknown')
            volume = edge.get('volume', 0)
            flow_type = edge.get('flow_type', 'recirculation')
            try:
                volume_str = f"{float(volume):,.2f}"
            except (ValueError, TypeError):
                volume_str = str(volume)
            display_text = f"  {self._format_node_name(from_node)} | {flow_type} | {volume_str} m³"
            listbox.insert(tk.END, display_text)

        selected_idx = [None]

        def on_select():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("No Selection", "Please select a recirculation loop")
                return
            selected_idx[0] = recirculation_edges[sel[0]][0]
            select_dialog.destroy()

        btn_frame = tk.Frame(select_dialog)
        btn_frame.pack(fill='x', padx=10, pady=10)
        tk.Button(btn_frame, text="✏️ Edit Selected", command=on_select,
                  bg='#9b59b6', fg='white', font=('Segoe UI', 11, 'bold'),
                  padx=25, pady=8).pack(side='right', padx=5)
        tk.Button(btn_frame, text="✖ Cancel", command=select_dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 11),
                  padx=25, pady=8).pack(side='right', padx=5)

        select_dialog.wait_window()

        if selected_idx[0] is None:
            return

        edge_idx = selected_idx[0]
        edge = edges[edge_idx]

        # Edit dialog
        dialog = self._create_styled_dialog("Edit Recirculation Properties", 550, 700)

        # Header
        header = tk.Frame(dialog, bg='#9b59b6', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="✏️ Edit Recirculation Loop",
                 font=('Segoe UI', 14, 'bold'), bg='#9b59b6', fg='white').pack(pady=18)

        # Form
        form = tk.Frame(dialog, bg='white', padx=25, pady=20)
        form.pack(fill='both', expand=True, padx=10, pady=10)

        # Component info (read-only)
        tk.Label(form, text="Component:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=8)
        tk.Label(form, text=self._format_node_name(edge['from']), font=('Segoe UI', 10), bg='#e8eef5',
                 relief='sunken', padx=10, pady=5).grid(row=0, column=1, pady=8, sticky='ew')

        # Flow type
        tk.Label(form, text="Flow Type:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=1, column=0, sticky='w', pady=8)
        type_var = tk.StringVar(value=edge.get('flow_type', 'recirculation'))
        type_combo = ttk.Combobox(form, textvariable=type_var,
                                  values=['recirculation', 'clean', 'evaporation', 'dirty', 'effluent',
                                         'runoff', 'ug_return', 'dewatering', 'outflow', 'inflow'],
                                  state='readonly', font=('Segoe UI', 10), width=18)
        type_combo.grid(row=1, column=1, sticky='w', pady=8)

        # Volume
        tk.Label(form, text="Volume (m³):", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=2, column=0, sticky='w', pady=8)
        vol_var = tk.StringVar(value=str(edge.get('volume', 0)))
        tk.Entry(form, textvariable=vol_var, font=('Segoe UI', 10), width=20).grid(row=2, column=1, sticky='w', pady=8)

        # Color
        tk.Label(form, text="Color (hex):", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=3, column=0, sticky='w', pady=8)
        color_var = tk.StringVar(value=edge.get('color', '#9b59b6'))
        color_entry = tk.Entry(form, textvariable=color_var, font=('Segoe UI', 10), width=20)
        color_entry.grid(row=3, column=1, sticky='w', pady=8)

        # Font size (supports decimals)
        tk.Label(form, text="Label Font Size:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=4, column=0, sticky='w', pady=8)
        font_size_var = tk.DoubleVar(value=edge.get('label_font_size', 8.0))
        font_size_frame = tk.Frame(form, bg='white')
        font_size_frame.grid(row=4, column=1, sticky='w', pady=8)
        font_size_spin = tk.Spinbox(font_size_frame, from_=4.0, to=36.0, increment=0.5,
                                   textvariable=font_size_var, font=('Segoe UI', 10), width=8)
        font_size_spin.pack(side='left', padx=2)
        tk.Label(font_size_frame, text="pt", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)

        # Size (width/height of rectangle)
        tk.Label(form, text="Box Width:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=5, column=0, sticky='w', pady=8)
        width_var = tk.IntVar(value=edge.get('box_width', 40))
        tk.Spinbox(form, from_=20, to=100, textvariable=width_var, font=('Segoe UI', 10), width=8).grid(row=5, column=1, sticky='w', pady=8)

        tk.Label(form, text="Box Height:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=6, column=0, sticky='w', pady=8)
        height_var = tk.IntVar(value=edge.get('box_height', 28))
        tk.Spinbox(form, from_=20, to=100, textvariable=height_var, font=('Segoe UI', 10), width=8).grid(row=6, column=1, sticky='w', pady=8)

        form.grid_columnconfigure(1, weight=1)

        def on_save():
            vol_str = vol_var.get().strip()
            if not vol_str:
                messagebox.showwarning("Invalid Input", "Volume cannot be empty")
                return

            # Parse volume
            try:
                if vol_str.replace('-', '').replace('.', '').isdigit() or (vol_str.startswith('-') and vol_str[1:].replace('.', '').isdigit()):
                    volume = float(vol_str)
                    label_str = f"{volume:,.2f}"
                else:
                    volume = vol_str
                    label_str = vol_str
            except ValueError:
                messagebox.showwarning("Invalid Volume", "Enter a valid number or text label")
                return

            # Update edge
            edge['flow_type'] = type_var.get()
            edge['volume'] = volume
            edge['label'] = label_str
            edge['color'] = color_var.get().strip() or '#9b59b6'
            edge['label_font_size'] = font_size_var.get()
            edge['box_width'] = width_var.get()
            edge['box_height'] = height_var.get()

            self._draw_diagram()
            dialog.destroy()
            messagebox.showinfo("Updated", f"Recirculation loop updated for {self._format_node_name(edge['from'])}")

        # Buttons
        btn_frame = tk.Frame(dialog, bg='#e8eef5')
        btn_frame.pack(fill='x', pady=15, padx=10)
        tk.Button(btn_frame, text="💾 Save Changes", command=on_save,
                  bg='#27ae60', fg='white', font=('Segoe UI', 11, 'bold'),
                  padx=25, pady=8).pack(side='right', padx=5)
        tk.Button(btn_frame, text="✖ Cancel", command=dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 11),
                  padx=25, pady=8).pack(side='right', padx=5)


    def _toggle_recirculation_lock(self):
        """Toggle lock state for selected recirculation"""
        edges = self.area_data.get('edges', [])
        recirculation_edges = [(i, e) for i, e in enumerate(edges) if e.get('is_recirculation', False)]
        
        if not recirculation_edges:
            messagebox.showwarning("No Recirculation", "No recirculation loops found in this diagram")
            return
        
        # Show dialog to select which recirculation to lock/unlock
        dialog = tk.Toplevel(self.canvas)
        dialog.title("Toggle Recirculation Lock")
        dialog.transient(self.canvas)
        dialog.grab_set()
        dialog.geometry("400x300")
        
        # Header
        header = tk.Label(dialog, text="🔒 Lock/Unlock Recirculation", 
                         font=('Segoe UI', 12, 'bold'), bg='#95a5a6', fg='white', pady=10)
        header.pack(fill='x')
        
        # Frame for listbox
        frame = tk.Frame(dialog)
        frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        tk.Label(frame, text="Select a recirculation to toggle lock:", 
                font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 5))
        
        # Listbox with scrollbar
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side='right', fill='y')
        
        listbox = tk.Listbox(frame, yscrollcommand=scrollbar.set, font=('Segoe UI', 9))
        listbox.pack(fill='both', expand=True)
        scrollbar.config(command=listbox.yview)
        enable_listbox_mousewheel(listbox)
        
        # Populate listbox
        for idx, (edge_idx, edge) in enumerate(recirculation_edges):
            from_node = edge.get('from', 'Unknown')
            volume = edge.get('value', 0)
            is_locked = self.recirculation_locked.get(edge_idx, False)
            lock_status = "🔒 LOCKED" if is_locked else "🔓 UNLOCKED"
            # Format volume (handle non-numeric values like '-')
            try:
                volume_str = f"{float(volume):,.2f}"
            except (ValueError, TypeError):
                volume_str = str(volume)
            display_text = f"{from_node}: {volume_str} m³ - {lock_status}"
            listbox.insert(tk.END, display_text)
        
        # Buttons
        button_frame = tk.Frame(dialog)
        button_frame.pack(fill='x', padx=10, pady=(10, 10))
        
        def toggle_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Select a recirculation to toggle")
                return
            
            selected_idx = selection[0]
            edge_idx, edge = recirculation_edges[selected_idx]
            
            # Toggle lock state
            current_state = self.recirculation_locked.get(edge_idx, False)
            self.recirculation_locked[edge_idx] = not current_state
            
            new_state = self.recirculation_locked[edge_idx]
            status = "🔒 LOCKED" if new_state else "🔓 UNLOCKED"
            messagebox.showinfo("Success", f"Recirculation {status}")
            
            self._draw_diagram()
            dialog.destroy()
        
        tk.Button(button_frame, text="✅ Toggle", command=toggle_selected, 
                 bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'), padx=20).pack(side='left', padx=5)
        tk.Button(button_frame, text="❌ Cancel", command=dialog.destroy,
                 bg='#e74c3c', fg='white', font=('Segoe UI', 10, 'bold'), padx=20).pack(side='left', padx=5)

    def _start_drawing(self):
        """Start manual flow line drawing"""
        self.drawing_mode = True
        self.drawing_segments = []
        self.drawing_from = None
        self.drawing_to = None
        self.canvas.config(cursor='tcross')  # Smaller, more precise cursor
        messagebox.showinfo("Draw Mode", 
                           "🔧 SNAP TO COMPONENT MODE\n\n"
                           "1. Click FROM component (snaps to nearest edge)\n"
                           "2. Move mouse - see orange anchors appear\n"
                           "3. Click TO component (snaps to nearest edge)\n"
                           "4. Add intermediate points by clicking canvas\n"
                           "5. Right-click to cancel\n\n"
                           "⭕ Orange anchors = snap points (within 15px)")


    def _on_canvas_click(self, event):
        """Handle click"""
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)

        # Drawing mode
        if self.drawing_mode:
            # Check if clicked on node
            clicked_node = self._get_node_at(canvas_x, canvas_y)
            if self.drawing_from is None:
                # First click - select FROM component and choose anchor point
                if clicked_node:
                    self.drawing_from = clicked_node
                    # Find nearest anchor point to snap start
                    snap_point = self._find_nearest_anchor(clicked_node, canvas_x, canvas_y)
                    if snap_point:
                        self.drawing_segments = [snap_point]
                        messagebox.showinfo("From Selected", 
                                          f"Drawing from: {clicked_node}\n"
                                          f"Start point snapped to component edge\n"
                                          f"Click to add path points or click target component")
                    else:
                        self.drawing_segments = [(canvas_x, canvas_y)]
                        messagebox.showinfo("From Selected", f"Drawing from: {clicked_node}\nClick to add path points or click target component")
                else:
                    # Allow starting from an existing flow line (junction start)
                    edges = self.area_data.get('edges', [])
                    min_dist = 1e9
                    threshold = 15.0
                    snap_x, snap_y = canvas_x, canvas_y
                    snapped = False
                    for edge_idx, e in enumerate(edges):
                        segs = e.get('segments', [])
                        for i in range(len(segs)-1):
                            x1, y1 = segs[i]
                            x2, y2 = segs[i+1]
                            dx = x2 - x1
                            dy = y2 - y1
                            length2 = dx*dx + dy*dy
                            if length2 < 1e-6:
                                continue
                            t = ((canvas_x - x1)*dx + (canvas_y - y1)*dy) / length2
                            t = max(0.0, min(1.0, t))
                            px = x1 + t*dx
                            py = y1 + t*dy
                            dist = ((canvas_x - px)**2 + (canvas_y - py)**2)**0.5
                            if dist < min_dist and dist <= threshold:
                                min_dist = dist
                                snap_x, snap_y = px, py
                                snapped = True
                    if snapped:
                        junction_id = f"junction_{len(edges)}_{snap_x:.0f}_{snap_y:.0f}"
                        self.drawing_from = junction_id
                        self.drawing_segments = [(snap_x, snap_y)]
                        messagebox.showinfo("From Selected", "Drawing from existing flow line (junction). Add waypoints or click target component/line.")
                    else:
                        messagebox.showwarning("Invalid", "Click on a component or near a flow line to start")
            elif clicked_node:
                # Clicked on a component - check if it's the same (recirculation loop) or different
                if clicked_node == self.drawing_from:
                    # Self-loop (recirculation) - must have intermediate waypoints
                    if len(self.drawing_segments) < 2:
                        messagebox.showwarning("Add Waypoints", 
                                             "♻️ RECIRCULATION LOOP\n\n"
                                             "To create a loop back to the same component:\n"
                                             "1. Add at least one waypoint by clicking on the canvas\n"
                                             "2. Then click the component again to complete the loop\n\n"
                                             "This creates a visible recirculation path.")
                        return
                    # Has waypoints - allow the loop
                    snap_point = self._find_nearest_anchor(clicked_node, canvas_x, canvas_y)
                    if snap_point:
                        self.drawing_segments.append(snap_point)
                    else:
                        self.drawing_segments.append((canvas_x, canvas_y))
                    # Create recirculation edge
                    self._finish_drawing(self.drawing_from, clicked_node)
                else:
                    # Different component - normal flow
                    snap_point = self._find_nearest_anchor(clicked_node, canvas_x, canvas_y)
                    if snap_point:
                        self.drawing_segments.append(snap_point)
                    else:
                        self.drawing_segments.append((canvas_x, canvas_y))
                    # Create edge with user-defined path
                    self._finish_drawing(self.drawing_from, clicked_node)
            else:
                # Clicked on canvas - check if clicking on a flow line to connect to it
                if len(self.drawing_segments) > 0:
                    # Check if click is near an existing flow line
                    edges = self.area_data.get('edges', [])
                    min_dist = 1e9
                    threshold = 15.0  # Slightly larger threshold for endpoint snap
                    snap_to_line = None
                    snap_x, snap_y = canvas_x, canvas_y
                    
                    for edge_idx, e in enumerate(edges):
                        segs = e.get('segments', [])
                        for i in range(len(segs)-1):
                            x1, y1 = segs[i]
                            x2, y2 = segs[i+1]
                            # Project point onto segment
                            dx = x2 - x1
                            dy = y2 - y1
                            length2 = dx*dx + dy*dy
                            if length2 < 1e-6:
                                continue
                            t = ((canvas_x - x1)*dx + (canvas_y - y1)*dy) / length2
                            t = max(0.0, min(1.0, t))
                            px = x1 + t*dx
                            py = y1 + t*dy
                            dist = ((canvas_x - px)**2 + (canvas_y - py)**2)**0.5
                            if dist < min_dist:
                                min_dist = dist
                                if dist <= threshold:
                                    snap_x, snap_y = px, py
                                    snap_to_line = {'edge_idx': edge_idx, 'point': (px, py)}
                    
                    # If snapped to a line, finish the drawing as line-to-line connection
                    if snap_to_line:
                        self.drawing_segments.append((snap_x, snap_y))
                        # Create a virtual junction ID for this connection point
                        junction_id = f"junction_{len(edges)}_{snap_x:.0f}_{snap_y:.0f}"
                        self._finish_drawing(self.drawing_from, junction_id, is_junction=True, junction_pos=(snap_x, snap_y))
                        return
                    
                    # Otherwise, add intermediate waypoint
                    self.drawing_segments.append((canvas_x, canvas_y))
                    self._draw_diagram()
                    # Add intermediate point; snap to nearby flow line if close
                    snap_x, snap_y = canvas_x, canvas_y
                    try:
                        edges = self.area_data.get('edges', [])
                        min_dist = 1e9
                        threshold = 8.0
                        for e in edges:
                            segs = e.get('segments', [])
                            for i in range(len(segs)-1):
                                x1, y1 = segs[i]
                                x2, y2 = segs[i+1]
                                # Project point onto segment
                                dx = x2 - x1
                                dy = y2 - y1
                                length2 = dx*dx + dy*dy
                                if length2 < 1e-6:
                                    continue
                                t = ((canvas_x - x1)*dx + (canvas_y - y1)*dy) / length2
                                t = max(0.0, min(1.0, t))
                                px = x1 + t*dx
                                py = y1 + t*dy
                                dist = ((canvas_x - px)**2 + (canvas_y - py)**2)**0.5
                                if dist < min_dist:
                                    min_dist = dist
                                    if dist <= threshold:
                                        snap_x, snap_y = px, py
                        # Use snapped point if within threshold
                    except Exception:
                        pass
                    self.drawing_segments.append((snap_x, snap_y))
                    self._draw_diagram()
            return

        # Check if clicked on a flow label or recirculation rectangle
        clicked_items = self.canvas.find_overlapping(canvas_x - 2, canvas_y - 2, canvas_x + 2, canvas_y + 2)
        logger.debug(f"Clicked items: {clicked_items}, label_items keys: {list(self.label_items.keys())}")
        for item in clicked_items:
            if item in self.label_items:
                edge_idx = self.label_items[item]
                # Check if this is a recirculation rectangle (not a regular label)
                edges = self.area_data.get('edges', [])
                if edge_idx < len(edges) and edges[edge_idx].get('is_recirculation', False):
                    # Mark this recirculation as selected for lock toggle
                    self._selected_recirculation_idx = edge_idx
                    self.selected_node = None  # Clear node selection
                    
                    # Check if recirculation is locked
                    is_locked = self.recirculation_locked.get(edge_idx, False)
                    if is_locked:
                        messagebox.showinfo("Locked", "🔒 This recirculation is locked. Unlock it first to move.")
                        return
                    # Dragging recirculation rectangle
                    self.dragging_recirculation = True
                    self.dragged_recirculation_edge_idx = edge_idx
                    self.recirculation_drag_start_x = canvas_x
                    self.recirculation_drag_start_y = canvas_y
                    logger.info(f"📦 Dragging recirculation rectangle for edge {edge_idx}")
                    return
                else:
                    # Regular label drag
                    self.dragging_label = True
                    self.dragged_label_edge_idx = edge_idx
                    self.drag_start_x = canvas_x
                    self.drag_start_y = canvas_y
                    logger.info(f"📝 Dragging label for edge {self.dragged_label_edge_idx}")
                    return

        # Normal mode - drag components
        clicked_node = self._get_node_at(canvas_x, canvas_y)
        if clicked_node:
            self.selected_node = clicked_node
            self._selected_recirculation_idx = None  # Clear recirculation selection
            self.dragging = True
            self.drag_start_x = canvas_x
            self.drag_start_y = canvas_y
            self._draw_diagram()

    def _on_canvas_drag(self, event):
        """Handle drag"""
        if self.drawing_mode:
            return
        
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)
        
        # Recirculation rectangle dragging - allow user to position it
        if self.dragging_recirculation and self.dragged_recirculation_edge_idx is not None:
            edges = self.area_data.get('edges', [])
            if self.dragged_recirculation_edge_idx < len(edges):
                edge = edges[self.dragged_recirculation_edge_idx]
                
                # Calculate the drag offset
                dx = canvas_x - self.recirculation_drag_start_x
                dy = canvas_y - self.recirculation_drag_start_y
                
                # Get current position
                loop_data = edge.get('recirculation_pos', {})
                cx = float(loop_data.get('x', 0))
                cy = float(loop_data.get('y', 0))
                
                # Update position
                new_cx = cx + dx
                new_cy = cy + dy
                
                # Update drag start for next frame
                self.recirculation_drag_start_x = canvas_x
                self.recirculation_drag_start_y = canvas_y
                
                # Update edge data
                edge['recirculation_pos'] = {'x': new_cx, 'y': new_cy}
                
                # Redraw to show new position
                self._draw_diagram()
            return
        
        # Label dragging - move canvas items directly for smooth dragging
        if self.dragging_label and self.dragged_label_edge_idx is not None:
            edges = self.area_data.get('edges', [])
            if self.dragged_label_edge_idx < len(edges):
                edge = edges[self.dragged_label_edge_idx]
                
                # Build path points from segments (works for all edges including junctions)
                segments = edge.get('segments', [])
                
                if len(segments) >= 2:
                    # Use segments directly - all manually drawn lines have segments
                    path_points = segments[:]
                else:
                    # Fallback: try to build from node centers (only for component-to-component)
                    from_id = edge.get('from')
                    to_id = edge.get('to')
                    if from_id in self.nodes_by_id and to_id in self.nodes_by_id:
                        from_node = self.nodes_by_id[from_id]
                        to_node = self.nodes_by_id[to_id]
                        from_x = from_node['x'] + from_node['width'] / 2
                        from_y = from_node['y'] + from_node['height'] / 2
                        to_x = to_node['x'] + to_node['width'] / 2
                        to_y = to_node['y'] + to_node['height'] / 2
                        path_points = [(from_x, from_y), (to_x, to_y)]
                    else:
                        # No valid path available
                        return
                
                # Find closest point on path to mouse
                min_dist = float('inf')
                best_offset = 0.5
                closest_x, closest_y = canvas_x, canvas_y
                
                for i in range(len(path_points) - 1):
                    x1, y1 = path_points[i]
                    x2, y2 = path_points[i + 1]
                    
                    segment_len = ((x2 - x1)**2 + (y2 - y1)**2)**0.5
                    if segment_len > 0:
                        t = max(0, min(1, ((canvas_x - x1) * (x2 - x1) + (canvas_y - y1) * (y2 - y1)) / segment_len**2))
                        cx = x1 + t * (x2 - x1)
                        cy = y1 + t * (y2 - y1)
                        dist = ((canvas_x - cx)**2 + (canvas_y - cy)**2)**0.5
                        
                        if dist < min_dist:
                            min_dist = dist
                            best_offset = (i + t) / (len(path_points) - 1)
                            closest_x, closest_y = cx, cy
                
                # Persist exact label coordinates for stable redraws
                edge['label_offset'] = best_offset
                edge['label_pos'] = {'x': closest_x, 'y': closest_y}
                
                # Move the label canvas items directly (smooth visual feedback)
                if self.dragged_label_edge_idx in self.edge_to_label_items:
                    box_id, text_id = self.edge_to_label_items[self.dragged_label_edge_idx]
                    # Update box position
                    self.canvas.coords(box_id, closest_x - 30, closest_y - 8, closest_x + 30, closest_y + 8)
                    # Update text position
                    self.canvas.coords(text_id, closest_x, closest_y)
            return
        
        # Component dragging
        if not self.dragging or not self.selected_node:
            return
        
        # Check if node is locked
        print(f"[DEBUG] Drag attempt: selected_node='{self.selected_node}', locked={self.locked_nodes.get(self.selected_node, False)}")
        if self.locked_nodes.get(self.selected_node, False):
            print(f"[DEBUG] Node '{self.selected_node}' is locked. Drag ignored.")
            return  # Don't move locked nodes
        canvas_x = self.canvas.canvasx(event.x)

        dx = canvas_x - self.drag_start_x
        dy = canvas_y - self.drag_start_y

        node = self.nodes_by_id[self.selected_node]
        old_x, old_y = node['x'], node['y']
        node['x'] += dx
        node['y'] += dy

        if self.snap_to_grid:
            node['x'] = round(node['x'] / self.grid_size) * self.grid_size
            node['y'] = round(node['y'] / self.grid_size) * self.grid_size

        # Update connected edges to follow the component
        actual_dx = node['x'] - old_x
        actual_dy = node['y'] - old_y
        self._update_connected_edges(self.selected_node, actual_dx, actual_dy)

        self.drag_start_x = canvas_x
        self.drag_start_y = canvas_y
        self._draw_diagram()

    def _on_canvas_release(self, event):
        """Handle release"""
        # If we were dragging a label, redraw once to finalize position
        was_dragging_label = self.dragging_label
        
        # If we were dragging recirculation box, save position
        was_dragging_recirculation = self.dragging_recirculation
        
        self.dragging = False
        self.dragging_label = False
        self.dragged_label_edge_idx = None
        self.dragging_recirculation = False
        
        # Redraw only once after drag is complete
        if was_dragging_label:
            self._draw_diagram()
            logger.info("✅ Label position saved")
        
        if was_dragging_recirculation:
            self._draw_diagram()
            logger.info("✅ Recirculation position saved")

    def _on_canvas_right_click(self, event):
        """Handle right-click - show context menu or cancel drawing"""
        if self.drawing_mode:
            self.drawing_mode = False
            self.drawing_segments = []
            self.drawing_from = None
            self.drawing_to = None
            self.canvas.config(cursor='hand2')
            messagebox.showinfo("Cancelled", "Drawing cancelled")
            return

        # Show context menu for component or empty space
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)
        clicked_node = self._get_node_at(canvas_x, canvas_y)
        
        if clicked_node:
            self.selected_node = clicked_node
            self._show_context_menu(event, clicked_node)
        else:
            # Right-clicked on empty space
            self._show_canvas_context_menu(event, canvas_x, canvas_y)
        
    def _show_context_menu(self, event, node_id):
        """Show context menu for component actions"""
        menu = tk.Menu(self.canvas, tearoff=0)
        
        # Get node info
        nodes = self.area_data.get('nodes', [])
        node_data = None
        for node in nodes:
            if node['id'] == node_id:
                node_data = node
                break
        
        node_label = node_data.get('label', node_id) if node_data else node_id
        is_locked = self.locked_nodes.get(node_id, False)
        
        # Menu title
        menu.add_command(label=f"📍 {node_label}", state='disabled',
                        background='#e8eef5', foreground='#2c3e50')
        menu.add_separator()
        
        # Edit component
        menu.add_command(label="✏️ Edit Properties", command=self._edit_node)
        
        # Lock/Unlock
        lock_text = "🔓 Unlock Position" if is_locked else "🔒 Lock Position"
        menu.add_command(label=lock_text, command=self._toggle_lock_selected)
        
        menu.add_separator()
        
        # Draw flow from/to this component
        menu.add_command(label="➡️ Draw Flow From Here",
                        command=lambda: self._start_drawing_from_node(node_id))
        menu.add_command(label="⬅️ Draw Flow To Here",
                        command=lambda: self._start_drawing_to_node(node_id))
        
        menu.add_separator()
        
        # Delete component
        menu.add_command(label="🗑️ Delete Component", command=self._delete_node,
                        foreground='#e74c3c')
        
        # Show menu at cursor position
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
    
    def _show_canvas_context_menu(self, event, canvas_x, canvas_y):
        """Show context menu for empty canvas space - add component at this location"""
        menu = tk.Menu(self.canvas, tearoff=0)
        
        # Menu title
        menu.add_command(label=f"📍 Canvas Position: ({int(canvas_x)}, {int(canvas_y)})", 
                        state='disabled', background='#e8eef5', foreground='#2c3e50')
        menu.add_separator()
        
        # Add component at this location
        menu.add_command(label="➕ Create Component Here",
                        command=lambda: self._add_component_at_position(canvas_x, canvas_y))
        
        # Show menu at cursor position
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()
    
    def _add_component_at_position(self, x, y):
        """Open add component dialog with position pre-filled"""
        # Create add dialog
        dialog = self._create_styled_dialog("Add Component at Position", 550, 750)
        
        # Title header
        header = tk.Frame(dialog, bg='#27ae60', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="➕ Add Component at Clicked Position",
                 font=('Segoe UI', 14, 'bold'), bg='#27ae60', fg='white').pack(pady=18)
        
        # Form frame
        form = tk.Frame(dialog, bg='white', padx=25, pady=20)
        form.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Position info (display only)
        tk.Label(form, text="Position:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=0, column=0, sticky='w', pady=8)
        position_text = f"X: {int(x)}, Y: {int(y)}"
        tk.Label(form, text=position_text, font=('Segoe UI', 10), bg='#e8eef5', relief='sunken', 
                 padx=10, pady=5).grid(row=0, column=1, pady=8, padx=5, sticky='ew')
        
        # Auto-generate unique component ID
        def generate_unique_id(base_label='component'):
            """Generate a unique ID based on label and counter"""
            nodes = self.area_data.get('nodes', [])
            existing_ids = {node['id'] for node in nodes}
            
            # Clean base label (remove spaces, special chars, convert to lowercase)
            base = base_label.strip().lower().replace(' ', '_').replace('-', '_')
            base = ''.join(c for c in base if c.isalnum() or c == '_')
            if not base:
                base = 'component'
            
            # Try base name first
            if base not in existing_ids:
                return base
            
            # Add counter suffix
            counter = 1
            while f"{base}_{counter}" in existing_ids:
                counter += 1
            return f"{base}_{counter}"
        
        # Component ID (auto-generated, read-only with refresh button)
        tk.Label(form, text="Component ID:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=1, column=0, sticky='w', pady=8)
        id_var = tk.StringVar(value=generate_unique_id())
        
        id_frame = tk.Frame(form, bg='white')
        id_frame.grid(row=1, column=1, pady=8, padx=5, sticky='ew')
        id_entry = tk.Entry(id_frame, textvariable=id_var, font=('Segoe UI', 10), state='readonly', width=28)
        id_entry.pack(side='left', fill='x', expand=True)
        
        def refresh_id():
            label_text = label_var.get() if 'label_var' in locals() else 'component'
            id_var.set(generate_unique_id(label_text))
        
        refresh_btn = tk.Button(id_frame, text="🔄", command=refresh_id, bg='#3498db', fg='white',
                               font=('Segoe UI', 9), padx=6, relief='flat', cursor='hand2')
        refresh_btn.pack(side='left', padx=(4, 0))
        
        tk.Label(form, text="(auto-generated, click 🔄 to refresh)", font=('Segoe UI', 8, 'italic'), bg='white', fg='#7f8c8d').grid(row=2, column=1, sticky='w', padx=5)
        
        # Label (display name)
        tk.Label(form, text="Label:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=3, column=0, sticky='w', pady=8)
        label_var = tk.StringVar(value='NEW COMPONENT')
        label_entry = tk.Entry(form, textvariable=label_var, font=('Segoe UI', 10), width=35)
        label_entry.grid(row=3, column=1, pady=8, padx=5, sticky='ew')
        
        # Auto-update ID when label changes
        def on_label_change(*args):
            refresh_id()
        label_var.trace_add('write', on_label_change)
        
        # Type
        tk.Label(form, text="Type:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=4, column=0, sticky='w', pady=8)
        type_var = tk.StringVar(value='process')
        type_combo = ttk.Combobox(form, textvariable=type_var,
                                  values=['source', 'process', 'storage', 'consumption', 'building', 
                                         'treatment', 'plant', 'tsf', 'reservoir', 'loss', 'discharge'],
                                  state='readonly', font=('Segoe UI', 10), width=15)
        type_combo.grid(row=4, column=1, sticky='w', pady=8, padx=5)
        
        # Shape
        tk.Label(form, text="Shape:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=5, column=0, sticky='w', pady=8)
        shape_var = tk.StringVar(value='rect')
        shape_combo = ttk.Combobox(form, textvariable=shape_var, values=['rect', 'oval', 'diamond'],
                                   state='readonly', font=('Segoe UI', 10), width=15)
        shape_combo.grid(row=5, column=1, sticky='w', pady=8, padx=5)
        
        # Font size (supports decimals: 7.5, 10.25, etc.)
        tk.Label(form, text="Font Size:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=6, column=0, sticky='w', pady=8)
        font_size_var = tk.DoubleVar(value=10.0)
        font_size_frame = tk.Frame(form, bg='white')
        font_size_frame.grid(row=6, column=1, sticky='w', pady=8, padx=5)
        font_size_spin = tk.Spinbox(font_size_frame, from_=4.0, to=36.0, increment=0.5, textvariable=font_size_var, font=('Segoe UI', 10), width=8)
        font_size_spin.pack(side='left', padx=2)
        tk.Label(font_size_frame, text="pt", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)
        
        # Font weight (Bold/Regular)
        tk.Label(form, text="Font Style:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=7, column=0, sticky='w', pady=8)
        font_weight_var = tk.StringVar(value='normal')
        font_weight_frame = tk.Frame(form, bg='white')
        font_weight_frame.grid(row=7, column=1, sticky='w', pady=8, padx=5)
        
        bold_btn = tk.Button(font_weight_frame, text="B  Bold", 
                            bg='#95a5a6', fg='white', font=('Segoe UI', 9, 'bold'), 
                            padx=10, relief='raised')
        regular_btn = tk.Button(font_weight_frame, text="Regular",
                               bg='#3498db', fg='white', font=('Segoe UI', 9),
                               padx=10, relief='sunken')
        
        def toggle_bold():
            current = font_weight_var.get()
            new_weight = 'normal' if current == 'bold' else 'bold'
            font_weight_var.set(new_weight)
            bold_btn.config(bg='#3498db' if new_weight == 'bold' else '#95a5a6', relief='sunken' if new_weight == 'bold' else 'raised')
            regular_btn.config(bg='#95a5a6' if new_weight != 'bold' else '#3498db', relief='sunken' if new_weight != 'bold' else 'raised')
        
        def set_regular():
            font_weight_var.set('normal')
            bold_btn.config(bg='#95a5a6', relief='raised')
            regular_btn.config(bg='#3498db', relief='sunken')
        
        bold_btn.config(command=toggle_bold)
        regular_btn.config(command=set_regular)
        bold_btn.pack(side='left', padx=2)
        regular_btn.pack(side='left', padx=2)
        
        # Size (with defaults based on type)
        tk.Label(form, text="Width:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=8, column=0, sticky='w', pady=8)
        width_var = tk.IntVar(value=120)
        width_frame = tk.Frame(form, bg='white')
        width_frame.grid(row=8, column=1, sticky='w', pady=8, padx=5)
        width_spin = tk.Spinbox(width_frame, from_=40, to=400, textvariable=width_var, font=('Segoe UI', 10), width=10)
        width_spin.pack(side='left', padx=2)
        tk.Label(width_frame, text="px", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)
        
        tk.Label(form, text="Height:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=9, column=0, sticky='w', pady=8)
        height_var = tk.IntVar(value=40)
        height_frame = tk.Frame(form, bg='white')
        height_frame.grid(row=9, column=1, sticky='w', pady=8, padx=5)
        height_spin = tk.Spinbox(height_frame, from_=20, to=200, textvariable=height_var, font=('Segoe UI', 10), width=10)
        height_spin.pack(side='left', padx=2)
        tk.Label(height_frame, text="px", font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(side='left', padx=2)
        
        # Fill color with picker
        tk.Label(form, text="Fill Color:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=10, column=0, sticky='w', pady=8)
        fill_var = tk.StringVar(value='#3498db')
        fill_frame = tk.Frame(form, bg='white')
        fill_frame.grid(row=10, column=1, sticky='w', pady=8, padx=5)
        fill_preview = tk.Canvas(fill_frame, width=30, height=25, bg=fill_var.get(), highlightthickness=1, highlightbackground='#95a5a6')
        fill_preview.pack(side='left', padx=2)
        def pick_fill_color():
            from tkinter.colorchooser import askcolor
            color = askcolor(color=fill_var.get(), title="Choose Fill Color")
            if color[1]:
                fill_var.set(color[1])
                fill_preview.config(bg=color[1])
        fill_btn = tk.Button(fill_frame, text="🎨 Pick", command=pick_fill_color, bg='#3498db', fg='white', font=('Segoe UI', 9), padx=8, relief='flat')
        fill_btn.pack(side='left', padx=2)
        fill_entry = tk.Entry(fill_frame, textvariable=fill_var, font=('Segoe UI', 9), width=10)
        fill_entry.pack(side='left', padx=2)
        
        # Outline color with picker
        tk.Label(form, text="Outline Color:", font=('Segoe UI', 10, 'bold'), bg='white').grid(row=11, column=0, sticky='w', pady=8)
        outline_var = tk.StringVar(value='#2c3e50')
        outline_frame = tk.Frame(form, bg='white')
        outline_frame.grid(row=11, column=1, sticky='w', pady=8, padx=5)
        outline_preview = tk.Canvas(outline_frame, width=30, height=25, bg=outline_var.get(), highlightthickness=1, highlightbackground='#95a5a6')
        outline_preview.pack(side='left', padx=2)
        def pick_outline_color():
            from tkinter.colorchooser import askcolor
            color = askcolor(color=outline_var.get(), title="Choose Outline Color")
            if color[1]:
                outline_var.set(color[1])
                outline_preview.config(bg=color[1])
        outline_btn = tk.Button(outline_frame, text="🎨 Pick", command=pick_outline_color, bg='#2c3e50', fg='white', font=('Segoe UI', 9), padx=8, relief='flat')
        outline_btn.pack(side='left', padx=2)
        outline_entry = tk.Entry(outline_frame, textvariable=outline_var, font=('Segoe UI', 9), width=10)
        outline_entry.pack(side='left', padx=2)
        
        form.grid_columnconfigure(1, weight=1)
        
        def create_component():
            # Auto-generate ID if empty (shouldn't happen, but safeguard)
            comp_id = id_var.get().strip()
            if not comp_id:
                comp_id = generate_unique_id(label_var.get())
                id_var.set(comp_id)
            
            # Check for duplicate ID (double-check safety)
            nodes = self.area_data.get('nodes', [])
            if any(node['id'] == comp_id for node in nodes):
                # Auto-regenerate with timestamp suffix
                import time
                comp_id = f"{comp_id}_{int(time.time() % 10000)}"
                id_var.set(comp_id)
            
            # Create new node with clicked position
            new_node = {
                'id': comp_id,
                'label': label_var.get(),
                'type': type_var.get(),
                'shape': shape_var.get(),
                'x': float(x),
                'y': float(y),
                'width': width_var.get(),
                'height': height_var.get(),
                'fill': fill_var.get(),
                'outline': outline_var.get(),
                'font_size': font_size_var.get(),
                'font_weight': font_weight_var.get(),
                'locked': False
            }
            
            # Add to area data
            self.area_data['nodes'].append(new_node)
            
            # Redraw
            self._draw_diagram()
            
            logger.info(f"✅ Added new component: {comp_id} at ({int(x)}, {int(y)})")
            messagebox.showinfo('Success', f'Component "{comp_id}" created at position ({int(x)}, {int(y)})!\n\nRemember to save your changes!')
            dialog.destroy()
        
        # Buttons
        btn_frame = tk.Frame(dialog, bg='#e8eef5')
        btn_frame.pack(fill='x', pady=15, padx=10)
        
        tk.Button(btn_frame, text="✅ Create", command=create_component,
                  bg='#27ae60', fg='white', font=('Segoe UI', 11, 'bold'),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        tk.Button(btn_frame, text="✖ Cancel", command=dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 11),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        dialog.wait_window()
    
    def _start_drawing_from_node(self, node_id):
        """Start drawing a flow line from the specified node"""
        self.drawing_mode = True
        self.drawing_from = node_id
        self.drawing_segments = []
        self.drawing_to = None
        self.canvas.config(cursor='crosshair')
        logger.info(f"✏️ Started drawing from {node_id} - click to add waypoints, click destination to finish")
        messagebox.showinfo("Draw Flow", f"Drawing from: {node_id}\n\nClick to add waypoints, then click destination component to finish")
    
    def _start_drawing_to_node(self, node_id):
        """Set this node as the destination for the next flow line"""
        messagebox.showinfo("Draw Flow To", f"Now click 'Draw Flow Line' button and select source.\nDestination will be: {node_id}")
        # This is informational - user still needs to use the Draw button

    def _on_canvas_motion(self, event):
        """Handle mouse motion - show preview line and anchor snap feedback when drawing"""
        canvas_x = self.canvas.canvasx(event.x)
        canvas_y = self.canvas.canvasy(event.y)
        
        # Check for hovered anchor points during drawing mode
        if self.drawing_mode:
            old_hover = self.hovered_anchor
            self.hovered_anchor = None
            
            # Find if hovering near any anchor point
            for node_id, anchors in self.snap_anchor_points.items():
                for anchor_name, (ax, ay) in anchors.items():
                    if anchor_name == 'center':
                        continue
                    dist = ((canvas_x - ax)**2 + (canvas_y - ay)**2)**0.5
                    if dist <= self.snap_distance:
                        self.hovered_anchor = {'node_id': node_id, 'name': anchor_name}
                        if old_hover != self.hovered_anchor:
                            self._draw_diagram()  # Redraw to show highlighted anchor
                        return
            
            # If we were hovering and now aren't, redraw
            if old_hover:
                self._draw_diagram()
        
        if not self.drawing_mode or len(self.drawing_segments) == 0:
            return

        # Only redraw if we have a preview line tag
        if hasattr(self, '_preview_line_id') and self._preview_line_id:
            self.canvas.delete(self._preview_line_id)
        
        last_x, last_y = self.drawing_segments[-1]
        
        # Snap preview to 90°
        dx = abs(canvas_x - last_x)
        dy = abs(canvas_y - last_y)
        
        if dx > dy:
            # Horizontal preview
            self._preview_line_id = self.canvas.create_line(last_x, last_y, canvas_x, last_y, 
                                   fill='#95a5a6', width=2, dash=(3, 3), tags='preview')
        else:
            # Vertical preview
            self._preview_line_id = self.canvas.create_line(last_x, last_y, last_x, canvas_y, 
                                   fill='#95a5a6', width=2, dash=(3, 3), tags='preview')

    def _zoom(self, factor):
        """Zoom canvas content by factor around origin"""
        try:
            self.canvas.scale('all', 0, 0, factor, factor)
            # Also scale scrollregion
            sr = self.canvas.cget('scrollregion')
            if sr:
                x1, y1, x2, y2 = map(float, sr.split())
                x2 *= factor
                y2 *= factor
                self.canvas.configure(scrollregion=(x1, y1, x2, y2))
        except Exception as e:
            logger.error(f"Zoom error: {e}")

    def _show_flow_type_dialog(self, from_id, to_id):
        """Show professional flow type selection dialog with scrollable options"""
        dialog = tk.Toplevel(self.canvas)
        dialog.title("Select Flow Type")
        dialog.transient(self.canvas)
        dialog.grab_set()
        
        # Center the dialog with reasonable height; scrolling keeps options visible on small screens
        dialog_width = 450
        dialog_height = 560
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - dialog_width) // 2
        y = (screen_height - dialog_height) // 2
        dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        dialog.minsize(dialog_width, dialog_height)
        
        # Header (fixed at top)
        header = tk.Label(dialog, text=f"Creating Flow Line", 
                         font=('Segoe UI', 12, 'bold'), bg='#3498db', fg='white', pady=10)
        header.pack(fill='x')
        
        # Connection info (fixed)
        info_frame = tk.Frame(dialog, bg='#e8eef5', pady=10)
        info_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(info_frame, text=f"From:", font=('Segoe UI', 9, 'bold'), 
                bg='#e8eef5').pack(anchor='w', padx=10)
        tk.Label(info_frame, text=from_id, font=('Segoe UI', 10), 
                bg='#e8eef5', fg='#2c3e50').pack(anchor='w', padx=20)
        
        tk.Label(info_frame, text=f"To:", font=('Segoe UI', 9, 'bold'), 
                bg='#e8eef5').pack(anchor='w', padx=10, pady=(5,0))
        tk.Label(info_frame, text=to_id, font=('Segoe UI', 10), 
                bg='#e8eef5', fg='#2c3e50').pack(anchor='w', padx=20)
        
        # Flow type selection label
        tk.Label(dialog, text="Select Flow Type:", 
                font=('Segoe UI', 10, 'bold'), pady=5).pack(anchor='w', padx=10)
        
        # Scrollable container for flow types
        scroll_container = tk.Frame(dialog)
        scroll_container.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Canvas for scrolling
        canvas = tk.Canvas(scroll_container, highlightthickness=0)
        scrollbar = tk.Scrollbar(scroll_container, orient="vertical", command=canvas.yview)
        
        # Frame inside canvas for content
        radio_frame = tk.Frame(canvas)
        
        # Configure canvas scrolling
        radio_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=radio_frame, anchor="nw", width=410)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Pack canvas and scrollbar
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        flow_types = [
            ("clean", "Clean Water", "Potable/treated water", "#4b78a8"),
            ("dirty", "Dirty Water", "Wastewater", "#e74c3c"),
            ("dewatering", "Dewatering", "Mine water/drainage", "#e74c3c"),
            ("ug_return", "Underground Return", "UG return water", "#e74c3c"),
            ("process_dirty", "Process Dirty", "Plant dirty water", "#e74c3c"),
            ("stormwater", "Stormwater", "Rainfall runoff", "#e74c3c"),
            ("recirculation", "Recirculation", "Internal return loop", "#9b59b6"),
            ("evaporation", "Evaporation", "Loss to air", "#000000")
        ]
        
        selected_type = tk.StringVar(value="clean")
        
        for value, label, desc, color in flow_types:
            frame = tk.Frame(radio_frame, relief='solid', borderwidth=1, bg='white')
            frame.pack(fill='x', pady=2)
            
            rb = tk.Radiobutton(frame, text=label, variable=selected_type, value=value,
                               font=('Segoe UI', 10, 'bold'), bg='white', 
                               activebackground='#e8eef5')
            rb.pack(anchor='w', padx=5, pady=2)
            
            desc_label = tk.Label(frame, text=desc, font=('Segoe UI', 8), 
                                 fg='#7f8c8d', bg='white')
            desc_label.pack(anchor='w', padx=25)
            
            color_label = tk.Label(frame, text="●", font=('Segoe UI', 16), 
                                  fg=color, bg='white')
            color_label.place(relx=0.95, rely=0.5, anchor='e')
        
        result = [None]
        
        def on_ok():
            result[0] = selected_type.get()
            canvas.unbind_all("<MouseWheel>")
            dialog.destroy()
        
        def on_cancel():
            canvas.unbind_all("<MouseWheel>")
            dialog.destroy()
        
        # Buttons (fixed at bottom)
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10, fill='x')

        ok_btn = tk.Button(btn_frame, text="OK", command=on_ok,
                  bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'),
                  padx=30, pady=6, width=12)
        ok_btn.pack(side='left', padx=10, expand=True)

        cancel_btn = tk.Button(btn_frame, text="Cancel", command=on_cancel,
                      bg='#95a5a6', fg='white', font=('Segoe UI', 10),
                      padx=30, pady=6, width=12)
        cancel_btn.pack(side='left', padx=10, expand=True)
        
        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        dialog.protocol("WM_DELETE_WINDOW", on_cancel)
        
        dialog.wait_window()
        return result[0]

    def _show_volume_dialog(self, from_id, to_id, flow_type):
        """Show professional volume input dialog"""
        dialog = tk.Toplevel(self.canvas)
        dialog.title("Enter Flow Volume")
        dialog.transient(self.canvas)
        dialog.grab_set()
        
        # Center the dialog
        dialog_width = 520
        dialog_height = 320
        screen_width = dialog.winfo_screenwidth()
        screen_height = dialog.winfo_screenheight()
        x = (screen_width - dialog_width) // 2
        y = (screen_height - dialog_height) // 2
        dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        dialog.minsize(dialog_width, dialog_height)
        
        # Header
        header = tk.Label(dialog, text=f"Flow Volume", 
                         font=('Segoe UI', 12, 'bold'), bg='#3498db', fg='white', pady=10)
        header.pack(fill='x')
        
        # Connection info
        info_frame = tk.Frame(dialog, bg='#e8eef5', pady=15)
        info_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(info_frame, text=f"From: {from_id}", font=('Segoe UI', 9), 
                bg='#e8eef5').pack(anchor='w', padx=10)
        tk.Label(info_frame, text=f"To: {to_id}", font=('Segoe UI', 9), 
                bg='#e8eef5').pack(anchor='w', padx=10)
        tk.Label(info_frame, text=f"Type: {flow_type.title()}", font=('Segoe UI', 9, 'bold'), 
                bg='#e8eef5', fg='#2980b9').pack(anchor='w', padx=10, pady=(5,0))
        
        # Volume input
        input_frame = tk.Frame(dialog)
        input_frame.pack(pady=15, padx=10, fill='x')
        
        tk.Label(input_frame, text="Volume or Label (m³/month):", 
            font=('Segoe UI', 10, 'bold')).pack(side='left', padx=5)
        
        volume_var = tk.StringVar(value='-')  # Default to '-' when no data
        volume_entry = tk.Entry(input_frame, textvariable=volume_var, 
                               font=('Segoe UI', 12), width=15)
        volume_entry.pack(side='left', padx=5)
        volume_entry.focus_set()
        volume_entry.select_range(0, tk.END)  # Select the '-' so user can type over it
        
        error_label = tk.Label(dialog, text="", font=('Segoe UI', 9), fg='#e74c3c')
        error_label.pack()
        
        result = [None]
        
        def validate_and_submit():
            val_str = volume_var.get().strip()
            if not val_str:
                error_label.config(text="Please enter a value or label")
                return
            
            # Allow any value: numbers (positive/negative), dashes, text, etc.
            result[0] = val_str
            dialog.destroy()
        
        def on_cancel():
            dialog.destroy()
        
        # Bind Enter key
        volume_entry.bind('<Return>', lambda e: validate_and_submit())
        
        # Buttons
        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=15, fill='x')

        ok_btn = tk.Button(btn_frame, text="OK", command=validate_and_submit,
                  bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'),
                  padx=30, pady=6, width=12)
        ok_btn.pack(side='left', padx=10, expand=True)

        cancel_btn = tk.Button(btn_frame, text="Cancel", command=on_cancel,
                      bg='#95a5a6', fg='white', font=('Segoe UI', 10),
                      padx=30, pady=6, width=12)
        cancel_btn.pack(side='left', padx=10, expand=True)
        
        dialog.wait_window()
        return result[0]

    def _finish_drawing(self, from_id, to_id, is_junction=False, junction_pos=None):
        """Finish drawing and create or update edge with smart edge connections
        
        Args:
            from_id: Source component ID
            to_id: Destination component ID or junction ID
            is_junction: True if connecting to a flow line junction point
            junction_pos: (x, y) tuple for junction position if is_junction=True
        """
        edges = self.area_data.get('edges', [])
        
        # Format display names for dialogs
        to_display = f"Flow Junction at ({junction_pos[0]:.0f}, {junction_pos[1]:.0f})" if is_junction else to_id
        # Detect if the source is a junction (started from an existing line)
        is_junction_start = False
        junction_start_pos = None
        try:
            if isinstance(from_id, str) and from_id.startswith("junction:"):
                parts = from_id.split(":")
                if len(parts) == 3:
                    jx = float(parts[1])
                    jy = float(parts[2])
                    is_junction_start = True
                    junction_start_pos = (jx, jy)
                    # Improve dialog display for clarity
                    from_id = f"Flow Junction at ({jx:.0f}, {jy:.0f})"
        except Exception:
            # If parsing fails, treat as normal component
            is_junction_start = False
            junction_start_pos = None
        
        # Get flow type using custom dialog
        flow_type = self._show_flow_type_dialog(from_id, to_display)
        if not flow_type:
            messagebox.showwarning("Cancelled", "Flow type required. Drawing cancelled.")
            self.drawing_mode = False
            self.drawing_segments = []
            self.drawing_from = None
            self.canvas.config(cursor='hand2')
            self._draw_diagram()
            return

        # Get volume using custom dialog
        volume = self._show_volume_dialog(from_id, to_display, flow_type)
        if volume is None:
            messagebox.showwarning("Cancelled", "Volume required. Drawing cancelled.")
            self.drawing_mode = False
            self.drawing_segments = []
            self.drawing_from = None
            self.canvas.config(cursor='hand2')
            self._draw_diagram()
            return

        # Determine color based on flow type
        color_map = {
            'clean': '#4b78a8',         # Blue
            'dirty': '#e74c3c',         # Red
            'dewatering': '#e74c3c',    # Red
            'ug_return': '#e74c3c',     # Red
            'process_dirty': '#e74c3c', # Red
            'stormwater': '#e74c3c',    # Red
            'recirculation': '#9b59b6', # Purple
            'evaporation': '#000000'    # Black
        }
        color = color_map.get(flow_type.lower(), '#3498db')

        # Format label with commas (handle non-numeric values like '-')
        try:
            label = f"{float(volume):,.2f}"
        except (ValueError, TypeError):
            # Keep non-numeric labels as-is (like '-', 'TBD', etc.)
            label = str(volume)

        # Create new edge with a unique id
        edge_id = str(uuid.uuid4())
        new_edge = {
            'id': edge_id,
            'from': from_id,
            'to': to_id,
            'segments': self.drawing_segments[:],
            'flow_type': flow_type,
            'volume': volume,
            'color': color,
            'label': label,
            'label_font_size': 8.0,  # Default edge label font size (supports decimals: 7.5, 10.25, etc.)
            'bidirectional': False,
            'is_junction': is_junction
        }
        if is_junction and junction_pos:
            new_edge['junction_pos'] = {'x': junction_pos[0], 'y': junction_pos[1]}
        if is_junction_start and junction_start_pos:
            new_edge['is_junction_start'] = True
            new_edge['junction_start_pos'] = {'x': junction_start_pos[0], 'y': junction_start_pos[1]}
        
        edges.append(new_edge)
        connection_type = "junction" if is_junction else "component"
        origin_type = "junction" if is_junction_start else "component"
        logger.info(f"✅ Created flow: {from_id} → {to_id} ({flow_type}, {label} m³) [from {origin_type} to {connection_type}]")

        self.area_data['edges'] = edges
        self.drawing_mode = False
        self.drawing_segments = []
        self.drawing_from = None
        self.canvas.config(cursor='hand2')
        
        self._draw_diagram()
        success_msg = f"Flow saved: {from_id} → {to_display}\n{flow_type.title()}: {label} m³"
        if is_junction:
            success_msg += "\n\n✅ Connected to existing flow line"
        if is_junction_start:
            success_msg += "\n\n✅ Started from existing flow line"
        messagebox.showinfo("Success", success_msg)

    def _get_node_at(self, x, y):
        """Get node at position"""
        items = self.canvas.find_overlapping(x-5, y-5, x+5, y+5)
        for item in items:
            if item in self.node_items:
                return self.node_items[item]
        return None

    def _edit_flow_unified(self):
        """Unified flow editor - handles all flow types including recirculation."""
        # Check if a flow line is selected
        if hasattr(self, 'selected_edge') and self.selected_edge:
            # Direct edit of selected flow
            self._edit_line()
        else:
            # Show picker dialog
            self._edit_line()
    
    def _open_excel_setup_unified(self):
        """Improved Excel setup with smart workflow."""
        dialog = self._create_styled_dialog("Excel Connection Setup", 950, 750)
        
        # Header
        header = tk.Frame(dialog, bg='#e67e22', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="🔧 Excel Connection Setup",
                 font=('Segoe UI', 14, 'bold'), bg='#e67e22', fg='white').pack(pady=18)
        
        # Content
        content = tk.Frame(dialog, bg='white')
        content.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Instructions
        tk.Label(content, text="Connect your flow diagram to Excel data columns",
                 font=('Segoe UI', 11), bg='white', fg='#2c3e50').pack(pady=(0,10))
        
        # Status Panel - shows current mapping status
        status_frame = tk.Frame(content, bg='#e8eef5', relief='solid', borderwidth=1)
        status_frame.pack(fill='x', pady=(0, 20), padx=5)
        
        status_inner = tk.Frame(status_frame, bg='#e8eef5')
        status_inner.pack(fill='both', padx=15, pady=12)
        
        # Calculate current status
        edges = self.area_data.get('edges', [])
        total_flows = len(edges)
        mapped_flows = sum(1 for e in edges if e.get('excel_mapping', {}).get('enabled') and 
                          e.get('excel_mapping', {}).get('sheet') and 
                          e.get('excel_mapping', {}).get('column'))
        unmapped_flows = total_flows - mapped_flows
        
        tk.Label(status_inner, text="Current Status:", font=('Segoe UI', 10, 'bold'),
                bg='#e8eef5', fg='#2c3e50').pack(anchor='w')
        
        stats_text = f"✅ {mapped_flows} flows connected  •  ⚠️ {unmapped_flows} flows need attention  •  📊 {total_flows} total flows"
        tk.Label(status_inner, text=stats_text, font=('Segoe UI', 9),
                bg='#e8eef5', fg='#34495e').pack(anchor='w', pady=(4, 0))
        
        # Action buttons with improved workflow
        btn_frame = tk.Frame(content, bg='white')
        btn_frame.pack(fill='x', pady=10)
        
        def smart_auto_map():
            """Auto-map with preview and confirmation."""
            # Show preview first
            preview_msg = (
                "This will automatically connect flows to Excel columns by:\n\n"
                "1. Matching flow names to column headers\n"
                "2. Using column aliases for renamed headers\n"
                "3. Applying intelligent pattern matching\n\n"
                "Current status:\n"
                f"• {unmapped_flows} flows need mapping\n"
                f"• {mapped_flows} flows already connected\n\n"
                "Continue with auto-mapping?"
            )
            
            if not messagebox.askyesno("Smart Auto-Map", preview_msg, icon='question'):
                return
            
            # Perform auto-map
            self._auto_map_excel_mappings()
            
            # Close dialog and show success
            dialog.destroy()
            messagebox.showinfo("Success", "Excel mappings have been automatically configured!\n\nUse 'Quick Fix' to handle any remaining unmapped flows.")
        
        def quick_fix_unmapped():
            """Show only unmapped flows for quick resolution."""
            dialog.destroy()
            self._open_quick_fix_dialog()
        
        # Primary action: Smart Auto-Map (recommended for most users)
        if unmapped_flows > 0:
            primary_color = '#27ae60'
            primary_text = f"🧭 Smart Auto-Map ({unmapped_flows} flows)"
            primary_desc = f"Automatically connect {unmapped_flows} unmapped flows to Excel columns"
        else:
            primary_color = '#16a085'
            primary_text = "🧭 Smart Auto-Map (All Connected)"
            primary_desc = "Re-run auto-mapping to refresh all connections"
        
        tk.Button(btn_frame, text=primary_text,
                  command=smart_auto_map,
                  bg=primary_color, fg='white', font=('Segoe UI', 11, 'bold'),
                  padx=20, pady=14, cursor='hand2', relief='flat').pack(pady=5, fill='x')
        
        tk.Label(btn_frame, text=primary_desc,
                 font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(pady=(0,15))
        
        # Secondary action: Quick Fix (for problem flows)
        if unmapped_flows > 0:
            fix_color = '#e67e22'
            fix_text = f"⚡ Quick Fix ({unmapped_flows} unmapped)"
            fix_desc = "Manually connect flows that couldn't be auto-mapped"
        else:
            fix_color = '#95a5a6'
            fix_text = "⚡ Quick Fix (Nothing to Fix)"
            fix_desc = "All flows are already connected"
        
        quick_fix_btn = tk.Button(btn_frame, text=fix_text,
                  command=quick_fix_unmapped,
                  bg=fix_color, fg='white', font=('Segoe UI', 10),
                  padx=20, pady=12, cursor='hand2', relief='flat',
                  state='normal' if unmapped_flows > 0 else 'disabled')
        quick_fix_btn.pack(pady=5, fill='x')
        
        tk.Label(btn_frame, text=fix_desc,
                 font=('Segoe UI', 9), bg='white', fg='#7f8c8d').pack(pady=(0,15))
        
        # Excel Column Management section
        manage_frame = tk.Frame(content, bg='#f5f6f7', relief='solid', borderwidth=1)
        manage_frame.pack(fill='x', pady=(0, 15), padx=5)
        
        manage_inner = tk.Frame(manage_frame, bg='#f5f6f7')
        manage_inner.pack(fill='both', padx=15, pady=12)
        
        tk.Label(manage_inner, text="📝 Excel Column Management:",
                font=('Segoe UI', 9, 'bold'), bg='#f5f6f7', fg='#34495e').pack(anchor='w')
        
        tk.Label(manage_inner, text="Add, rename, or delete Excel columns to match your flow diagram",
                font=('Segoe UI', 8), bg='#f5f6f7', fg='#7f8c8d').pack(anchor='w', pady=(2, 8))
        
        # Button row for column operations
        col_btn_row = tk.Frame(manage_inner, bg='#f5f6f7')
        col_btn_row.pack(fill='x')
        
        def open_column_manager():
            """Open the full column manager."""
            dialog.destroy()
            self._open_mapping_editor()
        
        tk.Button(col_btn_row, text="🔗 Manage Columns",
                 command=open_column_manager,
                 bg='#3498db', fg='white', font=('Segoe UI', 9),
                 padx=15, pady=8, cursor='hand2', relief='flat').pack(side='left')
        
        tk.Label(col_btn_row, text="Add, rename, delete, or link Excel columns",
                font=('Segoe UI', 8), bg='#f5f6f7', fg='#7f8c8d').pack(side='left', padx=(10, 0))
        
        # Help text
        help_frame = tk.Frame(content, bg='#e8f4f8', relief='solid', borderwidth=1)
        help_frame.pack(fill='x', pady=(10, 15), padx=5)
        
        help_inner = tk.Frame(help_frame, bg='#e8f4f8')
        help_inner.pack(fill='both', padx=15, pady=10)
        
        tk.Label(help_inner, text="💡 Quick Guide:",
                font=('Segoe UI', 9, 'bold'), bg='#e8f4f8', fg='#2980b9').pack(anchor='w')
        
        guide_text = (
            "• New users: Start with 'Smart Auto-Map' to automatically connect everything\n"
            "• If some flows don't auto-map: Use 'Quick Fix' to manually connect them"
        )
        tk.Label(help_inner, text=guide_text, font=('Segoe UI', 8),
                bg='#e8f4f8', fg='#34495e', justify='left').pack(anchor='w', pady=(4, 0))
        
        # Close button
        tk.Button(content, text="Close", command=dialog.destroy,
                  bg='#7f8c8d', fg='white', font=('Segoe UI', 10),
                  padx=30, pady=8, cursor='hand2', relief='flat').pack()

    def _open_quick_fix_dialog(self):
        """Streamlined dialog to fix only unmapped flows."""
        edges = self.area_data.get('edges', [])
        
        # Find unmapped flows
        unmapped = []
        for idx, edge in enumerate(edges):
            mapping = edge.get('excel_mapping', {}) or {}
            if not mapping.get('enabled') or not mapping.get('sheet') or not mapping.get('column'):
                unmapped.append((idx, edge))
        
        if not unmapped:
            messagebox.showinfo("All Connected", "✅ All flows are already connected to Excel columns!\n\nNothing to fix.")
            return
        
        dialog = self._create_styled_dialog("Quick Fix - Unmapped Flows", 900, 650)
        
        # Header
        header = tk.Frame(dialog, bg='#e67e22', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text=f"⚡ Quick Fix - {len(unmapped)} Unmapped Flows",
                 font=('Segoe UI', 14, 'bold'), bg='#e67e22', fg='white').pack(pady=18)
        
        # Content
        content = tk.Frame(dialog, bg='white')
        content.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Instructions
        instruction_text = (
            "These flows need to be connected to Excel columns. "
            "Select a flow, choose its sheet and column, then click 'Apply'."
        )
        tk.Label(content, text=instruction_text,
                 font=('Segoe UI', 10), bg='white', fg='#2c3e50', wraplength=850).pack(pady=(0,15))
        
        # Get available sheets and columns
        registry = get_excel_mapping_registry()
        loader = get_flow_volume_loader()
        
        try:
            loader.clear_cache()
        except Exception:
            pass
        
        available_sheets = set(loader.list_sheets())
        flow_sheets = sorted([s for s in available_sheets if s.startswith('Flows_')])
        if not flow_sheets:
            flow_sheets = [
                'Flows_UG2 North', 'Flows_Merensky North', 'Flows_Merensky South', 'Flows_Merensky Plant',
                'Flows_UG2 Main', 'Flows_UG2 Plant', 'Flows_Old TSF', 'Flows_Stockpile'
            ]
        
        sheet_columns = {}
        for sheet in flow_sheets:
            try:
                cols = loader.list_sheet_columns(sheet)
                sheet_columns[sheet] = [str(c).strip() for c in cols]
            except Exception:
                sheet_columns[sheet] = []
        
        # Split into left (list) and right (editor)
        main_split = tk.Frame(content, bg='white')
        main_split.pack(fill='both', expand=True)
        
        # Left panel - unmapped flows list
        left_panel = tk.Frame(main_split, bg='white', width=400)
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        tk.Label(left_panel, text="Unmapped Flows:", font=('Segoe UI', 10, 'bold'),
                bg='white', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        
        # Listbox with scrollbar
        list_frame = tk.Frame(left_panel, bg='white')
        list_frame.pack(fill='both', expand=True)
        
        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side='right', fill='y')
        
        flow_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set,
                                  font=('Consolas', 9), selectmode='single',
                                  bg='#f5f6f7', relief='solid', borderwidth=1)
        flow_listbox.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=flow_listbox.yview)
        enable_listbox_mousewheel(flow_listbox)
        
        # Populate listbox
        for idx, edge in unmapped:
            from_name = self._format_node_name(edge.get('from', ''))
            to_name = self._format_node_name(edge.get('to', ''))
            flow_listbox.insert('end', f"{idx:03d} | {from_name} → {to_name}")
        
        # Right panel - mapping editor
        right_panel = tk.Frame(main_split, bg='#f5f6f7', relief='solid', borderwidth=1)
        right_panel.pack(side='right', fill='both', expand=True)
        
        right_inner = tk.Frame(right_panel, bg='#f5f6f7')
        right_inner.pack(fill='both', expand=True, padx=15, pady=15)
        
        tk.Label(right_inner, text="Selected Flow:", font=('Segoe UI', 10, 'bold'),
                bg='#f5f6f7', fg='#2c3e50').pack(anchor='w', pady=(0, 5))
        
        flow_detail_var = tk.StringVar(value="Select a flow from the list")
        tk.Label(right_inner, textvariable=flow_detail_var, font=('Segoe UI', 9),
                bg='#f5f6f7', fg='#34495e', wraplength=400, justify='left').pack(anchor='w', pady=(0, 15))
        
        # Sheet selector
        tk.Label(right_inner, text="Excel Sheet:", font=('Segoe UI', 9, 'bold'),
                bg='#f5f6f7', fg='#2c3e50').pack(anchor='w', pady=(10, 2))
        
        sheet_var = tk.StringVar()
        sheet_combo = ttk.Combobox(right_inner, textvariable=sheet_var,
                                   values=sorted(sheet_columns.keys()), state='readonly',
                                   width=40, font=('Segoe UI', 9))
        sheet_combo.pack(anchor='w', fill='x', pady=(0, 10))
        
        # Column selector
        tk.Label(right_inner, text="Excel Column:", font=('Segoe UI', 9, 'bold'),
                bg='#f5f6f7', fg='#2c3e50').pack(anchor='w', pady=(10, 2))
        
        column_var = tk.StringVar()
        column_combo = ttk.Combobox(right_inner, textvariable=column_var,
                                    width=40, font=('Segoe UI', 9))
        column_combo.pack(anchor='w', fill='x', pady=(0, 15))
        
        # Preview
        preview_var = tk.StringVar(value="Select sheet and column to preview")
        preview_label = tk.Label(right_inner, textvariable=preview_var,
                                font=('Segoe UI', 8, 'italic'), bg='#f5f6f7',
                                fg='#7f8c8d', wraplength=400, justify='left')
        preview_label.pack(anchor='w', pady=(0, 15))
        
        # Status
        status_var = tk.StringVar(value="")
        status_label = tk.Label(right_inner, textvariable=status_var,
                               font=('Segoe UI', 9, 'bold'), bg='#f5f6f7',
                               fg='#27ae60', wraplength=400, justify='left')
        status_label.pack(anchor='w', pady=(10, 0))
        
        # Track current selection
        current_selection = {'idx': None, 'edge': None}
        fixed_count = [0]  # Use list to allow modification in nested function
        
        def update_columns(*_):
            """Update column dropdown when sheet changes."""
            selected_sheet = sheet_var.get()
            cols = sheet_columns.get(selected_sheet, [])
            column_combo['values'] = cols
            if cols:
                column_var.set('')  # Reset selection
        
        def update_preview(*_):
            """Update preview text."""
            sheet = sheet_var.get()
            col = column_var.get()
            if sheet and col:
                preview_var.set(f"✓ Will connect to: {sheet} → {col}")
            elif sheet:
                preview_var.set(f"Sheet selected: {sheet} (choose a column)")
            else:
                preview_var.set("Select sheet and column to preview")
        
        def on_flow_select(event):
            """Handle flow selection from list."""
            selection = flow_listbox.curselection()
            if not selection:
                return
            
            idx = selection[0]
            edge_idx, edge = unmapped[idx]
            current_selection['idx'] = edge_idx
            current_selection['edge'] = edge
            
            from_name = self._format_node_name(edge.get('from', ''))
            to_name = self._format_node_name(edge.get('to', ''))
            flow_detail_var.set(f"Flow: {from_name} → {to_name}\nIndex: {edge_idx}")
            
            # Try to pre-fill sheet based on source node
            mapping = edge.get('excel_mapping', {}) or {}
            if mapping.get('sheet'):
                sheet_var.set(mapping['sheet'])
            else:
                # Infer sheet from from_id
                from_id = edge.get('from', '').lower()
                for area_key, sheet_name in [
                    ('ug2n', 'Flows_UG2 North'),
                    ('mern', 'Flows_Merensky North'),
                    ('mers', 'Flows_Merensky South'),
                    ('merplant', 'Flows_Merensky Plant'),
                    ('ug2s', 'Flows_UG2 Main'),
                    ('ug2plant', 'Flows_UG2 Plant'),
                    ('oldtsf', 'Flows_Old TSF'),
                    ('stockpile', 'Flows_Stockpile')
                ]:
                    if area_key in from_id:
                        sheet_var.set(sheet_name)
                        break
            
            if mapping.get('column'):
                column_var.set(mapping['column'])
            else:
                column_var.set('')
            
            status_var.set("")
        
        def apply_mapping():
            """Apply the selected mapping to the current flow."""
            if not current_selection['edge']:
                messagebox.showwarning("No Selection", "Please select a flow from the list first.")
                return
            
            sheet = sheet_var.get()
            col = column_var.get()
            
            if not sheet or not col:
                messagebox.showwarning("Incomplete", "Please select both sheet and column.")
                return
            
            # Validate column exists in Excel
            available_cols = sheet_columns.get(sheet, [])
            if col not in available_cols:
                # Column doesn't exist - offer to create it
                create_response = messagebox.askyesno(
                    "Create Column?",
                    f"Column '{col}' doesn't exist in sheet '{sheet}'.\n\n"
                    "Would you like to create it automatically?"
                )
                
                if create_response:
                    try:
                        # Create the column in Excel
                        from openpyxl import load_workbook
                        wb = load_workbook(loader.excel_path)
                        ws = wb[sheet]
                        
                        # Append after the last header in row 3
                        header_row = 3
                        last_header_idx = 0
                        for cell in ws[header_row]:
                            val = cell.value
                            if val is not None and str(val).strip():
                                last_header_idx = cell.column
                        
                        target_col = (last_header_idx if last_header_idx > 0 else ws.max_column) + 1
                        ws.cell(row=header_row, column=target_col, value=col)
                        wb.save(loader.excel_path)
                        wb.close()
                        
                        # Refresh column cache
                        loader.clear_cache()
                        cols = loader.list_sheet_columns(sheet)
                        sheet_columns[sheet] = [str(c).strip() for c in cols]
                        update_columns()
                        
                        logger.info(f"✅ Created column '{col}' in sheet '{sheet}'")
                        status_var.set(f"✓ Column '{col}' created!")
                    except Exception as e:
                        messagebox.showerror("Error", f"Failed to create column:\n{e}")
                        return
                else:
                    # User declined to create column
                    messagebox.showinfo("Cancelled", "Mapping not saved. Please select or create a valid column.")
                    return
            
            # Apply mapping
            edge = current_selection['edge']
            edge['excel_mapping'] = {
                'enabled': True,
                'sheet': sheet,
                'column': col
            }
            
            # Update registry
            flow_id = self._flow_registry_id(edge)
            registry.link_column_to_flow(flow_id, sheet, col)
            
            # Remove from unmapped list
            current_idx = flow_listbox.curselection()[0]
            flow_listbox.delete(current_idx)
            unmapped.pop(current_idx)
            fixed_count[0] += 1
            
            # Update status
            status_var.set(f"✓ Connected! ({fixed_count[0]} fixed)")
            
            # Save to JSON
            self._save_to_json()
            
            # Select next item if available
            if unmapped:
                next_idx = min(current_idx, len(unmapped) - 1)
                flow_listbox.selection_set(next_idx)
                on_flow_select(None)
            else:
                # All done!
                dialog.destroy()
                messagebox.showinfo("All Fixed!", f"✅ Successfully connected {fixed_count[0]} flows to Excel!\n\nAll flows are now mapped.")
        
        # Wire up events
        sheet_var.trace_add('write', update_columns)
        sheet_var.trace_add('write', update_preview)
        column_var.trace_add('write', update_preview)
        flow_listbox.bind('<<ListboxSelect>>', on_flow_select)
        
        # Action buttons at bottom
        action_frame = tk.Frame(right_inner, bg='#f5f6f7')
        action_frame.pack(fill='x', pady=(20, 0))
        
        tk.Button(action_frame, text="✓ Apply Mapping",
                 command=apply_mapping,
                 bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'),
                 padx=25, pady=10, cursor='hand2', relief='flat').pack(side='left', padx=(0, 10))
        
        tk.Button(action_frame, text="Skip This Flow",
                 command=lambda: (flow_listbox.delete(flow_listbox.curselection()[0]) if flow_listbox.curselection() else None,
                                 unmapped.pop(flow_listbox.curselection()[0]) if flow_listbox.curselection() else None),
                 bg='#95a5a6', fg='white', font=('Segoe UI', 9),
                 padx=15, pady=8, cursor='hand2', relief='flat').pack(side='left')
        
        # Bottom buttons
        bottom_frame = tk.Frame(content, bg='white')
        bottom_frame.pack(fill='x', pady=(15, 0))
        
        tk.Button(bottom_frame, text=f"Save & Close ({fixed_count[0]} fixed)",
                 command=lambda: (self._save_to_json(), dialog.destroy()),
                 bg='#3498db', fg='white', font=('Segoe UI', 10),
                 padx=30, pady=10, cursor='hand2', relief='flat').pack(side='left', padx=(0, 10))
        
        tk.Button(bottom_frame, text="Cancel",
                 command=dialog.destroy,
                 bg='#7f8c8d', fg='white', font=('Segoe UI', 10),
                 padx=30, pady=10, cursor='hand2', relief='flat').pack(side='left')

    def _save_to_json(self):
        """Save to JSON"""
        try:
            # Save lock states to node data
            for node in self.area_data.get('nodes', []):
                node['locked'] = self.locked_nodes.get(node['id'], False)
            
            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump(self.area_data, f, indent=2, ensure_ascii=False)
            logger.info("✅ Saved diagram")
            messagebox.showinfo("Saved", "All changes saved!")
        except Exception as e:
            logger.error(f"❌ Save error: {e}")
            messagebox.showerror("Error", f"Failed to save: {e}")

    def _flow_registry_id(self, edge: dict) -> str:
        """Return a stable flow identifier for registry storage.
        Prefer a unique edge id when available to support parallel edges.
        """
        area_prefix = self.area_data.get('area_code') or self.area_code or 'unknown'
        edge_uid = edge.get('id')
        if edge_uid:
            return f"{area_prefix}::edge:{edge_uid}"
        # Fallback for legacy edges without id
        from_id = edge.get('from', '')
        to_id = edge.get('to', '')
        return f"{area_prefix}::{from_id}->{to_id}"

    def _auto_map_excel_mappings(self):
        """Auto-map flow excel_mapping based on Excel sheet columns."""
        try:
            registry = get_excel_mapping_registry()
            loader = get_flow_volume_loader()
            
            # Map area codes to correct sheet names
            area_to_sheet = {
                'ug2n': 'Flows_UG2 North',
                'mern': 'Flows_Merensky North',
                'mers': 'Flows_Merensky South',
                'merplant': 'Flows_Merensky Plant',
                'ug2s': 'Flows_UG2 Main',
                'ug2plant': 'Flows_UG2 Plant',
                'oldtsf': 'Flows_Old TSF',
                'stockpile': 'Flows_Stockpile',
                'newtsf': 'Flows_New TSF'
            }
            
            # Map old abbreviated sheet names to correct names (for registry compatibility)
            sheet_name_fixes = {
                'Flows_OLDTSF': 'Flows_Old TSF',
                'Flows_MERP': 'Flows_Merensky Plant',
                'Flows_NEWTSF': 'Flows_New TSF',
                'Flows_UG2P': 'Flows_UG2 Plant',
                'Flows_UG2S': 'Flows_UG2 Main',
                'Flows_MERS': 'Flows_Merensky South',
                'Flows_UG2N': 'Flows_UG2 North',
                'Flows_STOCKPILE': 'Flows_Stockpile1',
                'Flows_Stockpile': 'Flows_Stockpile1'
            }

            updated = 0
            skipped = 0
            area_data_changed = False

            for edge in self.area_data.get('edges', []):
                from_id = edge.get('from', '')
                to_id = edge.get('to', '')

                flow_id = self._flow_registry_id(edge)

                # Resolve sheet from existing mapping first
                mapping_existing = edge.get('excel_mapping', {}) or {}
                mapping = registry.resolve_mapping(flow_id, mapping_existing) or {}
                sheet = mapping.get('sheet')
                column_hint = mapping.get('column')

                if not sheet:
                    # Infer from from_id
                    from_id_lower = from_id.lower()
                    for key, sheet_name in area_to_sheet.items():
                        if key in from_id_lower:
                            sheet = sheet_name
                            break

                if not sheet:
                    skipped += 1
                    continue
                
                # Fix abbreviated sheet names from old registry data
                sheet = sheet_name_fixes.get(sheet, sheet)

                # Load sheet columns via loader (cached)
                df = loader._load_sheet(sheet)
                if df is None or df.empty:
                    skipped += 1
                    continue

                skip_cols = {'Date', 'Year', 'Month'}
                columns = [c for c in df.columns if c not in skip_cols]
                if not columns:
                    skipped += 1
                    continue

                target = f"{from_id}__TO__{to_id}".lower()

                found = None

                # If registry provided a column hint and it exists, use it directly
                if column_hint and column_hint in columns:
                    found = column_hint

                # Exact match first
                if not found:
                    for col in columns:
                        if col.lower() == target:
                            found = col
                            break

                # Check column aliases for renamed headers
                if not found:
                    alias_resolver = get_column_alias_resolver()
                    for col in columns:
                        alias_target = alias_resolver.resolve(col)
                        if alias_target and alias_target.lower() == target:
                            found = col
                            break

                # Partial match fallback (contains both parts)
                if not found:
                    from_part = from_id.lower()
                    to_part = to_id.lower()
                    for col in columns:
                        col_lower = col.lower()
                        if from_part in col_lower and to_part in col_lower:
                            found = col
                            break

                if found:
                    edge['excel_mapping'] = {
                        'enabled': True,
                        'sheet': sheet,
                        'column': found
                    }
                    registry.link_column_to_flow(flow_id, sheet, found)
                    updated += 1
                    area_data_changed = True
                else:
                    skipped += 1

            if area_data_changed:
                self._save_to_json()

            message = (
                f"Auto-map completed\n\n"
                f"✅ Updated: {updated}\n"
                f"⚠️ Skipped: {skipped}\n\n"
                "Note: Only columns present in Excel were mapped."
            )
            messagebox.showinfo("Auto-Map Excel", message)
            logger.info(f"🔁 Auto-map complete: updated={updated}, skipped={skipped}")
            self._draw_diagram()
        except Exception as e:
            logger.error(f"❌ Auto-map error: {e}")
            messagebox.showerror("Error", f"Auto-map failed: {e}")

    def _open_mapping_editor(self):
        """Open a simple dialog to edit excel_mapping for any flow."""
        # Reload latest diagram data from disk to reflect recent auto-map changes
        try:
            self._load_diagram_data()
        except Exception:
            pass
        edges = self.area_data.get('edges', [])
        if not edges:
            messagebox.showwarning("No Flows", "No flow lines to edit")
            return

        registry = get_excel_mapping_registry()

        # Prepare Excel columns per sheet using loader
        # Refresh Excel loader to pick up renamed columns
        try:
            from utils.flow_volume_loader import reset_flow_volume_loader
            reset_flow_volume_loader()
        except Exception:
            pass
        loader = get_flow_volume_loader()
        try:
            loader.clear_cache()
        except Exception:
            pass

        # Use actual available sheets from the workbook to avoid empty column lists
        available_sheets = set(loader.list_sheets())
        # Prefer Flows_* sheets for mapping
        flow_sheets = sorted([s for s in available_sheets if s.startswith('Flows_')])
        # Fallback to known area sheets only if workbook listing failed
        if not flow_sheets:
            flow_sheets = [
                'Flows_UG2 North', 'Flows_Merensky North', 'Flows_Merensky South', 'Flows_Merensky Plant',
                'Flows_UG2 Main', 'Flows_UG2 Plant', 'Flows_Old TSF', 'Flows_Stockpile'
            ]

        sheet_columns = {}
        for sheet in flow_sheets:
            try:
                cols = loader.list_sheet_columns(sheet)
                sheet_columns[sheet] = [str(c).strip() for c in cols]
            except Exception:
                sheet_columns[sheet] = []

        def _open_link_manager():
            candidates = []
            for idx, edge in enumerate(edges):
                mapping = edge.get('excel_mapping', {}) or {}
                if not mapping.get('sheet') or not mapping.get('column') or not mapping.get('enabled'):
                    candidates.append((idx, edge))

            if not candidates:
                messagebox.showinfo("All Flows Linked", "All flows already have sheet/column mappings.")
                return

            link_win = tk.Toplevel(self.canvas)
            link_win.title("Link Columns to Flows")
            link_win.transient(self.canvas)
            link_win.grab_set()
            self._center_window(link_win, 860, 540)

            header_frame = ttk.Frame(link_win, padding=(12, 10, 12, 6))
            header_frame.pack(fill='x')
            ttk.Label(header_frame, text="Link new or unmatched flows to Excel columns",
                      font=('Segoe UI', 11, 'bold')).pack(anchor='w')
            ttk.Label(header_frame, text="Select a flow, pick sheet/column, and save to persist the link",
                      font=('Segoe UI', 9)).pack(anchor='w', pady=(4, 0))

            body = ttk.Frame(link_win, padding=(12, 6, 12, 10))
            body.pack(fill='both', expand=True)

            tree_cols = ('flow', 'sheet', 'column', 'status')
            tree_link = ttk.Treeview(body, columns=tree_cols, show='headings', selectmode='browse')
            from ui.mouse_wheel_support import enable_treeview_mousewheel
            enable_treeview_mousewheel(tree_link)
            for col_name, heading, width in (
                ('flow', 'Flow', 360),
                ('sheet', 'Sheet', 140),
                ('column', 'Column', 260),
                ('status', 'Status', 80),
            ):
                tree_link.heading(col_name, text=heading)
                tree_link.column(col_name, width=width, anchor='w')
            scroll = ttk.Scrollbar(body, orient='vertical', command=tree_link.yview)
            tree_link.configure(yscrollcommand=scroll.set)
            scroll.pack(side='right', fill='y')
            tree_link.pack(side='left', fill='both', expand=True)

            row_map = {}
            for pos, (idx, edge) in enumerate(candidates):
                from_name = self._format_node_name(edge.get('from', ''))
                to_name = self._format_node_name(edge.get('to', ''))
                mapping = edge.get('excel_mapping', {}) or {}
                sheet = mapping.get('sheet', '-')
                column = mapping.get('column', '-')
                status = 'Unmapped'
                iid = str(pos)
                tree_link.insert('', 'end', iid=iid,
                                 values=(f"{idx:03d} | {from_name} → {to_name}", sheet, column, status))
                row_map[iid] = (idx, edge)

            form = ttk.Frame(link_win, padding=(12, 8, 12, 12))
            form.pack(fill='x')

            ttk.Label(form, text="Sheet:").grid(row=0, column=0, sticky='w')
            link_sheet_var = tk.StringVar(master=form)
            sheet_combo = ttk.Combobox(form, textvariable=link_sheet_var,
                                      values=sorted(sheet_columns.keys()), state='readonly', width=26)
            sheet_combo.grid(row=0, column=1, sticky='we', padx=(8, 0))

            ttk.Label(form, text="Column:").grid(row=1, column=0, sticky='w', pady=(10, 0))
            link_column_var = tk.StringVar(master=form)
            column_combo = ttk.Combobox(form, textvariable=link_column_var, width=36)
            column_combo.grid(row=1, column=1, sticky='we', padx=(8, 0), pady=(10, 0))

            persist_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(form, text="Persist link (recommended)", variable=persist_var).grid(
                row=2, column=1, sticky='w', pady=(10, 0)
            )

            status_var = tk.StringVar(value="Select a flow to link")
            ttk.Label(form, textvariable=status_var, foreground='#444').grid(
                row=3, column=0, columnspan=2, sticky='w', pady=(10, 0)
            )

            def _refresh_columns_for_sheet(*_args):
                cols = sheet_columns.get(link_sheet_var.get(), [])
                column_combo['values'] = cols

            link_sheet_var.trace_add('write', _refresh_columns_for_sheet)

            def _load_selection(*_):
                selection = tree_link.selection()
                if not selection:
                    return
                iid = selection[0]
                idx_val, edge_ref = row_map.get(iid)
                flow_id = self._flow_registry_id(edge_ref)
                mapping = edge_ref.get('excel_mapping', {}) or {}
                stored = registry.get_link(flow_id) or {}
                link_sheet_var.set(stored.get('sheet') or mapping.get('sheet', ''))
                link_column_var.set(stored.get('column') or mapping.get('column', ''))
                persist_var.set(True if stored else True)
                _refresh_columns_for_sheet()
                status_var.set(f"Editing flow {idx_val:03d}")

            tree_link.bind('<<TreeviewSelect>>', _load_selection)

            def _save_link():
                selection = tree_link.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Pick a flow to link")
                    return
                iid = selection[0]
                idx_val, edge_ref = row_map.get(iid)
                sheet_sel = link_sheet_var.get()
                col_sel = link_column_var.get()
                if not sheet_sel or not col_sel:
                    messagebox.showwarning("Missing Data", "Select both sheet and column")
                    return
                edge_ref['excel_mapping'] = {
                    'enabled': True,
                    'sheet': sheet_sel,
                    'column': col_sel
                }
                flow_id = self._flow_registry_id(edge_ref)
                if persist_var.get():
                    registry.link_column_to_flow(flow_id, sheet_sel, col_sel)
                else:
                    registry.remove_link(flow_id)
                tree_link.set(iid, 'sheet', sheet_sel)
                tree_link.set(iid, 'column', col_sel)
                tree_link.set(iid, 'status', 'Linked')
                status_var.set(f"Saved link for flow {idx_val:03d}")
                _populate_rows()

            ttk.Button(form, text="Save Link", command=_save_link, width=18).grid(
                row=4, column=1, sticky='e', pady=(12, 0)
            )

            def _close_link_manager():
                link_win.destroy()
                _populate_rows()

            ttk.Button(form, text="Close", command=_close_link_manager, width=12).grid(
                row=4, column=0, sticky='w', pady=(12, 0)
            )

        def _manual_mapper():
            """Interactive step-by-step column mapping with verification."""
            unmapped = []
            for idx, edge in enumerate(edges):
                mapping = edge.get('excel_mapping', {}) or {}
                sheet = mapping.get('sheet', '')
                column = mapping.get('column', '')
                enabled = mapping.get('enabled', False)
                
                # Include if: unmapped (no sheet/column), disabled, OR invalid (column not in sheet)
                if not sheet or not column or not enabled:
                    unmapped.append((idx, edge, "unmapped"))
                else:
                    cols = sheet_columns.get(sheet, [])
                    if column not in cols:
                        unmapped.append((idx, edge, "invalid"))
            
            if not unmapped:
                messagebox.showinfo("All Mapped", "All flows have valid Excel mappings!")
                return
            
            mapper_win = tk.Toplevel(self.canvas)
            mapper_win.title("Manual Column Mapper")
            mapper_win.transient(self.canvas)
            mapper_win.grab_set()
            self._center_window(mapper_win, 920, 680)
            
            # Header
            header_frame = ttk.Frame(mapper_win, padding=(12, 10, 12, 6))
            header_frame.pack(fill='x')
            ttk.Label(header_frame, text="Manual Column Mapper — Step-by-Step Verification",
                      font=('Segoe UI', 11, 'bold')).pack(anchor='w')
            ttk.Label(header_frame, text=f"Showing {len(unmapped)} flows with missing/invalid Excel links — verify each before saving",
                      font=('Segoe UI', 9)).pack(anchor='w', pady=(4, 0))
            
            # Navigation
            nav_frame = ttk.Frame(header_frame)
            nav_frame.pack(fill='x', pady=(8, 0))
            current_idx_var = tk.IntVar(value=0)
            progress_var = tk.StringVar(value="Flow 1 of 1")
            ttk.Label(nav_frame, textvariable=progress_var, font=('Segoe UI', 9, 'bold')).pack(side='left')
            
            # Main content
            content = ttk.Frame(mapper_win, padding=(12, 6, 12, 10))
            content.pack(fill='both', expand=True)
            
            # Flow info panel
            info_panel = ttk.LabelFrame(content, text="Current Flow", padding=(12, 8))
            info_panel.pack(fill='x', pady=(0, 10))
            
            flow_from_var = tk.StringVar(master=info_panel)
            flow_to_var = tk.StringVar(master=info_panel)
            flow_value_var = tk.StringVar(master=info_panel)
            ttk.Label(info_panel, text="From:").grid(row=0, column=0, sticky='w')
            ttk.Label(info_panel, textvariable=flow_from_var, font=('Segoe UI', 10, 'bold')).grid(row=0, column=1, sticky='w', padx=(8, 0))
            ttk.Label(info_panel, text="To:").grid(row=1, column=0, sticky='w', pady=(4, 0))
            ttk.Label(info_panel, textvariable=flow_to_var, font=('Segoe UI', 10, 'bold')).grid(row=1, column=1, sticky='w', padx=(8, 0), pady=(4, 0))
            ttk.Label(info_panel, text="Value (m³):").grid(row=2, column=0, sticky='w', pady=(4, 0))
            ttk.Label(info_panel, textvariable=flow_value_var, font=('Segoe UI', 10, 'bold')).grid(row=2, column=1, sticky='w', padx=(8, 0), pady=(4, 0))
            
            # Mapping panel
            map_panel = ttk.LabelFrame(content, text="Select Mapping", padding=(12, 8))
            map_panel.pack(fill='both', expand=True, pady=(0, 10))
            
            ttk.Label(map_panel, text="Sheet:").pack(anchor='w', pady=(0, 2))
            mapper_sheet_var = tk.StringVar(master=map_panel)
            mapper_sheet_combo = ttk.Combobox(map_panel, textvariable=mapper_sheet_var,
                                             values=sorted(sheet_columns.keys()), state='readonly', width=50)
            mapper_sheet_combo.pack(anchor='w', fill='x', pady=(0, 10))
            
            ttk.Label(map_panel, text="Column (from selected sheet):").pack(anchor='w', pady=(0, 2))
            mapper_col_var = tk.StringVar(master=map_panel)
            mapper_col_combo = ttk.Combobox(map_panel, textvariable=mapper_col_var, width=50)
            mapper_col_combo.pack(anchor='w', fill='x', pady=(0, 10))
            
            save_alias_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(map_panel, text="Save as alias (so auto-map finds renamed columns)",
                           variable=save_alias_var).pack(anchor='w', pady=(0, 8))
            
            # Preview panel
            preview_panel = ttk.LabelFrame(content, text="Preview", padding=(12, 8))
            preview_panel.pack(fill='x', pady=(0, 10))
            
            preview_text = tk.StringVar(value="Select sheet and column to preview")
            ttk.Label(preview_panel, textvariable=preview_text, foreground='#666', wraplength=600).pack(anchor='w')
            
            # Track mappings
            mapping_history = {}
            
            def _update_columns(*_):
                selected_sheet = mapper_sheet_var.get()
                cols = sheet_columns.get(selected_sheet, [])
                mapper_col_combo['values'] = cols
            
            def _update_preview(*_):
                sheet = mapper_sheet_var.get()
                col = mapper_col_var.get()
                if sheet and col:
                    preview_text.set(f"Will map to: {sheet} -> {col}")
                else:
                    preview_text.set("Select both sheet and column")
            
            mapper_sheet_var.trace_add('write', _update_columns)
            mapper_sheet_var.trace_add('write', _update_preview)
            mapper_col_var.trace_add('write', _update_preview)
            
            def _show_current():
                curr = current_idx_var.get()
                idx_val, edge_ref, status = unmapped[curr]
                from_id = edge_ref.get('from', '')
                to_id = edge_ref.get('to', '')
                # Use friendly names for display to avoid confusion
                from_name = self._format_node_name(from_id)
                to_name = self._format_node_name(to_id)
                value = edge_ref.get('volume', '-')
                
                flow_from_var.set(from_name)
                flow_to_var.set(to_name)
                flow_value_var.set(str(value) if value not in ['-', None, ''] else '-')
                progress_var.set(f"Flow {curr+1} of {len(unmapped)} ({status})")
                
                # Pre-fill if already mapped
                if idx_val in mapping_history:
                    sh, col = mapping_history[idx_val]
                    mapper_sheet_var.set(sh)
                    mapper_col_var.set(col)
                else:
                    mapper_sheet_var.set('')
                    mapper_col_var.set('')
                    _update_preview()
            
            def _save_current():
                curr = current_idx_var.get()
                idx_val, edge_ref, status = unmapped[curr]
                sheet = mapper_sheet_var.get()
                col = mapper_col_var.get()
                
                if not sheet or not col:
                    messagebox.showwarning("Incomplete", "Select both sheet and column")
                    return
                
                mapping_history[idx_val] = (sheet, col)
                if _next_flow():
                    return
                
                # Commit all mappings
                _commit_mappings()
            
            def _next_flow():
                curr = current_idx_var.get()
                if curr < len(unmapped) - 1:
                    current_idx_var.set(curr + 1)
                    _show_current()
                    return True
                return False
            
            def _prev_flow():
                curr = current_idx_var.get()
                if curr > 0:
                    current_idx_var.set(curr - 1)
                    _show_current()
            
            def _commit_mappings():
                from utils.column_alias_resolver import get_column_alias_resolver
                alias_resolver = get_column_alias_resolver()
                
                for idx_val, (sheet, col) in mapping_history.items():
                    edge_ref = edges[idx_val]
                    edge_ref['excel_mapping'] = {
                        'enabled': True,
                        'sheet': sheet,
                        'column': col
                    }
                    if save_alias_var.get():
                        from_id = edge_ref.get('from', '')
                        to_id = edge_ref.get('to', '')
                        target = f"{from_id}__TO__{to_id}"
                        alias_resolver.add_alias(col, target)
                
                self._save_to_json()
                messagebox.showinfo("Success", f"Mapped {len(mapping_history)} flows!\nAuto-map will now recognize these columns.")
                mapper_win.destroy()
                _populate_rows()
            
            # Buttons
            button_frame = ttk.Frame(mapper_win, padding=(12, 6, 12, 12))
            button_frame.pack(fill='x')
            
            ttk.Button(button_frame, text="< Previous", command=_prev_flow, width=14).pack(side='left')
            ttk.Button(button_frame, text="Save & Next", command=_save_current, width=14).pack(side='left', padx=(8, 0))
            ttk.Button(button_frame, text="Skip", 
                      command=_next_flow, width=10).pack(side='left', padx=(8, 0))
            ttk.Button(button_frame, text="Cancel", 
                      command=mapper_win.destroy, width=10).pack(side='right')
            
            _show_current()

        dialog = tk.Toplevel(self.canvas)
        dialog.title("💧 Excel Connection Manager")
        dialog.transient(self.canvas)
        dialog.grab_set()
        self._center_window(dialog, 1100, 700)
        dialog.configure(bg='#f5f7fa')

        # Modern header with gradient-like effect
        header = tk.Frame(dialog, bg='#2c3e50', height=80)
        header.pack(fill='x')
        header.pack_propagate(False)
        
        header_inner = tk.Frame(header, bg='#2c3e50')
        header_inner.pack(fill='both', expand=True, padx=20, pady=15)
        
        tk.Label(header_inner, text="💧 Excel Connection Manager", 
                font=('Segoe UI', 16, 'bold'), bg='#2c3e50', fg='white').pack(anchor='w')
        tk.Label(header_inner, text="Connect flow diagram elements to Excel data columns • Double-click to edit", 
                font=('Segoe UI', 10), bg='#2c3e50', fg='#c5d3e6').pack(anchor='w', pady=(4, 0))

        # Stats bar
        stats_bar = tk.Frame(dialog, bg='#e8eef5', height=50)
        stats_bar.pack(fill='x')
        stats_bar.pack_propagate(False)
        
        stats_inner = tk.Frame(stats_bar, bg='#e8eef5')
        stats_inner.pack(fill='x', padx=20, pady=10)
        
        # Count stats
        total_flows = len(edges)
        connected_flows = sum(1 for e in edges if e.get('excel_mapping', {}).get('enabled'))
        unmapped_flows = total_flows - connected_flows
        
        # Create a single label that can be updated
        stats_text = f"Total: {total_flows}  •  Connected: {connected_flows}  •  Unmapped: {unmapped_flows}"
        stats_label = tk.Label(stats_inner, text=stats_text, bg='#e8eef5', 
                              font=('Segoe UI', 10, 'bold'), fg='#34495e')
        stats_label.pack(side='left')

        # Action toolbar
        toolbar = tk.Frame(dialog, bg='white', height=60)
        toolbar.pack(fill='x', pady=(0, 0))
        toolbar.pack_propagate(False)
        
        toolbar_inner = tk.Frame(toolbar, bg='white')
        toolbar_inner.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Left side: Quick actions
        actions_left = tk.Frame(toolbar_inner, bg='white')
        actions_left.pack(side='left', fill='y')
        
        # Styled buttons
        btn_style = {
            'font': ('Segoe UI', 9, 'bold'),
            'cursor': 'hand2',
            'relief': 'flat',
            'bd': 0,
            'pady': 8,
            'padx': 15
        }
        
        # Fix Invalid and Edit All buttons removed - streamlined to just Auto-Map and Quick Fix
        # tk.Button(actions_left, text="⚠️ Fix Invalid", command=lambda: _open_edit_all_mappings(True),
        #          bg='#3498db', fg='white', activebackground='#2980b9', activeforeground='white',
        #          **btn_style).pack(side='left', padx=(0, 8))
        # tk.Button(actions_left, text="📋 Edit All", command=lambda: _open_edit_all_mappings(False),
        #          bg='#9b59b6', fg='white', activebackground='#8e44ad', activeforeground='white',
        #          **btn_style).pack(side='left', padx=(0, 8))
        
        tk.Button(actions_left, text="🔗 Link Columns", command=lambda: _open_link_manager(),
                 bg='#16a085', fg='white', activebackground='#138d75', activeforeground='white',
                 **btn_style).pack(side='left', padx=(0, 8))
        
        tk.Button(actions_left, text="↻ Refresh", command=lambda: _refresh_headers(),
                 bg='#34495e', fg='white', activebackground='#2c3e50', activeforeground='white',
                 **btn_style).pack(side='left')
        
        # Right side: Search/Filter
        search_frame = tk.Frame(toolbar_inner, bg='white')
        search_frame.pack(side='right', fill='y')
        
        tk.Label(search_frame, text="🔍", font=('Segoe UI', 12), bg='white', fg='#7f8c8d').pack(side='left', padx=(0, 5))
        
        filter_var = tk.StringVar(master=search_frame)
        filter_entry = tk.Entry(search_frame, textvariable=filter_var, 
                               font=('Segoe UI', 10), width=30,
                               bg='#e8eef5', fg='#2c3e50', relief='flat', bd=0)
        filter_entry.pack(side='left', ipady=6, ipadx=10)
        filter_entry.insert(0, "Search flows, sheets, or columns...")
        filter_entry.config(fg='#95a5a6')
        
        def on_focus_in(e):
            if filter_entry.get() == "Search flows, sheets, or columns...":
                filter_var.set('')  # Clear the variable too
                filter_entry.delete(0, 'end')
                filter_entry.config(fg='#2c3e50')
        
        def on_focus_out(e):
            if not filter_entry.get():
                filter_var.set("Search flows, sheets, or columns...")  # Restore placeholder in variable
                filter_entry.insert(0, "Search flows, sheets, or columns...")
                filter_entry.config(fg='#95a5a6')
        
        filter_entry.bind('<FocusIn>', on_focus_in)
        filter_entry.bind('<FocusOut>', on_focus_out)
        
        def _open_edit_all_mappings(invalid_only: bool = False):
            """View and edit flow mappings in a searchable table.
            Pass invalid_only=True to show only flows with missing/invalid Excel links.
            """
            edit_win = tk.Toplevel(self.canvas)
            edit_win.title("Edit All Flow Mappings")
            edit_win.transient(self.canvas)
            edit_win.grab_set()
            self._center_window(edit_win, 1100, 700)
            
            # Header (consistent color block)
            header = tk.Frame(edit_win, bg='#2c3e50', height=70)
            header.pack(fill='x')
            header.pack_propagate(False)
            title_inner = tk.Frame(header, bg='#2c3e50')
            title_inner.pack(fill='both', expand=True, padx=16, pady=12)
            tk.Label(title_inner, text=f"💧 Edit All Flow Mappings — {len(edges)} Flows",
                     font=('Segoe UI', 13, 'bold'), bg='#2c3e50', fg='white').pack(anchor='w')
            tk.Label(title_inner, text="Click a flow to edit its Excel mapping. Invalid mappings marked with ⚠️",
                     font=('Segoe UI', 9), bg='#2c3e50', fg='#e8eef5').pack(anchor='w')
            
            # Stats bar
            stats_bar = tk.Frame(edit_win, bg='#e8eef5', height=45)
            stats_bar.pack(fill='x')
            stats_bar.pack_propagate(False)
            stats_inner = tk.Frame(stats_bar, bg='#e8eef5')
            stats_inner.pack(fill='x', padx=16, pady=10)
            
            # Calculate stats
            total_flows = len(edges)
            connected_flows = sum(1 for e in edges if e.get('excel_mapping', {}).get('enabled') and e.get('excel_mapping', {}).get('sheet') and e.get('excel_mapping', {}).get('column'))
            unmapped_flows = total_flows - connected_flows
            
            tk.Label(stats_inner, text="✅", font=('Segoe UI', 11), bg='#e8eef5', fg='#27ae60').pack(side='left', padx=(0, 4))
            tk.Label(stats_inner, text=f"{connected_flows} connected", font=('Segoe UI', 9, 'bold'), bg='#e8eef5', fg='#27ae60').pack(side='left', padx=(0, 12))
            tk.Label(stats_inner, text="•", font=('Segoe UI', 9), bg='#e8eef5', fg='#7f8c8d').pack(side='left', padx=(0, 12))
            tk.Label(stats_inner, text="⚠️", font=('Segoe UI', 11), bg='#e8eef5', fg='#e74c3c').pack(side='left', padx=(0, 4))
            tk.Label(stats_inner, text=f"{unmapped_flows} need attention", font=('Segoe UI', 9, 'bold'), bg='#e8eef5', fg='#e74c3c').pack(side='left', padx=(0, 12))
            tk.Label(stats_inner, text="•", font=('Segoe UI', 9), bg='#e8eef5', fg='#7f8c8d').pack(side='left', padx=(0, 12))
            tk.Label(stats_inner, text="📊", font=('Segoe UI', 11), bg='#e8eef5', fg='#34495e').pack(side='left', padx=(0, 4))
            tk.Label(stats_inner, text=f"{total_flows} total flows", font=('Segoe UI', 9, 'bold'), bg='#e8eef5', fg='#34495e').pack(side='left')
            
            # Search bar
            search_frame = ttk.Frame(header, padding=(0, 8, 0, 0))
            search_frame.pack(fill='x')
            ttk.Label(search_frame, text="Search:").pack(side='left')
            search_var = tk.StringVar(master=search_frame)
            search_entry = ttk.Entry(search_frame, textvariable=search_var, width=40)
            search_entry.pack(side='left', padx=(8, 0), fill='x', expand=True)
            # Invalid-only toggle
            invalid_only_var = tk.BooleanVar(value=invalid_only)
            ttk.Checkbutton(search_frame, text="Show invalid/missing only", variable=invalid_only_var).pack(side='right')
            
            # Table frame
            table_frame = ttk.Frame(edit_win)
            table_frame.pack(fill='both', expand=True, padx=12, pady=6)
            
            # Headers for table
            headers_frame = ttk.Frame(table_frame)
            headers_frame.pack(fill='x', pady=(0, 4))
            ttk.Label(headers_frame, text="Flow", width=30, font=('Segoe UI', 9, 'bold')).pack(side='left')
            ttk.Label(headers_frame, text="Value (m³)", width=10, font=('Segoe UI', 9, 'bold')).pack(side='left')
            ttk.Label(headers_frame, text="Sheet", width=16, font=('Segoe UI', 9, 'bold')).pack(side='left')
            ttk.Label(headers_frame, text="Column", width=30, font=('Segoe UI', 9, 'bold')).pack(side='left')
            
            # Scrollable list
            list_canvas = tk.Canvas(table_frame, bg='white', highlightthickness=0)
            scrollbar = ttk.Scrollbar(table_frame, orient='vertical', command=list_canvas.yview)
            list_canvas.config(yscrollcommand=scrollbar.set)
            
            list_frame = ttk.Frame(list_canvas)
            list_canvas.create_window((0, 0), window=list_frame, anchor='nw')
            
            list_canvas.pack(side='left', fill='both', expand=True)
            scrollbar.pack(side='right', fill='y')
            
            def _populate_list():
                # Clear existing
                for widget in list_frame.winfo_children():
                    widget.destroy()
                
                search_term = search_var.get().lower()
                shown = 0
                
                for idx, edge in enumerate(edges):
                    # Show friendly names instead of internal IDs
                    flow_id = f"{self._format_node_name(edge.get('from', '?'))} → {self._format_node_name(edge.get('to', '?'))}"
                    mapping = edge.get('excel_mapping', {}) or {}
                    sheet = mapping.get('sheet', '')
                    column = mapping.get('column', '')
                    volume = edge.get('volume', '-')
                    
                    # Filter by search (matches flow, sheet, or column)
                    if search_term and (
                        search_term not in flow_id.lower() and
                        search_term not in str(sheet).lower() and
                        search_term not in str(column).lower()
                    ):
                        continue
                    
                    # Check if valid
                    is_valid = True
                    if not sheet or not column:
                        is_valid = False
                    else:
                        cols = sheet_columns.get(sheet, [])
                        if column not in cols:
                            is_valid = False
                    
                    # Skip valid rows when invalid-only is enabled
                    if invalid_only_var.get() and is_valid:
                        continue

                    row = ttk.Frame(list_frame)
                    row.pack(fill='x', pady=1)
                    
                    # Status indicator
                    status_text = "✅" if is_valid else "⚠️"
                    ttk.Label(row, text=status_text, width=2).pack(side='left', padx=(2, 4))
                    
                    # Flow name (clickable button)
                    def make_edit_handler(edge_idx):
                        def edit_this():
                            _edit_single_mapping(edge_idx)
                            _populate_list()  # Refresh
                        return edit_this
                    
                    btn = ttk.Button(row, text=flow_id[:27], command=make_edit_handler(idx), width=28)
                    btn.pack(side='left', padx=0)
                    
                    # Volume/Value display
                    volume_str = str(volume) if volume not in ['-', None, ''] else '-'
                    volume_color = '#000' if volume not in ['-', None, ''] else '#999'
                    ttk.Label(row, text=volume_str[:9], width=10, foreground=volume_color).pack(side='left', padx=2)
                    
                    # Sheet
                    ttk.Label(row, text=sheet if sheet else "(none)", width=16, 
                             foreground='#888' if not sheet else '#000').pack(side='left', padx=2)
                    
                    # Column (truncated)
                    col_display = column[:30] if column else "(none)"
                    if column and len(column) > 30:
                        col_display = column[:27] + "..."
                    ttk.Label(row, text=col_display, foreground='#888' if not column else '#000').pack(side='left', padx=2)
                    
                    shown += 1
                
                if shown == 0:
                    no_match = ttk.Label(list_frame, text="No flows match search", foreground='#999')
                    no_match.pack(fill='x', padx=8, pady=20)
                
                # Update canvas scroll region
                list_frame.update_idletasks()
                list_canvas.config(scrollregion=list_canvas.bbox('all'))
            
            def _edit_single_mapping(edge_idx):
                """Edit a single flow's mapping."""
                edge = edges[edge_idx]
                mapping = edge.get('excel_mapping', {}) or {}
                
                edit_single_win = tk.Toplevel(edit_win)
                edit_single_win.title("Edit Mapping")
                edit_single_win.transient(edit_win)
                edit_single_win.grab_set()
                self._center_window(edit_single_win, 600, 280)
                
                # Flow info
                info = ttk.Frame(edit_single_win, padding=(12, 10, 12, 6))
                info.pack(fill='x')
                flow_id = f"{self._format_node_name(edge.get('from', '?'))} → {self._format_node_name(edge.get('to', '?'))}"
                flow_value = edge.get('volume', '-')
                value_str = str(flow_value) if flow_value not in ['-', None, ''] else '-'
                info_text = f"Flow: {flow_id}  |  Value: {value_str} m³"
                ttk.Label(info, text=info_text, font=('Segoe UI', 10, 'bold')).pack(anchor='w')
                
                # Sheet selection
                body = ttk.Frame(edit_single_win, padding=(12, 6, 12, 10))
                body.pack(fill='both', expand=True)
                
                ttk.Label(body, text="Sheet:", font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(0, 4))
                sheet_var = tk.StringVar(value=mapping.get('sheet', ''))
                sheet_combo = ttk.Combobox(body, textvariable=sheet_var,
                                          values=sorted(sheet_columns.keys()), state='readonly', width=50)
                sheet_combo.pack(fill='x', pady=(0, 12))
                
                # Column selection
                ttk.Label(body, text="Column:", font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(0, 4))
                col_var = tk.StringVar(value=mapping.get('column', ''))
                col_combo = ttk.Combobox(body, textvariable=col_var, width=50)
                col_combo.pack(fill='x', pady=(0, 12))
                
                # Update columns when sheet changes
                def _on_sheet_change(*_):
                    sheet = sheet_var.get()
                    cols = sheet_columns.get(sheet, [])
                    col_combo['values'] = sorted(cols)
                
                sheet_var.trace_add('write', _on_sheet_change)
                _on_sheet_change()  # Initialize
                
                # Buttons
                btn_frame = ttk.Frame(body)
                btn_frame.pack(fill='x')
                
                def _save_edit():
                    sheet = sheet_var.get()
                    col = col_var.get()
                    if not sheet or not col:
                        messagebox.showwarning("Incomplete", "Select both sheet and column")
                        return
                    # Validate presence and offer auto-create if missing
                    cols = sheet_columns.get(sheet, [])
                    if col not in cols:
                        if not messagebox.askyesno("Column Missing",
                                                    f"Column '{col}' not found in sheet '{sheet}'.\n\nCreate it on row 3 and continue?"):
                            return
                        try:
                            excel_path = getattr(self.flow_loader, 'excel_path', None)
                            if not excel_path or not Path(excel_path).exists():
                                messagebox.showerror(
                                    "Excel File Missing",
                                    f"Flow Diagram Excel file not found:\n\n{excel_path}\n\n"
                                    "This file contains flow volumes for the diagram visualization.\n"
                                    "Please configure the path in Settings > Data Sources (timeseries_excel_path)."
                                )
                                return
                            wb = load_workbook(excel_path)
                            ws = wb[sheet]
                            header_row = 3
                            last_header_idx = 0
                            for cell in ws[header_row]:
                                val = cell.value
                                if val is not None and str(val).strip():
                                    last_header_idx = cell.column
                            target_col = (last_header_idx if last_header_idx > 0 else ws.max_column) + 1
                            ws.cell(row=header_row, column=target_col, value=col)
                            wb.save(excel_path)
                            wb.close()
                            # Update caches
                            sheet_columns[sheet] = cols + [col]
                            try:
                                self.flow_loader.clear_cache()
                            except Exception:
                                pass
                        except Exception as ce:
                            messagebox.showerror("Error", f"Failed to create column:\n{ce}")
                            return
                    
                    # Update edge
                    edges[edge_idx]['excel_mapping'] = {
                        'enabled': True,
                        'sheet': sheet,
                        'column': col
                    }
                    
                    # Persist diagram
                    try:
                        with open(diagram_file, 'w') as f:
                            json.dump(diagram, f, indent=2)
                        logger.info(f"✅ Updated mapping for {flow_id}: {sheet}.{col}")
                    except Exception as e:
                        logger.error(f"Failed to save mapping: {e}", exc_info=True)
                    
                    edit_single_win.destroy()
                
                ttk.Button(btn_frame, text="Save", command=_save_edit).pack(side='left', padx=(0, 4))
                ttk.Button(btn_frame, text="Cancel", command=edit_single_win.destroy).pack(side='left')
            
            # Footer
            footer = ttk.Frame(edit_win, padding=(12, 6, 12, 10))
            footer.pack(fill='x')
            ttk.Button(footer, text="Close", command=edit_win.destroy).pack(side='right')
            
            # Initial populate
            _populate_list()
            
            # Re-populate on search/toggle
            search_var.trace_add('write', lambda *_: _populate_list())
            invalid_only_var.trace_add('write', lambda *_: _populate_list())
        
        def _refresh_headers():
            try:
                from utils.flow_volume_loader import reset_flow_volume_loader
                reset_flow_volume_loader()
            except Exception:
                pass
            nonlocal sheet_columns
            loader_local = get_flow_volume_loader()
            try:
                loader_local.clear_cache()
            except Exception:
                pass
            for sheet in list(sheet_columns.keys()):
                df = loader_local._load_sheet(sheet)
                if df is not None and not df.empty:
                    cols = [str(c).strip() for c in df.columns if str(c).strip() not in {'Date', 'Year', 'Month'}]
                    sheet_columns[sheet] = cols
                else:
                    sheet_columns[sheet] = []
            messagebox.showinfo("✅ Headers Refreshed", "Excel column headers reloaded successfully!")
            _populate_rows()

        # Main content area with modern table
        content_frame = tk.Frame(dialog, bg='#f5f7fa')
        content_frame.pack(fill='both', expand=True, padx=20, pady=(10, 15))

        # Table container with shadow effect
        table_container = tk.Frame(content_frame, bg='white', relief='flat', bd=0)
        table_container.pack(fill='both', expand=True)
        
        # Modern styled treeview
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Modern.Treeview",
                       background="white",
                       foreground="#2c3e50",
                       rowheight=32,
                       fieldbackground="white",
                       borderwidth=0,
                       font=('Segoe UI', 9))
        style.map('Modern.Treeview',
                 background=[('selected', '#3498db')])
        style.configure("Modern.Treeview.Heading",
                       background="#34495e",
                       foreground="white",
                       relief="flat",
                       font=('Segoe UI', 10, 'bold'))
        style.map("Modern.Treeview.Heading",
                 background=[('active', '#2c3e50')])

        columns = ('id', 'flow', 'sheet', 'column', 'enabled')
        tree = ttk.Treeview(table_container, columns=columns, show='headings', 
                          selectmode='browse', style="Modern.Treeview")
        from ui.mouse_wheel_support import enable_treeview_mousewheel
        enable_treeview_mousewheel(tree)
        
        tree.heading('id', text='#')
        tree.heading('flow', text='Flow Connection')
        tree.heading('sheet', text='Excel Sheet')
        tree.heading('column', text='Excel Column')
        tree.heading('enabled', text='Status')
        
        tree.column('id', width=50, anchor='center')
        tree.column('flow', width=400, anchor='w')
        tree.column('sheet', width=150, anchor='w')
        tree.column('column', width=350, anchor='w')
        tree.column('enabled', width=100, anchor='center')

        scrollbar = ttk.Scrollbar(table_container, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side='right', fill='y')
        tree.pack(side='left', fill='both', expand=True)

        # Alternating row colors
        tree.tag_configure('connected', background='#e8f8f5', foreground='#27ae60')
        tree.tag_configure('disconnected', background='#fadbd8', foreground='#e74c3c')
        tree.tag_configure('odd_connected', background='#d5f4e6', foreground='#27ae60')
        tree.tag_configure('odd_disconnected', background='#f5b7b1', foreground='#c0392b')

        # Populate list
        def _populate_rows():
            """Populate tree with flows and update stats."""
            tree.delete(*tree.get_children())
            
            total_flows = len(edges)
            connected_flows = 0
            unmapped_flows = 0
            
            for idx, edge in enumerate(edges):
                from_name = self._format_node_name(edge.get('from', ''))
                to_name = self._format_node_name(edge.get('to', ''))
                mapping = edge.get('excel_mapping', {}) or {}
                sheet = mapping.get('sheet', '-')
                column = mapping.get('column', '-')
                enabled = mapping.get('enabled', False)
                
                # Track stats
                if sheet != '-' and column != '-' and enabled:
                    connected_flows += 1
                    status = '✅ Connected'
                    tag = 'connected' if idx % 2 == 0 else 'odd_connected'
                else:
                    unmapped_flows += 1
                    status = '❌ Disconnected'
                    tag = 'disconnected' if idx % 2 == 0 else 'odd_disconnected'
                
                flow_text = f"{from_name} → {to_name}"
                
                # Apply filter if set (ignore placeholder text)
                filter_text = filter_var.get()
                if filter_text and filter_text != "Search flows, sheets, or columns...":
                    q = filter_text.lower()
                    if q not in flow_text.lower() and q not in str(sheet).lower() and q not in str(column).lower():
                        continue
                
                tree.insert('', 'end', iid=str(idx), 
                          values=(f"{idx+1}", flow_text, sheet, column, status), 
                          tags=(tag,))
            
            # Update stats in header
            stats_label.config(text=f"Total: {total_flows}  •  Connected: {connected_flows}  •  Unmapped: {unmapped_flows}")
        
        _populate_rows()
        
        def _on_filter_change(*_):
            """Re-populate when filter changes."""
            _populate_rows()
        
        filter_var.trace_add('write', _on_filter_change)

        def edit_selected(event=None):
            """Edit a selected flow mapping with modern dialog."""
            selection = tree.selection()
            if not selection:
                messagebox.showwarning("⚠️ No Selection", "Please select a flow to edit")
                return
            item_id = selection[0]
            idx = int(item_id)
            edge = edges[idx]
            flow_id = self._flow_registry_id(edge)

            edit_win = tk.Toplevel(dialog)
            edit_win.title("Edit Flow Mapping")
            edit_win.configure(bg='#f5f7fa')
            self._center_window(edit_win, 580, 520)

            # Header
            header_frame = tk.Frame(edit_win, bg='#3498db', height=60)
            header_frame.pack(fill='x')
            header_frame.pack_propagate(False)
            
            tk.Label(header_frame, text="✏️ Edit Excel Mapping", 
                    bg='#3498db', fg='white', 
                    font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=20, pady=15)

            # Form container
            form = tk.Frame(edit_win, bg='white', padx=25, pady=20)
            form.pack(fill='both', expand=True, padx=20, pady=(15, 10))

            # Flow info (read-only)
            tk.Label(form, text="Flow Connection:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            flow_label = tk.Label(form, 
                                 text=f"{edge.get('from')} → {edge.get('to')}", 
                                 bg='#e8eef5', fg='#2c3e50',
                                 font=('Segoe UI', 10, 'bold'),
                                 anchor='w', padx=12, pady=8)
            flow_label.pack(fill='x', pady=(0, 15))

            # Enabled checkbox
            enabled_var = tk.BooleanVar(value=edge.get('excel_mapping', {}).get('enabled', False))
            enabled_check = tk.Checkbutton(form, text="✅ Enable Excel mapping for this flow", 
                                          variable=enabled_var, bg='white', 
                                          font=('Segoe UI', 9, 'bold'),
                                          fg='#27ae60', selectcolor='white',
                                          activebackground='white', activeforeground='#27ae60')
            enabled_check.pack(anchor='w', pady=(0, 15))

            # Sheet selection
            tk.Label(form, text="Excel Sheet:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            sheet_var = tk.StringVar(value=edge.get('excel_mapping', {}).get('sheet', ''))
            sheet_combo = ttk.Combobox(form, textvariable=sheet_var,
                                      values=sorted(sheet_columns.keys()), 
                                      state='readonly', font=('Segoe UI', 10))
            sheet_combo.pack(anchor='w', fill='x', pady=(0, 15))

            # Column selection
            tk.Label(form, text="Excel Column:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            
            column_row = tk.Frame(form, bg='white')
            column_row.pack(anchor='w', fill='x', pady=(0, 15))
            
            column_var = tk.StringVar(value=edge.get('excel_mapping', {}).get('column', ''))
            column_combo = ttk.Combobox(column_row, textvariable=column_var, font=('Segoe UI', 10))
            column_combo.pack(side='left', fill='x', expand=True, padx=(0, 8))
            
            def rename_current_column():
                """Rename the currently selected Excel column."""
                current_sheet = sheet_var.get()
                current_column = column_var.get()
                
                if not current_sheet:
                    messagebox.showwarning("No Sheet", "Please select a sheet first")
                    return
                if not current_column or current_column == '-':
                    messagebox.showwarning("No Column", "Please select a column to rename")
                    return
                if current_column in ['Date', 'Year', 'Month']:
                    messagebox.showwarning("Cannot Rename", "Standard columns (Date, Year, Month) cannot be renamed")
                    return
                
                # Create modern rename dialog
                rename_dialog = tk.Toplevel(edit_win)
                rename_dialog.title("Rename Excel Column")
                rename_dialog.configure(bg='#f5f7fa')
                self._center_window(rename_dialog, 520, 320)
                rename_dialog.transient(edit_win)
                rename_dialog.grab_set()
                
                # Header
                header = tk.Frame(rename_dialog, bg='#3498db', height=50)
                header.pack(fill='x')
                header.pack_propagate(False)
                tk.Label(header, text="✏️ Rename Excel Column", 
                        bg='#3498db', fg='white', 
                        font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=20, pady=15)
                
                # Body
                body = tk.Frame(rename_dialog, bg='white')
                body.pack(fill='both', expand=True, padx=25, pady=20)
                
                # Current name display
                tk.Label(body, text="Current Name:", bg='white', fg='#7f8c8d',
                        font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
                current_display = tk.Label(body, text=current_column, bg='#e8eef5', fg='#2c3e50',
                                          font=('Segoe UI', 10, 'bold'),
                                          anchor='w', padx=12, pady=8)
                current_display.pack(fill='x', pady=(0, 15))
                
                # New name input
                tk.Label(body, text="New Name:", bg='white', fg='#7f8c8d',
                        font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
                new_name_var = tk.StringVar(value=current_column)
                name_entry = tk.Entry(body, textvariable=new_name_var, 
                                     font=('Segoe UI', 11), 
                                     bg='#e8eef5', fg='#2c3e50',
                                     relief='flat', bd=0, insertbackground='#3498db')
                name_entry.pack(fill='x', pady=(0, 20), ipady=8, ipadx=10)
                name_entry.focus()
                name_entry.select_range(0, tk.END)  # Pre-select text
                
                result = [None]  # Mutable to capture result
                
                def apply_rename():
                    new_name = new_name_var.get().strip()
                    if not new_name:
                        messagebox.showwarning("Empty Name", "Column name cannot be empty", parent=rename_dialog)
                        return
                    if new_name == current_column:
                        messagebox.showinfo("No Change", "New name is the same as current name", parent=rename_dialog)
                        return
                    result[0] = new_name
                    rename_dialog.destroy()
                
                def cancel_rename():
                    rename_dialog.destroy()
                
                # Buttons
                button_frame = tk.Frame(body, bg='white')
                button_frame.pack(fill='x', padx=0, pady=0)
                
                cancel_btn = tk.Button(button_frame, text="Cancel", 
                                      command=cancel_rename,
                                      bg='#95a5a6', fg='white',
                                      font=('Segoe UI', 10),
                                      relief='flat', padx=20, pady=8,
                                      cursor='hand2')
                cancel_btn.pack(side='right', padx=(8, 0))
                
                rename_btn = tk.Button(button_frame, text="✏️ Rename", 
                                      command=apply_rename,
                                      bg='#27ae60', fg='white',
                                      font=('Segoe UI', 10, 'bold'),
                                      relief='flat', padx=20, pady=8,
                                      cursor='hand2')
                rename_btn.pack(side='right')
                
                # Allow Enter to confirm
                name_entry.bind('<Return>', lambda e: apply_rename())
                
                # Wait for dialog to close
                edit_win.wait_window(rename_dialog)
                
                if result[0] is None:
                    return  # User cancelled
                
                new_name = result[0]
                
                try:
                    from openpyxl import load_workbook
                    excel_path = self.flow_loader.excel_path
                    wb = load_workbook(excel_path)
                    ws = wb[current_sheet]
                    
                    # Find and rename in row 3 (actual header row)
                    for cell in ws[3]:
                        if cell.value == current_column:
                            cell.value = new_name
                            break
                    
                    wb.save(excel_path)
                    wb.close()
                    
                    # Update JSON mappings for all flows using this column
                    self._load_diagram_data()
                    edges_to_update = self.area_data.get('edges', [])
                    updated_count = 0
                    for edge_item in edges_to_update:
                        mapping = edge_item.get('excel_mapping', {}) or {}
                        if (mapping.get('enabled') and 
                            mapping.get('sheet') == current_sheet and 
                            mapping.get('column') == current_column):
                            mapping['column'] = new_name
                            edge_item['excel_mapping'] = mapping
                            updated_count += 1
                    
                    # Save updated JSON
                    with open(self.json_file, 'w', encoding='utf-8') as f:
                        json.dump(self.area_data, f, indent=2, ensure_ascii=False)
                    
                    # Clear cache and update UI
                    self.flow_loader.clear_cache()
                    column_var.set(new_name)
                    refresh_columns()
                    
                    msg = f"✅ Column renamed to '{new_name}'"
                    if updated_count > 0:
                        msg += f"\n({updated_count} flow mapping(s) updated)"
                    messagebox.showinfo("Success", msg)
                    
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to rename column:\n{e}")
            
            rename_col_btn = tk.Button(column_row, text="✏️", 
                                      command=rename_current_column,
                                      bg='#3498db', fg='white',
                                      font=('Segoe UI', 10),
                                      width=3,
                                      relief='flat', padx=8, pady=6,
                                      cursor='hand2',
                                      activebackground='#2980b9',
                                      activeforeground='white')
            rename_col_btn.pack(side='left')

            # Persist checkbox
            persist_var = tk.BooleanVar(value=bool(registry.get_link(flow_id)))
            persist_check = tk.Checkbutton(form, 
                                          text="🔗 Remember this link (survives column renames)", 
                                          variable=persist_var, bg='white',
                                          font=('Segoe UI', 9),
                                          fg='#34495e', selectcolor='white',
                                          activebackground='white', activeforeground='#2c3e50')
            persist_check.pack(anchor='w', pady=(0, 20))

            def refresh_columns(*_args):
                cols = sheet_columns.get(sheet_var.get(), [])
                column_combo['values'] = cols
            sheet_var.trace_add('write', refresh_columns)
            refresh_columns()

            def save_mapping():
                """Save mapping and update tree row."""
                edge['excel_mapping'] = {
                    'enabled': enabled_var.get(),
                    'sheet': sheet_var.get(),
                    'column': column_var.get()
                }
                if persist_var.get() and sheet_var.get() and column_var.get():
                    registry.link_column_to_flow(flow_id, sheet_var.get(), column_var.get())
                else:
                    registry.remove_link(flow_id)
                
                # Update stats and refresh tree
                _populate_rows()
                edit_win.destroy()
                messagebox.showinfo("✅ Saved", "Mapping updated successfully!")

            # Footer buttons
            button_frame = tk.Frame(edit_win, bg='#f5f7fa')
            button_frame.pack(fill='x', padx=20, pady=(0, 20))
            
            cancel_btn = tk.Button(button_frame, text="Cancel", 
                                   command=edit_win.destroy,
                                   bg='#95a5a6', fg='white', 
                                   font=('Segoe UI', 10),
                                   relief='flat', padx=20, pady=8,
                                   cursor='hand2')
            cancel_btn.pack(side='right', padx=(8, 0))
            
            save_btn = tk.Button(button_frame, text="💾 Save Mapping", 
                                command=save_mapping,
                                bg='#27ae60', fg='white', 
                                font=('Segoe UI', 10, 'bold'),
                                relief='flat', padx=20, pady=8,
                                cursor='hand2')
            save_btn.pack(side='right')

        tree.bind('<Double-1>', edit_selected)

        def save_all_and_close():
            """Save all mappings to JSON and close dialog."""
            self._save_to_json()
            messagebox.showinfo("✅ Saved", "All flow mappings saved successfully!")
            dialog.destroy()

        # Modern footer
        footer = tk.Frame(dialog, bg='#e8eef5', height=70)
        footer.pack(fill='x')
        footer.pack_propagate(False)
        
        footer_inner = tk.Frame(footer, bg='#e8eef5')
        footer_inner.pack(fill='both', expand=True, padx=20, pady=15)
        
        # Export button (left)
        export_btn = tk.Button(footer_inner, text="📤 Export Mappings", 
                              command=lambda: messagebox.showinfo("Export", "Export feature coming soon!"),
                              bg='#95a5a6', fg='white', 
                              font=('Segoe UI', 9),
                              relief='flat', padx=15, pady=8,
                              cursor='hand2')
        export_btn.pack(side='left')
        
        # Close button (right, secondary)
        close_btn = tk.Button(footer_inner, text="Close", 
                             command=dialog.destroy,
                             bg='#7f8c8d', fg='white', 
                             font=('Segoe UI', 10),
                             relief='flat', padx=20, pady=8,
                             cursor='hand2')
        close_btn.pack(side='right', padx=(8, 0))
        
        # Save All button (right, primary)
        save_all_btn = tk.Button(footer_inner, text="💾 Save All & Close", 
                                command=save_all_and_close,
                                bg='#27ae60', fg='white', 
                                font=('Segoe UI', 10, 'bold'),
                                relief='flat', padx=25, pady=8,
                                cursor='hand2')
        save_all_btn.pack(side='right')

    def _open_excel_structure_manager(self):
        """Open Excel Structure Manager - manage sheets and columns"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Missing Package", "openpyxl package is required.\nInstall: pip install openpyxl")
            return
        
        # Get Excel path
        excel_path = self.flow_loader.excel_path
        if not excel_path or not Path(excel_path).exists():
            messagebox.showerror(
                "Excel File Missing",
                f"Flow Diagram Excel file not found:\n\n{excel_path}\n\n"
                "This file contains flow volumes for the diagram visualization.\n"
                "Please configure the path in Settings > Data Sources (timeseries_excel_path)."
            )
            return
        
        dialog = self._create_styled_dialog("Excel Structure Manager", 950, 700)
        
        # Header
        header = tk.Frame(dialog, bg='#e67e22', height=60)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="🔧 Excel Structure Manager",
                 font=('Segoe UI', 14, 'bold'), bg='#e67e22', fg='white').pack(side='left', padx=20, pady=18)
        tk.Label(header, text=f"📁 {Path(excel_path).name}",
                 font=('Segoe UI', 9), bg='#e67e22', fg='#fff3cd').pack(side='right', padx=20)
        
        # Notebook (tabs)
        notebook = ttk.Notebook(dialog)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # ===== TAB 1: SHEETS MANAGER =====
        sheets_frame = ttk.Frame(notebook, padding=10)
        notebook.add(sheets_frame, text='📋 Sheets')
        
        tk.Label(sheets_frame, text="Manage Excel Worksheets",
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))
        
        sheets_list_frame = tk.Frame(sheets_frame)
        sheets_list_frame.pack(fill='both', expand=True)
        
        sheets_scrollbar = tk.Scrollbar(sheets_list_frame)
        sheets_scrollbar.pack(side='right', fill='y')
        
        sheets_listbox = tk.Listbox(sheets_list_frame, yscrollcommand=sheets_scrollbar.set,
                                     font=('Segoe UI', 10), height=20)
        sheets_listbox.pack(side='left', fill='both', expand=True)
        sheets_scrollbar.config(command=sheets_listbox.yview)
        enable_listbox_mousewheel(sheets_listbox)
        
        def refresh_sheets():
            sheets_listbox.delete(0, tk.END)
            try:
                wb = load_workbook(excel_path)
                for sheet_name in wb.sheetnames:
                    sheets_listbox.insert(tk.END, f"📄 {sheet_name}")
                wb.close()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load sheets:\n{e}")
        
        def add_sheet():
            # Modern add sheet dialog
            add_dialog = tk.Toplevel(dialog)
            add_dialog.title("Add New Sheet")
            add_dialog.configure(bg='#f5f7fa')
            self._center_window(add_dialog, 500, 280)
            add_dialog.transient(dialog)
            add_dialog.grab_set()
            
            # Header
            header = tk.Frame(add_dialog, bg='#27ae60', height=50)
            header.pack(fill='x')
            header.pack_propagate(False)
            tk.Label(header, text="➕ Add New Sheet", 
                    bg='#27ae60', fg='white', 
                    font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=20, pady=15)
            
            # Body
            body = tk.Frame(add_dialog, bg='white')
            body.pack(fill='both', expand=True, padx=25, pady=20)
            
            tk.Label(body, text="Sheet Name:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            name_var = tk.StringVar()
            entry = tk.Entry(body, textvariable=name_var, 
                           font=('Segoe UI', 11), 
                           bg='#e8eef5', fg='#2c3e50',
                           relief='flat', bd=0, insertbackground='#27ae60')
            entry.pack(fill='x', pady=(0, 20), ipady=8, ipadx=10)
            entry.focus()
            
            result = [None]
            
            def apply_add():
                new_name = name_var.get().strip()
                if not new_name:
                    messagebox.showwarning("Empty Name", "Sheet name cannot be empty", parent=add_dialog)
                    return
                result[0] = new_name
                add_dialog.destroy()
            
            def cancel_add():
                add_dialog.destroy()
            
            # Buttons
            btn_frame = tk.Frame(body, bg='white')
            btn_frame.pack(fill='x')
            
            tk.Button(btn_frame, text="Cancel", command=cancel_add,
                     bg='#95a5a6', fg='white', font=('Segoe UI', 10),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right', padx=(8, 0))
            tk.Button(btn_frame, text="➕ Add", command=apply_add,
                     bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right')
            
            entry.bind('<Return>', lambda e: apply_add())
            
            dialog.wait_window(add_dialog)
            
            if result[0] is None:
                return
            
            new_name = result[0]
            
            try:
                wb = load_workbook(excel_path)
                if new_name in wb.sheetnames:
                    messagebox.showwarning("Duplicate", f"Sheet '{new_name}' already exists!")
                    wb.close()
                    return
                ws = wb.create_sheet(new_name)
                # Add standard headers
                ws['A1'] = 'Date'
                ws['B1'] = 'Year'
                ws['C1'] = 'Month'
                wb.save(excel_path)
                wb.close()
                refresh_sheets()
                refresh_sheets_combo()
                messagebox.showinfo("✅ Success", f"Sheet '{new_name}' created with standard headers!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add sheet:\n{e}")
        
        def rename_sheet():
            selection = sheets_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a sheet to rename")
                return
            old_name = sheets_listbox.get(selection[0]).replace("📄 ", "")
            
            # Modern rename dialog
            rename_dialog = tk.Toplevel(dialog)
            rename_dialog.title("Rename Sheet")
            rename_dialog.configure(bg='#f5f7fa')
            self._center_window(rename_dialog, 520, 320)
            rename_dialog.transient(dialog)
            rename_dialog.grab_set()
            
            # Header
            header = tk.Frame(rename_dialog, bg='#3498db', height=50)
            header.pack(fill='x')
            header.pack_propagate(False)
            tk.Label(header, text="✏️ Rename Sheet", 
                    bg='#3498db', fg='white', 
                    font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=20, pady=15)
            
            # Body
            body = tk.Frame(rename_dialog, bg='white')
            body.pack(fill='both', expand=True, padx=25, pady=20)
            
            # Current name
            tk.Label(body, text="Current Name:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            current_label = tk.Label(body, text=old_name, bg='#e8eef5', fg='#2c3e50',
                                    font=('Segoe UI', 10, 'bold'),
                                    anchor='w', padx=12, pady=8)
            current_label.pack(fill='x', pady=(0, 15))
            
            # New name
            tk.Label(body, text="New Name:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            new_var = tk.StringVar(value=old_name)
            entry = tk.Entry(body, textvariable=new_var, 
                           font=('Segoe UI', 11), 
                           bg='#e8eef5', fg='#2c3e50',
                           relief='flat', bd=0, insertbackground='#3498db')
            entry.pack(fill='x', pady=(0, 20), ipady=8, ipadx=10)
            entry.focus()
            entry.select_range(0, tk.END)
            
            result = [None]
            
            def apply_rename():
                new_name = new_var.get().strip()
                if not new_name:
                    messagebox.showwarning("Empty Name", "Sheet name cannot be empty", parent=rename_dialog)
                    return
                if new_name == old_name:
                    messagebox.showinfo("No Change", "New name is the same as current name", parent=rename_dialog)
                    return
                result[0] = new_name
                rename_dialog.destroy()
            
            def cancel_rename():
                rename_dialog.destroy()
            
            # Buttons
            btn_frame = tk.Frame(body, bg='white')
            btn_frame.pack(fill='x')
            
            tk.Button(btn_frame, text="Cancel", command=cancel_rename,
                     bg='#95a5a6', fg='white', font=('Segoe UI', 10),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right', padx=(8, 0))
            tk.Button(btn_frame, text="✏️ Rename", command=apply_rename,
                     bg='#3498db', fg='white', font=('Segoe UI', 10, 'bold'),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right')
            
            entry.bind('<Return>', lambda e: apply_rename())
            
            dialog.wait_window(rename_dialog)
            
            if result[0] is None:
                return
            
            new_name = result[0]
            
            try:
                wb = load_workbook(excel_path)
                if old_name not in wb.sheetnames:
                    messagebox.showerror("Error", f"Sheet '{old_name}' not found!")
                    wb.close()
                    return
                if new_name in wb.sheetnames:
                    messagebox.showwarning("Duplicate", f"Sheet '{new_name}' already exists!")
                    wb.close()
                    return
                wb[old_name].title = new_name
                wb.save(excel_path)
                wb.close()
                refresh_sheets()
                refresh_sheets_combo()
                messagebox.showinfo("✅ Success", f"Sheet renamed to '{new_name}'!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to rename sheet:\n{e}")
        
        def delete_sheet():
            selection = sheets_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a sheet to delete")
                return
            sheet_name = sheets_listbox.get(selection[0]).replace("📄 ", "")
            if not messagebox.askyesno("Confirm Delete", f"Delete sheet '{sheet_name}'?\n\n⚠️ This cannot be undone!"):
                return
            try:
                wb = load_workbook(excel_path)
                if sheet_name not in wb.sheetnames:
                    messagebox.showerror("Error", f"Sheet '{sheet_name}' not found!")
                    wb.close()
                    return
                if len(wb.sheetnames) == 1:
                    messagebox.showwarning("Cannot Delete", "Cannot delete the last sheet in the workbook!")
                    wb.close()
                    return
                wb.remove(wb[sheet_name])
                wb.save(excel_path)
                wb.close()
                refresh_sheets()
                messagebox.showinfo("Deleted", f"Sheet '{sheet_name}' deleted!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete sheet:\n{e}")
        
        sheets_btn_frame = tk.Frame(sheets_frame)
        sheets_btn_frame.pack(fill='x', pady=10)
        
        tk.Button(sheets_btn_frame, text="➕ Add Sheet", command=add_sheet,
                  bg='#27ae60', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(sheets_btn_frame, text="✏️ Rename Sheet", command=rename_sheet,
                  bg='#3498db', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(sheets_btn_frame, text="🗑️ Delete Sheet", command=delete_sheet,
                  bg='#e74c3c', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(sheets_btn_frame, text="🔄 Refresh", command=refresh_sheets,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='right', padx=5)
        
        # ===== TAB 2: COLUMNS MANAGER =====
        columns_frame = ttk.Frame(notebook, padding=10)
        notebook.add(columns_frame, text='📊 Columns')
        
        tk.Label(columns_frame, text="Manage Columns in Selected Sheet",
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))
        
        # Sheet selector
        sheet_select_frame = tk.Frame(columns_frame)
        sheet_select_frame.pack(fill='x', pady=(0, 10))
        tk.Label(sheet_select_frame, text="Select Sheet:", font=('Segoe UI', 10, 'bold')).pack(side='left', padx=(0, 10))
        
        sheet_var = tk.StringVar(master=sheet_select_frame)
        sheet_combo = ttk.Combobox(sheet_select_frame, textvariable=sheet_var, state='readonly', font=('Segoe UI', 10), width=30)
        sheet_combo.pack(side='left', padx=5)
        
        columns_list_frame = tk.Frame(columns_frame)
        columns_list_frame.pack(fill='both', expand=True)
        
        columns_scrollbar = tk.Scrollbar(columns_list_frame)
        columns_scrollbar.pack(side='right', fill='y')
        
        columns_listbox = tk.Listbox(columns_list_frame, yscrollcommand=columns_scrollbar.set,
                                      font=('Segoe UI', 10), height=18)
        columns_listbox.pack(side='left', fill='both', expand=True)
        columns_scrollbar.config(command=columns_listbox.yview)
        enable_listbox_mousewheel(columns_listbox)
        
        def refresh_sheets_combo():
            try:
                wb = load_workbook(excel_path)
                sheet_combo['values'] = wb.sheetnames
                if wb.sheetnames:
                    sheet_combo.current(0)
                    # Trigger column refresh after combo value is set
                    sheet_combo.event_generate('<<ComboboxSelected>>')
                wb.close()
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load sheets:\n{e}")
        
        def refresh_columns():
            columns_listbox.delete(0, tk.END)
            selected_sheet = sheet_var.get()
            if not selected_sheet:
                return
            try:
                # Use FlowVolumeLoader to list columns consistently with loader logic
                loader = get_flow_volume_loader()
                cols = loader.list_sheet_columns(selected_sheet)
                for col in cols:
                    columns_listbox.insert(tk.END, f"📊 {col}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load columns:\n{e}")
        
        sheet_combo.bind('<<ComboboxSelected>>', lambda e: refresh_columns())
        
        def add_column():
            selected_sheet = sheet_var.get()
            if not selected_sheet:
                messagebox.showwarning("No Sheet", "Please select a sheet first")
                return
            
            # Modern add column dialog
            add_dialog = tk.Toplevel(dialog)
            add_dialog.title("Add New Column")
            add_dialog.configure(bg='#f5f7fa')
            self._center_window(add_dialog, 500, 280)
            add_dialog.transient(dialog)
            add_dialog.grab_set()
            
            # Header
            header = tk.Frame(add_dialog, bg='#27ae60', height=50)
            header.pack(fill='x')
            header.pack_propagate(False)
            tk.Label(header, text="➕ Add New Column", 
                    bg='#27ae60', fg='white', 
                    font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=20, pady=15)
            
            # Body
            body = tk.Frame(add_dialog, bg='white')
            body.pack(fill='both', expand=True, padx=25, pady=20)
            
            tk.Label(body, text="Column Name:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            col_var = tk.StringVar()
            entry = tk.Entry(body, textvariable=col_var, 
                           font=('Segoe UI', 11), 
                           bg='#e8eef5', fg='#2c3e50',
                           relief='flat', bd=0, insertbackground='#27ae60')
            entry.pack(fill='x', pady=(0, 20), ipady=8, ipadx=10)
            entry.focus()
            
            result = [None]
            
            def apply_add():
                new_col = col_var.get().strip()
                if not new_col:
                    messagebox.showwarning("Empty Name", "Column name cannot be empty", parent=add_dialog)
                    return
                result[0] = new_col
                add_dialog.destroy()
            
            def cancel_add():
                add_dialog.destroy()
            
            # Buttons
            btn_frame = tk.Frame(body, bg='white')
            btn_frame.pack(fill='x')
            
            tk.Button(btn_frame, text="Cancel", command=cancel_add,
                     bg='#95a5a6', fg='white', font=('Segoe UI', 10),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right', padx=(8, 0))
            tk.Button(btn_frame, text="➕ Add", command=apply_add,
                     bg='#27ae60', fg='white', font=('Segoe UI', 10, 'bold'),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right')
            
            entry.bind('<Return>', lambda e: apply_add())
            
            dialog.wait_window(add_dialog)
            
            if result[0] is None:
                return
            
            new_col = result[0]
            
            try:
                wb = load_workbook(excel_path)
                ws = wb[selected_sheet]
                # Append after the last header in row 3 (actual header row)
                header_row = 3
                last_header_idx = 0
                for cell in ws[header_row]:
                    val = cell.value
                    if val is not None and str(val).strip():
                        last_header_idx = cell.column
                # Fallback to ws.max_column if row 3 yields nothing
                target_col = (last_header_idx if last_header_idx > 0 else ws.max_column) + 1
                ws.cell(row=header_row, column=target_col, value=new_col)
                wb.save(excel_path)
                wb.close()
                refresh_columns()
                # Clear loader cache to pick up new column
                self.flow_loader.clear_cache()
                messagebox.showinfo("✅ Success", f"Column '{new_col}' added to sheet '{selected_sheet}'!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to add column:\n{e}")
        
        def rename_column():
            selected_sheet = sheet_var.get()
            if not selected_sheet:
                messagebox.showwarning("No Sheet", "Please select a sheet first")
                return
            selection = columns_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a column to rename")
                return
            old_name = columns_listbox.get(selection[0]).replace("📊 ", "")
            if old_name in ['Date', 'Year', 'Month']:
                messagebox.showwarning("Cannot Rename", "Standard columns (Date, Year, Month) cannot be renamed")
                return
            
            # Modern rename dialog
            rename_dialog = tk.Toplevel(dialog)
            rename_dialog.title("Rename Column")
            rename_dialog.configure(bg='#f5f7fa')
            self._center_window(rename_dialog, 520, 320)
            rename_dialog.transient(dialog)
            rename_dialog.grab_set()
            
            # Header
            header = tk.Frame(rename_dialog, bg='#3498db', height=50)
            header.pack(fill='x')
            header.pack_propagate(False)
            tk.Label(header, text="✏️ Rename Column", 
                    bg='#3498db', fg='white', 
                    font=('Segoe UI', 12, 'bold')).pack(anchor='w', padx=20, pady=15)
            
            # Body
            body = tk.Frame(rename_dialog, bg='white')
            body.pack(fill='both', expand=True, padx=25, pady=20)
            
            # Current name
            tk.Label(body, text="Current Name:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            current_label = tk.Label(body, text=old_name, bg='#e8eef5', fg='#2c3e50',
                                    font=('Segoe UI', 10, 'bold'),
                                    anchor='w', padx=12, pady=8)
            current_label.pack(fill='x', pady=(0, 15))
            
            # New name
            tk.Label(body, text="New Name:", bg='white', fg='#7f8c8d',
                    font=('Segoe UI', 9)).pack(anchor='w', pady=(0, 3))
            new_var = tk.StringVar(value=old_name)
            entry = tk.Entry(body, textvariable=new_var, 
                           font=('Segoe UI', 11), 
                           bg='#e8eef5', fg='#2c3e50',
                           relief='flat', bd=0, insertbackground='#3498db')
            entry.pack(fill='x', pady=(0, 20), ipady=8, ipadx=10)
            entry.focus()
            entry.select_range(0, tk.END)
            
            result = [None]
            
            def apply_rename():
                new_name = new_var.get().strip()
                if not new_name:
                    messagebox.showwarning("Empty Name", "Column name cannot be empty", parent=rename_dialog)
                    return
                if new_name == old_name:
                    messagebox.showinfo("No Change", "New name is the same as current name", parent=rename_dialog)
                    return
                result[0] = new_name
                rename_dialog.destroy()
            
            def cancel_rename():
                rename_dialog.destroy()
            
            # Buttons
            btn_frame = tk.Frame(body, bg='white')
            btn_frame.pack(fill='x')
            
            tk.Button(btn_frame, text="Cancel", command=cancel_rename,
                     bg='#95a5a6', fg='white', font=('Segoe UI', 10),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right', padx=(8, 0))
            tk.Button(btn_frame, text="✏️ Rename", command=apply_rename,
                     bg='#3498db', fg='white', font=('Segoe UI', 10, 'bold'),
                     relief='flat', padx=20, pady=8, cursor='hand2').pack(side='right')
            
            entry.bind('<Return>', lambda e: apply_rename())
            
            dialog.wait_window(rename_dialog)
            
            if result[0] is None:
                return
            
            new_name = result[0]
            
            try:
                wb = load_workbook(excel_path)
                ws = wb[selected_sheet]
                # Find the column in row 3 (actual header row)
                for cell in ws[3]:
                    if cell.value == old_name:
                        cell.value = new_name
                        break
                wb.save(excel_path)
                wb.close()

                # Sync: update excel_mapping.column in diagram JSON for all edges
                self._load_diagram_data()
                edges = self.area_data.get('edges', [])
                updated = 0
                for edge in edges:
                    mapping = edge.get('excel_mapping', {}) or {}
                    if mapping.get('enabled') and mapping.get('sheet') == selected_sheet and mapping.get('column') == old_name:
                        mapping['column'] = new_name
                        edge['excel_mapping'] = mapping
                        updated += 1

                if updated > 0:
                    try:
                        with open(self.json_file, 'w', encoding='utf-8') as f:
                            json.dump(self.area_data, f, indent=2, ensure_ascii=False)
                    except Exception as je:
                        messagebox.showerror("Warning", f"Column renamed in Excel, but failed to update diagram mappings:\n{je}")

                refresh_columns()
                self.flow_loader.clear_cache()
                if updated:
                    messagebox.showinfo("✅ Success", f"Column renamed to '{new_name}'.\nUpdated {updated} flow mapping(s).")
                else:
                    messagebox.showinfo("✅ Success", f"Column renamed to '{new_name}'.\nNo existing flow mappings referenced this column.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to rename column:\n{e}")
        
        def delete_column():
            selected_sheet = sheet_var.get()
            if not selected_sheet:
                messagebox.showwarning("No Sheet", "Please select a sheet first")
                return
            selection = columns_listbox.curselection()
            if not selection:
                messagebox.showwarning("No Selection", "Please select a column to delete")
                return
            col_name = columns_listbox.get(selection[0]).replace("📊 ", "")
            if col_name in ['Date', 'Year', 'Month']:
                messagebox.showwarning("Cannot Delete", "Cannot delete standard columns (Date, Year, Month)")
                return
            if not messagebox.askyesno("Confirm Delete", f"Delete column '{col_name}' from sheet '{selected_sheet}'?\n\n⚠️ This cannot be undone!"):
                return
            try:
                wb = load_workbook(excel_path)
                ws = wb[selected_sheet]
                # Find and delete the column from row 3 (actual header row)
                for idx, cell in enumerate(ws[3], start=1):
                    if cell.value == col_name:
                        ws.delete_cols(idx)
                        break
                wb.save(excel_path)
                wb.close()
                refresh_columns()
                self.flow_loader.clear_cache()
                messagebox.showinfo("Deleted", f"Column '{col_name}' deleted!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to delete column:\n{e}")
        
        columns_btn_frame = tk.Frame(columns_frame)
        columns_btn_frame.pack(fill='x', pady=10)
        
        tk.Button(columns_btn_frame, text="➕ Add Column", command=add_column,
                  bg='#27ae60', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(columns_btn_frame, text="✏️ Rename Column", command=rename_column,
                  bg='#3498db', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(columns_btn_frame, text="🗑️ Delete Column", command=delete_column,
                  bg='#e74c3c', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        
        def sync_columns_to_labels():
            """Rename flow columns to reflect current component labels and sync diagram mappings."""
            selected_sheet = sheet_var.get()
            if not selected_sheet:
                messagebox.showwarning("No Sheet", "Please select a sheet first")
                return
            if not messagebox.askyesno(
                "Sync Column Names",
                "Update flow column names to match current component labels?\n\n"
                "This will rename columns in Excel and update linked flow mappings in the diagram."
            ):
                return
            try:
                # Build id→label map from current diagram data
                self._load_diagram_data()
                nodes = self.area_data.get('nodes', [])
                id_to_label = {n.get('id'): (n.get('label') or n.get('id')) for n in nodes}

                # Open workbook and target sheet
                wb = load_workbook(excel_path)
                if selected_sheet not in wb.sheetnames:
                    wb.close()
                    messagebox.showerror("Error", f"Sheet '{selected_sheet}' not found in Excel")
                    return
                ws = wb[selected_sheet]

                # Existing headers in row 1
                headers = {cell.value: cell for cell in ws[1] if cell.value}

                updated = 0
                for edge in self.area_data.get('edges', []):
                    mapping = edge.get('excel_mapping', {}) or {}
                    if not mapping.get('enabled'):
                        continue
                    if mapping.get('sheet') != selected_sheet:
                        continue

                    old_col = mapping.get('column')
                    # Compute new name from labels
                    from_label = id_to_label.get(edge.get('from', ''), edge.get('from', ''))
                    to_label = id_to_label.get(edge.get('to', ''), edge.get('to', ''))
                    new_col = f"{from_label} → {to_label}"

                    # If unchanged or header not present, skip safely
                    header_cell = headers.get(old_col)
                    if not header_cell:
                        # Column might already have been renamed previously; try to find by current value
                        header_cell = headers.get(new_col)
                        if not header_cell:
                            continue
                    if header_cell.value == new_col:
                        # Already synced
                        mapping['column'] = new_col
                        edge['excel_mapping'] = mapping
                        continue

                    # Ensure uniqueness if another column already uses new_col
                    candidate = new_col
                    suffix = 2
                    while candidate in headers and headers[candidate] is not header_cell:
                        candidate = f"{new_col} ({suffix})"
                        suffix += 1
                    new_col = candidate

                    # Apply rename in Excel
                    header_cell.value = new_col
                    headers[new_col] = header_cell
                    if old_col in headers:
                        del headers[old_col]

                    # Update mapping in diagram
                    mapping['column'] = new_col
                    edge['excel_mapping'] = mapping
                    updated += 1

                # Persist workbook and diagram
                wb.save(excel_path)
                wb.close()
                with open(self.json_file, 'w', encoding='utf-8') as f:
                    json.dump(self.area_data, f, indent=2, ensure_ascii=False)

                refresh_columns()
                self.flow_loader.clear_cache()
                messagebox.showinfo("Synced", f"Updated {updated} column name(s) to match current labels.")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to sync columns:\n{e}")

        tk.Button(columns_btn_frame, text="🔁 Sync Names to Labels", command=sync_columns_to_labels,
                  bg='#f39c12', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        tk.Button(columns_btn_frame, text="🔄 Refresh", command=refresh_columns,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='right', padx=5)
        
        # ===== TAB 3: UNMAPPED FLOWS =====
        unmapped_frame = ttk.Frame(notebook, padding=10)
        notebook.add(unmapped_frame, text='⚠️ Unmapped Flows')
        
        tk.Label(unmapped_frame, text="Flows Without Excel Mappings",
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))
        
        tk.Label(unmapped_frame, text="These flows need Excel columns to load volume data:",
                 font=('Segoe UI', 9), fg='#7f8c8d').pack(anchor='w', pady=(0, 10))
        
        unmapped_list_frame = tk.Frame(unmapped_frame)
        unmapped_list_frame.pack(fill='both', expand=True)
        
        unmapped_scrollbar = tk.Scrollbar(unmapped_list_frame)
        unmapped_scrollbar.pack(side='right', fill='y')
        
        unmapped_listbox = tk.Listbox(unmapped_list_frame, yscrollcommand=unmapped_scrollbar.set,
                                       font=('Segoe UI', 10), height=20)
        unmapped_listbox.pack(side='left', fill='both', expand=True)
        unmapped_scrollbar.config(command=unmapped_listbox.yview)
        enable_listbox_mousewheel(unmapped_listbox)
        
        def refresh_unmapped():
            unmapped_listbox.delete(0, tk.END)
            edges = self.area_data.get('edges', [])
            count = 0
            for edge in edges:
                mapping = edge.get('excel_mapping', {})
                if not mapping or not mapping.get('enabled'):
                    from_id = edge.get('from', 'Unknown')
                    to_id = edge.get('to', 'Unknown')
                    suggested_col = f"{from_id}_{to_id}".upper()
                    unmapped_listbox.insert(tk.END, f"⚠️ {from_id} → {to_id}  (Suggest: {suggested_col})")
                    count += 1
            if count == 0:
                unmapped_listbox.insert(tk.END, "✅ All flows are mapped!")
        
        unmapped_btn_frame = tk.Frame(unmapped_frame)
        unmapped_btn_frame.pack(fill='x', pady=10)
        
        tk.Button(unmapped_btn_frame, text="🔄 Refresh", command=refresh_unmapped,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 10), padx=15, pady=5).pack(side='left', padx=5)
        tk.Label(unmapped_btn_frame, text="💡 Tip: Use Auto-Map button to quickly map flows",
                 font=('Segoe UI', 9), fg='#7f8c8d').pack(side='left', padx=20)
        
        # Initialize tabs
        refresh_sheets()
        refresh_sheets_combo()
        refresh_unmapped()
        
        # Footer
        footer = tk.Frame(dialog, bg='#e8eef5')
        footer.pack(fill='x', pady=10, padx=10)
        
        tk.Button(footer, text="✖ Close", command=dialog.destroy,
                  bg='#95a5a6', fg='white', font=('Segoe UI', 11),
                  padx=25, pady=8, relief='flat', cursor='hand2').pack(side='right', padx=5)
        
        dialog.wait_window()

    def _create_styled_dialog(self, title: str, width: int, height: int) -> tk.Toplevel:
        """Create a professionally styled, centered dialog window."""
        dialog = tk.Toplevel(self.canvas)
        dialog.title(title)
        dialog.transient(self.canvas)
        dialog.grab_set()
        
        # Professional styling
        dialog.configure(bg='#e8eef5')
        
        # Center the window
        self._center_window(dialog, width, height)
        
        return dialog
    
    def _center_window(self, window: tk.Toplevel, width: int, height: int) -> None:
        """Center a toplevel window over the main application window."""
        window.update_idletasks()
        try:
            parent_x = self.parent.winfo_rootx()
            parent_y = self.parent.winfo_rooty()
            parent_w = self.parent.winfo_width() or window.winfo_screenwidth()
            parent_h = self.parent.winfo_height() or window.winfo_screenheight()
        except Exception:
            parent_x, parent_y = 0, 0
            parent_w, parent_h = window.winfo_screenwidth(), window.winfo_screenheight()

        x = parent_x + max((parent_w - width) // 2, 0)
        y = parent_y + max((parent_h - height) // 2, 0)
        window.geometry(f"{width}x{height}+{x}+{y}")

    def _reload_from_json(self):
        """Reload from file"""
        if messagebox.askyesno("Reload", "Discard unsaved changes?"):
            self._load_diagram_data()
            messagebox.showinfo("Reloaded", "Reloaded from file")

    def _load_volumes_from_excel(self):
        """Load all flow volumes from Excel for selected month."""
        try:
            # Get selected month/year
            year = int(self.year_var.get())
            
            # Get month number from combobox value (1-12)
            months = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December']
            month_name = self.month_var.get()
            month = months.index(month_name) + 1 if month_name in months else self.current_month
            
            self.current_year = year
            self.current_month = month
            
            # Determine area code from JSON
            if not self.area_code:
                self.area_code = self.area_data.get('area_code', 'UG2N')
            
            logger.info(f"📥 Loading flows from Excel: {self.area_code} for {year}-{month:02d}")

            # Refresh loader instance in case Settings changed the path
            self.flow_loader = get_flow_volume_loader()
            
            # Always clear loader cache so recent Excel edits are picked up
            self.flow_loader.clear_cache()
            
            # Update all edges with volumes from Excel
            self.area_data = self.flow_loader.update_diagram_edges(
                self.area_data,
                self.area_code,
                year,
                month
            )
            
            # Redraw diagram with updated volumes
            self._draw_diagram()
            
            # Show brief status message instead of popup
            logger.info(f"✅ Flow volumes loaded for {month_name} {year}")
        
        except Exception as e:
            logger.error(f"❌ Error loading from Excel: {e}")
            messagebox.showerror("Error", f"Failed to load volumes from Excel:\n{e}")
    
    def refresh_flow_loader(self):
        """Refresh flow loader instance (call after Settings path change)."""
        self.flow_loader = get_flow_volume_loader()
        logger.info("🔄 Flow loader refreshed in diagram")
    
    def _get_area_code_from_title(self):
        """Extract area code from diagram title."""
        title = self.area_data.get('title', '')
        area_code_map = {
            'UG2 North': 'UG2N',
            'Merensky North': 'MERN',
            'Merensky South': 'MERENSKY_SOUTH',
            'UG2 South': 'UG2S',
            'Stockpile': 'STOCKPILE',
            'Old TSF': 'OLDTSF',
            'UG2 Plant': 'UG2PLANT',
            'Merensky Plant': 'MERPLANT',
        }
        
        for key, code in area_code_map.items():
            if key.lower() in title.lower():
                return code
        
        return 'UG2N'  # Default
    
    def _open_balance_check(self):
        """Open the balance check dialog for flow categorization"""
        try:
            if not self.area_data.get('edges'):
                messagebox.showwarning("No Flows", "No flowlines found in the diagram. Draw some flows first!")
                return
            
            # Ensure area_code is initialized before opening the dialog
            # so category loading picks the correct area-specific set.
            if not self.area_code:
                # Prefer explicit area_code from JSON, else infer from title.
                inferred = self.area_data.get('area_code') or self._get_area_code_from_title()
                self.area_code = inferred
                logger.info(f"📍 Initialized area_code for balance check: {self.area_code}")

            # Create and show balance check dialog
            dialog = FlowDiagramBalanceCheckDialog(self.parent, self)
            dialog.show()
            
        except Exception as e:
            logger.error(f"Error opening balance check: {e}", exc_info=True)
            messagebox.showerror("Error", f"Failed to open balance check:\n{e}")


# For compatibility
FlowDiagramDashboard = DetailedNetworkFlowDiagram

