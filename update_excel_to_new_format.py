#!/usr/bin/env python
"""
Update Excel template to NEW naming convention (UPPERCASE with arrow separator).

The current template has lowercase names with '__TO__' separator.
We need to convert them to UPPERCASE with ' → ' separator.
"""

from openpyxl import load_workbook
from pathlib import Path

EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

def convert_column_name(old_name):
    """Convert old format to new format."""
    if not old_name or not isinstance(old_name, str):
        return old_name
    
    # Skip if already in new format
    if ' → ' in old_name:
        return old_name
    
    # Convert __TO__ to → and uppercase
    if '__TO__' in old_name:
        return old_name.replace('__TO__', ' → ').upper()
    
    return old_name

def update_template():
    """Update Excel template with new naming convention."""
    print(f"Opening {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH)
    
    sheets_updated = 0
    columns_updated = 0
    
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith('Flows_'):
            continue
        
        ws = wb[sheet_name]
        print(f"\n📋 Processing sheet: {sheet_name}")
        
        # Headers are in row 3
        headers_row = 3
        print(f"   Converting headers in row {headers_row}...")
        
        sheet_updates = 0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(headers_row, col_idx)
            old_name = cell.value
            new_name = convert_column_name(old_name)
            
            if old_name != new_name and new_name:
                cell.value = new_name
                print(f"   ✅ {old_name:50} → {new_name}")
                sheet_updates += 1
                columns_updated += 1
        
        if sheet_updates > 0:
            sheets_updated += 1
            print(f"   Updated {sheet_updates} columns in {sheet_name}")
    
    if columns_updated > 0:
        print(f"\n💾 Saving workbook...")
        wb.save(EXCEL_PATH)
        print(f"✅ Successfully updated {columns_updated} columns across {sheets_updated} sheets!")
    else:
        print("\n⚠️  No columns needed updating")
    
    # Verify
    print("\n🔍 Verification:")
    wb2 = load_workbook(EXCEL_PATH)
    for sheet_name in ['Flows_UG2N', 'Flows_STOCKPILE']:
        if sheet_name in wb2.sheetnames:
            ws2 = wb2[sheet_name]
            headers = [c.value for c in ws2[3]]
            print(f"\n{sheet_name} headers (sample):")
            for i, h in enumerate(headers[:10], 1):
                if h and h != 'Date' and h != 'Year' and h != 'Month':
                    print(f"  {i}: {h}")

if __name__ == '__main__':
    update_template()
