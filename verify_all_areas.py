"""
Comprehensive verification: Excel vs JSON for all areas.
"""
from pathlib import Path
from openpyxl import load_workbook
import json

# Find latest Excel
test_templates = Path('test_templates')
xlsx_files = [f for f in test_templates.glob('Water_Balance_TimeSeries_Template_*.xlsx') 
              if f.stem.split('_')[-1].isdigit()]
latest_file = sorted(xlsx_files, key=lambda x: int(x.stem.split('_')[-1]))[-1]

print(f'📊 COMPREHENSIVE VERIFICATION')
print(f'=' * 100)
print(f'Excel File: {latest_file.name}')
print()

# Load Excel and JSON
wb = load_workbook(str(latest_file))
with open('data/diagrams/ug2_north_decline.json') as f:
    diagram = json.load(f)

# Map areas to their patterns in node IDs - updated for correct categorization
area_patterns = {
    'UG2N': {
        'prefixes': ['rainfall__', 'ndcd__', 'bh_ndgwa__', 'softening__', 'reservoir__', 'offices__', 
                    'guest_house__', 'septic__', 'sewage__', 'north_decline__', 'north_shaft__', 
                    'losses__', 'losses2__', 'consumption__', 'consumption2__', 'evaporation__',
                    'dust_suppression__', 'spill__', 'junction_127__', 'junction_128__'],
        'contains': ['__TO__rainfall', '__TO__ndcd', '__TO__bh_ndgwa', '__TO__softening', '__TO__reservoir', 
                    '__TO__offices', '__TO__guest_house', '__TO__septic', '__TO__sewage', '__TO__north_decline', 
                    '__TO__north_shaft', '__TO__losses', '__TO__losses2', '__TO__consumption', '__TO__consumption2',
                    '__TO__evaporation', '__TO__dust_suppression', '__TO__spill', '__TO__junction_127', '__TO__junction_128']
    },
    'MERN': {
        'prefixes': ['rainfall_merensky__', 'ndcd_merensky__', 'bh_mcgwa__', 'softening_merensky__', 
                    'offices_merensky__', 'merensky_north__', 'evaporation_merensky__', 'dust_suppression_merensky__',
                    'spill_merensky__', 'consumption_merensky__', 'losses_merensky__'],
        'contains': ['__TO__rainfall_merensky', '__TO__ndcd_merensky', '__TO__bh_mcgwa', '__TO__softening_merensky', 
                    '__TO__offices_merensky', '__TO__merensky_north', '__TO__evaporation_merensky', '__TO__dust_suppression_merensky',
                    '__TO__spill_merensky', '__TO__consumption_merensky', '__TO__losses_merensky']
    },
    'UG2S': {
        'prefixes': ['ug2s_'],
        'contains': ['__TO__ug2s_']
    },
    'UG2P': {
        'prefixes': ['ug2plant_', 'cprwsd'],
        'contains': ['__TO__ug2plant_', '__TO__cprwsd']
    },
    'MERP': {
        'prefixes': ['merplant_', 'mprwsd'],
        'contains': ['__TO__merplant_', '__TO__mprwsd']
    },
    'MERS': {
        'prefixes': ['mers_'],
        'contains': ['__TO__mers_']
    },
    'OLDTSF': {
        'prefixes': ['oldtsf_', 'nt_rwd', 'new_tsf', 'trtd'],
        'contains': ['__TO__oldtsf_', '__TO__nt_rwd', '__TO__new_tsf', '__TO__trtd']
    },
    'STOCKPILE': {
        'prefixes': ['stockpile_', 'spcd1'],
        'contains': ['__TO__stockpile_', '__TO__spcd1']
    }
}

# Extract flows per area from JSON
json_flows_by_area = {}
for area, patterns_dict in area_patterns.items():
    json_flows_by_area[area] = []
    for edge in diagram.get('edges', []):
        from_id = edge.get('from')
        to_id = edge.get('to')
        flow_str = f'{from_id}__TO__{to_id}'
        
        matches = False
        
        # Check prefixes
        for prefix in patterns_dict.get('prefixes', []):
            if flow_str.startswith(prefix):
                matches = True
                break
        
        # Check contains patterns
        if not matches:
            for contains in patterns_dict.get('contains', []):
                if contains in flow_str:
                    matches = True
                    break
        
        if matches:
            json_flows_by_area[area].append(flow_str)

# Extract flows per area from Excel
excel_flows_by_area = {}
sheet_mapping = {
    'Flows_UG2N': 'UG2N',
    'Flows_UG2S': 'UG2S',
    'Flows_UG2P': 'UG2P',
    'Flows_MERN': 'MERN',
    'Flows_MERP': 'MERP',
    'Flows_MERENSKY_SOUTH': 'MERS',
    'Flows_OLDTSF': 'OLDTSF',
    'Flows_STOCKPILE': 'STOCKPILE'
}

for sheet_name, area in sheet_mapping.items():
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Flow headers are in row 3, skip Date/Year/Month columns (A-C)
        cols = [cell.value for cell in ws[3][3:] if cell.value and '__TO__' in str(cell.value)]
        excel_flows_by_area[area] = cols

print('📋 AREA-BY-AREA VERIFICATION')
print('=' * 100)

total_json = 0
total_excel = 0
mismatches = []

for area in sorted(area_patterns.keys()):
    json_flows = json_flows_by_area.get(area, [])
    excel_flows = excel_flows_by_area.get(area, [])
    
    total_json += len(json_flows)
    total_excel += len(excel_flows)
    
    json_set = set(json_flows)
    excel_set = set(excel_flows)
    
    only_in_json = json_set - excel_set
    only_in_excel = excel_set - json_set
    
    status = '✅' if only_in_json == set() and only_in_excel == set() else '⚠️'
    
    print(f'\n{status} {area}')
    print(f'   JSON flows: {len(json_flows)}')
    print(f'   Excel flows: {len(excel_flows)}')
    
    if len(json_flows) == len(excel_flows):
        print(f'   ✅ COUNT MATCH')
    else:
        print(f'   ❌ COUNT MISMATCH (Diff: {len(json_flows) - len(excel_flows)})')
        mismatches.append((area, len(json_flows), len(excel_flows)))
    
    if only_in_json:
        print(f'   ⚠️ In JSON but NOT Excel ({len(only_in_json)}):')
        for flow in sorted(list(only_in_json))[:3]:
            print(f'      • {flow}')
        if len(only_in_json) > 3:
            print(f'      ... and {len(only_in_json) - 3} more')
    
    if only_in_excel:
        print(f'   ⚠️ In Excel but NOT JSON ({len(only_in_excel)}):')
        for flow in sorted(list(only_in_excel))[:3]:
            print(f'      • {flow}')
        if len(only_in_excel) > 3:
            print(f'      ... and {len(only_in_excel) - 3} more')

print()
print('=' * 100)
print('📊 OVERALL SUMMARY')
print('-' * 100)
print(f'Total JSON flows: {total_json}')
print(f'Total Excel flows: {total_excel}')

if total_json == total_excel:
    print(f'\n✅ PERFECT MATCH - All flows accounted for!')
else:
    print(f'\n⚠️ DIFFERENCE: {abs(total_json - total_excel)} flows')

if mismatches:
    print(f'\n⚠️ Areas with count mismatch:')
    for area, json_count, excel_count in mismatches:
        diff = json_count - excel_count
        print(f'   • {area:20} JSON: {json_count:2}  Excel: {excel_count:2}  Diff: {diff:+3}')
else:
    print(f'\n✅ All areas have matching flow counts!')

# Detailed comparison for each area
print()
print('=' * 100)
print('📋 DETAILED AREA BREAKDOWN')
print('=' * 100)

for area in sorted(area_patterns.keys()):
    json_flows = json_flows_by_area.get(area, [])
    excel_flows = excel_flows_by_area.get(area, [])
    
    print(f'\n🔍 {area} ({len(excel_flows)} flows in Excel):')
    print('-' * 100)
    
    if excel_flows:
        for idx, flow in enumerate(sorted(excel_flows), 1):
            status = '✅' if flow in json_flows else '⚠️'
            print(f'   {idx:2}. {status} {flow}')
    else:
        print(f'   (No flows in Excel)')
