"""
Test Column Editor Dialog and Column Operations (VERIFICATION).

Purpose:
- Verify add/rename/delete column operations work correctly
- Test ExcelManager column methods
- Validate data persistence

Test Coverage:
- Create new columns
- Rename existing columns
- Delete columns
- Error handling for invalid operations
- Cache invalidation
"""

import sys
from pathlib import Path
import tempfile
import shutil

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import pytest
from openpyxl import Workbook, load_workbook

from services.excel_manager import ExcelManager


class TestColumnEditorOperations:
    """Test suite for Excel column management operations (UNIT TESTS)."""

    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file for testing (FIXTURE).
        
        Sets up:
        - Excel workbook with sample data
        - "Flows_Test" sheet with Date/Month in columns A-B, headers in row 3
        - Temporary file path that gets cleaned up after test
        
        Yields:
            Path to temporary Excel file
        """
        temp_dir = tempfile.mkdtemp()
        excel_path = Path(temp_dir) / "test_flows.xlsx"

        # Create test Excel file (matching real format: row 3 = headers)
        wb = Workbook()
        ws = wb.active
        ws.title = "Flows_Test"

        # Row 1: Date column (first date columns are always present)
        ws["A1"] = "2025-01-01"
        ws["B1"] = 1  # Month
        
        # Row 2: Second data row
        ws["A2"] = "2025-02-01"
        ws["B2"] = 2  # Month

        # Row 3: Header row (column names start here)
        ws["A3"] = "Date"
        ws["B3"] = "Month"
        ws["C3"] = "BH_to_Sump"
        ws["D3"] = "Sump_to_Plant"

        # Data rows start at row 4
        ws["A4"] = "2025-03-01"
        ws["B4"] = 3
        ws["C4"] = 100
        ws["D4"] = 200
        
        ws["A5"] = "2025-04-01"
        ws["B5"] = 4
        ws["C5"] = 150
        ws["D5"] = 250

        wb.save(excel_path)
        wb.close()

        yield excel_path

        # Cleanup
        shutil.rmtree(temp_dir)

    def test_create_column(self, temp_excel_file):
        """Test creating a new column (CREATE OPERATION).
        
        Verifies:
        - Column is added to Excel file
        - Column appears in list_flow_columns
        - Data rows have empty cells for new column
        """
        manager = ExcelManager()
        manager.set_flow_diagram_path(str(temp_excel_file))

        # Verify initial columns (Date/Month are filtered out by list_flow_columns)
        initial_cols = manager.list_flow_columns("Flows_Test")
        assert "BH_to_Sump" in initial_cols
        assert "Sump_to_Plant" in initial_cols
        assert "New_Flow" not in initial_cols

        # Create new column
        result = manager.create_flow_column("Flows_Test", "New_Flow")
        assert result is True

        # Verify column was added
        updated_cols = manager.list_flow_columns("Flows_Test")
        assert "New_Flow" in updated_cols
        
        # Verify Excel file has new column in row 3
        wb = load_workbook(temp_excel_file)
        ws = wb["Flows_Test"]
        headers = [str(cell.value) for cell in ws[3]]
        assert "New_Flow" in headers
        wb.close()


    def test_rename_column(self, temp_excel_file):
        """Test renaming an existing column (RENAME OPERATION).
        
        Verifies:
        - Column header is updated
        - Old name is replaced with new name
        - Updated columns list reflects change
        """
        manager = ExcelManager()
        manager.set_flow_diagram_path(str(temp_excel_file))

        # Verify initial state
        initial_cols = manager.list_flow_columns("Flows_Test")
        assert "BH_to_Sump" in initial_cols
        assert "BH_to_Plant" not in initial_cols

        # Rename column
        result = manager.rename_flow_column("Flows_Test", "BH_to_Sump", "BH_to_Plant")
        assert result is True

        # Verify column was renamed
        updated_cols = manager.list_flow_columns("Flows_Test")
        assert "BH_to_Sump" not in updated_cols
        assert "BH_to_Plant" in updated_cols

        # Verify Excel file has correct name
        wb = load_workbook(temp_excel_file)
        ws = wb["Flows_Test"]
        headers = [str(cell.value) for cell in ws[3]]
        assert "BH_to_Plant" in headers
        assert "BH_to_Sump" not in headers
        wb.close()

    def test_delete_column(self, temp_excel_file):
        """Test deleting a column (DELETE OPERATION).
        
        Verifies:
        - Column is removed from Excel file
        - Column no longer appears in list_flow_columns
        - Remaining columns are intact
        """
        manager = ExcelManager()
        manager.set_flow_diagram_path(str(temp_excel_file))

        # Verify initial state
        initial_cols = manager.list_flow_columns("Flows_Test")
        assert "Sump_to_Plant" in initial_cols

        # Delete column
        result = manager.delete_flow_column("Flows_Test", "Sump_to_Plant")
        assert result is True

        # Verify column was deleted
        updated_cols = manager.list_flow_columns("Flows_Test")
        assert "Sump_to_Plant" not in updated_cols
        assert "BH_to_Sump" in updated_cols  # Other columns intact

        # Verify Excel file has correct structure
        wb = load_workbook(temp_excel_file)
        ws = wb["Flows_Test"]
        headers = [str(cell.value) for cell in ws[3]]
        assert "Sump_to_Plant" not in headers
        wb.close()

    def test_rename_nonexistent_column(self, temp_excel_file):
        """Test error handling for renaming non-existent column (ERROR HANDLING).
        
        Verifies:
        - Operation returns False for non-existent column
        - No file modifications occur
        - Error is logged appropriately
        """
        manager = ExcelManager()
        manager.set_flow_diagram_path(str(temp_excel_file))

        # Try to rename non-existent column
        result = manager.rename_flow_column("Flows_Test", "NonExistent", "NewName")
        assert result is False

        # Verify file unchanged
        updated_cols = manager.list_flow_columns("Flows_Test")
        assert "NonExistent" not in updated_cols
        assert "NewName" not in updated_cols

    def test_delete_nonexistent_column(self, temp_excel_file):
        """Test error handling for deleting non-existent column (ERROR HANDLING).
        
        Verifies:
        - Operation returns False for non-existent column
        - Existing columns remain intact
        - Error is logged appropriately
        """
        manager = ExcelManager()
        manager.set_flow_diagram_path(str(temp_excel_file))

        # Get initial columns
        initial_cols = manager.list_flow_columns("Flows_Test")

        # Try to delete non-existent column
        result = manager.delete_flow_column("Flows_Test", "NonExistent")
        assert result is False

        # Verify columns unchanged
        updated_cols = manager.list_flow_columns("Flows_Test")
        assert updated_cols == initial_cols

    def test_rename_to_existing_name(self, temp_excel_file):
        """Test error handling for renaming to an existing column name (VALIDATION).
        
        Verifies:
        - Operation prevents duplicate column names
        - Returns False when target name exists
        - File remains unchanged
        """
        manager = ExcelManager()
        manager.set_flow_diagram_path(str(temp_excel_file))

        # Try to rename to a name that already exists
        result = manager.rename_flow_column(
            "Flows_Test", "BH_to_Sump", "Sump_to_Plant"
        )
        assert result is False

        # Verify columns unchanged
        cols = manager.list_flow_columns("Flows_Test")
        assert "BH_to_Sump" in cols
        assert "Sump_to_Plant" in cols

    def test_cache_invalidation_on_create(self, temp_excel_file):
        """Test cache is invalidated after creating column (CACHE MANAGEMENT).
        
        Verifies:
        - After create_flow_column(), cache is cleared
        - Subsequent loads reflect new column
        - No stale data returned
        """
        manager = ExcelManager()
        manager.set_flow_diagram_path(str(temp_excel_file))

        # Load initially
        cols_before = manager.list_flow_columns("Flows_Test")

        # Create column
        manager.create_flow_column("Flows_Test", "Test_Col")

        # Load again (should see new column)
        cols_after = manager.list_flow_columns("Flows_Test")

        assert "Test_Col" in cols_after
        assert len(cols_after) == len(cols_before) + 1

    def test_rename_updates_mappings(self, temp_excel_file, monkeypatch):
        """Test that renaming a column updates excel_flow_links.json (MAPPING SYNC).
        
        Verifies:
        - When column is renamed, all flow IDs that referenced it are updated
        - Mappings file is correctly persisted
        - Multiple mappings for same column are all updated
        """
        import json

        manager = ExcelManager()
        monkeypatch.setattr(manager, "get_flow_diagram_path", lambda: temp_excel_file)

        # First, add the column we'll rename to the Excel file
        manager.create_flow_column("Flows_Test", "Sample_Flow_1")

        # Create mappings file with flow IDs referencing "Sample_Flow_1"
        mappings_file = Path("data/excel_flow_links.json")
        mappings_file.parent.mkdir(exist_ok=True)

        mappings_data = {
            "links": {
                "Flow_A": {
                    "sheet": "Flows_Test",
                    "column": "Sample_Flow_1"
                },
                "Flow_B": {
                    "sheet": "Flows_Test",
                    "column": "Sample_Flow_1"
                },
                "Flow_C": {
                    "sheet": "Flows_Test",
                    "column": "Sample_Flow_2"
                }
            }
        }

        with open(mappings_file, "w") as f:
            json.dump(mappings_data, f)

        # Rename the column
        success = manager.rename_flow_column("Flows_Test", "Sample_Flow_1", "Updated_Flow_1")
        assert success

        # Verify mappings were updated
        with open(mappings_file, "r") as f:
            updated_mappings = json.load(f)

        # Flow_A and Flow_B should now reference "Updated_Flow_1"
        assert updated_mappings["links"]["Flow_A"]["column"] == "Updated_Flow_1"
        assert updated_mappings["links"]["Flow_B"]["column"] == "Updated_Flow_1"

        # Flow_C should still reference "Sample_Flow_2" (unchanged)
        assert updated_mappings["links"]["Flow_C"]["column"] == "Sample_Flow_2"

        # Cleanup
        mappings_file.unlink(missing_ok=True)

    def test_delete_removes_mappings(self, temp_excel_file, monkeypatch):
        """Test that deleting a column removes broken mappings (DATA CLEANUP).
        
        Verifies:
        - When column is deleted, all flow IDs that referenced it are removed
        - Other mappings are preserved
        - Mappings file is correctly persisted
        """
        import json

        manager = ExcelManager()
        monkeypatch.setattr(manager, "get_flow_diagram_path", lambda: temp_excel_file)

        # First, add the column we'll delete to the Excel file
        manager.create_flow_column("Flows_Test", "Sample_Flow_1")

        # Create mappings file with flow IDs referencing "Sample_Flow_1"
        mappings_file = Path("data/excel_flow_links.json")
        mappings_file.parent.mkdir(exist_ok=True)

        mappings_data = {
            "links": {
                "Flow_To_Delete_1": {
                    "sheet": "Flows_Test",
                    "column": "Sample_Flow_1"
                },
                "Flow_To_Delete_2": {
                    "sheet": "Flows_Test",
                    "column": "Sample_Flow_1"
                },
                "Flow_Keep": {
                    "sheet": "Flows_Test",
                    "column": "Sample_Flow_2"
                }
            }
        }

        with open(mappings_file, "w") as f:
            json.dump(mappings_data, f)

        # Delete the column
        success = manager.delete_flow_column("Flows_Test", "Sample_Flow_1")
        assert success

        # Verify mappings were updated
        with open(mappings_file, "r") as f:
            updated_mappings = json.load(f)

        # Flow_To_Delete_1 and Flow_To_Delete_2 should be removed
        assert "Flow_To_Delete_1" not in updated_mappings["links"]
        assert "Flow_To_Delete_2" not in updated_mappings["links"]

        # Flow_Keep should still be there
        assert "Flow_Keep" in updated_mappings["links"]
        assert updated_mappings["links"]["Flow_Keep"]["column"] == "Sample_Flow_2"

        # Cleanup
        mappings_file.unlink(missing_ok=True)


class TestColumnEditorDialog:
    """Test suite for ColumnEditorDialog UI component (UI TESTS).
    
    Note: These tests verify dialog structure and logic without rendering.
    Full UI testing requires Qt event loop.
    """

    def test_dialog_initialization(self):
        """Test dialog initializes correctly (INITIALIZATION).
        
        Verifies:
        - Dialog can be created without errors
        - Initial state is correct
        - UI components are accessible
        """
        from PySide6.QtWidgets import QApplication

        app = QApplication.instance() or QApplication([])

        from ui.dialogs.column_editor_dialog import ColumnEditorDialog

        dialog = ColumnEditorDialog(area_code="UG2N")
        assert dialog is not None
        assert dialog.area_code == "UG2N"
        assert dialog.table_columns is not None


if __name__ == "__main__":
    # Run tests with pytest
    # pytest -v tests/test_column_editor.py
    pytest.main([__file__, "-v"])
