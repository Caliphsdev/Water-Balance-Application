"""Tests for ExcelManager flow diagram column operations.

Covers:
- Auto-mapping flow IDs to existing columns
- Creating new flow columns without duplication
- Reading a flow volume for a given month
"""

from __future__ import annotations

from pathlib import Path
from typing import Tuple
import sys

# Ensure src/ is on sys.path for test imports (pytest runs from repo root)
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "src"))

import yaml
from openpyxl import Workbook, load_workbook

from core.config_manager import config
from services.excel_manager import ExcelManager


def _write_flow_excel(path: Path) -> None:
    """Create a minimal Flow Diagram Excel file for tests.

    Args:
        path: Destination path for the workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Flows_UG2 North"
    ws.append(["Year", "Month", "bh_ndgwa_to_sump_1"])
    ws.append([2026, 2, 123.45])
    wb.save(path)
    wb.close()


def _load_test_manager(tmp_path: Path, excel_path: Path) -> Tuple[ExcelManager, Path]:
    """Create a ConfigManager-backed ExcelManager for tests.

    Args:
        tmp_path: Pytest temporary directory.
        excel_path: Path to the test Excel workbook.

    Returns:
        Tuple of (ExcelManager instance, config file path).
    """
    config_dir = tmp_path / "config"
    config_dir.mkdir(parents=True, exist_ok=True)
    cfg_path = config_dir / "app_config.yaml"

    cfg = {
        "data_sources": {
            "timeseries_excel_path": str(excel_path),
        }
    }
    cfg_path.write_text(yaml.safe_dump(cfg), encoding="utf-8")

    # Force config to use the test YAML file.
    config.load_config(config_path=str(cfg_path))

    return ExcelManager(config), cfg_path


def test_auto_map_flow_column(tmp_path: Path) -> None:
    """Auto-map should resolve the flow column based on IDs."""
    excel_path = tmp_path / "flow_diagram.xlsx"
    _write_flow_excel(excel_path)

    manager, _ = _load_test_manager(tmp_path, excel_path)
    mapping = manager.auto_map_flow_column(
        from_id="BH_NDGWA",
        to_id="SUMP_1",
        area_code_or_sheet="Flows_UG2 North",
    )

    assert mapping is not None
    assert mapping["sheet"] == "Flows_UG2 North"
    assert mapping["column"] == "bh_ndgwa_to_sump_1"


def test_create_flow_column_no_duplicate(tmp_path: Path) -> None:
    """Creating an existing column should not duplicate headers."""
    excel_path = tmp_path / "flow_diagram.xlsx"
    _write_flow_excel(excel_path)

    manager, _ = _load_test_manager(tmp_path, excel_path)
    sheet_name = "Flows_UG2 North"
    existing_headers_before = manager.list_flow_columns(sheet_name)

    # Attempt to create a column that already exists
    created = manager.create_flow_column(sheet_name, "bh_ndgwa_to_sump_1")
    existing_headers_after = manager.list_flow_columns(sheet_name)

    assert created is True
    assert existing_headers_before == existing_headers_after


def test_create_flow_column_adds_header(tmp_path: Path) -> None:
    """Creating a new column should append a header to the sheet."""
    excel_path = tmp_path / "flow_diagram.xlsx"
    _write_flow_excel(excel_path)

    manager, _ = _load_test_manager(tmp_path, excel_path)
    sheet_name = "Flows_UG2 North"
    new_column = manager.suggest_flow_column_name("BH_NDGWA", "PLANT")

    created = manager.create_flow_column(sheet_name, new_column)
    assert created is True

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    headers = [str(cell.value).strip() for cell in ws[1] if cell.value is not None]
    wb.close()

    assert new_column in headers


def test_get_flow_volume_for_month(tmp_path: Path) -> None:
    """Flow volume lookup should return the numeric value for year/month."""
    excel_path = tmp_path / "flow_diagram.xlsx"
    _write_flow_excel(excel_path)

    manager, _ = _load_test_manager(tmp_path, excel_path)
    volume = manager.get_flow_volume("Flows_UG2 North", "bh_ndgwa_to_sump_1", 2026, 2)

    assert volume == 123.45
