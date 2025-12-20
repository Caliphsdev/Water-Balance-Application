from openpyxl import load_workbook

wb = load_workbook("test_templates/Water_Balance_TimeSeries_Template.xlsx")
ws = wb["Flows_MERS"]

print("Flows_MERS row 4 (data), last 15 columns:")
for col in range(ws.max_column - 14, ws.max_column + 1):
    header = ws.cell(3, col).value
    data = ws.cell(4, col).value
    print(f"  Col {col} ({header if header else 'None'}): {data}")
