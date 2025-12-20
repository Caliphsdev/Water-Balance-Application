from openpyxl import load_workbook

wb = load_workbook("test_templates/Water_Balance_TimeSeries_Template.xlsx")
ws = wb["Flows_MERS"]

print("Headers in Flows_MERS (row 3):")
headers = [ws.cell(3, c).value for c in range(1, ws.max_column+1)]
for i, h in enumerate(headers, start=1):
    print(f"  {i:>3}: {h}")

print("\nRow 4 numeric flags for MERS_* columns:")
for i, h in enumerate(headers, start=1):
    if isinstance(h, str) and h.startswith("MERS"):
        v = ws.cell(4, i).value
        is_num = isinstance(v, (int,float))
        print(f"  {i:>3}: {h} -> {v} (num={is_num})")
