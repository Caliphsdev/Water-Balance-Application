"""Automatically create and format the professional License Manager Excel sheet."""

import sys
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_professional_license_sheet(output_path: str):
    """Create professionally formatted license manager Excel sheet."""
    
    print("🎨 Creating Professional License Manager Sheet...")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "licenses"
    
    # Define headers
    headers = [
        "license_key",
        "status", 
        "expiry_date",
        "hw_component_1",
        "hw_component_2",
        "hw_component_3",
        "transfer_count",
        "licensee_name",
        "licensee_email",
        "license_tier",
        "last_validated"
    ]
    
    # Column widths (in characters)
    column_widths = {
        'A': 25,  # license_key
        'B': 12,  # status
        'C': 15,  # expiry_date
        'D': 35,  # hw_component_1
        'E': 35,  # hw_component_2
        'F': 35,  # hw_component_3
        'G': 13,  # transfer_count
        'H': 25,  # licensee_name
        'I': 30,  # licensee_email
        'J': 12,  # license_tier
        'K': 20   # last_validated
    }
    
    # Set column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Header styling
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add data validation for status column (B2:B1000)
    status_validation = DataValidation(
        type="list",
        formula1='"active,revoked,expired,pending"',
        allow_blank=False,
        showDropDown=True,
        showErrorMessage=True,
        error='Please select a valid status',
        errorTitle='Invalid Status'
    )
    status_validation.add(f'B2:B1000')
    ws.add_data_validation(status_validation)
    
    # Add data validation for license_tier column (J2:J1000)
    tier_validation = DataValidation(
        type="list",
        formula1='"trial,standard,premium"',
        allow_blank=False,
        showDropDown=True,
        showErrorMessage=True,
        error='Please select a valid tier',
        errorTitle='Invalid License Tier'
    )
    tier_validation.add(f'J2:J1000')
    ws.add_data_validation(tier_validation)
    
    # Add sample data row
    sample_data = [
        "ABC-123-XYZ",                                          # A: license_key
        "active",                                                # B: status
        "2026-12-31",                                           # C: expiry_date
        "2241367234182d6dc76c5",                                # D: hw_component_1
        "9b4943be60223d8e56f7",                                 # E: hw_component_2
        "9bd943be0822",                                         # F: hw_component_3
        0,                                                       # G: transfer_count
        "Test Company",                                         # H: licensee_name
        "test@example.com",                                     # I: licensee_email
        "standard",                                             # J: license_tier
        ""                                                       # K: last_validated (auto-filled by webhook)
    ]
    
    # Data row styling
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    # Write sample data
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = value
        cell.alignment = data_alignment
        cell.border = data_border
        
        # Special formatting for specific columns
        if col_idx == 7:  # transfer_count
            cell.number_format = '0'  # Integer format
        elif col_idx == 3:  # expiry_date
            cell.number_format = 'YYYY-MM-DD'  # Date format
    
    # Add alternate row shading for next 10 rows (light gray)
    light_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    for row_idx in range(3, 13):
        if row_idx % 2 == 1:  # Odd rows get light fill
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = light_fill
                cell.border = data_border
    
    # Add instructions/notes in a separate sheet
    ws_notes = wb.create_sheet("Setup Instructions")
    ws_notes.column_dimensions['A'].width = 80
    
    instructions = [
        "📋 SETUP INSTRUCTIONS",
        "",
        "1. Upload this file to Google Drive",
        "2. Right-click → Open with Google Sheets (converts to Google Sheets format)",
        "3. In Google Sheets, go to Extensions → Apps Script",
        "4. Copy code from: docs/Google_Apps_Script_License_Webhook.js",
        "5. Deploy as Web App (Execute as: Me, Access: Anyone)",
        "6. Copy the webhook URL",
        "7. Update config/app_config.yaml with:",
        "   - sheet_url: (the Google Sheets URL)",
        "   - webhook_url: (the Apps Script deployment URL)",
        "8. Make sheet public (Share → Anyone with link → Viewer)",
        "",
        "✅ CONDITIONAL FORMATTING (Apply in Google Sheets):",
        "",
        "After uploading to Google Sheets, apply these formatting rules:",
        "",
        "Status Column (B) - 4 Rules:",
        "  • active → Green background (#C8E6C9), Green text (#2E7D32)",
        "  • revoked → Red background (#FFCDD2), Red text (#D32F2F)",
        "  • expired → Orange background (#FFE0B2), Orange text (#F57C00)",
        "  • pending → Blue background (#BBDEFB), Blue text (#1976D2)",
        "",
        "Expiry Date Column (C) - 1 Rule:",
        "  • Custom formula: =AND(C2<>\"\", C2>=TODAY(), C2<=TODAY()+7)",
        "  • Yellow background (#FFF9C4), Orange text (#F57C00)",
        "  • (Highlights licenses expiring within 7 days)",
        "",
        "📊 COLUMN REFERENCE:",
        "",
        "A - license_key: Unique license identifier (e.g., ABC-123-XYZ)",
        "B - status: License status (active/revoked/expired/pending) - DROPDOWN",
        "C - expiry_date: Expiration date (YYYY-MM-DD format)",
        "D - hw_component_1: MAC address hash (auto-filled by app)",
        "E - hw_component_2: CPU ID hash (auto-filled by app)",
        "F - hw_component_3: Motherboard hash (auto-filled by app)",
        "G - transfer_count: Number of hardware transfers (auto-increments)",
        "H - licensee_name: Customer/company name",
        "I - licensee_email: Contact email",
        "J - license_tier: License tier (trial/standard/premium) - DROPDOWN",
        "K - last_validated: Last online validation timestamp (auto-updated)",
        "",
        "🔒 AUTO-UPDATED COLUMNS (Don't edit manually):",
        "  • D, E, F: Hardware components (filled on activation)",
        "  • G: Transfer count (increments on each transfer)",
        "  • K: Last validated (updates every time app validates)",
        "",
        "✏️ MANUAL ENTRY COLUMNS:",
        "  • A: license_key (you create and enter)",
        "  • B: status (use dropdown, or revoke by changing to 'revoked')",
        "  • C: expiry_date (set when creating license)",
        "  • H, I: licensee info (optional, or auto-filled from app)",
        "  • J: license_tier (affects validation frequency)",
        "",
        "📞 Support: support@water-balance.com | +27 123 456 7890"
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        cell = ws_notes.cell(row=row_idx, column=1)
        cell.value = instruction
        if row_idx == 1:
            cell.font = Font(name='Segoe UI', size=14, bold=True)
        elif instruction.startswith(('✅', '📋', '📊', '🔒', '✏️', '📞')):
            cell.font = Font(name='Segoe UI', size=12, bold=True)
        else:
            cell.font = Font(name='Segoe UI', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Save file
    output_file = Path(output_path)
    wb.save(output_file)
    
    print(f"✅ Sheet created successfully!")
    print(f"📁 Location: {output_file}")
    print(f"📊 Structure: 11 columns with professional formatting")
    print(f"🎨 Features:")
    print(f"   ✅ Header: Bold, dark blue (#0D47A1), white text, frozen")
    print(f"   ✅ Dropdowns: Status (active/revoked/expired/pending)")
    print(f"   ✅ Dropdowns: License Tier (trial/standard/premium)")
    print(f"   ✅ Sample data: 1 test license row")
    print(f"   ✅ Instructions: See 'Setup Instructions' sheet")
    print(f"   ✅ Column widths: Optimized for readability")
    print(f"\n🚀 Next Steps:")
    print(f"   1. Upload to Google Drive")
    print(f"   2. Open with Google Sheets")
    print(f"   3. Apply conditional formatting (see Instructions sheet)")
    print(f"   4. Deploy Apps Script webhook")
    print(f"   5. Update config/app_config.yaml")
    print(f"\n📖 Full guide: docs/PROFESSIONAL_GOOGLE_SHEETS_SETUP_GUIDE.md")


if __name__ == "__main__":
    # Output to Downloads folder
    output_path = r"D:\Downloads\Water Balance License Manager.xlsx"
    
    try:
        create_professional_license_sheet(output_path)
    except Exception as e:
        print(f"❌ Error: {e}")
        print(f"\nTrying alternate location...")
        # Fallback to project directory if Downloads folder fails
        output_path = "Water_Balance_License_Manager.xlsx"
        create_professional_license_sheet(output_path)
