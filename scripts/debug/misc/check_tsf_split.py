import json

with open("data/diagrams/ug2_north_decline.json") as f:
    diagram = json.load(f)

print("=== SEARCHING FOR TSF AREAS ===\n")

# Look for TSF nodes
tsf_nodes = [n for n in diagram.get("nodes", []) if "tsf" in n.get("id", "").lower()]
print(f"Nodes with TSF: {len(tsf_nodes)}")
for node in sorted(tsf_nodes):
    print(f"  - {node.get(\"id\")}")

# Look for TSF edges
tsf_edges = []
for edge in diagram.get("edges", []):
    from_id = edge.get("from", "").lower()
    to_id = edge.get("to", "").lower()
    if "tsf" in from_id or "tsf" in to_id:
        tsf_edges.append(edge)

print(f"\n\nEdges involving TSF: {len(tsf_edges)}")
for edge in tsf_edges[:10]:
    print(f"  {edge.get(\"from\")} → {edge.get(\"to\")}")
    print(f"    Sheet: {edge.get(\"excel_mapping\", {}).get(\"sheet\")}")
    print(f"    Column: {edge.get(\"excel_mapping\", {}).get(\"column\")[:50] if edge.get(\"excel_mapping\", {}).get(\"column\") else None}")

# Check template sheets
from openpyxl import load_workbook
wb = load_workbook("test_templates/Water_Balance_TimeSeries_Template.xlsx")
print(f"\n\n=== TEMPLATE SHEETS ===")
print(f"Total sheets: {len(wb.sheetnames)}")
for sheet in wb.sheetnames:
    if "tsf" in sheet.lower() or "old" in sheet.lower() or "new" in sheet.lower():
        print(f"  ✓ {sheet}")

# Check specifically for OLD TSF vs NEW TSF
print("\n\nLooking for split TSF sheets:")
tsf_sheets = [s for s in wb.sheetnames if "tsf" in s.lower()]
print(f"TSF sheets found: {len(tsf_sheets)}")
for s in tsf_sheets:
    print(f"  - {s}")
