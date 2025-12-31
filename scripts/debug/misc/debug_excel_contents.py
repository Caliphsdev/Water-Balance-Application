"""
Debug: Check what's in the latest Excel file and why verification isn't finding flows.
"""
from pathlib import Path
from openpyxl import load_workbook

# Find latest Excel
test_templates = Path('test_templates')
xlsx_files = sorted([f for f in test_templates.glob('Water_Balance_TimeSeries_Template_FIXED_*.xlsx')], 
                    key=lambda x: x.stat().st_mtime, reverse=True)

print(f"Latest Excel files:")
for f in xlsx_files[:3]:
    print(f"  - {f.name}")

if xlsx_files:
    latest = xlsx_files[0]
    print(f"\n✅ Using: {latest.name}\n")
    
    wb = load_workbook(str(latest))
    
    print("Sheets in workbook:")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name != 'Reference Guide':
            # Get flow columns (skip Date, Year, Month)
            flows = [cell.value for cell in ws[3][3:] if cell.value]
            print(f"  {sheet_name:20} → {len(flows):3} flow columns")
        else:
            rows = sum(1 for row in ws.iter_rows() if any(cell.value for cell in row))
            print(f"  {sheet_name:20} → {rows:3} rows")
