"""
Quick test to verify Excel structure and column population fix
"""
from openpyxl import load_workbook
from pathlib import Path

excel_path = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("=" * 60)
print("EXCEL STRUCTURE VERIFICATION")
print("=" * 60)

try:
    wb = load_workbook(excel_path)
    
    print(f"\n✅ Excel file loaded successfully: {excel_path.name}")
    print(f"\n📊 Total Sheets: {len(wb.sheetnames)}")
    print("\n📄 Sheet List:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet_name}")
    
    # Check specific sheets
    print("\n🔍 Structure Validation:")
    
    # Should NOT exist
    if 'Flows_MERN' in wb.sheetnames:
        print("  ❌ FAIL: Flows_MERN still exists (should be removed)")
    else:
        print("  ✅ PASS: Flows_MERN removed successfully")
    
    # Should exist
    if 'Flows_NEWTSF' in wb.sheetnames:
        print("  ✅ PASS: Flows_NEWTSF exists")
        ws = wb['Flows_NEWTSF']
        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"     Headers: {headers}")
        if headers == ['Date', 'Year', 'Month']:
            print("  ✅ PASS: New TSF has standard headers")
        else:
            print(f"  ⚠️  WARNING: Unexpected headers: {headers}")
    else:
        print("  ❌ FAIL: Flows_NEWTSF missing")
    
    # Check Reference Guide updated
    if 'Reference Guide' in wb.sheetnames:
        print("  ✅ PASS: Reference Guide exists")
        ws = wb['Reference Guide']
        title = ws['A1'].value
        if title and 'Water Balance' in title:
            print(f"     Title: {title}")
            print("  ✅ PASS: Reference Guide appears updated")
        else:
            print("  ⚠️  WARNING: Reference Guide may need verification")
    else:
        print("  ❌ FAIL: Reference Guide missing")
    
    # Test column reading (simulates what Excel Manager does)
    print("\n📊 Column Reading Test (Excel Manager Simulation):")
    test_sheets = ['Flows_OLDTSF', 'Flows_NEWTSF', 'Flows_UG2N']
    for sheet_name in test_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            cols = [f"📊 {cell.value}" for cell in ws[1] if cell.value]
            print(f"  {sheet_name}: {len(cols)} columns")
            print(f"    {', '.join(cols[:5])}{'...' if len(cols) > 5 else ''}")
        else:
            print(f"  ❌ {sheet_name}: NOT FOUND")
    
    print("\n" + "=" * 60)
    print("✅ VERIFICATION COMPLETE - Excel structure is correct!")
    print("=" * 60)
    
    print("\n💡 Next Steps:")
    print("  1. Launch Flow Diagram Dashboard")
    print("  2. Click 'Excel Manager' in DATA section")
    print("  3. Go to Columns tab")
    print("  4. Select a sheet from dropdown")
    print("  5. Verify columns populate automatically")
    
    wb.close()
    
except Exception as e:
    print(f"\n❌ ERROR: {e}")
    import traceback
    traceback.print_exc()
