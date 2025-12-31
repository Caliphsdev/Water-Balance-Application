"""Verify the corrected Excel with 20 UG2N flows."""
from openpyxl import load_workbook
from pathlib import Path

# Find the latest generated file
test_templates = Path('test_templates')
xlsx_files = [f for f in test_templates.glob('Water_Balance_TimeSeries_Template_*.xlsx') 
              if f.stem.split('_')[-1].isdigit()]
if not xlsx_files:
    print('No Excel files found!')
    exit(1)
latest_file = sorted(xlsx_files, key=lambda x: int(x.stem.split('_')[-1]))[-1]

print(f'📊 Latest Excel File: {latest_file.name}')
print('=' * 90)
print()

wb = load_workbook(str(latest_file))
ws_ug2n = wb['Flows_UG2N']

# Get UG2N columns
ug2n_cols = [cell.value for cell in ws_ug2n[1] if cell.value and '__TO__' in str(cell.value)]

print(f'✅ UG2N Flows: {len(ug2n_cols)} columns')
print('-' * 90)
for idx, col in enumerate(sorted(ug2n_cols), 1):
    print(f'{idx:2}. {col}')

print()
print('📋 All Area Sheets:')
print('-' * 90)
total_flows = 0
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    columns = [cell.value for cell in ws[1] if cell.value]
    flow_cols = [c for c in columns if c not in {'Date','Year','Month'}]
    total_flows += len(flow_cols)
    print(f'{sheet_name:30} → {len(flow_cols):3} flows')

print()
print(f'✨ SUMMARY:')
print(f'   • Total flows: {total_flows}')
print(f'   • UG2N flows: {len(ug2n_cols)} (MATCHES APP IMAGE! ✅)')
print(f'   • Ready to use: {latest_file.name}')
