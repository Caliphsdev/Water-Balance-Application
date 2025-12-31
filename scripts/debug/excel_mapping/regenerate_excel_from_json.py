"""
Regenerate Excel workbook using JSON diagrams as source of truth.
Extracts all flows from JSON diagrams instead of database connections.
"""

import sys
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent / 'src'))

# Parse all JSON diagrams
diagrams_dir = Path(__file__).parent / 'data' / 'diagrams'
all_edges = []
all_nodes = {}
area_edges = {}

# Find all JSON diagram files
json_files = list(diagrams_dir.glob('*.json'))
print(f'📁 Found {len(json_files)} diagram files:')

for json_file in json_files:
    if 'test_' in json_file.name:  # Skip test files
        continue
    
    try:
        with open(json_file) as f:
            diagram = json.load(f)
        
        filename = json_file.stem
        print(f'   • {json_file.name}')
        
        # Extract nodes
        for node in diagram.get('nodes', []):
            node_id = node.get('id')
            node_label = node.get('label', node_id)
            all_nodes[node_id] = node_label
        
        # Extract edges
        edges = diagram.get('edges', [])
        print(f'     ├─ {len(edges)} edges')
        
        for edge in edges:
            from_node = edge.get('from')
            to_node = edge.get('to')
            edge_data = {
                'from': from_node,
                'to': to_node,
                'column': f'{from_node}__TO__{to_node}',
                'file': filename
            }
            all_edges.append(edge_data)
            
            # Categorize by area - check both from and to nodes
            area = 'OTHER'
            node_to_check = from_node or to_node
            
            if node_to_check:
                # Check for specific UG2N nodes (without suffix)
                if node_to_check in ['rainfall', 'ndcd', 'bh_ndgwa', 'softening', 'reservoir', 'offices', 
                                     'guest_house', 'septic', 'sewage', 'north_decline', 'north_shaft', 
                                     'losses', 'losses2', 'consumption', 'consumption2', 'evaporation',
                                     'dust_suppression', 'spill']:
                    area = 'UG2N'
                # Merensky North
                elif node_to_check in ['rainfall_merensky', 'ndcd_merensky', 'bh_mcgwa', 'softening_merensky', 
                                       'offices_merensky', 'merensky_north_decline', 'merensky_north_shaft',
                                       'losses_merensky', 'consumption_merensky', 'evaporation_merensky',
                                       'dust_suppression_merensky', 'spill_merensky']:
                    area = 'MERN'
                # UG2 South
                elif any(x in node_to_check for x in ['ug2s_', 'ug2s']):
                    area = 'UG2S'
                # UG2 Plant
                elif any(x in node_to_check for x in ['ug2plant_', 'ug2plant']):
                    area = 'UG2P'
                # UG2 (general/uncategorized)
                elif node_to_check.startswith('ug2_'):
                    area = 'UG2'
                # Merensky Plant
                elif any(x in node_to_check for x in ['merplant_', 'merplant']):
                    area = 'MERP'
                # Merensky South
                elif any(x in node_to_check for x in ['mers_', 'mers']):
                    area = 'MERS'
                # Old TSF / Tailings
                elif any(x in node_to_check for x in ['oldtsf_', 'nt_rwd', 'new_tsf', 'trtd', 'spcd1', 'stockpile_']):
                    if 'stockpile' in node_to_check:
                        area = 'STOCKPILE'
                    else:
                        area = 'OLDTSF'
            
            if area not in area_edges:
                area_edges[area] = []
            area_edges[area].append(edge_data)
    
    except Exception as e:
        print(f'   ❌ Error reading {json_file.name}: {e}')

print(f'\n📊 Total edges extracted: {len(all_edges)}')
print(f'🔤 Total unique nodes: {len(all_nodes)}')
print(f'🗂️ Areas found: {len(area_edges)}')

for area, edges in sorted(area_edges.items()):
    print(f'   • {area:15} → {len(edges):3} edges')

# Create workbook
print(f'\n📝 Creating Excel workbook...\n')
wb = Workbook()
wb.remove(wb.active)

# Reference Guide sheet
print('📖 Creating Reference Guide...')
ref_sheet = wb.create_sheet('Reference Guide', 0)

ref_sheet['A1'] = 'Flow Diagram Node Reference'
ref_sheet['A1'].font = Font(bold=True, size=14)
ref_sheet.merge_cells('A1:D1')

# Nodes section
ref_sheet['A3'] = 'All Nodes in Diagrams'
ref_sheet['A3'].font = Font(bold=True, size=11, color='FFFFFF')
ref_sheet['A3'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ref_sheet.merge_cells('A3:B3')

ref_sheet['A4'] = 'Node ID'
ref_sheet['B4'] = 'Label'
for col in ['A', 'B']:
    ref_sheet[f'{col}4'].font = Font(bold=True)
    ref_sheet[f'{col}4'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

row = 5
for node_id in sorted(all_nodes.keys()):
    ref_sheet[f'A{row}'] = node_id
    ref_sheet[f'B{row}'] = all_nodes[node_id]
    row += 1

ref_sheet.column_dimensions['A'].width = 25
ref_sheet.column_dimensions['B'].width = 40

# Flows section
flows_start_row = row + 3
ref_sheet[f'A{flows_start_row}'] = 'All Flow Connections'
ref_sheet[f'A{flows_start_row}'].font = Font(bold=True, size=11, color='FFFFFF')
ref_sheet[f'A{flows_start_row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ref_sheet.merge_cells(f'A{flows_start_row}:D{flows_start_row}')

flow_header_row = flows_start_row + 1
ref_sheet[f'A{flow_header_row}'] = 'Column Name'
ref_sheet[f'B{flow_header_row}'] = 'From'
ref_sheet[f'C{flow_header_row}'] = 'To'
ref_sheet[f'D{flow_header_row}'] = 'Area'

for col in ['A', 'B', 'C', 'D']:
    ref_sheet[f'{col}{flow_header_row}'].font = Font(bold=True)
    ref_sheet[f'{col}{flow_header_row}'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

flow_row = flow_header_row + 1
for edge in sorted(all_edges, key=lambda e: e['column']):
    ref_sheet[f'A{flow_row}'] = edge['column']
    ref_sheet[f'B{flow_row}'] = edge['from']
    ref_sheet[f'C{flow_row}'] = edge['to']
    # Find area for this edge
    for area, area_edges_list in area_edges.items():
        if edge in area_edges_list:
            ref_sheet[f'D{flow_row}'] = area
            break
    flow_row += 1

ref_sheet.column_dimensions['A'].width = 35
ref_sheet.column_dimensions['B'].width = 25
ref_sheet.column_dimensions['C'].width = 25
ref_sheet.column_dimensions['D'].width = 15

print(f'   ✅ Added {len(all_nodes)} nodes')
print(f'   ✅ Added {len(all_edges)} flows\n')

# Area-specific sheets
area_map = {
    'UG2N': 'Flows_UG2N',
    'UG2S': 'Flows_UG2S',
    'UG2P': 'Flows_UG2P',
    'MERN': 'Flows_MERN',
    'MERP': 'Flows_MERP',
    'MERS': 'Flows_MERENSKY_SOUTH',
    'OLDTSF': 'Flows_OLDTSF',
    'STOCKPILE': 'Flows_STOCKPILE',
    'OTHER': 'Flows_OTHER'
}

for area, sheet_name in area_map.items():
    if area in area_edges and area_edges[area]:
        print(f'📄 Creating {sheet_name} sheet...')
        ws = wb.create_sheet(sheet_name)
        
        # Headers: Date/Year/Month + flow columns
        ws['A1'] = 'Date'
        ws['B1'] = 'Year'
        ws['C1'] = 'Month'
        
        # Add flow columns (starting from column D)
        from openpyxl.utils import get_column_letter
        col_idx = 4  # D is the 4th column
        for edge in sorted(area_edges[area], key=lambda e: e['column']):
            col_letter = get_column_letter(col_idx)
            ws[f'{col_letter}1'] = edge['column']
            col_idx += 1
        
        # Format header row
        for cell in ws[1]:
            if cell.value:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Adjust column widths
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 10
        for col_idx in range(4, 4 + len(area_edges[area])):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
        
        ws.row_dimensions[1].height = 40
        
        print(f'   ✅ Sheet created with {len(area_edges[area])} flows\n')

# Save workbook
import time
timestamp = int(time.time())
output_path = Path(__file__).parent / 'test_templates' / f'Water_Balance_TimeSeries_Template_{timestamp}.xlsx'
wb.save(str(output_path))

print(f'✅ Excel regenerated successfully!')
print(f'📊 Total flows from JSON diagrams: {len(all_edges)}')
print(f'📁 Output: {output_path}')
print(f'\n✨ Excel now uses JSON diagrams as source of truth')
print(f'✨ All {len(area_edges)} area sheets created')
