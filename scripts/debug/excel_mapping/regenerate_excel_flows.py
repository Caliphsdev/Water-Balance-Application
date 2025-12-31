"""
Regenerate Excel sheets with source->destination flow naming.
Much clearer for users to understand where each flow comes from and goes.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Flow reference: (Sheet Name, Column Name, From Component, To Component, Description, Type)
FLOW_DEFINITIONS = [
    # UG2 North Decline Area
    ('Flows_UG2N', 'BOREHOLE_TO_SOFTENING', 'BOREHOLE', 'SOFTENING_PLANT', 'Water from borehole to softening plant', 'Inflow'),
    ('Flows_UG2N', 'SOFTENING_TO_RESERVOIR', 'SOFTENING_PLANT', 'RESERVOIR', 'Softened water to storage reservoir', 'Inflow'),
    ('Flows_UG2N', 'RESERVOIR_TO_OFFICES', 'RESERVOIR', 'OFFICES', 'Water for offices and ablutions', 'Inflow'),
    ('Flows_UG2N', 'RESERVOIR_TO_GUEST_HOUSE', 'RESERVOIR', 'GUEST_HOUSE', 'Water for guest accommodation', 'Inflow'),
    ('Flows_UG2N', 'SEWAGE_TO_TREATMENT', 'OFFICES', 'SEWAGE_TREATMENT', 'Wastewater from offices', 'Dirty'),
    ('Flows_UG2N', 'SEWAGE_TO_NDCD', 'SEWAGE_TREATMENT', 'NDCD', 'Treated effluent to underground', 'Dirty'),
    ('Flows_UG2N', 'RAINFALL_COLLECTION', 'DIRECT_RAINFALL', 'RESERVOIR', 'Direct rainwater collected', 'Inflow'),
    ('Flows_UG2N', 'NDCD_DEWATERING', 'NDCD_SHAFTS', 'NORTH_DECLINE', 'Dewatering from north decline shafts 1-2', 'Inflow'),
    ('Flows_UG2N', 'NDCD_TO_SEPTIC', 'NORTH_DECLINE', 'SEPTIC_TANK', 'Effluent to septic tank', 'Outflow'),
    ('Flows_UG2N', 'NDCD_TO_CONSUMPTION', 'NORTH_DECLINE', 'CONSUMPTION', 'Water for mining consumption', 'Outflow'),
    ('Flows_UG2N', 'NDCD_EVAPORATION', 'NORTH_DECLINE', 'LOSSES', 'Evaporation loss from shaft area', 'Loss'),
    ('Flows_UG2N', 'NDCD_SPILL', 'NORTH_DECLINE', 'SPILL', 'Spillage loss from underground', 'Loss'),
    ('Flows_UG2N', 'NDCD_DUST_SUPPRESSION', 'NORTH_DECLINE', 'DUST_SUPPRESSION', 'Dust suppression water', 'Outflow'),
    
    # Merensky North
    ('Flows_MERN', 'RAINFALL_MERN_COLLECTION', 'DIRECT_RAINFALL', 'MERN_AREA', 'Rainwater in Merensky North', 'Inflow'),
    ('Flows_MERN', 'MERN_TO_TREATMENT', 'MERN_STORAGE', 'TREATMENT_PLANT', 'Water to treatment plant', 'Inflow'),
    ('Flows_MERN', 'TREATMENT_TO_PLANT', 'TREATMENT_PLANT', 'MERENSKY_PLANT', 'Treated water to processing plant', 'Inflow'),
    ('Flows_MERN', 'PLANT_OUTFLOW', 'MERENSKY_PLANT', 'DISCHARGE', 'Discharge from processing', 'Outflow'),
    ('Flows_MERN', 'CONSERVATION_POOL_VOLUME', 'STORAGE', 'CONSERVATION_POOL', 'Water in conservation storage', 'Storage'),
    ('Flows_MERN', 'EVAPORATION_MERN', 'MERN_STORAGE', 'LOSSES', 'Evaporation loss Merensky North', 'Loss'),
    
    # Merensky South
    ('Flows_MERENSKY_SOUTH', 'RAINFALL_SOUTH_COLLECTION', 'DIRECT_RAINFALL', 'MERENSKY_SOUTH', 'Rainwater in Merensky South', 'Inflow'),
    ('Flows_MERENSKY_SOUTH', 'SOUTH_STORAGE', 'MERENSKY_SOUTH', 'STORAGE', 'Water in storage facilities', 'Storage'),
    ('Flows_MERENSKY_SOUTH', 'TREATMENT_OUTFLOW_SOUTH', 'TREATMENT', 'DISCHARGE', 'Treated water discharge', 'Outflow'),
    ('Flows_MERENSKY_SOUTH', 'EVAPORATION_SOUTH', 'STORAGE', 'LOSSES', 'Evaporation loss Merensky South', 'Loss'),
    ('Flows_MERENSKY_SOUTH', 'RECIRCULATION_SOUTH', 'STORAGE', 'PROCESS', 'Water recirculated to process', 'Recirculation'),
    
    # UG2 South
    ('Flows_UG2S', 'RAINFALL_UG2S_COLLECTION', 'DIRECT_RAINFALL', 'UG2S_AREA', 'Rainwater in UG2 South', 'Inflow'),
    ('Flows_UG2S', 'UG2S_DEWATERING', 'UG2S_SHAFTS', 'UG2S_PROCESSING', 'Dewatering from UG2 South', 'Inflow'),
    ('Flows_UG2S', 'PROCESSING_OUTFLOW', 'UG2S_PROCESSING', 'OUTFLOW_POINT', 'Water leaving plant', 'Outflow'),
    ('Flows_UG2S', 'SEEPAGE_LOSS_UG2S', 'UG2S_AREA', 'LOSSES', 'Seepage loss UG2 South', 'Loss'),
    
    # Stockpile
    ('Flows_STOCKPILE', 'RAINFALL_STOCKPILE_COLLECTION', 'DIRECT_RAINFALL', 'STOCKPILE', 'Rainwater on stockpile', 'Inflow'),
    ('Flows_STOCKPILE', 'STOCKPILE_RUNOFF', 'STOCKPILE', 'RUNOFF_POINT', 'Runoff from stockpile surface', 'Outflow'),
    ('Flows_STOCKPILE', 'STOCKPILE_INFILTRATION', 'STOCKPILE', 'INFILTRATION_LOSS', 'Water infiltrating into ground', 'Loss'),
    ('Flows_STOCKPILE', 'PUMP_EXTRACTION', 'STOCKPILE', 'PUMP', 'Water pumped from stockpile', 'Outflow'),
    ('Flows_STOCKPILE', 'PUMP_TO_TREATMENT', 'PUMP', 'TREATMENT', 'Pumped water to treatment', 'Inflow'),
    
    # Old TSF
    ('Flows_OLDTSF', 'RAINFALL_TSF_COLLECTION', 'DIRECT_RAINFALL', 'TSF', 'Rainwater on TSF', 'Inflow'),
    ('Flows_OLDTSF', 'INFLOW_TO_TSF', 'SOURCES', 'TSF', 'Water fed into TSF', 'Inflow'),
    ('Flows_OLDTSF', 'DECANT_FROM_TSF', 'TSF', 'DECANT_POINT', 'Clarified water decanted from TSF', 'Outflow'),
    ('Flows_OLDTSF', 'SEEPAGE_TSF', 'TSF', 'SEEPAGE_LOSS', 'Seepage through TSF walls/base', 'Loss'),
    ('Flows_OLDTSF', 'EVAPORATION_TSF', 'TSF', 'EVAPORATION_LOSS', 'Evaporation from TSF surface', 'Loss'),
    ('Flows_OLDTSF', 'SPILLWAY_TSF', 'TSF', 'SPILLWAY_POINT', 'Emergency spillway flow', 'Outflow'),
    
    # UG2 Plant
    ('Flows_UG2PLANT', 'INFLOW_UG2PLANT', 'SOURCES', 'PLANT', 'Total inflow to UG2 plant', 'Inflow'),
    ('Flows_UG2PLANT', 'PROCESS_CONSUMPTION', 'PLANT', 'PROCESS', 'Water used in processing', 'Outflow'),
    ('Flows_UG2PLANT', 'RECYCLED_TO_PROCESS', 'RECYCLING', 'PROCESS', 'Recycled water back to process', 'Inflow'),
    ('Flows_UG2PLANT', 'PLANT_OUTFLOW_UG2', 'PLANT', 'DISCHARGE', 'Total outflow from UG2 plant', 'Outflow'),
    ('Flows_UG2PLANT', 'EVAPORATION_UG2PLANT', 'PLANT', 'EVAPORATION_LOSS', 'Evaporation at plant site', 'Loss'),
    ('Flows_UG2PLANT', 'TREATMENT_DISCHARGE_UG2', 'TREATMENT', 'DISCHARGE', 'Treated discharge UG2', 'Outflow'),
    
    # Merensky Plant
    ('Flows_MERPLANT', 'INFLOW_MERPLANT', 'SOURCES', 'PLANT', 'Total inflow to Merensky plant', 'Inflow'),
    ('Flows_MERPLANT', 'PROCESS_CONSUMPTION_MER', 'PLANT', 'PROCESS', 'Water used in Merensky processing', 'Outflow'),
    ('Flows_MERPLANT', 'RECYCLED_TO_MERPLANT', 'RECYCLING', 'PROCESS', 'Recycled water at Merensky', 'Inflow'),
    ('Flows_MERPLANT', 'PLANT_OUTFLOW_MER', 'PLANT', 'DISCHARGE', 'Total outflow from Merensky plant', 'Outflow'),
    ('Flows_MERPLANT', 'EVAPORATION_MERPLANT', 'PLANT', 'EVAPORATION_LOSS', 'Evaporation at Merensky plant', 'Loss'),
    ('Flows_MERPLANT', 'THICKENER_PROCESS', 'THICKENER', 'PROCESS', 'Tailings thickener water', 'Process'),
]

# Abbreviations reference
ABBREVIATIONS = [
    ('NDCD', 'North Decline (Shafts 1-2, NDCD1-4)'),
    ('NDSWD', 'North Decline Sewage Water Disposal'),
    ('TSF', 'Tailings Storage Facility'),
    ('UG2N', 'UG2 North Decline Area'),
    ('UG2S', 'UG2 South Decline Area'),
    ('UG2PLANT', 'UG2 Processing Plant'),
    ('MERN', 'Merensky North Area'),
    ('MERPLANT', 'Merensky Processing Plant'),
    ('MERENSKY_SOUTH', 'Merensky South Mining Area'),
    ('STP', 'Sewage Treatment Plant'),
    ('SEPTIC', 'Septic Tank (wastewater)'),
    ('RESERVOIR', 'Water Storage Reservoir'),
    ('STORAGE', 'General Storage Facility'),
    ('INFILTRATION', 'Water soaking into ground'),
    ('SEEPAGE', 'Water leaking through soil'),
    ('DECANT', 'Clarified water poured off (from TSF)'),
    ('EVAPORATION', 'Water loss through atmospheric evaporation'),
    ('SPILL/SPILLAGE', 'Unplanned water overflow or loss'),
    ('RUNOFF', 'Water flowing over ground surface'),
    ('RECIRCULATION', 'Water reused in internal loop'),
    ('INFLOW', 'Water entering an area or component'),
    ('OUTFLOW', 'Water leaving an area or component'),
    ('EFFLUENT', 'Treated wastewater discharge'),
    ('CONSUMPTION', 'Water used in mining/processing operations'),
]

def regenerate_excel_with_flow_columns():
    """Regenerate Excel with source->destination column names."""
    excel_path = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')
    
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Group flows by sheet
    flows_by_sheet = {}
    for sheet, col, from_comp, to_comp, desc, ftype in FLOW_DEFINITIONS:
        if sheet not in flows_by_sheet:
            flows_by_sheet[sheet] = []
        flows_by_sheet[sheet].append({
            'column': col,
            'from': from_comp,
            'to': to_comp,
            'description': desc,
            'type': ftype
        })
    
    # Regenerate each flow sheet
    for sheet_name in flows_by_sheet.keys():
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        ws = wb.create_sheet(sheet_name)
        flows = flows_by_sheet[sheet_name]
        
        # Headers
        headers = ['Date'] + [f['column'] for f in flows]
        ws.append(headers)
        
        # Format header row
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        for i, f in enumerate(flows, 2):
            col_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[col_letter].width = 22
        
        # Add 24 months of sample data (zeros)
        from datetime import datetime, timedelta
        start_date = datetime(2025, 1, 1)
        for row_num in range(2, 26):  # 24 months
            date_val = start_date + timedelta(days=30 * (row_num - 2))
            ws.cell(row=row_num, column=1).value = date_val
            # Leave all flow columns as 0
            for col_num in range(2, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).value = 0
        
        # Format date column
        for row in range(2, 26):
            ws.cell(row=row, column=1).number_format = 'YYYY-MM-DD'
        
        print(f"✅ Regenerated sheet: {sheet_name} ({len(flows)} flow columns)")
    
    # Update/create Reference Guide with abbreviations
    if 'Reference Guide' in wb.sheetnames:
        del wb['Reference Guide']
    
    ref_ws = wb.create_sheet('Reference Guide', 0)
    ref_ws.column_dimensions['A'].width = 25
    ref_ws.column_dimensions['B'].width = 60
    ref_ws.column_dimensions['C'].width = 15
    ref_ws.column_dimensions['D'].width = 45
    
    # Part 1: Column Descriptions
    ref_ws['A1'] = 'FLOW COLUMN REFERENCE (Source → Destination)'
    ref_ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ref_ws['A1'].fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    ref_ws.merge_cells('A1:D1')
    
    headers = ['Column Name', 'From → To', 'Type', 'Description']
    for col_num, header in enumerate(headers, 1):
        cell = ref_ws.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    type_colors = {
        'Inflow': 'D4EDDA',
        'Outflow': 'F8D7DA',
        'Loss': 'FFF3CD',
        'Storage': 'D1ECF1',
        'Recirculation': 'E7D4F5',
        'Process': 'E2E3E5',
        'Dirty': 'FFD7D7',
    }
    
    row = 3
    for sheet, col, from_comp, to_comp, desc, ftype in FLOW_DEFINITIONS:
        ref_ws.cell(row=row, column=1).value = col
        ref_ws.cell(row=row, column=2).value = f"{from_comp} → {to_comp}"
        ref_ws.cell(row=row, column=3).value = ftype
        ref_ws.cell(row=row, column=4).value = desc
        
        # Color by type
        bg = type_colors.get(ftype, 'FFFFFF')
        for col_num in range(1, 5):
            cell = ref_ws.cell(row=row, column=col_num)
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_num == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # Part 2: Abbreviations legend (separate section below)
    abbrev_start_row = row + 2
    ref_ws.cell(row=abbrev_start_row, column=1).value = 'ABBREVIATIONS & COMPONENT CODES'
    ref_ws.cell(row=abbrev_start_row, column=1).font = Font(bold=True, size=12, color='FFFFFF')
    ref_ws.cell(row=abbrev_start_row, column=1).fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    ref_ws.merge_cells(f'A{abbrev_start_row}:B{abbrev_start_row}')
    
    abbrev_header_row = abbrev_start_row + 1
    ref_ws.cell(row=abbrev_header_row, column=1).value = 'Abbreviation'
    ref_ws.cell(row=abbrev_header_row, column=2).value = 'Meaning'
    for col_num in (1, 2):
        cell = ref_ws.cell(row=abbrev_header_row, column=col_num)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    abbrev_row = abbrev_header_row + 1
    for abbrev, meaning in ABBREVIATIONS:
        ref_ws.cell(row=abbrev_row, column=1).value = abbrev
        ref_ws.cell(row=abbrev_row, column=2).value = meaning
        ref_ws.cell(row=abbrev_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        abbrev_row += 1
    
    ref_ws.freeze_panes = f'A3'
    
    # Save
    wb.save(excel_path)
    print(f"\n✅ Excel regenerated with flow-specific columns")
    print(f"📋 {len(FLOW_DEFINITIONS)} flow columns defined (Source → Destination)")
    print(f"📖 Reference Guide includes {len(ABBREVIATIONS)} abbreviations")

if __name__ == '__main__':
    regenerate_excel_with_flow_columns()
