import json
from openpyxl import load_workbook

# Load JSON diagram
with open('data/diagrams/ug2_north_decline.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

edges = data.get('edges', [])

# Check which edges have mappings
with_mapping = 0
without_mapping = 0
enabled_mapping = 0
disabled_mapping = 0

print("Checking edge mappings:\n")

# Group by status
enabled_edges = []
disabled_edges = []
no_mapping_edges = []

for edge in edges:
    edge_id = f"{edge.get('from')} → {edge.get('to')}"
    mapping = edge.get('excel_mapping', {})
    
    if not mapping:
        without_mapping += 1
        no_mapping_edges.append(edge_id)
    else:
        with_mapping += 1
        if mapping.get('enabled'):
            enabled_mapping += 1
            enabled_edges.append({
                'id': edge_id,
                'sheet': mapping.get('sheet'),
                'column': mapping.get('column'),
                'label': edge.get('label', '-')
            })
        else:
            disabled_mapping += 1
            disabled_edges.append(edge_id)

print(f"Total edges: {len(edges)}")
print(f"  With mapping: {with_mapping}")
print(f"    - Enabled: {enabled_mapping}")
print(f"    - Disabled: {disabled_mapping}")
print(f"  Without mapping: {without_mapping}")
print()

# Check if Excel columns exist for enabled mappings
print("Checking Excel columns for enabled mappings:\n")
wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')

sheets_data = {}
for sheet_name in ['Flows_UG2P', 'Flows_MERP', 'Flows_UG2N', 'Flows_STOCKPILE']:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[3] if cell.value]
        sheets_data[sheet_name] = set(headers)

found_in_excel = 0
not_found_in_excel = 0

for edge in enabled_edges:
    sheet = edge['sheet']
    column = edge['column']
    
    if sheet in sheets_data:
        if column in sheets_data[sheet]:
            found_in_excel += 1
        else:
            not_found_in_excel += 1
            print(f"❌ Missing in Excel: {edge['id']}")
            print(f"   Sheet: {sheet}, Column: {column}")

print()
print(f"Enabled mappings: {enabled_mapping}")
print(f"  Found in Excel: {found_in_excel}")
print(f"  Not found in Excel: {not_found_in_excel}")
print()

# Sample some edges without data
print("\nSample edges without mappings (first 10):")
for edge_id in no_mapping_edges[:10]:
    print(f"  - {edge_id}")

print()
print("\nSample disabled edges (first 10):")
for edge_id in disabled_edges[:10]:
    print(f"  - {edge_id}")
