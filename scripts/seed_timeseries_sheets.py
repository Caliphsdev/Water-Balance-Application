"""Seed or update TimeSeries sheets in the configured template Excel.

- Reads path from config: data_sources.template_excel_path
- Creates sheets if missing: Environmental, Consumption, Discharge, Storage_Facilities
- Does not remove or modify existing sheets or data beyond adding headers when empty
- Safe to run multiple times (idempotent)

Usage:
  .venv\Scripts\python scripts/seed_timeseries_sheets.py

"""
from __future__ import annotations

from pathlib import Path
from typing import List, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Import shim
import sys
from pathlib import Path as _P
sys.path.insert(0, str((_P(__file__).parent.parent / 'src').resolve()))

from utils.config_manager import config


SHEET_COLUMNS: Dict[str, List[str]] = {
    "Environmental": [
        "Date",
        "Rainfall_mm",
        "Custom_Evaporation_mm",
        "Pan_Coefficient",
    ],
    "Consumption": [
        "Date",
        "Dust_Suppression_m3",
        "Mining_m3",
        "Domestic_m3",
        "Irrigation_m3",
        "Other_m3",
    ],
    "Discharge": [
        "Date",
        "Facility_Code",
        "Discharge_Volume_m3",
        "Discharge_Type",
        "Reason",
        "Approval_Reference",
    ],
    "Storage_Facilities": [
        "Date",
        "Facility_Code",
        "Inflow_m3",
        "Outflow_m3",
        # Optional: Abstraction_m3 is supported by reader if present
        # "Abstraction_m3",
    ],
    "Production": [
        "Date",
        "Concentrate_Produced_t",
        "Concentrate_Moisture_Percent",
        "Slurry_Density_t_per_m3",
        "Tailings_Moisture_Percent",
    ],
    "Seepage_Losses": [
        "Date",
        "Seepage_Loss_m3",
        "Seepage_Gain_m3",
        "Unaccounted_Losses_m3",
    ],
}


def ensure_sheet_headers(wb: Workbook, sheet_name: str, columns: List[str]) -> None:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row >= 1:
            # If headers missing in first row, set them
            if all(ws.cell(row=1, column=i + 1).value is None for i in range(len(columns))):
                for idx, col in enumerate(columns, start=1):
                    c = ws.cell(row=1, column=idx, value=col)
                    c.font = Font(bold=True)
            return
    else:
        ws = wb.create_sheet(sheet_name)
        for idx, col in enumerate(columns, start=1):
            c = ws.cell(row=1, column=idx, value=col)
            c.font = Font(bold=True)
        # Set basic widths
        for i in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = max(16, min(len(columns[i-1]) + 2, 38))


def main() -> int:
    # Resolve template path from config
    template_path = config.get('data_sources.template_excel_path', 'templates/Water_Balance_TimeSeries_Template.xlsx')
    path = Path(template_path)
    if not path.is_absolute():
        path = Path(__file__).resolve().parents[1] / template_path

    path.parent.mkdir(parents=True, exist_ok=True)

    # Load or create workbook
    if path.exists():
        wb = load_workbook(path)
        created = False
    else:
        wb = Workbook()
        # Remove default sheet for cleanliness
        try:
            wb.remove(wb.active)
        except Exception:
            pass
        created = True

    # Ensure each expected sheet exists with headers
    for sheet_name, cols in SHEET_COLUMNS.items():
        ensure_sheet_headers(wb, sheet_name, cols)

    wb.save(path)
    print(f"✅ {'Created' if created else 'Updated'}: {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
