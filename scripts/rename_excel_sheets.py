#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Rename Excel sheets to match area names.
Updates sheet names in the Water Balance TimeSeries Template.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# Sheet name mapping (old -> new)
SHEET_RENAMES = {
    'Flows_UG2N': 'Flows_UG2 North',
    'Flows_MERN': 'Flows_Merensky North',
    'Flows_MERS': 'Flows_Merensky South',
    'Flows_UG2S': 'Flows_UG2 Main',
    'Flows_STOCKPILE': 'Flows_Stockpile',
    'Flows_OLDTSF': 'Flows_Old TSF',
    'Flows_NEWTSF': 'Flows_New TSF',
    'Flows_UG2P': 'Flows_UG2 Plant',
    'Flows_MERP': 'Flows_Merensky Plant',
}

def rename_sheets(excel_path: Path, dry_run: bool = False):
    """Rename sheets in the Excel workbook."""
    
    if not excel_path.exists():
        print(f"❌ File not found: {excel_path}")
        return False
    
    print(f"📂 Opening: {excel_path}")
    wb = load_workbook(excel_path)
    
    print(f"\n📋 Current sheets ({len(wb.sheetnames)}):")
    for sheet_name in wb.sheetnames:
        print(f"   - {sheet_name}")
    
    print(f"\n{'🔍 DRY RUN - ' if dry_run else ''}Renaming sheets...")
    renamed = 0
    skipped = 0
    
    for old_name, new_name in SHEET_RENAMES.items():
        if old_name in wb.sheetnames:
            if dry_run:
                print(f"   Would rename: '{old_name}' → '{new_name}'")
            else:
                ws = wb[old_name]
                ws.title = new_name
                print(f"   ✓ Renamed: '{old_name}' → '{new_name}'")
            renamed += 1
        else:
            print(f"   ⊘ Not found: '{old_name}'")
            skipped += 1
    
    print(f"\n📊 Summary:")
    print(f"   Renamed: {renamed}")
    print(f"   Skipped: {skipped}")
    
    if not dry_run and renamed > 0:
        print(f"\n💾 Saving changes to: {excel_path}")
        wb.save(excel_path)
        print("✅ Done!")
        
        # Verify
        wb = load_workbook(excel_path)
        print(f"\n📋 Updated sheets ({len(wb.sheetnames)}):")
        for sheet_name in sorted(wb.sheetnames):
            print(f"   - {sheet_name}")
    else:
        print("\n⚠️  No changes saved (dry run mode)")
    
    return True


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Rename Excel sheets to match area names')
    parser.add_argument('--file', type=str, 
                       default='test_templates/Water_Balance_TimeSeries_Template.xlsx',
                       help='Path to Excel file (default: test_templates/Water_Balance_TimeSeries_Template.xlsx)')
    parser.add_argument('--dry-run', action='store_true',
                       help='Preview changes without modifying the file')
    
    args = parser.parse_args()
    
    excel_path = Path(args.file)
    rename_sheets(excel_path, dry_run=args.dry_run)
