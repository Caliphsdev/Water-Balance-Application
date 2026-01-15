"""Generate a fresh TimeSeries Excel template with flow sheets and Reference Guide.

- Reads diagram JSONs under data/diagrams to build the Reference Guide (Node ID, Label, Area).
- Builds flow sheets with predefined column headers matching the current diagrams.
- No hard-coded output path: pass --out <path>. Does not touch settings/config.
- Light formatting using openpyxl (title row fill, bold headers, column widths).
"""

import argparse
from pathlib import Path
from typing import Dict, List, Tuple

import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Column definitions per flow sheet (kept local; no path dependencies)
FLOW_SHEETS: Dict[str, List[str]] = {
    "Flows_New TSF": [
        "Date",
        "Year",
        "Month",
        "JUNCTION_132_288_2496 → OLDTSF_NEW_TSF",
        "OLDTSF_NEW_TSF → OLDTSF_NEW_TSF",
        "OLDTSF_NEW_TSF → OLDTSF_NEW_TSF_EVAP",
        "OLDTSF_NEW_TSF → OLDTSF_NEW_TSF_INTERSTITIAL",
        "OLDTSF_NEW_TSF → OLDTSF_NEW_TSF_SEEPAGE",
        "OLDTSF_NEW_TSF → OLDTSF_NT_RWD",
        "OLDTSF_NEW_TSF_RAINRUN → OLDTSF_NEW_TSF",
        "OLDTSF_NT_RWD → OLDTSF_NEW_TSF",
    ],
    "Flows_Old TSF": [
        "Date",
        "Year",
        "Month",
        "OLDTSF_GWA_BOREHOLES → OLDTSF_OFFICES",
        "OLDTSF_NT_RWD → OLDTSF_NT_RWD",
        "OLDTSF_NT_RWD → OLDTSF_NT_RWD_EVAP",
        "OLDTSF_NT_RWD_RAIN → OLDTSF_NT_RWD",
        "OLDTSF_OFFICES → OLDTSF_OFF_CONSUMPTION",
        "OLDTSF_OFFICES → OLDTSF_OFF_SEPTIC",
        "OLDTSF_OLD_TSF → OLDTSF_OLD_TSF",
        "OLDTSF_OLD_TSF → OLDTSF_OLD_TSF_EVAP",
        "OLDTSF_OLD_TSF → OLDTSF_OLD_TSF_INTERSTITIAL",
        "OLDTSF_OLD_TSF → OLDTSF_OLD_TSF_SEEPAGE",
        "OLDTSF_OLD_TSF → OLDTSF_TRTD",
        "OLDTSF_OLD_TSF_RAINRUN → OLDTSF_OLD_TSF",
        "OLDTSF_TRTD → OLDTSF_OLD_TSF",
        "OLDTSF_TRTD → OLDTSF_TRTD",
        "OLDTSF_TRTD → OLDTSF_TRTD_EVAP",
        "OLDTSF_TRTD_RAIN → OLDTSF_TRTD",
        "OLDTSF_NT_RWD → OLDTSF_NT_RWD_SPILL",
        "OLDTSF_TRTD → OLDTSF_TRTD_SPILL",
    ],
    "Flows_UG2 Plant": [
        "Date",
        "Year",
        "Month",
        "JUNCTION_134_1193_2880 → UG2PLANT_UG2P_PLANT",
        "UG2PLANT_CPGWA_BOREHOLES → UG2PLANT_UG2P_SOFT",
        "UG2PLANT_CPRWSD1 → UG2PLANT_UG2P_PLANT",
        "UG2PLANT_UG2P_OFF → UG2PLANT_UG2P_OFF_CONSUMPTION",
        "UG2PLANT_UG2P_OFF → UG2PLANT_UG2P_STP",
        "UG2PLANT_UG2P_ORE_WATER → UG2PLANT_UG2P_PLANT",
        "UG2PLANT_UG2P_PLANT → JUNCTION_133_288_2922",
        "UG2PLANT_UG2P_PLANT → JUNCTION_135_1229_2857",
        "UG2PLANT_UG2P_PLANT → UG2PLANT_UG2P_PLANT_LOSSES",
        "UG2PLANT_UG2P_PLANT → UG2PLANT_UG2P_PLANT_PRODUCT",
        "UG2PLANT_UG2P_PLANT → UG2PLANT_UG2PCD1",
        "UG2PCD1_to_Plant",
        "UG2PLANT_UG2P_RES → UG2PLANT_UG2P_OFF",
        "UG2PLANT_UG2P_RIVERS → UG2PLANT_CPRWSD1",
        "UG2PLANT_UG2P_SOFT → UG2PLANT_UG2P_RES",
        "UG2PLANT_UG2P_SOFT → UG2PLANT_UG2P_STP_LOSSES",
        "UG2PLANT_UG2P_STP → UG2PLANT_UG2P_PLANT",
        "UG2PLANT_UG2P_STP → UG2PLANT_UG2P_SOFT_LOSSES",
        "UG2PLANT_UG2PCD1 → UG2PLANT_UG2PCD1",
        "UG2PLANT_UG2PCD1 → UG2PLANT_UG2PCD1_EVAP",
        "UG2PLANT_UG2PCD1 → UG2PLANT_UG2PCD1_SPILL",
        "UG2PLANT_UG2PCD1_RAIN → UG2PLANT_UG2PCD1",
    ],
    "Flows_UG2 Main": [
        "Date",
        "Year",
        "Month",
        "UG2S_BOREHOLE → UG2S_SOFT",
        "UG2S_MDCDG → JUNCTION_130_1229_1793",
        "UG2S_MDCDG → UG2S_MDCD_DUST",
        "UG2S_MDCDG → UG2S_MDCD_EVAP",
        "UG2S_MDCDG → UG2S_MDCD_SPILL",
        "UG2S_MDCDG → UG2S_MDCDG",
        "UG2S_MDCDG → UG2S_SD",
        "UG2S_OFFICES → UG2S_CONSUMPTION",
        "UG2S_OFFICES → UG2S_STP",
        "UG2S_RAINFALL → UG2S_MDCDG",
        "UG2S_RES → UG2S_OFFICES",
        "UG2S_SD → UG2S_MDCDG",
        "UG2S_SDSA → UG2S_MDCDG",
        "UG2S_SOFT → UG2S_RES",
        "UG2S_SOFT → UG2S_STP_LOSSES",
        "UG2S_STP → UG2S_MDCDG",
        "UG2S_STP → UG2S_SOFT_LOSSES",
    ],
    "Flows_UG2 North": [
        "Date",
        "Year",
        "Month",
        "BH_NDGWA → SOFTENING",
        "NDCD → EVAPORATION",
        "NDCD → NDCD",
        "NDCD → NORTH_DECLINE",
        "NDCD → SPILL",
        "NORTH_DECLINE → NDCD",
        "NORTH_SHAFT → NDCD",
        "OFFICES → CONSUMPTION2",
        "OFFICES → SEWAGE",
        "RAINFALL → NDCD",
        "RESERVOIR → OFFICES",
        "SEWAGE → JUNCTION_127_1208_365",
        "SEWAGE → LOSSES2",
        "SOFTENING → LOSSES",
        "SOFTENING → RESERVOIR",
        "RAINFALL_INFLOW → NDCD",
        "NDCD → DUST_SUPPRESSION",
        "SOFTENING → TRP_CLINIC",
        "TRP_CLINIC → SEPTIC",
        "TRP_CLINIC → CONSUMPTION",
    ],
    "Flows_Merensky Plant": [
        "Date",
        "Year",
        "Month",
        "MERPLANT_MERP_OFF → MERPLANT_MERP_OFF_CONSUMPTION",
        "MERPLANT_MERP_OFF → MERPLANT_MERP_STP",
        "MERPLANT_MERP_ORE_WATER → MERPLANT_MERP_PLANT",
        "MERPLANT_MERP_PLANT → MERPLANT_MERP_PLANT_LOSSES",
        "MERPLANT_MERP_PLANT → MERPLANT_MERP_PLANT_PRODUCT",
        "MERPLANT_MERP_PLANT → MERPLANT_MPSWD12",
        "MERPLANT_MERP_PLANT → OLDTSF_OLD_TSF",
        "MERPLANT_MERP_RES → MERPLANT_MERP_OFF",
        "MERPLANT_MERP_SOFT → MERPLANT_MERP_OFF_LOSSES",
        "MERPLANT_MERP_SOFT → MERPLANT_MERP_RES",
        "MERPLANT_MERP_STP → MERPLANT_MERP_STP_LOSSES",
        "MERPLANT_MERP_STP → MERPLANT_MPSWD12",
        "MERPLANT_MPGWA_BOREHOLES → MERPLANT_MERP_SOFT",
        "MERPLANT_MPRWSD1 → MERPLANT_MERP_PLANT",
        "MERPLANT_MPRWSD1 → OLDTSF_TRTD",
        "MERPLANT_MPRWSD1_RIVERS → MERPLANT_MPRWSD1",
        "MERPLANT_MPSWD12 → MERPLANT_MERP_PLANT",
        "MERPLANT_MPSWD12 → MERPLANT_MPSWD12",
        "MERPLANT_MPSWD12 → MERPLANT_MPSWD12_DUST",
        "MERPLANT_MPSWD12 → MERPLANT_MPSWD12_EVAP",
        "MERPLANT_MPSWD12 → MERPLANT_MPSWD12_SPILL",
        "MERPLANT_MPSWD12_RAIN → MERPLANT_MPSWD12",
        "NDCD → MERPLANT_MPRWSD1",
    ],
    "Flows_Merensky South": [
        "Date",
        "Year",
        "Month",
        "MERS_BOREHOLE → MERS_SOFT",
        "MERS_MDCDG → JUNCTION_131_1229_2072",
        "MERS_MDCDG → MERS_MDCD_DUST",
        "MERS_MDCDG → MERS_MDCD_EVAP",
        "MERS_MDCDG → MERS_SD",
        "MERS_MDCDG → UG2S_MDCDG",
        "MERS_OFFICES → MERS_CONSUMPTION",
        "MERS_OFFICES → UG2S_STP",
        "MERS_RAINFALL → MERS_MDCDG",
        "MERS_SD → MERS_MDCDG",
        "MERS_SOFT → MERS_OFFICES",
        "MERS_SOFT → MERS_SOFT_LOSSES",
        "MERS_MDCDG → MERS_MDCD_SPILL",
        "MERS_SDSA → UG2S_MDCDG",
        "MERS_MDCDG → MERS_MDCDG",
    ],
    "Flows_Stockpile": [
        "Date",
        "Year",
        "Month",
        "SPCD1 → SPCD1",
        "SPCD1 → STOCKPILE_DUST_SUPPRESSION",
        "SPCD1 → STOCKPILE_EVAPORATION2",
        "STOCKPILE_AREA → SPCD1",
        "STOCKPILE_AREA → STOCKPILE_AREA",
        "STOCKPILE_AREA → STOCKPILE_EVAPORATION",
        "STOCKPILE_BH_GWA → STOCKPILE_OFFICES",
        "STOCKPILE_OFFICES → STOCKPILE_CONSUMPTION",
        "STOCKPILE_OFFICES → STOCKPILE_SEPTIC",
        "STOCKPILE_RAINFALL → STOCKPILE_AREA",
        "STOCKPILE_RAINFALL_SPCD1 → SPCD1",
        "SPCD1 → STOCKPILE_SPILL",
        "SPCD1 → JUNCTION_129_1140_1242",
    ],
}


TITLE_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def extract_reference_guide() -> List[Tuple[str, str, str]]:
    """Extract (node_id, label, area) from diagram JSONs."""
    diagrams_dir = Path("data/diagrams")
    rows: List[Tuple[str, str, str]] = []
    for p in sorted(diagrams_dir.glob("*.json")):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        area = data.get("area_code")
        if not area or area == "UNKNOWN":
            continue
        for node in data.get("nodes", []):
            node_id = node.get("id")
            label = node.get("label")
            if node_id and label:
                rows.append((node_id, label, area))
    # Deduplicate while preserving order
    seen = set()
    deduped = []
    for r in rows:
        if r not in seen:
            deduped.append(r)
            seen.add(r)
    return deduped


def add_reference_sheet(wb: Workbook, rows: List[Tuple[str, str, str]]):
    ws = wb.create_sheet("Reference Guide")
    ws.append(["Node ID", "Label", "Area"])
    for node_id, label, area in rows:
        ws.append([node_id, label, area])
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = max(12, min(max_len + 2, 40))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def add_flow_sheet(wb: Workbook, name: str, columns: List[str]):
    ws = wb.create_sheet(name)
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=f"{name} Water Balance Template")
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Header row (row 3 for compatibility with existing loaders)
    header_row = 3
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=col)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        # Set column width based on text length
        ws.column_dimensions[cell.column_letter].width = max(12, min(len(col) + 2, 45))

    # Placeholder data row (row 4) with empty cells
    ws.cell(row=4, column=1, value="")


def build_workbook() -> Workbook:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Reference Guide
    ref_rows = extract_reference_guide()
    add_reference_sheet(wb, ref_rows)

    # Flow sheets
    for sheet_name, cols in FLOW_SHEETS.items():
        add_flow_sheet(wb, sheet_name, cols)

    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate TimeSeries Excel template")
    parser.add_argument("--out", required=True, help="Output .xlsx path to create")
    args = parser.parse_args()

    out_path = Path(args.out).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(out_path)
    print(f"✅ Template created: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
