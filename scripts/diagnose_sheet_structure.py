"""
Diagnose Excel sheet structure for failing sheets.

Usage:
  .venv\Scripts\python scripts\diagnose_sheet_structure.py
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

import pandas as pd
from utils.config_manager import config
from openpyxl import load_workbook

# Get Excel path from config
config.load_config()
excel_path = config.get(
    'data_sources.timeseries_excel_path',
    config.get('data_sources.template_excel_path',
               'test_templates/Water_Balance_TimeSeries_Template.xlsx')
)
base_dir = Path(__file__).parent.parent
if not Path(excel_path).is_absolute():
    excel_path = base_dir / excel_path

print(f"Excel file: {excel_path}")
print(f"Exists: {excel_path.exists()}")
print()

# Sheets that are failing
failing_sheets = ['Flows_MERP', 'Flows_OLDTSF', 'Flows_UG2P', 'Flows_NEWTSF', 'Flows_MERS']

for sheet_name in failing_sheets:
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    # Read with openpyxl to see raw structure
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
            print("\nFirst 6 rows (raw):")
            for r in range(1, min(7, ws.max_row + 1)):
                row_vals = [cell.value for cell in ws[r]]
                print(f"  Row {r}: {row_vals[:10]}")  # First 10 cells
        wb.close()
    except Exception as e:
        print(f"  openpyxl error: {e}")
    
    # Try pandas with different headers
    print("\nPandas reading attempts:")
    for header in [0, 1, 2, 3, [0,1], [0,1,2]]:
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=header, engine='openpyxl')
            cols = list(df.columns)
            print(f"  header={header}: {len(cols)} columns, first 5: {cols[:5]}")
            print(f"    First data row: {df.iloc[0].to_dict() if not df.empty else 'EMPTY'}")
        except Exception as e:
            print(f"  header={header}: ERROR - {e}")
