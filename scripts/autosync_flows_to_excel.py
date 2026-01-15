"""
Auto-sync Flow Lines from Diagram to Excel with Smart Column Naming
This creates:
1. Clear, descriptive column names from flow line labels
2. Columns added to appropriate area sheets
3. Mapping reference sheet for documentation
4. Makes monthly data entry fast and intuitive
"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from datetime import datetime
import shutil

# Paths
diagram_path = Path('data/diagrams/ug2_north_decline.json')
excel_path = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("=" * 70)
print("AUTO-SYNC: DIAGRAM FLOWS → EXCEL COLUMNS WITH SMART NAMING")
print("=" * 70)

# Load diagram
with open(diagram_path, 'r') as f:
    diagram = json.load(f)

print(f"\n📊 Reading diagram: {diagram_path.name}")
edges = diagram.get('edges', [])
print(f"   Found {len(edges)} flow lines")

# Create backup
backup_path = excel_path.with_suffix(f'.backup_pre_autosync_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy(excel_path, backup_path)
print(f"\n✅ Backup created: {backup_path.name}")

# Load Excel
wb = load_workbook(excel_path)
print(f"\n📄 Reading Excel: {excel_path.name}")

# Map zones to sheets
zone_to_sheet = {
    'UG2N': 'Flows_UG2N',
    'STOCKPILE': 'Flows_STOCKPILE',
    'UG2S': 'Flows_UG2S',
    'MERS': 'Flows_MERS',
    'OLDTSF': 'Flows_OLDTSF',
    'NEWTSF': 'Flows_NEWTSF',
    'UG2P': 'Flows_UG2P',
    'MERP': 'Flows_MERP',
}

# Extract and organize flows
flows_by_sheet = {}
mapping_data = []

print(f"\n🔍 Processing {len(edges)} flow lines...")

for idx, edge in enumerate(edges, 1):
    from_node = edge.get('from', '')
    to_node = edge.get('to', '')
    label = edge.get('label', '')
    
    # Check if edge has excel_mapping configured
    excel_mapping = edge.get('excel_mapping', {})
    
    if not excel_mapping.get('enabled'):
        continue
    
    sheet_name = excel_mapping.get('sheet', '')
    col_name = excel_mapping.get('column', '')
    
    if not sheet_name or not col_name:
        continue
    
    # Clean up column name for clarity
    col_name = col_name.strip().upper()
    # Replace underscores between component names with arrow for clarity
    col_name = col_name.replace('__TO__', ' → ')
    
    if sheet_name not in flows_by_sheet:
        flows_by_sheet[sheet_name] = []
    
    flows_by_sheet[sheet_name].append(col_name)
    
    # Store mapping info
    mapping_data.append({
        'column': col_name,
        'sheet': sheet_name,
        'from': from_node,
        'to': to_node,
        'label': label,
        'edge_id': edge.get('id', f'edge_{idx}'),
    })

print(f"\n✅ Extracted flows from {len(flows_by_sheet)} sheets")

# Add columns to Excel sheets
print(f"\n🔧 Adding columns to Excel sheets...\n")

standard_cols = ['Date', 'Year', 'Month']

for sheet_name, columns in sorted(flows_by_sheet.items()):
    if sheet_name not in wb.sheetnames:
        print(f"   ⚠️  {sheet_name}: NOT FOUND")
        continue
    
    ws = wb[sheet_name]
    
    # Get existing columns (skip standard ones)
    existing_cols = []
    for cell in ws[1]:
        if cell.value and cell.value not in standard_cols:
            existing_cols.append(cell.value)
    
    # Find next available column
    next_col_idx = len(standard_cols) + 1
    
    # Add new columns
    added = 0
    for col_name in columns:
        if col_name not in existing_cols:
            ws.cell(row=1, column=next_col_idx).value = col_name
            
            # Format header
            cell = ws.cell(row=1, column=next_col_idx)
            cell.font = Font(bold=True, size=11, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set column width
            ws.column_dimensions[cell.column_letter].width = max(15, len(col_name))
            
            next_col_idx += 1
            added += 1
    
    total_cols = next_col_idx - 1
    print(f"   ✅ {sheet_name}")
    print(f"      Standard: {', '.join(standard_cols)}")
    print(f"      Flow columns: {added} added, {total_cols - len(standard_cols)} total")

# Create Mapping Reference sheet
print(f"\n📋 Creating Mapping Reference sheet...")

if 'Column Mapping Reference' in wb.sheetnames:
    ref_ws = wb['Column Mapping Reference']
    ref_ws.delete_rows(1, ref_ws.max_row)
else:
    ref_ws = wb.create_sheet('Column Mapping Reference', 1)

# Headers
headers = ['Excel Column', 'Sheet', 'From Component', 'To Component', 'Original Label', 'Description']
for col_idx, header in enumerate(headers, 1):
    cell = ref_ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill(start_color='203864', end_color='203864', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Add mapping data
row = 2
for mapping in sorted(mapping_data, key=lambda x: (x['sheet'], x['column'])):
    ref_ws.cell(row=row, column=1).value = mapping['column']
    ref_ws.cell(row=row, column=2).value = mapping['sheet']
    ref_ws.cell(row=row, column=3).value = mapping['from']
    ref_ws.cell(row=row, column=4).value = mapping['to']
    ref_ws.cell(row=row, column=5).value = mapping['label']
    
    # Description: explain what this column represents
    desc = f"Flow from {mapping['from']} to {mapping['to']}"
    ref_ws.cell(row=row, column=6).value = desc
    
    # Alternate row colors for readability
    if row % 2 == 0:
        fill = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
        for col in range(1, 7):
            ref_ws.cell(row=row, column=col).fill = fill
    
    row += 1

# Set column widths for Mapping Reference
ref_ws.column_dimensions['A'].width = 30
ref_ws.column_dimensions['B'].width = 20
ref_ws.column_dimensions['C'].width = 25
ref_ws.column_dimensions['D'].width = 25
ref_ws.column_dimensions['E'].width = 30
ref_ws.column_dimensions['F'].width = 40

print(f"   ✅ Added {len(mapping_data)} flow mappings")

# Add instructions sheet
print(f"\n📖 Adding Monthly Data Entry Instructions...")

if 'Instructions' in wb.sheetnames:
    inst_ws = wb['Instructions']
    inst_ws.delete_rows(1, inst_ws.max_row)
else:
    inst_ws = wb.create_sheet('Instructions', 0)

instructions = [
    ('MONTHLY DATA ENTRY GUIDE', 14, 'FFFFFF', '0070C0'),
    ('', 11, '000000', 'FFFFFF'),
    ('How to Update Monthly Flow Values:', 12, '000000', 'FFFFFF'),
    ('', 11, '000000', 'FFFFFF'),
    ('1. Open the Excel file', 11, '000000', 'FFFFFF'),
    ('2. Navigate to the area sheet (e.g., Flows_UG2 North, Flows_Stockpile, etc.)', 11, '000000', 'FFFFFF'),
    ('3. Find the month row based on Date column (day) and Month column', 11, '000000', 'FFFFFF'),
    ('4. Enter the flow volume (in m³) in the corresponding flow column', 11, '000000', 'FFFFFF'),
    ('5. Save the file', 11, '000000', 'FFFFFF'),
    ('', 11, '000000', 'FFFFFF'),
    ('Column Naming Convention:', 12, '000000', 'FFFFFF'),
    ('', 11, '000000', 'FFFFFF'),
    ('Column names follow the pattern: SOURCE_TO_DESTINATION or SOURCE_PROCESS', 11, '000000', 'FFFFFF'),
    ('Examples:', 11, '000000', 'FFFFFF'),
    ('  • UG2N_TO_STOCKPILE = Flow from UG2 North to Stockpile', 11, '000000', 'FFFFFF'),
    ('  • STOCKPILE_TO_OLDTSF = Flow from Stockpile to Old TSF', 11, '000000', 'FFFFFF'),
    ('  • PUMP_RECOVERY = Water recovery/recirculation flow', 11, '000000', 'FFFFFF'),
    ('', 11, '000000', 'FFFFFF'),
    ('Understanding the Sheets:', 12, '000000', 'FFFFFF'),
    ('', 11, '000000', 'FFFFFF'),
    ('• Flows_UG2N = UG2 North Decline mining area', 11, '000000', 'FFFFFF'),
    ('• Flows_Stockpile = Ore stockpile facility', 11, '000000', 'FFFFFF'),
    ('• Flows_UG2S = UG2 South Decline mining area', 11, '000000', 'FFFFFF'),
    ('• Flows_MERS = Merensky South mining area', 11, '000000', 'FFFFFF'),
    ('• Flows_OLDTSF = Old Tailings Storage Facility', 11, '000000', 'FFFFFF'),
    ('• Flows_New TSF = New Tailings Storage Facility (split from Old TSF)', 11, '000000', 'FFFFFF'),
    ('• Flows_UG2P = UG2 Plant processing facility', 11, '000000', 'FFFFFF'),
    ('• Flows_MERP = Merensky Plant processing facility', 11, '000000', 'FFFFFF'),
    ('', 11, '000000', 'FFFFFF'),
    ('Reference Guide:', 12, '000000', 'FFFFFF'),
    ('', 11, '000000', 'FFFFFF'),
    ('For detailed descriptions of all columns and their mappings, see the "Column Mapping Reference" sheet.', 11, '000000', 'FFFFFF'),
    ('This sheet lists every flow column, what it represents, and which components it connects.', 11, '000000', 'FFFFFF'),
]

row = 1
for text, size, font_color, bg_color in instructions:
    cell = inst_ws.cell(row=row, column=1)
    cell.value = text
    
    if size == 14:  # Title
        cell.font = Font(bold=True, size=size, color=font_color)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        inst_ws.merge_cells(f'A{row}:D{row}')
        inst_ws.row_dimensions[row].height = 25
    elif size == 12:  # Section header
        cell.font = Font(bold=True, size=size)
        cell.fill = PatternFill(start_color='D9E8F5', end_color='D9E8F5', fill_type='solid')
    else:
        cell.font = Font(size=size)
    
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    row += 1

inst_ws.column_dimensions['A'].width = 80
inst_ws.row_dimensions[1].height = 25

print(f"   ✅ Added instructions for monthly data entry")

# Save
wb.save(excel_path)
print(f"\n✅ Excel file updated successfully!")

print(f"\n" + "=" * 70)
print("📊 SUMMARY")
print("=" * 70)
print(f"\n✅ Columns Added:")
total_flows = sum(len(cols) for cols in flows_by_sheet.values())
print(f"   Total flow columns added: {total_flows}")
for sheet_name in sorted(flows_by_sheet.keys()):
    print(f"   • {sheet_name}: {len(flows_by_sheet[sheet_name])} columns")

print(f"\n✅ New Reference Sheets Created:")
print(f"   • Column Mapping Reference: {len(mapping_data)} flow definitions")
print(f"   • Instructions: Monthly data entry guide")

print(f"\n✅ Files:")
print(f"   • Excel: {excel_path.name}")
print(f"   • Backup: {backup_path.name}")

print(f"\n💡 Next Steps:")
print(f"   1. Open Excel Manager in Flow Diagram Dashboard")
print(f"   2. Go to Columns tab and select a sheet")
print(f"   3. See all flow columns with clear names")
print(f"   4. Each month, just update the values in these columns")
print(f"   5. Refer to 'Column Mapping Reference' sheet for definitions")

print(f"\n🎯 Monthly Workflow:")
print(f"   1. Open test_templates/Water_Balance_TimeSeries_Template.xlsx")
print(f"   2. Select the area sheet (Flows_UG2 North, Flows_Stockpile, etc.)")
print(f"   3. Find the row for your date/month")
print(f"   4. Enter values in each flow column (e.g., UG2N_TO_STOCKPILE)")
print(f"   5. Save and run dashboard calculations")

wb.close()
