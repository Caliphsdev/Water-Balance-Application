#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Update Excel Column Names to Friendly Format

Converts technical column names in Excel flow sheets to human-friendly labels.
Creates a backup before modifying the file.
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from src.utils.friendly_names import convert_column_to_friendly
from datetime import datetime


def update_excel_column_names(excel_path: Path, dry_run: bool = False):
    """
    Update column names in Excel flow sheets to friendly format.
    
    Args:
        excel_path: Path to Excel file
        dry_run: If True, only preview changes without saving
    """
    
    if not excel_path.exists():
        print(f"❌ File not found: {excel_path}")
        return False
    
    # Create backup
    if not dry_run:
        backup_path = excel_path.with_suffix(f'.backup-{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx')
        import shutil
        shutil.copy2(excel_path, backup_path)
        print(f"💾 Backup created: {backup_path}")
    
    print(f"\n📂 Opening: {excel_path}")
    wb = load_workbook(excel_path)
    
    # Find flow sheets
    flow_sheets = [s for s in wb.sheetnames if s.startswith('Flows_')]
    print(f"\n📋 Found {len(flow_sheets)} flow sheets")
    
    total_changes = 0
    
    for sheet_name in flow_sheets:
        print(f"\n{'🔍 ' if dry_run else '📝 '}Processing: {sheet_name}")
        ws = wb[sheet_name]
        
        # Assume row 3 contains column headers (after Date, Year, Month)
        header_row = 3
        changes_in_sheet = 0
        
        # Get all columns
        max_col = ws.max_column
        
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            original_value = cell.value
            
            if not original_value or original_value in ['Date', 'Year', 'Month']:
                continue
            
            # Convert to friendly name
            friendly_value = convert_column_to_friendly(str(original_value))
            
            if friendly_value != original_value:
                if dry_run:
                    print(f"   Would change col {col_idx}:")
                    print(f"     FROM: {original_value}")
                    print(f"     TO:   {friendly_value}")
                else:
                    cell.value = friendly_value
                    # Apply formatting (bold, centered)
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    print(f"   ✓ Col {col_idx}: {original_value[:40]}... → {friendly_value[:40]}...")
                
                changes_in_sheet += 1
        
        print(f"   Changes: {changes_in_sheet}")
        total_changes += changes_in_sheet
    
    print(f"\n📊 Total changes: {total_changes}")
    
    if not dry_run and total_changes > 0:
        print(f"\n💾 Saving changes...")
        wb.save(excel_path)
        print("✅ Excel file updated successfully!")
    elif dry_run:
        print("\n⚠️  No changes saved (dry run mode)")
    else:
        print("\n✓ No changes needed")
    
    return True


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Update Excel column names to friendly format'
    )
    parser.add_argument(
        '--file',
        type=str,
        default='test_templates/Water_Balance_TimeSeries_Template.xlsx',
        help='Path to Excel file'
    )
    parser.add_argument(
        '--dry-run',
        action='store_true',
        help='Preview changes without modifying the file'
    )
    
    args = parser.parse_args()
    
    excel_path = Path(args.file)
    update_excel_column_names(excel_path, dry_run=args.dry_run)
