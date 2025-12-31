"""
Fix categorization: Assign each flow to exactly ONE area based on source node primary area.
Regenerate Excel with correct categorization.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

print("🔧 FIXING CATEGORIZATION & EXCEL")
print("=" * 100)

# Load JSON
with open('data/diagrams/ug2_north_decline.json') as f:
    diagram = json.load(f)

edges = diagram.get('edges', [])
print(f"✅ Loaded {len(edges)} edges from JSON")

# Define area categorization by SOURCE node prefix
# Each node belongs to ONE area, so we categorize flows by their source area
area_definitions = {
    'UG2N': ['rainfall', 'ndcd', 'bh_ndgwa', 'softening', 'reservoir', 'offices', 
             'guest_house', 'septic', 'sewage', 'north_decline', 'north_shaft', 
             'losses', 'losses2', 'consumption', 'consumption2', 'evaporation',
             'dust_suppression', 'spill', 'junction_127', 'junction_128'],
    'UG2S': ['ug2s_'],
    'UG2P': ['ug2plant_', 'cprwsd'],
    'MERN': ['rainfall_merensky', 'ndcd_merensky', 'bh_mcgwa', 'softening_merensky', 
             'offices_merensky', 'merensky_north', 'evaporation_merensky', 'dust_suppression_merensky',
             'spill_merensky', 'consumption_merensky', 'losses_merensky', 'junction_128'],
    'MERP': ['merplant_', 'mprwsd'],
    'MERS': ['mers_'],
    'OLDTSF': ['oldtsf_', 'nt_rwd', 'new_tsf', 'trtd'],
    'STOCKPILE': ['stockpile_', 'spcd1'],
}

def find_area_for_node(node_id):
    """Find which area a node belongs to. Check in order of specificity."""
    if not node_id:
        return None
    
    node_lower = node_id.lower()
    
    # Check each area's patterns
    for area, patterns in area_definitions.items():
        for pattern in patterns:
            if pattern in node_lower:
                return area
    
    return None

# Categorize flows by SOURCE node
flows_by_area = {area: [] for area in area_definitions.keys()}
flows_by_area['OTHER'] = []
uncategorized_flows = []

for edge in edges:
    from_id = edge.get('from')
    to_id = edge.get('to')
    
    # Categorize by SOURCE node (from_id)
    source_area = find_area_for_node(from_id)
    
    flow_key = f'{from_id}__TO__{to_id}'
    
    if source_area:
        flows_by_area[source_area].append(flow_key)
    else:
        # Try destination node as fallback
        dest_area = find_area_for_node(to_id)
        if dest_area:
            flows_by_area[dest_area].append(flow_key)
        else:
            flows_by_area['OTHER'].append(flow_key)
            uncategorized_flows.append((from_id, to_id))

print("\n📊 FLOW CATEGORIZATION BY AREA")
print("=" * 100)

total = 0
for area in ['UG2N', 'UG2P', 'UG2S', 'MERN', 'MERP', 'MERS', 'OLDTSF', 'STOCKPILE', 'OTHER']:
    count = len(flows_by_area[area])
    total += count
    status = '✅' if count > 0 else '⚠️'
    print(f'{status} {area:15} → {count:3} flows')

print(f"\n{'=' * 100}")
print(f"{'TOTAL':15} → {total:3} flows")

if uncategorized_flows:
    print(f"\n⚠️  UNCATEGORIZED FLOWS ({len(uncategorized_flows)}):")
    for from_id, to_id in uncategorized_flows:
        print(f"   {from_id} → {to_id}")

# Now generate Excel with correct categorization
print(f"\n📝 GENERATING EXCEL WORKBOOK")
print("=" * 100)

wb = Workbook()
wb.remove(wb.active)

# Style templates
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
area_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
area_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Create Reference Guide sheet
ws_ref = wb.create_sheet('Reference Guide', 0)
ws_ref['A1'] = 'Water Balance Template - Node & Flow Reference'
ws_ref['A1'].font = Font(bold=True, size=14)

row = 3
ws_ref[f'A{row}'] = 'Node ID'
ws_ref[f'B{row}'] = 'Label'
ws_ref[f'C{row}'] = 'Area'

for cell in [ws_ref[f'A{row}'], ws_ref[f'B{row}'], ws_ref[f'C{row}']]:
    cell.fill = header_fill
    cell.font = header_font

row += 1
nodes = diagram.get('nodes', [])
node_list = nodes if isinstance(nodes, list) else [{'id': k, 'label': v.get('label', '')} for k, v in nodes.items()]

for node in sorted(node_list, key=lambda x: x.get('id', '')):
    node_id = node.get('id')
    area = find_area_for_node(node_id)
    
    ws_ref[f'A{row}'] = node_id
    ws_ref[f'B{row}'] = node.get('label', '')
    ws_ref[f'C{row}'] = area if area else 'OTHER'
    row += 1

ws_ref.column_dimensions['A'].width = 25
ws_ref.column_dimensions['B'].width = 40
ws_ref.column_dimensions['C'].width = 15

# Create area-specific sheets
sheet_order = ['UG2N', 'UG2P', 'UG2S', 'MERN', 'MERP', 'MERS', 'OLDTSF', 'STOCKPILE', 'OTHER']

for sheet_idx, area in enumerate(sheet_order, start=1):
    flows = flows_by_area[area]
    if not flows:
        continue
    
    ws = wb.create_sheet(f'Flows_{area}', sheet_idx)
    
    # Header
    ws['A1'] = f'{area} Water Balance Template'
    ws['A1'].font = Font(bold=True, size=12)
    
    # Column headers
    ws['A3'] = 'Date'
    ws['B3'] = 'Year'
    ws['C3'] = 'Month'
    
    for cell in [ws['A3'], ws['B3'], ws['C3']]:
        cell.fill = area_fill
        cell.font = area_font
        cell.border = border
    
    # Flow columns
    from openpyxl.utils import get_column_letter
    col = 4  # Start from column D (4)
    for flow in sorted(flows):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}3'] = flow
        ws[f'{col_letter}3'].fill = header_fill
        ws[f'{col_letter}3'].font = header_font
        ws[f'{col_letter}3'].border = border
        ws[f'{col_letter}3'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        
        # Set column width
        ws.column_dimensions[col_letter].width = 25
        
        col += 1
    
    # Set standard column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    
    # Set header row height
    ws.row_dimensions[3].height = 30

# Save
timestamp = int(datetime.now().timestamp() * 1000)
output_file = f'test_templates/Water_Balance_TimeSeries_Template_FIXED_{timestamp}.xlsx'
wb.save(output_file)

print(f"✅ Excel saved: {Path(output_file).name}")
print(f"\n📊 SUMMARY:")
print(f"   Total flows distributed: {total}")
print(f"   Reference Guide: {len(node_list)} nodes")
print(f"   Area sheets: {len(sheet_order)}")
print(f"\n📋 FLOW DISTRIBUTION:")
for area in sheet_order:
    count = len(flows_by_area[area])
    if count > 0:
        print(f"   {area:15} {count:3} flows")
