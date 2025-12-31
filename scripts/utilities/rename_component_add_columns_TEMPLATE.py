#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
TEMPLATE: Add New Component Columns to Excel

Copy this script and customize for adding any new component columns.
Example: rename_component_add_columns.py

Instructions:
1. Copy this file to a new name, e.g., add_new_columns.py
2. Replace OLD_COMPONENT and NEW_COMPONENT with your names
3. Update COLUMNS_TO_ADD with your new column names
4. Run: python add_new_columns.py
5. Then run: python rename_component.py old_name new_name
"""

from openpyxl import load_workbook
from pathlib import Path

# ===== CONFIGURATION =====
OLD_COMPONENT = "guest_house"      # Old component name
NEW_COMPONENT = "trp_clinic"       # New component name
TARGET_SHEET = "Flows_UG2N"        # Which sheet to update

# New column names (in Excel format: UPPERCASE → UPPERCASE)
COLUMNS_TO_ADD = [
    'SOFTENING → TRP_CLINIC',
    'TRP_CLINIC → SEPTIC',
    'TRP_CLINIC → CONSUMPTION',
]

# ===== END CONFIGURATION =====

EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("="*70)
print(f"Adding {NEW_COMPONENT.upper()} columns to Excel")
print("="*70)

wb = load_workbook(EXCEL_PATH)
ws = wb[TARGET_SHEET]

# Get current headers from row 3
current_headers = [c.value for c in ws[3] if c.value]
print(f"\nCurrent {TARGET_SHEET} headers: {len(current_headers)} columns")

# Find last populated column
last_col = 0
for col in range(1, ws.max_column + 1):
    if ws.cell(3, col).value:
        last_col = col

print(f"Last data column: {last_col}")

# Add new columns
print(f"\nAdding {len(COLUMNS_TO_ADD)} new columns:")
for idx, col_name in enumerate(COLUMNS_TO_ADD, 1):
    new_col = last_col + idx
    ws.cell(3, new_col).value = col_name
    print(f"  Column {new_col}: {col_name}")
    
    # Add placeholder '-' to data rows (rows 4-13)
    for row in range(4, 14):
        if not ws.cell(row, new_col).value:
            ws.cell(row, new_col).value = '-'

wb.save(EXCEL_PATH)
print(f"\n✅ Excel updated with {NEW_COMPONENT} columns")

# Verify
wb2 = load_workbook(EXCEL_PATH)
ws2 = wb2[TARGET_SHEET]
headers = [c.value for c in ws2[3] if c.value]
print(f"\nVerification:")
print(f"  Total headers: {len(headers)}")
for h in COLUMNS_TO_ADD:
    if h in headers:
        col_idx = [c.value for c in ws2[3]].index(h) + 1
        print(f"  ✅ Column {col_idx}: {h}")
    else:
        print(f"  ❌ {h} NOT FOUND")

wb2.close()

print(f"\n" + "="*70)
print(f"Next steps:")
print(f"  1. Run: python rename_component.py {OLD_COMPONENT} {NEW_COMPONENT} --dry-run")
print(f"  2. Review the dry run output")
print(f"  3. Run: python rename_component.py {OLD_COMPONENT} {NEW_COMPONENT}")
print(f"="*70)
