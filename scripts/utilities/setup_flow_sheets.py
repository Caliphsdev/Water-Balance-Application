"""
Setup script: Add Flow Lines sheets to Excel template.

This creates area-specific sheets for flow line volumes:
- Flows_UG2N (UG2 North flows)
- Flows_MERN (Merensky North flows)
- Flows_MERENSKY_SOUTH (Merensky South flows)
- Flows_UG2S (UG2 South flows)
- Flows_STOCKPILE (Stockpile flows)
- Flows_OLDTSF (Old TSF flows)
- Flows_UG2PLANT (UG2 Plant flows)
- Flows_MERPLANT (Merensky Plant flows)
"""

import sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd


def setup_flow_sheets():
    """Add flow lines sheets to Excel template."""
    
    excel_path = Path(__file__).resolve().parent / 'test_templates' / 'Water_Balance_TimeSeries_Template.xlsx'
    
    if not excel_path.exists():
        print(f"❌ Excel file not found: {excel_path}")
        return False
    
    print(f"📂 Opening: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    
    # Define area sheets and their flow columns
    # These map to the diagram JSON flow segments
    area_flows = {
        'Flows_UG2N': {
            'area': 'UG2 North Decline Area',
            'flows': [
                'BOREHOLE_ABSTRACTION',
                'RAINFALL_UG2N',
                'OFFICES',
                'ACCOMMODATION',
                'NDCD1_INFLOW',
                'NDCD2_INFLOW',
                'RECOVERY_PLANT',
                'RECYCLED_WATER',
                'SPILLAGE_LOSS',
                'SEEPAGE_LOSS',
            ]
        },
        'Flows_MERN': {
            'area': 'Merensky North Area',
            'flows': [
                'RAINFALL_MERN',
                'MERENSKY_PLANT_INFLOW',
                'TREATMENT_PLANT',
                'DISCHARGE_POINT',
                'EVAPORATION_LOSS',
                'CONSERVATION_POOL',
            ]
        },
        'Flows_MERENSKY_SOUTH': {
            'area': 'Merensky South Area',
            'flows': [
                'RAINFALL_MERENSKY_SOUTH',
                'STORAGE_VOLUME',
                'TREATMENT_OUTFLOW',
                'EVAPORATION_MERENSKY_SOUTH',
                'RECIRCULATION_LOOP',
            ]
        },
        'Flows_UG2S': {
            'area': 'UG2 South Decline Area',
            'flows': [
                'RAINFALL_UG2S',
                'UG2S_INFLOW',
                'OUTFLOW_PLANT',
                'SEEPAGE_UG2S',
            ]
        },
        'Flows_STOCKPILE': {
            'area': 'Stockpile Area',
            'flows': [
                'RAINFALL_STOCKPILE',
                'STOCKPILE_RUNOFF',
                'INFILTRATION_LOSS',
                'PUMP_EXTRACTION',
                'TREATMENT_INPUT',
            ]
        },
        'Flows_OLDTSF': {
            'area': 'Old TSF Area',
            'flows': [
                'RAINFALL_OLDTSF',
                'TSF_INFLOW',
                'DECANT_OUTFLOW',
                'SEEPAGE_OLDTSF',
                'EVAPORATION_OLDTSF',
                'SPILLWAY_FLOW',
            ]
        },
        'Flows_UG2PLANT': {
            'area': 'UG2 Plant Area',
            'flows': [
                'PLANT_INFLOW',
                'PROCESS_WATER_USAGE',
                'RECYCLED_PROCESS_WATER',
                'PLANT_OUTFLOW',
                'EVAPORATION_UG2PLANT',
                'TREATMENT_DISCHARGE',
            ]
        },
        'Flows_MERPLANT': {
            'area': 'Merensky Plant Area',
            'flows': [
                'PLANT_INFLOW_MER',
                'PROCESS_WATER_MER',
                'RECYCLED_MER',
                'OUTFLOW_MER',
                'EVAPORATION_MER',
                'TAILING_THICKENER',
            ]
        },
    }
    
    # Colors for styling
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Generate 12 months of sample data (2025-2026)
    dates = []
    for month in range(1, 13):
        dates.append(datetime(2025, month, 1))
    for month in range(1, 13):
        dates.append(datetime(2026, month, 1))
    
    added_count = 0
    
    for sheet_name, sheet_config in area_flows.items():
        # Skip if sheet already exists
        if sheet_name in wb.sheetnames:
            print(f"⏭️  Sheet '{sheet_name}' already exists, skipping")
            continue
        
        print(f"➕ Creating sheet: {sheet_name}")
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        headers = ['Date'] + sheet_config['flows']
        ws.append(headers)
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add sample data (all zeros - user fills in actual values)
        for date_val in dates:
            row_data = [date_val] + [0.0] * len(sheet_config['flows'])
            ws.append(row_data)
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=len(dates) + 1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                # Format date column
                if cell.column == 1:
                    cell.number_format = 'YYYY-MM-DD'
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.number_format = '0.0'
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
        
        # Add area description (optional, for documentation)
        # Skip comment as it requires author parameter
        
        added_count += 1
        print(f"  ✅ Added {len(sheet_config['flows'])} flow columns")
        print(f"  📊 Sample: {', '.join(sheet_config['flows'][:3])}...")
    
    # Save workbook
    try:
        wb.save(excel_path)
        print(f"\n✅ Excel template updated!")
        print(f"📁 Added {added_count} flow sheets")
        print(f"📝 File: {excel_path}")
        return True
    except Exception as e:
        print(f"❌ Error saving workbook: {e}")
        return False


if __name__ == '__main__':
    success = setup_flow_sheets()
    sys.exit(0 if success else 1)
