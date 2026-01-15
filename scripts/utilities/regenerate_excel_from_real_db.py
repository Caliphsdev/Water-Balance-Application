"""
Regenerate Excel workbook with REAL database connections (59 flows).
Uses actual wb_flow_connections table as source of truth.
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).parent / 'src'))

from database.db_manager import DatabaseManager

db = DatabaseManager()

# Query all real connections from database
query = '''
SELECT 
    fs.structure_code as from_code,
    fs.structure_name as from_name,
    ts.structure_code as to_code,
    ts.structure_name as to_name
FROM wb_flow_connections fc
JOIN wb_structures fs ON fc.from_structure_id = fs.structure_id
JOIN wb_structures ts ON fc.to_structure_id = ts.structure_id
ORDER BY from_code, to_code
'''

connections = db.execute_query(query)
print(f'📊 Found {len(connections)} real connections in database\n')

# Map connections by area prefix
area_connections = {}
for conn in connections:
    from_code = conn.get('from_code')
    from_name = conn.get('from_name')
    to_code = conn.get('to_code')
    to_name = conn.get('to_name')
    
    # Determine area from source code
    if from_code.startswith('UG2N'):
        area = 'UG2N'
    elif from_code.startswith('UG2S'):
        area = 'UG2S'
    elif from_code.startswith('UG2P'):
        area = 'UG2P'
    elif from_code.startswith('MERN'):
        area = 'MERN'
    elif from_code.startswith('MERP'):
        area = 'MERP'
    elif from_code.startswith('MERS'):
        area = 'MERS'
    elif from_code.startswith('OT'):
        area = 'OLDTSF'
    elif from_code.startswith('SP'):
        area = 'STOCKPILE'
    else:
        area = 'OTHER'
    
    if area not in area_connections:
        area_connections[area] = []
    
    area_connections[area].append({
        'from_code': from_code,
        'from_name': from_name,
        'to_code': to_code,
        'to_name': to_name,
        'column_name': f'{from_code}__TO__{to_code}'
    })

# Create workbook with data structure
print('📝 Creating Excel workbook structure...\n')
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Add Reference Guide sheet
print('📖 Creating Reference Guide sheet...')
ref_sheet = wb.create_sheet('Reference Guide', 0)

# Reference sheet headers
ref_sheet['A1'] = 'Flow Component Mapping'
ref_sheet['A1'].font = Font(bold=True, size=14)
ref_sheet.merge_cells('A1:D1')

# Component codes section
ref_sheet['A3'] = 'Component Code Abbreviations'
ref_sheet['A3'].font = Font(bold=True, size=11, color='FFFFFF')
ref_sheet['A3'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ref_sheet.merge_cells('A3:B3')

ref_sheet['A4'] = 'Code'
ref_sheet['B4'] = 'Full Name'
ref_sheet['A4'].font = Font(bold=True)
ref_sheet['B4'].font = Font(bold=True)
ref_sheet['A4'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
ref_sheet['B4'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

# Get all structures and add to reference
structures = db.execute_query('SELECT structure_code, structure_name FROM wb_structures ORDER BY structure_code')
row = 5
for struct in structures:
    ref_sheet[f'A{row}'] = struct.get('structure_code')
    ref_sheet[f'B{row}'] = struct.get('structure_name')
    row += 1

ref_sheet.column_dimensions['A'].width = 15
ref_sheet.column_dimensions['B'].width = 40

# Add flows section
flows_start_row = row + 3
ref_sheet[f'A{flows_start_row}'] = 'Flow Columns (Source → Destination)'
ref_sheet[f'A{flows_start_row}'].font = Font(bold=True, size=11, color='FFFFFF')
ref_sheet[f'A{flows_start_row}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
ref_sheet.merge_cells(f'A{flows_start_row}:D{flows_start_row}')

flow_header_row = flows_start_row + 1
ref_sheet[f'A{flow_header_row}'] = 'Column Name'
ref_sheet[f'B{flow_header_row}'] = 'From (Source)'
ref_sheet[f'C{flow_header_row}'] = 'To (Destination)'
ref_sheet[f'D{flow_header_row}'] = 'Description'

for col in ['A', 'B', 'C', 'D']:
    ref_sheet[f'{col}{flow_header_row}'].font = Font(bold=True)
    ref_sheet[f'{col}{flow_header_row}'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

flow_row = flow_header_row + 1
for conn in connections:
    col_name = f'{conn.get("from_code")}__TO__{conn.get("to_code")}'
    ref_sheet[f'A{flow_row}'] = col_name
    ref_sheet[f'B{flow_row}'] = conn.get('from_name')
    ref_sheet[f'C{flow_row}'] = conn.get('to_name')
    ref_sheet[f'D{flow_row}'] = ''  # User can fill in
    flow_row += 1

ref_sheet.column_dimensions['A'].width = 30
ref_sheet.column_dimensions['B'].width = 35
ref_sheet.column_dimensions['C'].width = 35
ref_sheet.column_dimensions['D'].width = 40

print(f'   ✅ Added {len(structures)} component codes')
print(f'   ✅ Added {len(connections)} real flows\n')

# Create area-specific sheets
area_map = {
    'UG2N': 'Flows_UG2N',
    'UG2S': 'Flows_UG2S',
    'UG2P': 'Flows_UG2P',
    'MERN': 'Flows_Merensky North',
    'MERP': 'Flows_MERP',
    'MERS': 'Flows_MERENSKY_SOUTH',
    'OLDTSF': 'Flows_OLDTSF',
    'STOCKPILE': 'Flows_STOCKPILE',
    'OTHER': 'Flows_OTHER'
}

for area, sheet_name in area_map.items():
    if area in area_connections:
        print(f'📄 Creating {sheet_name} sheet...')
        ws = wb.create_sheet(sheet_name)
        
        # Headers: Date/Year/Month + flow columns
        ws['A1'] = 'Date'
        ws['B1'] = 'Year'
        ws['C1'] = 'Month'
        
        # Add flow columns
        col_idx = 4
        for conn in area_connections[area]:
            col_letter = chr(64 + col_idx)  # Convert to letter (D, E, F, ...)
            ws[f'{col_letter}1'] = conn['column_name']
            col_idx += 1
        
        # Format header row
        for cell in ws[1]:
            if cell.value:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 10
        for col_idx in range(4, 4 + len(area_connections[area])):
            col_letter = chr(64 + col_idx)
            ws.column_dimensions[col_letter].width = 20
        
        # Set row height for headers
        ws.row_dimensions[1].height = 40
        
        print(f'   ✅ Sheet created with {len(area_connections[area])} real connections\n')

# Save workbook
output_path = Path(__file__).parent / 'test_templates' / 'Water_Balance_TimeSeries_Template.xlsx'
wb.save(str(output_path))
print(f'✅ Excel regenerated successfully!')
print(f'📊 Total real connections: {len(connections)}')
print(f'📁 Output: {output_path}')
print(f'\n✨ Excel now uses ONLY actual database connections (no invented flows)')
print(f'✨ All {len(area_connections)} area sheets created with real flows')
