"""
Rebuild Excel sheets with proper column structure for flow data
Based on the diagram structure with proper Date, Year, Month headers
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
import shutil
from datetime import datetime

excel_path = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

# Create backup
backup_path = excel_path.with_suffix(f'.backup_before_restructure_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy(excel_path, backup_path)
print(f"✅ Backup created: {backup_path.name}\n")

# Load workbook
wb = load_workbook(excel_path)

# Define standard headers for all flow sheets
standard_headers = ['Date', 'Year', 'Month']

# Define header formatting
header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

print("🔧 Restructuring Excel sheets...\n")

# Process each sheet except Reference Guide
for sheet_name in wb.sheetnames:
    if sheet_name == 'Reference Guide':
        continue
    
    ws = wb[sheet_name]
    print(f"Processing {sheet_name}...")
    
    # Clear all existing data
    ws.delete_rows(1, ws.max_row)
    
    # Add standard headers
    for col_idx, header_name in enumerate(standard_headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Set column widths
    ws.column_dimensions['A'].width = 12  # Date
    ws.column_dimensions['B'].width = 8   # Year
    ws.column_dimensions['C'].width = 10  # Month
    
    print(f"  ✅ Added headers: {standard_headers}")

# Update Reference Guide with clear instructions
print(f"\nUpdating Reference Guide...")
ref_sheet = wb['Reference Guide']
ref_sheet.delete_rows(1, ref_sheet.max_row)

# Title
ref_sheet['A1'] = 'Water Balance Excel Structure - Reference Guide'
ref_sheet['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ref_sheet['A1'].fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
ref_sheet.merge_cells('A1:D1')
ref_sheet.row_dimensions[1].height = 25

# Last Updated
ref_sheet['A2'] = f'Last Updated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
ref_sheet['A2'].font = Font(italic=True, size=9)
ref_sheet.row_dimensions[2].height = 20

# Overview
ref_sheet['A4'] = '📋 OVERVIEW'
ref_sheet['A4'].font = Font(bold=True, size=12, color='FFFFFF')
ref_sheet['A4'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
ref_sheet['A5'] = 'Each flow sheet (Flows_AREA) contains monthly flow volume data for that area/facility.'
ref_sheet['A6'] = 'Standard columns: Date (day), Year, Month'

# Sheet Structure
ref_sheet['A8'] = '📄 SHEET STRUCTURE'
ref_sheet['A8'].font = Font(bold=True, size=12, color='FFFFFF')
ref_sheet['A8'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

ref_sheet['A10'] = 'Sheet Name'
ref_sheet['B10'] = 'Area/Facility'
ref_sheet['C10'] = 'Description'
ref_sheet['A10'].font = Font(bold=True, color='FFFFFF')
ref_sheet['B10'].font = Font(bold=True, color='FFFFFF')
ref_sheet['C10'].font = Font(bold=True, color='FFFFFF')
ref_sheet['A10'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ref_sheet['B10'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ref_sheet['C10'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

sheets_info = [
    ('Flows_UG2N', 'UG2 North Decline', 'Underground 2 North mining area flows'),
    ('Flows_STOCKPILE', 'Stockpile Area', 'Ore stockpile facility flows'),
    ('Flows_UG2S', 'UG2 South Decline', 'Underground 2 South mining area flows'),
    ('Flows_MERS', 'Merensky South', 'Merensky South mining area flows'),
    ('Flows_OLDTSF', 'Old TSF', 'Old Tailings Storage Facility (original facility)'),
    ('Flows_NEWTSF', 'New TSF', 'New Tailings Storage Facility (split from Old TSF) ⭐ NEW'),
    ('Flows_UG2P', 'UG2 Plant', 'Underground 2 processing plant flows'),
    ('Flows_MERP', 'Merensky Plant', 'Merensky processing plant flows'),
]

row = 11
for sheet_name, area, desc in sheets_info:
    ref_sheet[f'A{row}'] = sheet_name
    ref_sheet[f'B{row}'] = area
    ref_sheet[f'C{row}'] = desc
    ref_sheet[f'A{row}'].font = Font(name='Courier New', size=9)
    row += 1

# How to use
ref_sheet[f'A{row+1}'] = '📚 HOW TO USE'
ref_sheet[f'A{row+1}'].font = Font(bold=True, size=12, color='FFFFFF')
ref_sheet[f'A{row+1}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

tips_row = row + 3
tips = [
    '1. Select a sheet tab (e.g., Flows_UG2N)',
    '2. Enter data in rows starting from row 2',
    '3. Required columns (Date, Year, Month) in columns A, B, C',
    '4. Add flow volume columns after Month column (D onwards)',
    '5. Use Excel Manager in Flow Diagram Dashboard to manage columns',
    '',
    '💡 Column Naming Convention:',
    '   Format: COMPONENT1_TO_COMPONENT2 or COMPONENT1_COMPONENT2',
    '   Example: PROCESSING_TO_STORAGE, UG2N_TO_PLANT',
    '',
    '⚙️ Data Format:',
    '   • Date: Day of month (1-31)',
    '   • Year: 4-digit year (2024, 2025, etc.)',
    '   • Month: Month number (1-12) or month name',
    '   • Flows: Numeric values in m³ (cubic meters)',
    '',
    '✨ Recent Changes:',
    '   • Merensky North Area removed (Flows_MERN)',
    '   • Old TSF split into Old TSF + New TSF',
    '   • All sheets now have standard header structure',
    '',
    '📝 Tips:',
    '   • Use Auto-Map feature to link columns to flow diagram',
    '   • Always backup Excel before major changes',
    '   • Export diagrams to refresh flow calculations',
]

for tip in tips:
    ref_sheet[f'A{tips_row}'] = tip
    ref_sheet[f'A{tips_row}'].font = Font(size=10)
    if tip.startswith('💡') or tip.startswith('⚙️') or tip.startswith('✨') or tip.startswith('📝'):
        ref_sheet[f'A{tips_row}'].font = Font(size=10, bold=True)
    tips_row += 1

# Set column widths for Reference Guide
ref_sheet.column_dimensions['A'].width = 25
ref_sheet.column_dimensions['B'].width = 30
ref_sheet.column_dimensions['C'].width = 60
ref_sheet.column_dimensions['D'].width = 20

# Save all changes
wb.save(excel_path)
print(f"  ✅ Updated Reference Guide with complete instructions")

print(f"\n✅ Excel restructuring complete!")
print(f"   File: {excel_path.name}")
print(f"\n📊 Current structure:")
print(f"   Total sheets: {len(wb.sheetnames)}")
print(f"   All sheets now have standard headers: {standard_headers}")

wb.close()

print(f"\n✅ Backup: {backup_path.name}")
print(f"\n🎯 Next: Launch Flow Diagram Dashboard and test Excel Manager Columns tab")
