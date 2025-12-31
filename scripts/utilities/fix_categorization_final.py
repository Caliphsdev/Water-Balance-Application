"""
Fix categorization with better pattern ordering - check specific patterns first.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

print("🔧 FIXING CATEGORIZATION WITH SPECIFIC PATTERNS")
print("=" * 100)

# Load JSON
with open('data/diagrams/ug2_north_decline.json') as f:
    diagram = json.load(f)

edges = diagram.get('edges', [])
print(f"✅ Loaded {len(edges)} edges from JSON")

# Define area categorization with SPECIFIC patterns checked FIRST
# Use more specific patterns that avoid overlaps
area_definitions = {
    # UG2N - specific to UG2 North (NO merensky)
    'UG2N': {
        'prefixes': ['rainfall__', 'ndcd__', 'bh_ndgwa__', 'softening__', 'reservoir__', 'offices__', 
                    'guest_house__', 'septic__', 'sewage__', 'north_decline__', 'north_shaft__', 
                    'losses__', 'losses2__', 'consumption__', 'consumption2__', 'evaporation__',
                    'dust_suppression__', 'spill__', 'junction_127__', 'junction_128__'],
        'contains': ['__TO__rainfall', '__TO__ndcd', '__TO__bh_ndgwa', '__TO__softening', '__TO__reservoir', 
                    '__TO__offices', '__TO__guest_house', '__TO__septic', '__TO__sewage', '__TO__north_decline', 
                    '__TO__north_shaft', '__TO__losses', '__TO__losses2', '__TO__consumption', '__TO__consumption2',
                    '__TO__evaporation', '__TO__dust_suppression', '__TO__spill', '__TO__junction_127', '__TO__junction_128']
    },
    # MERN - specific to Merensky North (ndcd_merensky, etc.)
    'MERN': {
        'prefixes': ['rainfall_merensky__', 'ndcd_merensky__', 'bh_mcgwa__', 'softening_merensky__', 
                    'offices_merensky__', 'merensky_north__', 'evaporation_merensky__', 'dust_suppression_merensky__',
                    'spill_merensky__', 'consumption_merensky__', 'losses_merensky__'],
        'contains': ['__TO__rainfall_merensky', '__TO__ndcd_merensky', '__TO__bh_mcgwa', '__TO__softening_merensky', 
                    '__TO__offices_merensky', '__TO__merensky_north', '__TO__evaporation_merensky', '__TO__dust_suppression_merensky',
                    '__TO__spill_merensky', '__TO__consumption_merensky', '__TO__losses_merensky']
    },
    # UG2S - specific to UG2 South
    'UG2S': {
        'prefixes': ['ug2s_'],
        'contains': ['__TO__ug2s_']
    },
    # UG2P - specific to UG2 Plant
    'UG2P': {
        'prefixes': ['ug2plant_', 'cprwsd'],
        'contains': ['__TO__ug2plant_', '__TO__cprwsd']
    },
    # MERP - specific to Merensky Plant
    'MERP': {
        'prefixes': ['merplant_', 'mprwsd'],
        'contains': ['__TO__merplant_', '__TO__mprwsd']
    },
    # MERS - specific to Merensky South
    'MERS': {
        'prefixes': ['mers_'],
        'contains': ['__TO__mers_']
    },
    # OLDTSF - specific to Old TSF
    'OLDTSF': {
        'prefixes': ['oldtsf_', 'nt_rwd', 'new_tsf', 'trtd'],
        'contains': ['__TO__oldtsf_', '__TO__nt_rwd', '__TO__new_tsf', '__TO__trtd']
    },
    # STOCKPILE - specific to Stockpile
    'STOCKPILE': {
        'prefixes': ['stockpile_', 'spcd1'],
        'contains': ['__TO__stockpile_', '__TO__spcd1']
    }
}

def find_area_for_flow(from_id, to_id):
    """Find which area a flow belongs to by checking patterns in order."""
    if not from_id or not to_id:
        return None
    
    flow_str = f'{from_id}__TO__{to_id}'
    
    # Check in order: more specific areas first
    for area in ['MERN', 'MERP', 'MERS', 'UG2N', 'UG2S', 'UG2P', 'OLDTSF', 'STOCKPILE']:
        patterns = area_definitions.get(area, {})
        
        # Check prefixes
        for prefix in patterns.get('prefixes', []):
            if flow_str.startswith(prefix):
                return area
        
        # Check contains patterns
        for contains in patterns.get('contains', []):
            if contains in flow_str:
                return area
    
    return None

# Categorize flows
flows_by_area = {area: [] for area in area_definitions.keys()}
flows_by_area['OTHER'] = []

for edge in edges:
    from_id = edge.get('from')
    to_id = edge.get('to')
    
    flow_key = f'{from_id}__TO__{to_id}'
    area = find_area_for_flow(from_id, to_id)
    
    if area:
        flows_by_area[area].append(flow_key)
    else:
        flows_by_area['OTHER'].append(flow_key)

print("\n📊 FLOW CATEGORIZATION BY AREA")
print("=" * 100)

total = 0
for area in ['UG2N', 'UG2S', 'UG2P', 'MERN', 'MERP', 'MERS', 'OLDTSF', 'STOCKPILE', 'OTHER']:
    count = len(flows_by_area[area])
    total += count
    status = '✅' if count > 0 else '⚠️'
    print(f'{status} {area:15} → {count:3} flows')

print(f"\n{'=' * 100}")
print(f"{'TOTAL':15} → {total:3} flows")

# Generate Excel
print(f"\n📝 GENERATING EXCEL WORKBOOK")
print("=" * 100)

wb = Workbook()
wb.remove(wb.active)

# Style templates
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
area_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
area_font = Font(bold=True, size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Reference Guide
ws_ref = wb.create_sheet('Reference Guide', 0)
ws_ref['A1'] = 'Water Balance Template - Node & Flow Reference'
ws_ref['A1'].font = Font(bold=True, size=14)

def find_area_for_node(node_id):
    """Find which area a node belongs to."""
    if not node_id:
        return None
    node_lower = node_id.lower()
    # Simple area detection for nodes
    for area in ['MERN', 'MERP', 'MERS', 'UG2N', 'UG2S', 'UG2P', 'OLDTSF', 'STOCKPILE']:
        patterns = area_definitions.get(area, {})
        for prefix in patterns.get('prefixes', []):
            if prefix.rstrip('_').replace('__', '_') in node_lower:
                return area
    return None

row = 3
ws_ref[f'A{row}'] = 'Node ID'
ws_ref[f'B{row}'] = 'Label'
ws_ref[f'C{row}'] = 'Area'

for cell in [ws_ref[f'A{row}'], ws_ref[f'B{row}'], ws_ref[f'C{row}']]:
    cell.fill = header_fill
    cell.font = header_font

row += 1
nodes = diagram.get('nodes', [])
node_list = nodes if isinstance(nodes, list) else [{'id': k, 'label': v.get('label', '')} for k, v in nodes.items()]

for node in sorted(node_list, key=lambda x: x.get('id', '')):
    node_id = node.get('id')
    area = find_area_for_node(node_id)
    
    ws_ref[f'A{row}'] = node_id
    ws_ref[f'B{row}'] = node.get('label', '')
    ws_ref[f'C{row}'] = area if area else 'OTHER'
    row += 1

ws_ref.column_dimensions['A'].width = 25
ws_ref.column_dimensions['B'].width = 40
ws_ref.column_dimensions['C'].width = 15

# Create area sheets
sheet_order = ['UG2N', 'UG2S', 'UG2P', 'MERN', 'MERP', 'MERS', 'OLDTSF', 'STOCKPILE', 'OTHER']

for sheet_idx, area in enumerate(sheet_order, start=1):
    flows = flows_by_area[area]
    if not flows:
        continue
    
    ws = wb.create_sheet(f'Flows_{area}', sheet_idx)
    
    # Header
    ws['A1'] = f'{area} Water Balance Template'
    ws['A1'].font = Font(bold=True, size=12)
    
    # Column headers
    ws['A3'] = 'Date'
    ws['B3'] = 'Year'
    ws['C3'] = 'Month'
    
    for cell in [ws['A3'], ws['B3'], ws['C3']]:
        cell.fill = area_fill
        cell.font = area_font
        cell.border = border
    
    # Flow columns
    col = 4
    for flow in sorted(flows):
        col_letter = get_column_letter(col)
        ws[f'{col_letter}3'] = flow
        ws[f'{col_letter}3'].fill = header_fill
        ws[f'{col_letter}3'].font = header_font
        ws[f'{col_letter}3'].border = border
        ws[f'{col_letter}3'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws.column_dimensions[col_letter].width = 25
        col += 1
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.row_dimensions[3].height = 30

# Save
timestamp = int(datetime.now().timestamp() * 1000)
output_file = f'test_templates/Water_Balance_TimeSeries_Template_FIXED_{timestamp}.xlsx'
wb.save(output_file)

print(f"✅ Excel saved: {Path(output_file).name}")
print(f"\n📊 SUMMARY:")
print(f"   Total flows distributed: {total}")
print(f"   Reference Guide: {len(node_list)} nodes")
print(f"   Area sheets: {len(sheet_order)}")
print(f"\n📋 FLOW DISTRIBUTION:")
for area in sheet_order:
    count = len(flows_by_area[area])
    if count > 0:
        print(f"   {area:15} {count:3} flows")
