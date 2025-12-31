#!/usr/bin/env python
"""Quick validation test of JSON to Excel mappings."""

import json
from pathlib import Path
from openpyxl import load_workbook

DIAGRAM_PATH = Path('data/diagrams/ug2_north_decline.json')
EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("Validation Test: JSON Excel Mappings")
print("=" * 60)

# Load Excel headers
wb = load_workbook(EXCEL_PATH)
excel_headers = {}
for sheet_name in wb.sheetnames:
    if sheet_name.startswith('Flows_'):
        ws = wb[sheet_name]
        headers = [c.value for c in ws[3] if c.value]
        excel_headers[sheet_name] = headers
wb.close()

# Load JSON
with open(DIAGRAM_PATH) as f:
    diagram = json.load(f)

# Check each edge
errors = []
warnings = []
valid = 0

for edge in diagram['edges']:
    mapping = edge.get('excel_mapping', {})
    if not mapping or not mapping.get('enabled'):
        continue
    
    sheet = mapping.get('sheet')
    column = mapping.get('column')
    edge_id = f"{edge.get('from')} -> {edge.get('to')}"
    
    if sheet not in excel_headers:
        errors.append(f"Sheet '{sheet}' not found for {edge_id}")
        continue
    
    if column not in excel_headers[sheet]:
        errors.append(f"Column '{column}' not in {sheet} for {edge_id}")
        continue
    
    valid += 1

print(f"\nResults:")
print(f"  Valid mappings: {valid}")
print(f"  Errors: {len(errors)}")

if errors:
    print(f"\nErrors found:")
    for err in errors[:10]:
        print(f"  - {err}")
    if len(errors) > 10:
        print(f"  ... and {len(errors) - 10} more")

print("\n" + "=" * 60)
if errors:
    print("VALIDATION FAILED - See errors above")
else:
    print("VALIDATION PASSED!")
