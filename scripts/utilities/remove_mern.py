from openpyxl import load_workbook

print("Checking Excel template sheets...\n")

wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("BEFORE:")
print(f"  Total sheets: {len(wb.sheetnames)}")
for sheet in sorted(wb.sheetnames):
    print(f"    - {sheet}")

# Remove MERN if it exists
if 'Flows_MERN' in wb.sheetnames:
    print("\n  Removing Flows_MERN...")
    del wb['Flows_MERN']
    wb.save('test_templates/Water_Balance_TimeSeries_Template.xlsx')
    print("  ✓ Saved")
else:
    print("\n  Flows_MERN not found - may already be removed")

# Verify
wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')
print("\nAFTER:")
print(f"  Total sheets: {len(wb.sheetnames)}")
for sheet in sorted(wb.sheetnames):
    print(f"    - {sheet}")

if 'Flows_MERN' not in wb.sheetnames:
    print("\n✓ SUCCESS: Flows_MERN removed!")
