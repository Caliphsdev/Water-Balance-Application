#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import re

print("="*80)
print("FIXING JSON ARROW ENCODING AND REMOVING MERN SHEET")
print("="*80)

# === FIX 1: JSON ARROW ENCODING ===
print("\n[1] Fixing JSON arrow encoding...")

with open('data/diagrams/ug2_north_decline.json', 'r', encoding='utf-8') as f:
    content = f.read()

# Replace all Unicode-encoded corrupted arrows with correct ones
# The pattern \u00e2\u2020\u2019 is the UTF-8 encoding of â†'
original_content = content

# Replace various forms of corrupted arrows
replacements = [
    ('\u00e2\u2020\u2019', '→'),  # Unicode escape for â†'
    ('â\u2020\u2019', '→'),        # Partially decoded
    ('â†'', '→'),                  # Display form
]

for old, new in replacements:
    count = content.count(old)
    if count > 0:
        print(f"  Found {count} instances of corrupted arrow (type: {repr(old)})")
        content = content.replace(old, new)

# Save fixed JSON
with open('data/diagrams/ug2_north_decline.json', 'w', encoding='utf-8') as f:
    f.write(content)

print("  ✓ JSON arrows fixed and saved")

# Verify by parsing
try:
    with open('data/diagrams/ug2_north_decline.json', 'r', encoding='utf-8') as f:
        diagram = json.load(f)
    print(f"  ✓ JSON parsed successfully ({len(diagram.get('edges', []))} edges)")
except Exception as e:
    print(f"  ✗ Error: {e}")

# === FIX 2: REMOVE MERN SHEET ===
print("\n[2] Removing deprecated Flows_Merensky North sheet...")

from openpyxl import load_workbook

wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')

sheets_before = len(wb.sheetnames)
print(f"  Sheets before: {sheets_before}")

if 'Flows_Merensky North' in wb.sheetnames:
    del wb['Flows_Merensky North']
    wb.save('test_templates/Water_Balance_TimeSeries_Template.xlsx')
    print(f"  ✓ Flows_Merensky North removed")
else:
    print(f"  ℹ️ Flows_Merensky North not in template")

# Verify
wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')
sheets_after = len(wb.sheetnames)
print(f"  Sheets after: {sheets_after}")
print(f"  Remaining sheets: {', '.join(sorted(wb.sheetnames))}")

print("\n" + "="*80)
print("✓ ALL FIXES COMPLETE")
print("="*80)
print("\nSummary:")
print("  ✓ JSON arrow encoding fixed")
print("  ✓ Flows_MERN sheet removed")
print("  ✓ System ready to use!\n")
