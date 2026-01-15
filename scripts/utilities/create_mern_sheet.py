import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

excel_file = Path('test_templates/Water_Balance_TimeSeries_Template_POPULATED.xlsx')

# Create sample MERN flow data structure
data = {
    'Date': [1, 1, 1],
    'Year': [2024, 2024, 2024],
    'Month': [10, 11, 12],
    'merensky_north_decline__TO__ndcd': [50000, 51000, 52000],
    'merensky_north_shaft__TO__ndcd': [25000, 26000, 27000],
    'merensky_offices__TO__consumption': [2000, 2100, 2200],
    # Add more flows as needed
}

df = pd.DataFrame(data)

# Load existing workbook
book = load_workbook(excel_file)
writer = pd.ExcelWriter(excel_file, engine='openpyxl')
writer.book = book

# Remove old MERN sheet if exists
if 'Flows_MERN' in book.sheetnames:
    del book['Flows_MERN']

# Write new MERN sheet
df.to_excel(writer, sheet_name='Flows_MERN', index=False)

writer.save()
writer.close()

print('✅ Created Flows_Merensky North sheet with sample data')
print(f'   Columns: {list(df.columns)}')
print(f'   Rows: {len(df)}')
