#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Sync JSON diagram mappings to Excel headers (fixed version)."""

import json
from openpyxl import load_workbook
from pathlib import Path

DIAGRAM_PATH = Path('data/diagrams/ug2_north_decline.json')
EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print("=" * 75)
print("SYNC JSON DIAGRAM - EXCEL HEADERS")
print("=" * 75)

# Load Excel headers
wb = load_workbook(EXCEL_PATH)
excel_headers_by_sheet = {}

for sheet_name in wb.sheetnames:
    if sheet_name.startswith('Flows_'):
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[3] if cell.value]  # Row 3 has headers
        excel_headers_by_sheet[sheet_name] = set(headers)
        print(f"[OK] {sheet_name}: {len(headers)} headers")

wb.close()

# Load diagram
with open(DIAGRAM_PATH, 'r', encoding='utf-8') as f:
    diagram = json.load(f)

edges = diagram.get('edges', [])
print(f"\n[OK] Found {len(edges)} edges in diagram\n")

# Process edges
stats = {'updated': 0, 'exact_match': 0, 'fuzzy_match': 0, 'failed': 0, 'skipped': 0}

for edge in edges:
    edge_id = f"{edge.get('from', '')} -> {edge.get('to', '')}"
    mapping = edge.get('excel_mapping', {})
    
    # Skip disabled edges
    if not mapping or not mapping.get('enabled'):
        stats['skipped'] += 1
        continue
    
    sheet = mapping.get('sheet', '')
    old_column = mapping.get('column', '')
    
    if sheet not in excel_headers_by_sheet:
        stats['failed'] += 1
        print(f"[FAIL] {edge_id}")
        print(f"       Sheet not found: {sheet}")
        continue
    
    available = excel_headers_by_sheet[sheet]
    
    # Exact match
    if old_column in available:
        stats['exact_match'] += 1
        continue
    
    # Fuzzy match - convert old format to new format
    # Old: lowercase with __TO__
    # New: UPPERCASE with ' -> '
    fuzzy_search = old_column.upper().replace('__TO__', ' -> ').replace('_', ' ')
    
    best_match = None
    best_score = 0
    
    for header in available:
        header_norm = header.replace('_', ' ')
        # Score based on matching words
        fuzzy_words = set(w for w in fuzzy_search.split() if w and w != '->')
        header_words = set(w for w in header_norm.split() if w and w != '->')
        overlap = len(fuzzy_words & header_words)
        
        if overlap > best_score:
            best_score = overlap
            best_match = header
    
    if best_match and best_score >= 2:  # At least 2 words match
        mapping['column'] = best_match
        edge['excel_mapping'] = mapping
        stats['fuzzy_match'] += 1
        print(f"[OK] {edge_id}")
        print(f"     Old: {old_column}")
        print(f"     New: {best_match}")
    else:
        stats['failed'] += 1
        print(f"[FAIL] {edge_id}")
        print(f"       Looked for: {old_column}")
        print(f"       Available: {list(available)[:3]}...")

# Save diagram
with open(DIAGRAM_PATH, 'w', encoding='utf-8') as f:
    json.dump(diagram, f, indent=2)

print("\n" + "=" * 75)
print("SUMMARY")
print("=" * 75)
print(f"Total edges: {len(edges)}")
print(f"Exact matches: {stats['exact_match']}")
print(f"Fuzzy matches: {stats['fuzzy_match']}")
print(f"Failed: {stats['failed']}")
print(f"Skipped: {stats['skipped']}")
print(f"TOTAL UPDATED: {stats['exact_match'] + stats['fuzzy_match']}/{len(edges)}")
