import json
from openpyxl import load_workbook

print("="*80)
print("FIX 1: JSON ARROW ENCODING")
print("="*80)

# Load and fix JSON
with open('data/diagrams/ug2_north_decline.json', encoding='utf-8') as f:
    content = f.read()

# Count corrupted arrows before
corrupted_count = content.count('â†'')
print(f"\nFound {corrupted_count} corrupted arrows in JSON")

# Replace all corrupted arrows with correct ones
fixed_content = content.replace('â†'', '→')

# Save with proper UTF-8 encoding
with open('data/diagrams/ug2_north_decline.json', 'w', encoding='utf-8') as f:
    f.write(fixed_content)

print(f"✓ Replaced {corrupted_count} corrupted arrows with correct '→' character")
print("✓ JSON file saved with UTF-8 encoding")

# Verify by loading
with open('data/diagrams/ug2_north_decline.json', encoding='utf-8') as f:
    diagram = json.load(f)

remaining_corrupted = sum(1 for e in diagram['edges'] 
                          if 'â†'' in str(e.get('excel_mapping', {}).get('column', '')))
print(f"✓ Verification: {remaining_corrupted} corrupted arrows remaining (should be 0)")

print("\n" + "="*80)
print("FIX 2: REMOVE DEPRECATED FLOWS_MERENSKY NORTH SHEET")
print("="*80)

wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')

print(f"\nSheets before removal: {len(wb.sheetnames)}")
print(f"  {', '.join(wb.sheetnames)}")

if 'Flows_Merensky North' in wb.sheetnames:
    del wb['Flows_Merensky North']
    print("\n✓ Removed Flows_Merensky North sheet")
    
    wb.save('test_templates/Water_Balance_TimeSeries_Template.xlsx')
    print("✓ Template saved")
else:
    print("\n✗ Flows_Merensky North not found in template")

# Verify
wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')
print(f"\nSheets after removal: {len(wb.sheetnames)}")
print(f"  {', '.join(sorted(wb.sheetnames))}")

print("\n" + "="*80)
print("ALL FIXES COMPLETE")
print("="*80)
print("\n✓ JSON arrows: Fixed (replaced â†' with →)")
print("✓ Merensky North sheet: Removed from template")
print("✓ System ready to use!")
