import json
from openpyxl import load_workbook

print("="*80)
print("FINAL JSON ARROW FIX AND VERIFICATION")
print("="*80)

# Load files
with open('data/diagrams/ug2_north_decline.json', encoding='utf-8') as f:
    diagram = json.load(f)

wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')

# Get all correct headers from Excel
excel_headers = {}
for sheet_name in wb.sheetnames:
    if 'Flows_' in sheet_name:
        ws = wb[sheet_name]
        headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
        excel_headers[sheet_name] = [h for h in headers if h]

print(f"\nLoaded {sum(len(h) for h in excel_headers.values())} headers from Excel")

# Fix all edges
fixed_count = 0
for edge in diagram['edges']:
    mapping = edge.get('excel_mapping', {})
    sheet = mapping.get('sheet', '')
    column = mapping.get('column', '')
    
    if not sheet or not column:
        continue
    
    if sheet not in excel_headers:
        continue
    
    # If column not in headers, try to find it in Excel
    if column not in excel_headers[sheet]:
        # Try to find similar (might be encoding issue)
        for excel_col in excel_headers[sheet]:
            # Compare ignoring arrow encoding issues
            if excel_col:
                # Normalize both: replace any arrow variants
                normalized_json = column.replace('â†'', '→').replace('→', 'ARROW')
                normalized_excel = str(excel_col).replace('→', 'ARROW')
                
                if normalized_json == normalized_excel:
                    mapping['column'] = excel_col  # Use the correct Excel version
                    fixed_count += 1
                    break

print(f"Fixed {fixed_count} edges with corrupted encodings")

# Final verification
print("\nFINAL VERIFICATION:")
print("-" * 80)

total_valid = 0
total_invalid = 0

for edge in diagram['edges']:
    mapping = edge.get('excel_mapping', {})
    sheet = mapping.get('sheet', '')
    column = mapping.get('column', '')
    
    if not sheet or not column:
        total_invalid += 1
        continue
    
    if sheet not in excel_headers:
        total_invalid += 1
        continue
    
    if column in excel_headers[sheet]:
        total_valid += 1
    else:
        total_invalid += 1

print(f"  Valid mappings: {total_valid}")
print(f"  Invalid/missing: {total_invalid}")

# Save
with open('data/diagrams/ug2_north_decline.json', 'w', encoding='utf-8') as f:
    json.dump(diagram, f, indent=2, ensure_ascii=False)

print("\n✓ JSON saved with UTF-8 encoding and corrected arrows")
print("="*80)
