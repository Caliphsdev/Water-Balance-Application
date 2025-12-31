import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Load JSON diagram
with open('data/diagrams/ug2_north_decline.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Load Excel
wb = load_workbook('test_templates/Water_Balance_TimeSeries_Template.xlsx')

# Get all enabled mappings
edges = data.get('edges', [])
enabled_mappings = []
for edge in edges:
    mapping = edge.get('excel_mapping', {})
    if mapping and mapping.get('enabled'):
        enabled_mappings.append({
            'from': edge.get('from'),
            'to': edge.get('to'),
            'sheet': mapping.get('sheet'),
            'column': mapping.get('column')
        })

print(f'Found {len(enabled_mappings)} enabled mappings')

# Group by sheet
by_sheet = {}
for m in enabled_mappings:
    sheet = m['sheet']
    if sheet not in by_sheet:
        by_sheet[sheet] = []
    by_sheet[sheet].append(m)

# Add missing columns to each sheet
added_total = 0
for sheet_name, mappings in by_sheet.items():
    if sheet_name not in wb.sheetnames:
        print(f'⚠️  Sheet {sheet_name} not found in Excel')
        continue
    
    ws = wb[sheet_name]
    
    # Get existing headers from row 3
    existing_headers = [cell.value for cell in ws[3] if cell.value]
    
    # Find missing columns
    missing = []
    for m in mappings:
        col = m['column']
        if col and col not in existing_headers:
            missing.append(col)
    
    # Remove duplicates
    missing = list(set(missing))
    
    if not missing:
        print(f'✓ {sheet_name}: All columns exist ({len(existing_headers)} columns)')
        continue
    
    print(f'\n{sheet_name}: Adding {len(missing)} missing columns')
    
    # Find next available column
    next_col = len(existing_headers) + 1
    
    # Add missing columns
    for col_name in sorted(missing):
        # Add to row 3 (header row)
        cell = ws.cell(row=3, column=next_col, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add placeholder data in row 4
        ws.cell(row=4, column=next_col, value='-')
        
        print(f'  + Column {next_col}: {col_name}')
        next_col += 1
        added_total += 1

# Save
wb.save('test_templates/Water_Balance_TimeSeries_Template.xlsx')
print(f'\n✅ Added {added_total} columns total')
print(f'Excel file saved: test_templates/Water_Balance_TimeSeries_Template.xlsx')
