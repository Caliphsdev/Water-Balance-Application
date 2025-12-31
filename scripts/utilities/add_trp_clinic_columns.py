#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Add TRP_CLINIC columns to Excel (replacing old GUEST_HOUSE references)."""

from openpyxl import load_workbook
from pathlib import Path

EXCEL_PATH = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')

# Columns to add to Flows_UG2N sheet
COLUMNS_TO_ADD = [
    'SOFTENING → TRP_CLINIC',
    'TRP_CLINIC → SEPTIC',
    'TRP_CLINIC → CONSUMPTION',
]

print("="*70)
print("Adding TRP_CLINIC columns to Excel")
print("="*70)

wb = load_workbook(EXCEL_PATH)
ws = wb['Flows_UG2N']

# Get current headers from row 3
current_headers = [c.value for c in ws[3]]
print(f"\nCurrent Flows_UG2N headers: {len([h for h in current_headers if h])} columns")

# Find last populated column
last_col = 0
for col in range(1, ws.max_column + 1):
    if ws.cell(3, col).value:
        last_col = col

print(f"Last data column: {last_col}")

# Add new columns
print(f"\nAdding {len(COLUMNS_TO_ADD)} new columns:")
for idx, col_name in enumerate(COLUMNS_TO_ADD, 1):
    new_col = last_col + idx
    ws.cell(3, new_col).value = col_name
    print(f"  Column {new_col}: {col_name}")
    
    # Add placeholder '-' to data rows
    for row in range(4, 14):
        if not ws.cell(row, new_col).value:
            ws.cell(row, new_col).value = '-'

wb.save(EXCEL_PATH)
print(f"\n✅ Excel updated with TRP_CLINIC columns")

# Verify
wb2 = load_workbook(EXCEL_PATH)
ws2 = wb2['Flows_UG2N']
headers = [c.value for c in ws2[3] if c.value]
print(f"\nVerification:")
print(f"  Total headers: {len(headers)}")
for h in COLUMNS_TO_ADD:
    if h in headers:
        col_idx = headers.index(h) + 1
        print(f"  ✅ {h}")
    else:
        print(f"  ❌ {h} NOT FOUND")

wb2.close()
