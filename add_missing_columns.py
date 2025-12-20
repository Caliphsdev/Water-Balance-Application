"""
Add missing Excel columns for the 23 remaining disabled edges.
This adds new columns to the Excel template and enables them in JSON.
"""

import json
from openpyxl import load_workbook

# Load categorized data
with open("disabled_edges_categorized.json", "r", encoding="utf-8") as f:
    categorized = json.load(f)

needs_column = categorized["needs_column"]

print(f"Adding {len(needs_column)} new Excel columns...\n")

# Group by sheet
columns_by_sheet = {}
for item in needs_column:
    sheet = item["sheet"]
    if sheet not in columns_by_sheet:
        columns_by_sheet[sheet] = []
    columns_by_sheet[sheet].append(item)

# Load Excel
excel_path = "test_templates/Water_Balance_TimeSeries_Template.xlsx"
wb = load_workbook(excel_path)

added_columns = {}

for sheet_name, items in columns_by_sheet.items():
    if sheet_name not in wb.sheetnames:
        print(f"[SKIP] Sheet {sheet_name} not found in Excel")
        continue
    
    ws = wb[sheet_name]
    
    # Find the last column with data in row 3 (headers)
    last_col = 3  # Start after Date, Year, Month
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(3, col_idx).value:
            last_col = col_idx
    
    next_col = last_col + 1
    
    print(f"\n[{sheet_name}]")
    print(f"  Last column: {last_col}, adding from column {next_col}")
    
    added_columns[sheet_name] = []
    
    for item in items:
        column_name = item["expected_column"]
        from_id = item["from"]
        to_id = item["to"]
        
        # Add column header in row 3
        ws.cell(3, next_col).value = column_name
        
        # Add placeholder data in row 4 (first data row)
        ws.cell(4, next_col).value = 0
        
        added_columns[sheet_name].append(column_name)
        
        print(f"  [+] Column {next_col}: {column_name.replace(chr(8594), '->')}")
        
        next_col += 1

# Save Excel
wb.save(excel_path)
print(f"\n[SAVE] Saved Excel to {excel_path}")

# Now update JSON to enable these edges
print(f"\n[UPDATE] Updating JSON diagram...")

diagram_path = "data/diagrams/ug2_north_decline.json"
with open(diagram_path, "r", encoding="utf-8") as f:
    diagram = json.load(f)

# Build edge lookup
edge_lookup = {}
for i, edge in enumerate(diagram["edges"]):
    from_id = edge.get("from", "")
    to_id = edge.get("to", "")
    key = f"{from_id}|{to_id}"
    edge_lookup[key] = i

enabled_count = 0

for item in needs_column:
    from_id = item["from"]
    to_id = item["to"]
    expected_column = item["expected_column"]
    sheet = item["sheet"]
    
    key = f"{from_id}|{to_id}"
    
    if key in edge_lookup:
        edge_idx = edge_lookup[key]
        edge = diagram["edges"][edge_idx]
        mapping = edge.get("excel_mapping", {})
        
        # Update mapping
        mapping["enabled"] = True
        mapping["sheet"] = sheet
        mapping["column"] = expected_column
        
        enabled_count += 1

# Save updated diagram
with open(diagram_path, "w", encoding="utf-8") as f:
    json.dump(diagram, f, indent=2, ensure_ascii=False)

print(f"[OK] Enabled {enabled_count} edges")
print(f"[SAVE] Saved to {diagram_path}")

# Final count
enabled_total = sum(1 for e in diagram["edges"] if e.get("excel_mapping", {}).get("enabled"))
disabled_total = sum(1 for e in diagram["edges"] if not e.get("excel_mapping", {}).get("enabled"))

print(f"\n" + "="*80)
print(f"FINAL STATUS")
print(f"="*80)
print(f"  [OK] Enabled: {enabled_total}")
print(f"  [X] Disabled: {disabled_total}")
print(f"  [#] Total: {enabled_total + disabled_total}")
print(f"\n  [INFO] Added {sum(len(cols) for cols in added_columns.values())} new Excel columns")
print(f"  [INFO] {enabled_total} flow labels will now show data!")
