#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Component Rename Automation Tool

This script handles renaming components throughout the entire system:
- Updates JSON diagram node IDs
- Updates JSON edge references  
- Updates Excel column headers
- Updates edge mappings

Usage:
    python rename_component.py old_name new_name

Example:
    python rename_component.py guest_house trp_clinic
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook

def normalize_name(name: str) -> str:
    """Convert component name to standard formats."""
    return name.lower().replace(' ', '_').replace('-', '_')

def to_uppercase_format(name: str) -> str:
    """Convert to Excel format: UPPERCASE with underscores."""
    return name.upper().replace('_', '_')

def rename_component(old_name: str, new_name: str, dry_run: bool = False):
    """
    Rename a component throughout the system.
    
    Args:
        old_name: Old component name (e.g., 'guest_house')
        new_name: New component name (e.g., 'trp_clinic')
        dry_run: If True, only show what would change, don't actually change
    """
    
    old_normalized = normalize_name(old_name)
    new_normalized = normalize_name(new_name)
    old_upper = to_uppercase_format(old_normalized)
    new_upper = to_uppercase_format(new_normalized)
    
    print(f"{'='*70}")
    print(f"COMPONENT RENAME: {old_name.upper()} → {new_name.upper()}")
    print(f"{'='*70}")
    print(f"  Normalized: {old_normalized} → {new_normalized}")
    print(f"  Excel format: {old_upper} → {new_upper}")
    if dry_run:
        print(f"  MODE: DRY RUN (no changes will be made)")
    print()
    
    changes = {
        'json_nodes': 0,
        'json_edges': 0,
        'excel_headers': 0,
        'edge_mappings': 0,
    }
    
    # 1. Update JSON diagram
    json_path = Path('data/diagrams/ug2_north_decline.json')
    if json_path.exists():
        print(f"[1] Updating JSON diagram: {json_path.name}")
        with open(json_path, 'r', encoding='utf-8') as f:
            diagram = json.load(f)
        
        # Update node IDs
        for node in diagram.get('nodes', []):
            if node.get('id') == old_normalized:
                print(f"    Node ID: {old_normalized} → {new_normalized}")
                if not dry_run:
                    node['id'] = new_normalized
                changes['json_nodes'] += 1
        
        # Update edge references
        for edge in diagram.get('edges', []):
            old_from = edge.get('from')
            old_to = edge.get('to')
            new_from = old_from if old_from != old_normalized else new_normalized
            new_to = old_to if old_to != old_normalized else new_normalized
            
            if new_from != old_from or new_to != old_to:
                print(f"    Edge: {old_from} → {old_to}")
                print(f"           {new_from} → {new_to}")
                if not dry_run:
                    edge['from'] = new_from
                    edge['to'] = new_to
                changes['json_edges'] += 1
            
            # Update excel_mapping columns that reference old name
            mapping = edge.get('excel_mapping', {})
            old_column = mapping.get('column', '')
            if old_column and old_upper in old_column:
                new_column = old_column.replace(old_upper, new_upper)
                print(f"    Mapping: {old_column} → {new_column}")
                if not dry_run:
                    mapping['column'] = new_column
                    edge['excel_mapping'] = mapping
                changes['edge_mappings'] += 1
        
        if not dry_run:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(diagram, f, indent=2)
            print(f"    ✅ JSON diagram saved")
        else:
            print(f"    (no changes - dry run)")
        print()
    
    # 2. Update Excel headers
    excel_path = Path('test_templates/Water_Balance_TimeSeries_Template.xlsx')
    if excel_path.exists():
        print(f"[2] Updating Excel headers: {excel_path.name}")
        wb = load_workbook(excel_path)
        
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith('Flows_'):
                continue
            
            ws = wb[sheet_name]
            updated = False
            
            # Row 3 has headers
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(3, col_idx)
                old_header = cell.value
                
                if old_header and old_upper in str(old_header):
                    new_header = str(old_header).replace(old_upper, new_upper)
                    print(f"    {sheet_name}: {old_header} → {new_header}")
                    if not dry_run:
                        cell.value = new_header
                    changes['excel_headers'] += 1
                    updated = True
        
        if not dry_run:
            wb.save(excel_path)
            print(f"    ✅ Excel workbook saved")
        else:
            print(f"    (no changes - dry run)")
        wb.close()
        print()
    
    # Summary
    print(f"{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    print(f"  JSON nodes updated: {changes['json_nodes']}")
    print(f"  JSON edges updated: {changes['json_edges']}")
    print(f"  Excel headers updated: {changes['excel_headers']}")
    print(f"  Edge mappings updated: {changes['edge_mappings']}")
    print(f"  Total changes: {sum(changes.values())}")
    
    if dry_run:
        print(f"\n  ✓ Dry run complete. Re-run without --dry-run to apply changes.")
    else:
        print(f"\n  ✓ All changes applied successfully!")

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    
    old_name = sys.argv[1]
    new_name = sys.argv[2]
    dry_run = '--dry-run' in sys.argv
    
    rename_component(old_name, new_name, dry_run=dry_run)
