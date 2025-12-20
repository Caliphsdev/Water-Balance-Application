#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Component Rename Manager - Comprehensive Automation

This script manages component name changes across the entire system.
It reads from a configuration file and automatically updates:
- JSON diagrams (node IDs, edge references)
- Excel columns
- Edge mappings
- All dependent areas

Usage:
    python component_rename_manager.py                    # Apply all pending renames
    python component_rename_manager.py --dry-run          # Preview changes
    python component_rename_manager.py --config my_renames.json  # Custom config
    python component_rename_manager.py --list             # Show pending renames
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Tuple

class ComponentRenameManager:
    """Manages component renaming across the system."""
    
    def __init__(self, config_path: str = 'component_rename_config.json'):
        """Initialize the manager with configuration."""
        self.config_path = Path(config_path)
        self.config = self._load_config()
        self.changes = {'json': 0, 'excel': 0, 'mappings': 0}
        self.dry_run = False
    
    def _load_config(self) -> Dict:
        """Load configuration from file."""
        if not self.config_path.exists():
            print(f"[WARN] Config file not found: {self.config_path}")
            print(f"[WARN] Creating default config with example...")
            self._create_default_config()
            return self._read_config()
        
        return self._read_config()
    
    def _read_config(self) -> Dict:
        """Read and parse config file."""
        with open(self.config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    def _create_default_config(self):
        """Create default configuration file with example."""
        default_config = {
            "component_renames": [
                {
                    "old_name": "guest_house",
                    "new_name": "trp_clinic",
                    "excel_columns": [
                        "SOFTENING → TRP_CLINIC",
                        "TRP_CLINIC → SEPTIC",
                        "TRP_CLINIC → CONSUMPTION"
                    ],
                    "description": "Renamed from guest_house to TRP_CLINIC (renamed building)"
                }
            ],
            "files": {
                "json_diagram": "data/diagrams/ug2_north_decline.json",
                "excel_template": "test_templates/Water_Balance_TimeSeries_Template.xlsx"
            },
            "settings": {
                "auto_backup": True,
                "validate_after_rename": True
            }
        }
        
        with open(self.config_path, 'w', encoding='utf-8') as f:
            json.dump(default_config, f, indent=2)
        
        print(f"[OK] Created default config: {self.config_path}")
    
    def normalize_name(self, name: str) -> str:
        """Convert name to standard format (lowercase_with_underscores)."""
        return name.lower().replace(' ', '_').replace('-', '_')
    
    def to_uppercase(self, name: str) -> str:
        """Convert to Excel format (UPPERCASE_WITH_UNDERSCORES)."""
        return self.normalize_name(name).upper()
    
    def list_pending_renames(self):
        """Show all pending component renames."""
        renames = self.config.get('component_renames', [])
        
        if not renames:
            print("No component renames configured.")
            return
        
        print("\n" + "="*70)
        print("PENDING COMPONENT RENAMES")
        print("="*70)
        
        for idx, rename in enumerate(renames, 1):
            old = rename.get('old_name')
            new = rename.get('new_name')
            desc = rename.get('description', '')
            
            print(f"\n[{idx}] {old.upper()} → {new.upper()}")
            if desc:
                print(f"    {desc}")
            
            excel_cols = rename.get('excel_columns', [])
            if excel_cols:
                print(f"    Excel columns ({len(excel_cols)}):")
                for col in excel_cols[:3]:
                    print(f"      - {col}")
                if len(excel_cols) > 3:
                    print(f"      ... and {len(excel_cols) - 3} more")
    
    def apply_renames(self):
        """Apply all pending renames."""
        renames = self.config.get('component_renames', [])
        
        if not renames:
            print("[INFO] No renames to process.")
            return
        
        print("\n" + "="*70)
        print("COMPONENT RENAME MANAGER")
        print("="*70)
        
        for rename in renames:
            old_name = rename.get('old_name')
            new_name = rename.get('new_name')
            
            if not old_name or not new_name:
                print(f"[SKIP] Invalid rename config: {rename}")
                continue
            
            print(f"\n>>> Processing: {old_name.upper()} → {new_name.upper()}")
            self._apply_single_rename(old_name, new_name, rename)
        
        self._print_summary()
    
    def _apply_single_rename(self, old_name: str, new_name: str, config: Dict):
        """Apply a single component rename."""
        old_norm = self.normalize_name(old_name)
        new_norm = self.normalize_name(new_name)
        old_upper = self.to_uppercase(old_norm)
        new_upper = self.to_uppercase(new_norm)
        
        # Update JSON diagram
        json_path = Path(self.config['files']['json_diagram'])
        if json_path.exists():
            self._update_json(json_path, old_norm, new_norm, old_upper, new_upper)
        
        # Update Excel columns
        excel_path = Path(self.config['files']['excel_template'])
        if excel_path.exists():
            self._update_excel(excel_path, config.get('excel_columns', []))
    
    def _update_json(self, json_path: Path, old_norm: str, new_norm: str, old_upper: str, new_upper: str):
        """Update JSON diagram with component rename."""
        with open(json_path, 'r', encoding='utf-8') as f:
            diagram = json.load(f)
        
        json_changes = 0
        
        # Update node IDs
        for node in diagram.get('nodes', []):
            if node.get('id') == old_norm:
                if not self.dry_run:
                    node['id'] = new_norm
                print(f"    [JSON] Node ID: {old_norm} → {new_norm}")
                json_changes += 1
        
        # Update edge references
        for edge in diagram.get('edges', []):
            old_from = edge.get('from')
            old_to = edge.get('to')
            new_from = old_from if old_from != old_norm else new_norm
            new_to = old_to if old_to != old_norm else new_norm
            
            if new_from != old_from or new_to != old_to:
                if not self.dry_run:
                    edge['from'] = new_from
                    edge['to'] = new_to
                print(f"    [JSON] Edge: {old_from} → {old_to} = {new_from} → {new_to}")
                json_changes += 1
            
            # Update mapping columns
            mapping = edge.get('excel_mapping', {})
            old_column = mapping.get('column', '')
            if old_column and old_upper in old_column:
                new_column = old_column.replace(old_upper, new_upper)
                if not self.dry_run:
                    mapping['column'] = new_column
                    edge['excel_mapping'] = mapping
                print(f"    [JSON] Mapping: {old_column} → {new_column}")
                json_changes += 1
        
        if not self.dry_run:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(diagram, f, indent=2)
        
        self.changes['json'] += json_changes
    
    def _update_excel(self, excel_path: Path, excel_columns: List[str]):
        """Update Excel with new component columns."""
        wb = load_workbook(excel_path)
        excel_changes = 0
        
        # Add new columns to appropriate sheets
        for col_name in excel_columns:
            # Determine which sheet based on column name pattern
            sheet = self._determine_sheet(col_name)
            if sheet not in wb.sheetnames:
                continue
            
            ws = wb[sheet]
            
            # Find last populated column in row 3
            last_col = 0
            for col in range(1, ws.max_column + 1):
                if ws.cell(3, col).value:
                    last_col = col
            
            # Add new column
            existing_headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
            if col_name not in existing_headers:
                new_col = last_col + 1
                if not self.dry_run:
                    ws.cell(3, new_col).value = col_name
                    # Add numeric placeholder data to ensure loader shows 0 m³
                    for row in range(4, 14):
                        if ws.cell(row, new_col).value in (None, '-'):
                            ws.cell(row, new_col).value = 0
                
                print(f"    [EXCEL] {sheet}: Added column {col_name}")
                excel_changes += 1
        
        if not self.dry_run:
            wb.save(excel_path)
        
        wb.close()
        self.changes['excel'] += excel_changes
    
    def _determine_sheet(self, column_name: str) -> str:
        """Determine which sheet a column belongs to based on naming."""
        col_upper = column_name.upper()
        
        # Map column prefixes to sheets
        if col_upper.startswith('SOFTENING') or col_upper.startswith('RAINFALL') or col_upper.startswith('NDCD'):
            return 'Flows_UG2N'
        elif col_upper.startswith('UG2PLANT'):
            return 'Flows_UG2P'
        elif col_upper.startswith('OLDTSF'):
            return 'Flows_OLDTSF'
        elif col_upper.startswith('MERPLANT'):
            return 'Flows_MERP'
        elif col_upper.startswith('MERS'):
            return 'Flows_MERS'
        elif col_upper.startswith('UG2S'):
            return 'Flows_UG2S'
        elif col_upper.startswith('SPCD') or col_upper.startswith('STOCKPILE'):
            return 'Flows_STOCKPILE'
        
        return 'Flows_UG2N'  # Default
    
    def _print_summary(self):
        """Print summary of changes."""
        print("\n" + "="*70)
        print("SUMMARY")
        print("="*70)
        print(f"JSON changes: {self.changes['json']}")
        print(f"Excel changes: {self.changes['excel']}")
        print(f"Total changes: {sum(self.changes.values())}")
        
        if self.dry_run:
            print("\n[DRY RUN] No changes were applied.")
            print("Run without --dry-run to apply changes.")

def main():
    """Main entry point."""
    manager = ComponentRenameManager()
    
    if '--list' in sys.argv:
        manager.list_pending_renames()
    elif '--dry-run' in sys.argv:
        manager.dry_run = True
        manager.apply_renames()
    else:
        manager.apply_renames()

if __name__ == '__main__':
    main()
